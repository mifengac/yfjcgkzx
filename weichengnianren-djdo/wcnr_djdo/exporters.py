from __future__ import annotations

import io
import json
from datetime import datetime
from decimal import Decimal
from typing import Any, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .service import MetricResult


def _safe_filename_part(val: str) -> str:
    s = str(val or "").strip()
    return s.replace(":", "-").replace("/", "-").replace("\\", "-").replace(" ", "_")


def _excel_value(val: Any) -> Any:
    """
    openpyxl 单元格不支持 dict/list 等复杂类型，这里统一做可写入转换。
    """
    if val is None:
        return None
    if isinstance(val, (str, int, float, bool, datetime)):
        return val
    if isinstance(val, Decimal):
        try:
            return float(val)
        except Exception:
            return str(val)
    if isinstance(val, (dict, list, tuple)):
        try:
            return json.dumps(val, ensure_ascii=False, default=str)
        except Exception:
            return str(val)
    # psycopg2 jsonb 可能是 list/dict；其余对象兜底为字符串
    return str(val)


def build_metric_xlsx(result: MetricResult, start_time: datetime, end_time: datetime) -> tuple[bytes, str]:
    wb = Workbook()
    ws = wb.active
    ws.title = result.title[:31]

    ws.append([result.title])
    ws.append([f"{start_time:%Y-%m-%d %H:%M:%S} - {end_time:%Y-%m-%d %H:%M:%S}"])
    ws.append([])

    ws.append(["柱状图数据"])
    ws.append(["地区"] + result.series)
    for r in result.chart_rows:
        ws.append([_excel_value(r.get("地区"))] + [_excel_value(r.get(s)) for s in result.series])
    ws.append([])

    ws.append(["明细表"])
    if result.detail_rows:
        cols = list(result.detail_rows[0].keys())
        ws.append(cols)
        for row in result.detail_rows:
            ws.append([_excel_value(row.get(c)) for c in cols])
        for idx, col in enumerate(cols, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(40, max(10, len(str(col)) + 2))

    buf = io.BytesIO()
    wb.save(buf)
    ts = int(datetime.now().timestamp())
    filename = (
        f"{_safe_filename_part(start_time.strftime('%Y-%m-%d_%H-%M-%S'))}-"
        f"{_safe_filename_part(end_time.strftime('%Y-%m-%d_%H-%M-%S'))}{result.title}{ts}.xlsx"
    )
    return buf.getvalue(), filename


def build_details_xlsx(results: List[MetricResult], start_time: datetime, end_time: datetime) -> tuple[bytes, str]:
    wb = Workbook()
    wb.remove(wb.active)

    for result in results:
        ws = wb.create_sheet(title=result.title[:31])
        ws.append([result.title])
        ws.append([f"{start_time:%Y-%m-%d %H:%M:%S} - {end_time:%Y-%m-%d %H:%M:%S}"])
        ws.append([])

        if result.detail_rows:
            cols = list(result.detail_rows[0].keys())
            ws.append(cols)
            for row in result.detail_rows:
                ws.append([_excel_value(row.get(c)) for c in cols])
            for idx, col in enumerate(cols, start=1):
                ws.column_dimensions[get_column_letter(idx)].width = min(40, max(10, len(str(col)) + 2))

    buf = io.BytesIO()
    wb.save(buf)
    ts = int(datetime.now().timestamp())
    filename = (
        f"{_safe_filename_part(start_time.strftime('%Y-%m-%d_%H-%M-%S'))}-"
        f"{_safe_filename_part(end_time.strftime('%Y-%m-%d_%H-%M-%S'))}导出详情{ts}.xlsx"
    )
    return buf.getvalue(), filename


def build_overview_pdf(results: List[MetricResult], start_time: datetime, end_time: datetime) -> tuple[bytes, str]:
    try:
        from reportlab.lib import colors  # type: ignore
        from reportlab.lib.pagesizes import A4, landscape  # type: ignore
        from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
        from reportlab.pdfbase import pdfmetrics  # type: ignore
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont  # type: ignore
        from reportlab.platypus import (  # type: ignore
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
        from reportlab.graphics.charts.barcharts import VerticalBarChart  # type: ignore
        from reportlab.graphics.shapes import Drawing  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"缺少依赖 reportlab，无法导出 PDF：{exc}") from exc

    # ReportLab 默认字体 Helvetica 不支持中文，必须切换到 CJK 字体，否则会出现乱码/方块。
    # 使用 CIDFont（不依赖本机字体文件），优先保证跨环境可用。
    cn_font = "STSong-Light"
    try:
        pdfmetrics.registerFont(UnicodeCIDFont(cn_font))
    except Exception:
        # 重复注册或部分环境已预置时，忽略即可。
        pass

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    for style_name in ("Title", "Normal", "Heading2"):
        if style_name in styles.byName:
            styles[style_name].fontName = cn_font
    story: List[Any] = []

    story.append(Paragraph("全市未成年人打架斗殴六项指标监测", styles["Title"]))
    story.append(Paragraph(f"{start_time:%Y-%m-%d %H:%M:%S} - {end_time:%Y-%m-%d %H:%M:%S}", styles["Normal"]))
    story.append(Spacer(1, 12))

    palette = [colors.HexColor("#60a5fa"), colors.HexColor("#34d399"), colors.HexColor("#f59e0b")]

    for idx, result in enumerate(results):
        story.append(Paragraph(result.title, styles["Heading2"]))

        # 画一个“柱状图框”，并把图例纳入框内（框上方预留图例区）。
        drawing = Drawing(740, 290)
        frame_x, frame_y, frame_w, frame_h = 24, 24, 692, 242
        chart = VerticalBarChart()
        chart.x = frame_x + 36
        chart.y = frame_y + 28
        chart.height = frame_h - 70  # 顶部留给图例+间距
        chart.width = frame_w - 60
        chart.strokeColor = colors.black

        cats = [r["地区"] for r in result.chart_rows]
        data = []
        label_array = []
        for series_name in result.series:
            row_vals = [float(r.get(series_name) or 0) for r in result.chart_rows]
            data.append(row_vals)
            if "率" in str(series_name):
                label_array.append([f"{v:.2f}%" for v in row_vals])
            else:
                label_array.append([f"{int(v)}" if abs(v - int(v)) < 1e-9 else f"{v:g}" for v in row_vals])
        chart.data = data
        chart.categoryAxis.categoryNames = cats
        chart.categoryAxis.labels.boxAnchor = "ne"
        chart.categoryAxis.labels.angle = 45
        chart.categoryAxis.labels.fontName = cn_font
        chart.valueAxis.valueMin = 0
        chart.valueAxis.labels.fontName = cn_font
        # 避免最大柱子顶住框线：为 Y 轴上限增加留白
        try:
            import math

            vmax = max((v for row in data for v in row), default=0.0)
            if vmax <= 0:
                chart.valueAxis.valueMax = 1
            else:
                pad = max(1.0, math.ceil(vmax * 0.1))
                chart.valueAxis.valueMax = vmax + pad
        except Exception:
            pass

        for i in range(len(result.series)):
            chart.bars[i].fillColor = palette[i % len(palette)]

        # 柱子数值标签（“xx率”加 %）
        # reportlab 4.4.x 下 VerticalBarChart 的 barLabels/barLabelArray 行为不稳定，
        # 这里改为手动画 String，保证必定显示在柱子上方。
        bar_gap = 2.0
        try:
            chart.barSpacing = bar_gap
        except Exception:
            pass
        try:
            chart.groupSpacing = 10.0
        except Exception:
            pass
        try:
            n_cats = max(1, len(cats))
            n_series = max(1, len(data))
            group_w = float(chart.width) / float(n_cats)
            usable_w = group_w * 0.8
            bar_w = (usable_w - bar_gap * (n_series - 1)) / float(n_series)
            bar_w = max(6.0, min(18.0, bar_w))
            chart.barWidth = bar_w
        except Exception:
            bar_w = 10.0

        try:
            from reportlab.graphics.shapes import String  # type: ignore

            value_min = float(getattr(chart.valueAxis, "valueMin", 0) or 0)
            value_max = float(getattr(chart.valueAxis, "valueMax", 0) or 0)
            if value_max <= value_min:
                vmax2 = max((v for row in data for v in row), default=0.0)
                value_max = vmax2 if vmax2 > value_min else (value_min + 1.0)
            scale = float(chart.height) / (value_max - value_min)

            n_cats = len(cats)
            n_series = len(data)
            if n_cats > 0 and n_series > 0:
                group_spacing = float(getattr(chart, "groupSpacing", 0) or 0)
                if n_cats == 1:
                    group_w = float(chart.width)
                else:
                    group_w = (float(chart.width) - group_spacing * float(n_cats - 1)) / float(n_cats)
                total_bar_w = bar_w * n_series + bar_gap * (n_series - 1)
                for c_idx in range(n_cats):
                    start_x = (
                        float(chart.x)
                        + c_idx * (group_w + group_spacing)
                        + (group_w - total_bar_w) / 2.0
                    )
                    for s_idx in range(n_series):
                        try:
                            v = float(data[s_idx][c_idx])
                        except Exception:
                            v = 0.0
                        bar_h = max(0.0, (v - value_min) * scale)
                        bar_top = float(chart.y) + bar_h
                        label = ""
                        try:
                            label = str(label_array[s_idx][c_idx])
                        except Exception:
                            label = f"{v:g}"
                        x_center = start_x + s_idx * (bar_w + bar_gap) + (bar_w / 2.0)
                        y_text = min(bar_top + 3.0, float(chart.y) + float(chart.height) - 2.0)
                        drawing.add(
                            String(
                                x_center,
                                y_text,
                                label,
                                fontName=cn_font,
                                fontSize=8,
                                textAnchor="middle",
                            )
                        )
        except Exception:
            pass

        # 框线（包含图例区 + 图表区）
        try:
            from reportlab.graphics.shapes import Rect, String  # type: ignore

            drawing.add(Rect(frame_x, frame_y, frame_w, frame_h, strokeColor=colors.black, fillColor=None))

            # 图例：居中、横向排列（不使用 Legend，避免裁切/换行差异）
            legend_font_size = 9
            box_w = 8
            box_h = 8
            pad = 4
            gap = 14
            names = [str(x or "") for x in result.series]
            name_widths = []
            for name in names:
                try:
                    name_widths.append(float(pdfmetrics.stringWidth(name, cn_font, legend_font_size)))
                except Exception:
                    name_widths.append(float(len(name)) * legend_font_size)
            item_widths = [box_w + pad + w for w in name_widths]
            total_w = sum(item_widths) + gap * max(0, len(item_widths) - 1)
            x_cursor = frame_x + max(0.0, (frame_w - total_w) / 2.0)
            # 让图例位于框内顶部，并留出少量边距
            legend_top_y = frame_y + frame_h - 10
            box_y = legend_top_y - box_h
            text_y = legend_top_y - legend_font_size
            for i, name in enumerate(names):
                color = palette[i % len(palette)]
                drawing.add(Rect(x_cursor, box_y, box_w, box_h, strokeColor=None, fillColor=color))
                drawing.add(String(x_cursor + box_w + pad, text_y, name, fontName=cn_font, fontSize=legend_font_size))
                x_cursor += item_widths[i] + gap
        except Exception:
            pass

        drawing.add(chart)
        story.append(drawing)
        story.append(Spacer(1, 8))

        table_data = [["地区"] + result.series]
        for r in result.chart_rows:
            row_out = [r.get("地区")]
            for s in result.series:
                v = r.get(s)
                if "率" in str(s):
                    try:
                        row_out.append(f"{float(v or 0):.2f}%")
                    except Exception:
                        row_out.append(f"{v}%")
                else:
                    row_out.append(v)
            table_data.append(row_out)
        t = Table(table_data, repeatRows=1)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("FONTNAME", (0, 0), (-1, -1), cn_font),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        story.append(t)

        if idx != len(results) - 1:
            story.append(PageBreak())

    doc.build(story)
    filename = (
        f"{_safe_filename_part(start_time.strftime('%Y-%m-%d_%H-%M-%S'))}-"
        f"{_safe_filename_part(end_time.strftime('%Y-%m-%d_%H-%M-%S'))}全市未成年人打架斗殴六项指标监测.pdf"
    )
    return buf.getvalue(), filename
