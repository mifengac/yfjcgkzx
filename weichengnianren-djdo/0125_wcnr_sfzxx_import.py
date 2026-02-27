from __future__ import annotations

import argparse
import logging
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Optional


# 让脚本在任意工作目录运行时，也能 import 到项目代码
_REPO_ROOT = Path(__file__).resolve().parents[1]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))


SHEET_NAME_DEFAULT = "累计招生"
HEADER_ROW_INDEX = 3  # 第3行为列名
DB_SCHEMA = "ywdata"
DB_TABLE = "zq_wcnr_sfzxx"


EXPECTED_HEADERS = [
    "序号",
    "编号",
    "姓名",
    "性别",
    "民族",
    "出生日期",
    "身份证号码",
    "户籍地区",
    "户籍地址",
    "监护人",
    "联系电话",
    "原学校",
    "年级",
    "送生部门",
    "矫治原因",
    "危害等级",
    "入学时间",
    "矫治时间",
    "离校时间",
    "备注",
]


def _norm_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip()
    text = text.replace("\u3000", " ")  # 全角空格
    text = re.sub(r"\s+", "", text)  # 去除所有空白（常见 Excel 列名空格问题）
    text = text.rstrip(",").rstrip("，").strip()
    return text


def _to_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text if text != "" else None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip() or None


def _to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = _to_text(value)
    if not text:
        return None
    try:
        return int(text)
    except Exception:
        return None


def _to_months(value: Any) -> Optional[int]:
    text = _to_text(value)
    if not text:
        return None
    m = re.search(r"(\d+)", text)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _to_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # pandas.Timestamp 等
    if hasattr(value, "to_pydatetime"):
        try:
            dt = value.to_pydatetime()
            if isinstance(dt, datetime):
                return dt.date()
        except Exception:
            pass
    if isinstance(value, (int, float)):
        try:
            # 优先使用 openpyxl 的 from_excel（考虑 Excel 1900 闰年 bug 等）
            from openpyxl.utils.datetime import from_excel  # type: ignore

            return from_excel(value).date()
        except Exception:
            # 兜底：按 Excel 1900 体系粗略换算（大多数表格可用）
            try:
                base = datetime(1899, 12, 30)
                return (base + timedelta(days=float(value))).date()
            except Exception:
                return None
    text = _to_text(value)
    if not text:
        return None
    # 兼容 "2026-01-25 00:00:00" / "2026-01-25T00:00:00"
    text = text.split("T", 1)[0].split(" ", 1)[0]
    text = text.replace("/", "-").replace(".", "-")
    text = re.sub(r"[年月]", "-", text)
    text = text.replace("日", "")
    for fmt in ("%Y-%m-%d", "%Y-%m", "%Y%m%d"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.date()
        except Exception:
            continue
    return None


@dataclass(frozen=True)
class RowData:
    xh: Optional[int]
    bh: Optional[str]
    xm: Optional[str]
    xb: Optional[str]
    mz: Optional[str]
    csrq: Optional[date]
    sfzhm: Optional[str]
    hjdq: Optional[str]
    hjdz: Optional[str]
    jhr: Optional[str]
    lxdh: Optional[str]
    yxx: Optional[str]
    nj: Optional[str]
    ssbm: Optional[str]
    jzyy: Optional[str]
    whdj: Optional[str]
    rx_time: Optional[date]
    jz_time: Optional[int]
    lx_time: Optional[date]
    bz: Optional[str]


def read_xls_rows(xls_path: Path, sheet_name: str) -> list[RowData]:
    try:
        import pandas as pd  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"缺少依赖 pandas，无法读取 xls：{exc}") from exc

    try:
        df = pd.read_excel(str(xls_path), sheet_name=sheet_name, header=None, engine="xlrd")
    except ValueError as exc:
        # 常见：sheet 不存在
        raise RuntimeError(f'读取 xls 失败：{exc}') from exc
    except Exception as exc:
        raise RuntimeError(
            f'读取 xls 失败（请确认已安装 xlrd>=2.0.1 且文件为 .xls）：{exc}'
        ) from exc

    if df is None or getattr(df, "empty", False):
        return []

    if len(df.index) < HEADER_ROW_INDEX:
        raise RuntimeError(f"xls 行数不足，无法读取第{HEADER_ROW_INDEX}行表头")

    header_row = df.iloc[HEADER_ROW_INDEX - 1].tolist()
    header_values = [_norm_header(v) for v in header_row]

    header_to_index: dict[str, int] = {}
    for idx, header in enumerate(header_values):
        if header:
            header_to_index[header] = idx

    missing = [h for h in EXPECTED_HEADERS if _norm_header(h) not in header_to_index]
    if missing:
        raise RuntimeError(
            "Excel 表头不匹配，缺少列: "
            + ", ".join(missing)
            + f"；第{HEADER_ROW_INDEX}行读取到的列名为: {header_values}"
        )

    def cell_value(row_values: list[Any], header: str) -> Any:
        idx = header_to_index[_norm_header(header)]
        if idx >= len(row_values):
            return None
        val = row_values[idx]
        try:
            if pd.isna(val):
                return None
        except Exception:
            pass
        return val

    results: list[RowData] = []
    for _, series in df.iloc[HEADER_ROW_INDEX:].iterrows():
        row_values = series.tolist()
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_values):
            continue

        results.append(
            RowData(
                xh=_to_int(cell_value(row_values, "序号")),
                bh=_to_text(cell_value(row_values, "编号")),
                xm=_to_text(cell_value(row_values, "姓名")),
                xb=_to_text(cell_value(row_values, "性别")),
                mz=_to_text(cell_value(row_values, "民族")),
                csrq=_to_date(cell_value(row_values, "出生日期")),
                sfzhm=_to_text(cell_value(row_values, "身份证号码")),
                hjdq=_to_text(cell_value(row_values, "户籍地区")),
                hjdz=_to_text(cell_value(row_values, "户籍地址")),
                jhr=_to_text(cell_value(row_values, "监护人")),
                lxdh=_to_text(cell_value(row_values, "联系电话")),
                yxx=_to_text(cell_value(row_values, "原学校")),
                nj=_to_text(cell_value(row_values, "年级")),
                ssbm=_to_text(cell_value(row_values, "送生部门")),
                jzyy=_to_text(cell_value(row_values, "矫治原因")),
                whdj=_to_text(cell_value(row_values, "危害等级")),
                rx_time=_to_date(cell_value(row_values, "入学时间")),
                jz_time=_to_months(cell_value(row_values, "矫治时间")),
                lx_time=_to_date(cell_value(row_values, "离校时间")),
                bz=_to_text(cell_value(row_values, "备注")),
            )
        )

    return results


def read_xlsx_rows(xlsx_path: Path, sheet_name: str) -> list[RowData]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"缺少依赖 openpyxl，无法读取 xlsx：{exc}") from exc

    workbook = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise RuntimeError(f'未找到 sheet "{sheet_name}"，实际为: {workbook.sheetnames}')

        worksheet = workbook[sheet_name]
        header_row = next(
            worksheet.iter_rows(
                min_row=HEADER_ROW_INDEX, max_row=HEADER_ROW_INDEX, values_only=True
            )
        )
        header_values = [_norm_header(v) for v in header_row]
        header_to_index: dict[str, int] = {}
        for idx, header in enumerate(header_values):
            if header:
                header_to_index[header] = idx

        missing = [h for h in EXPECTED_HEADERS if _norm_header(h) not in header_to_index]
        if missing:
            raise RuntimeError(
                "Excel 表头不匹配，缺少列: "
                + ", ".join(missing)
                + f"；第{HEADER_ROW_INDEX}行读取到的列名为: {header_values}"
            )

        def cell_value(row: tuple[Any, ...], header: str) -> Any:
            return row[header_to_index[_norm_header(header)]]

        results: list[RowData] = []
        for row in worksheet.iter_rows(min_row=HEADER_ROW_INDEX + 1, values_only=True):
            if row is None:
                continue
            if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                continue

            results.append(
                RowData(
                    xh=_to_int(cell_value(row, "序号")),
                    bh=_to_text(cell_value(row, "编号")),
                    xm=_to_text(cell_value(row, "姓名")),
                    xb=_to_text(cell_value(row, "性别")),
                    mz=_to_text(cell_value(row, "民族")),
                    csrq=_to_date(cell_value(row, "出生日期")),
                    sfzhm=_to_text(cell_value(row, "身份证号码")),
                    hjdq=_to_text(cell_value(row, "户籍地区")),
                    hjdz=_to_text(cell_value(row, "户籍地址")),
                    jhr=_to_text(cell_value(row, "监护人")),
                    lxdh=_to_text(cell_value(row, "联系电话")),
                    yxx=_to_text(cell_value(row, "原学校")),
                    nj=_to_text(cell_value(row, "年级")),
                    ssbm=_to_text(cell_value(row, "送生部门")),
                    jzyy=_to_text(cell_value(row, "矫治原因")),
                    whdj=_to_text(cell_value(row, "危害等级")),
                    rx_time=_to_date(cell_value(row, "入学时间")),
                    jz_time=_to_months(cell_value(row, "矫治时间")),
                    lx_time=_to_date(cell_value(row, "离校时间")),
                    bz=_to_text(cell_value(row, "备注")),
                )
            )

        return results
    finally:
        try:
            workbook.close()
        except Exception:
            pass


def read_excel_rows(excel_path: Path, sheet_name: str) -> list[RowData]:
    suffix = excel_path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx_rows(excel_path, sheet_name)
    if suffix == ".xls":
        return read_xls_rows(excel_path, sheet_name)
    raise RuntimeError(f"不支持的文件类型: {excel_path.suffix}（仅支持 .xls/.xlsx）")


def ensure_table(cursor) -> None:
    cursor.execute(f'CREATE SCHEMA IF NOT EXISTS "{DB_SCHEMA}";')
    cursor.execute(
        f"""
        CREATE TABLE IF NOT EXISTS "{DB_SCHEMA}"."{DB_TABLE}" (
            id SERIAL,
            xh INTEGER,
            bh VARCHAR(100) NOT NULL,
            xm VARCHAR(100),
            xb VARCHAR(20),
            mz VARCHAR(50),
            csrq DATE,
            sfzhm VARCHAR(30),
            hjdq VARCHAR(100),
            hjdz VARCHAR(255),
            jhr VARCHAR(100),
            lxdh VARCHAR(50),
            yxx VARCHAR(255),
            nj VARCHAR(50),
            ssbm VARCHAR(255),
            jzyy TEXT,
            whdj VARCHAR(50),
            rx_time DATE,
            jz_time INTEGER,
            lx_time DATE,
            bz TEXT,
            PRIMARY KEY (bh)
        );
        """
    )
    # 兼容：表已存在但主键不在 bh 的情况，尝试迁移为 bh 主键
    cursor.execute(
        f"""
        DO $$
        DECLARE
            pk_name text;
            pk_cols text[];
            jz_time_udt text;
        BEGIN
            SELECT c.conname,
                   ARRAY(
                        SELECT a.attname
                          FROM unnest(c.conkey) WITH ORDINALITY AS k(attnum, ord)
                          JOIN pg_attribute a ON a.attrelid = c.conrelid AND a.attnum = k.attnum
                         ORDER BY k.ord
                   )
              INTO pk_name, pk_cols
              FROM pg_constraint c
              JOIN pg_class t ON t.oid = c.conrelid
              JOIN pg_namespace n ON n.oid = t.relnamespace
             WHERE n.nspname = '{DB_SCHEMA}'
               AND t.relname = '{DB_TABLE}'
               AND c.contype = 'p'
             LIMIT 1;

            -- 如果当前主键不是 (bh)，则移除后再添加 bh 主键
            IF pk_name IS NOT NULL AND (pk_cols IS NULL OR pk_cols <> ARRAY['bh']) THEN
                EXECUTE format('ALTER TABLE %I.%I DROP CONSTRAINT %I', '{DB_SCHEMA}', '{DB_TABLE}', pk_name);
            END IF;

            -- 确保 bh 不为空
            EXECUTE format('ALTER TABLE %I.%I ALTER COLUMN bh SET NOT NULL', '{DB_SCHEMA}', '{DB_TABLE}');

            -- 如果当前没有主键，则添加 bh 主键
            IF NOT EXISTS (
                SELECT 1
                  FROM pg_constraint c2
                  JOIN pg_class t2 ON t2.oid = c2.conrelid
                  JOIN pg_namespace n2 ON n2.oid = t2.relnamespace
                 WHERE n2.nspname = '{DB_SCHEMA}'
                   AND t2.relname = '{DB_TABLE}'
                   AND c2.contype = 'p'
            ) THEN
                EXECUTE format('ALTER TABLE %I.%I ADD CONSTRAINT %I PRIMARY KEY (bh)', '{DB_SCHEMA}', '{DB_TABLE}', '{DB_TABLE}_pkey');
            END IF;

            -- 矫治时间改为月数整数；旧 DATE 数据语义不兼容，统一置空
            SELECT c.udt_name
              INTO jz_time_udt
              FROM information_schema.columns c
             WHERE c.table_schema = '{DB_SCHEMA}'
               AND c.table_name = '{DB_TABLE}'
               AND c.column_name = 'jz_time'
             LIMIT 1;

            IF jz_time_udt IS NOT NULL AND jz_time_udt <> 'int4' THEN
                EXECUTE format('ALTER TABLE %I.%I ALTER COLUMN jz_time TYPE INTEGER USING NULL', '{DB_SCHEMA}', '{DB_TABLE}');
            END IF;
        END $$;
        """
    )


def insert_rows(cursor, rows: Iterable[RowData], truncate: bool) -> dict:
    try:
        from psycopg2.extras import execute_values  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"缺少依赖 psycopg2，无法批量写入数据库：{exc}") from exc

    if truncate:
        cursor.execute(f'TRUNCATE TABLE "{DB_SCHEMA}"."{DB_TABLE}";')

    # 主键为 bh（编号）：空编号跳过；同一批次重复编号保留最后一条，避免 ON CONFLICT 同批次重复报错
    rows_list = list(rows)
    dedup: dict[str, RowData] = {}
    skipped_no_bh = 0
    for r in rows_list:
        if not r.bh:
            skipped_no_bh += 1
            continue
        if r.bh in dedup:
            del dedup[r.bh]
        dedup[r.bh] = r
    if skipped_no_bh:
        logging.warning(f"发现 {skipped_no_bh} 行编号为空，已跳过（bh 为主键必填）")

    cols = [
        "xh",
        "bh",
        "xm",
        "xb",
        "mz",
        "csrq",
        "sfzhm",
        "hjdq",
        "hjdz",
        "jhr",
        "lxdh",
        "yxx",
        "nj",
        "ssbm",
        "jzyy",
        "whdj",
        "rx_time",
        "jz_time",
        "lx_time",
        "bz",
    ]
    sql = (
        f'INSERT INTO "{DB_SCHEMA}"."{DB_TABLE}" ('
        + ", ".join(f'"{c}"' for c in cols)
        + ') VALUES %s ON CONFLICT ("bh") DO UPDATE SET '
        + ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in cols if c != "bh")
    )
    values = [
        (
            r.xh,
            r.bh,
            r.xm,
            r.xb,
            r.mz,
            r.csrq,
            r.sfzhm,
            r.hjdq,
            r.hjdz,
            r.jhr,
            r.lxdh,
            r.yxx,
            r.nj,
            r.ssbm,
            r.jzyy,
            r.whdj,
            r.rx_time,
            r.jz_time,
            r.lx_time,
            r.bz,
        )
        for r in dedup.values()
    ]
    if not values:
        return {
            "input_rows": len(rows_list),
            "dedup_rows": len(dedup),
            "skipped_no_bh": skipped_no_bh,
            "inserted": 0,
        }

    execute_values(cursor, sql, values, page_size=500)
    return {
        "input_rows": len(rows_list),
        "dedup_rows": len(dedup),
        "skipped_no_bh": skipped_no_bh,
        "inserted": len(values),
    }


def import_sfzxx_file(excel_path: Path, sheet_name: str = SHEET_NAME_DEFAULT, truncate: bool = False) -> dict:
    """
    复用本脚本逻辑：读取 xls/xlsx 并导入到 ywdata.zq_wcnr_sfzxx。
    - 会校验 sheet 与第3行表头
    - bh(编号) 为主键：重复则更新，不重复则新增
    """
    rows = read_excel_rows(excel_path=excel_path, sheet_name=sheet_name)

    connection = None
    cursor = None
    try:
        from gonggong.config.database import get_database_connection  # noqa: E402

        connection = get_database_connection()
        cursor = connection.cursor()
        ensure_table(cursor)
        stats = insert_rows(cursor, rows, truncate=truncate)
        connection.commit()
        stats.update(
            {
                "file": str(excel_path),
                "sheet": sheet_name,
            }
        )
        return stats
    except Exception:
        if connection:
            connection.rollback()
        raise
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description='通过 xls/xlsx 导入人大金仓（PostgreSQL/Kingbase 兼容）表 "ywdata.zq_wcnr_sfzxx"'
    )
    parser.add_argument("xlsx", type=str, help="xls/xlsx 文件路径")
    parser.add_argument("--sheet", type=str, default=SHEET_NAME_DEFAULT, help="sheet 名称")
    parser.add_argument(
        "--truncate",
        action="store_true",
        help="导入前清空表（TRUNCATE）",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="仅解析 xls/xlsx，不写入数据库",
    )
    parser.add_argument("--log-level", type=str, default="INFO", help="日志级别：DEBUG/INFO/WARN/ERROR")
    args = parser.parse_args()

    logging.basicConfig(
        level=getattr(logging, args.log_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(message)s",
    )

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        logging.error(f"xlsx 不存在: {xlsx_path}")
        return 2

    rows = read_excel_rows(excel_path=xlsx_path, sheet_name=args.sheet)
    logging.info(f"解析完成: {xlsx_path.name} / sheet={args.sheet} / 行数={len(rows)} (从第{HEADER_ROW_INDEX+1}行开始)")

    if args.dry_run:
        return 0

    connection = None
    cursor = None
    try:
        from gonggong.config.database import get_database_connection  # noqa: E402

        connection = get_database_connection()
        cursor = connection.cursor()

        ensure_table(cursor)
        stats = insert_rows(cursor, rows, truncate=args.truncate)
        connection.commit()

        logging.info(
            f'写入完成: "{DB_SCHEMA}"."{DB_TABLE}" 插入 {stats.get("inserted", 0)} 行，'
            f'跳过编号为空 {stats.get("skipped_no_bh", 0)} 行（输入 {stats.get("input_rows", 0)} 行）'
        )
        return 0
    except Exception as exc:
        if connection:
            connection.rollback()
        logging.exception(f"导入失败: {exc}")
        return 1
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()


if __name__ == "__main__":
    raise SystemExit(main())
