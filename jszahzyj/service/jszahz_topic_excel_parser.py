from __future__ import annotations

import re
from dataclasses import dataclass
from typing import Any, BinaryIO, Dict, List

from openpyxl import load_workbook

from jszahzyj.jszahz_topic_constants import PERSON_TYPE_OPTIONS, RISK_OPTIONS


PARSE_EMPTY_ROW_BREAK_THRESHOLD = 200

PERSON_TYPE_RULES = {
    "服药情况": {
        "不规律服药": "不规律服药",
    },
    "监护情况": {
        "弱监护": "弱监护",
        "无监护": "无监护",
    },
    "既往有自杀或严重伤人": {
        "是": "既往有严重自杀或伤人行为",
    },
    "列为重点关注人员": {
        "是": "列为重点关注人员",
    },
}


@dataclass(frozen=True)
class ParsedImportResult:
    imported_row_count: int
    generated_tag_count: int
    tagged_person_count: int
    rows: List[Dict[str, Any]]
    all_zjhms: List[str]


def _normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _normalize_id_card(value: Any) -> str:
    text = _normalize_text(value).replace(" ", "")
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text.upper()


def parse_person_type_workbook(file_obj: BinaryIO) -> ParsedImportResult:
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        if sheet.title != "汇总":
            raise ValueError("Excel 第一个 sheet 名称必须为“汇总”")

        imported_row_count = 0
        rows: List[Dict[str, Any]] = []
        seen = set()
        tagged_people = set()
        all_zjhms_set: set = set()
        empty_row_streak = 0

        for row_no, row in enumerate(sheet.iter_rows(min_row=4, max_col=10, values_only=True), start=4):
            row_values = list(row or ())
            zjhm = _normalize_id_card(row_values[4] if len(row_values) > 4 else "")
            medicine_value = _normalize_text(row_values[6] if len(row_values) > 6 else "")
            guardian_value = _normalize_text(row_values[7] if len(row_values) > 7 else "")
            history_value = _normalize_text(row_values[8] if len(row_values) > 8 else "")
            focus_value = _normalize_text(row_values[9] if len(row_values) > 9 else "")

            if not any((zjhm, medicine_value, guardian_value, history_value, focus_value)):
                empty_row_streak += 1
                if empty_row_streak >= PARSE_EMPTY_ROW_BREAK_THRESHOLD:
                    break
                continue

            empty_row_streak = 0
            if not zjhm:
                continue

            imported_row_count += 1
            all_zjhms_set.add(zjhm)
            matched_labels: List[str] = []

            if medicine_value in PERSON_TYPE_RULES["服药情况"]:
                matched_labels.append(PERSON_TYPE_RULES["服药情况"][medicine_value])
            if guardian_value in PERSON_TYPE_RULES["监护情况"]:
                matched_labels.append(PERSON_TYPE_RULES["监护情况"][guardian_value])
            if history_value in PERSON_TYPE_RULES["既往有自杀或严重伤人"]:
                matched_labels.append(PERSON_TYPE_RULES["既往有自杀或严重伤人"][history_value])
            if focus_value in PERSON_TYPE_RULES["列为重点关注人员"]:
                matched_labels.append(PERSON_TYPE_RULES["列为重点关注人员"][focus_value])

            if not matched_labels:
                continue
            for label in matched_labels:
                key = (zjhm, label)
                if key in seen:
                    continue
                seen.add(key)
                tagged_people.add(zjhm)
                rows.append(
                    {
                        "zjhm": zjhm,
                        "person_type": label,
                        "source_row_no": row_no,
                    }
                )

        return ParsedImportResult(
            imported_row_count=imported_row_count,
            generated_tag_count=len(rows),
            tagged_person_count=len(tagged_people),
            rows=rows,
            all_zjhms=sorted(all_zjhms_set),
        )
    finally:
        workbook.close()