from __future__ import annotations

from dataclasses import dataclass
from typing import Any, BinaryIO, Dict, List

from openpyxl import load_workbook

from jszahzyj.jszahz_topic_constants import TOPIC_BRANCH_SHEET_CODE_MAP, TOPIC_BRANCH_SHEETS
from jszahzyj.service.jszahz_topic_excel_parser import (
    PARSE_EMPTY_ROW_BREAK_THRESHOLD,
    _normalize_id_card,
    _normalize_text,
)


BASE_EXCEL_HEADERS = ("序号", "姓名", "证件号码")


@dataclass(frozen=True)
class ParsedBaseImportResult:
    imported_row_count: int
    deduplicated_person_count: int
    rows: List[Dict[str, Any]]
    all_zjhms: List[str]


def _validate_sheet_names(sheet_names: List[str]) -> None:
    expected = list(TOPIC_BRANCH_SHEETS)
    actual = [str(name or "").strip() for name in sheet_names]
    if actual != expected:
        raise ValueError(
            "基础数据 Excel 必须且只能包含 5 个 sheet，且顺序为：" + "、".join(expected)
        )


def _validate_headers(sheet) -> None:
    header_values = tuple(_normalize_text(sheet.cell(row=1, column=index).value) for index in range(1, 4))
    if header_values != BASE_EXCEL_HEADERS:
        raise ValueError(
            f"sheet【{sheet.title}】表头必须为：{'、'.join(BASE_EXCEL_HEADERS)}"
        )


def parse_base_person_workbook(file_obj: BinaryIO) -> ParsedBaseImportResult:
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        _validate_sheet_names(workbook.sheetnames)

        imported_row_count = 0
        rows: List[Dict[str, Any]] = []
        seen_zjhms = set()

        for sheet_name in TOPIC_BRANCH_SHEETS:
            sheet = workbook[sheet_name]
            _validate_headers(sheet)
            empty_row_streak = 0
            ssfjdm = TOPIC_BRANCH_SHEET_CODE_MAP[sheet_name]

            for row_no, row in enumerate(sheet.iter_rows(min_row=2, max_col=3, values_only=True), start=2):
                row_values = list(row or ())
                source_seq_no = _normalize_text(row_values[0] if len(row_values) > 0 else "")
                xm = _normalize_text(row_values[1] if len(row_values) > 1 else "")
                zjhm = _normalize_id_card(row_values[2] if len(row_values) > 2 else "")

                if not any((source_seq_no, xm, zjhm)):
                    empty_row_streak += 1
                    if empty_row_streak >= PARSE_EMPTY_ROW_BREAK_THRESHOLD:
                        break
                    continue

                empty_row_streak = 0
                if not zjhm:
                    continue

                imported_row_count += 1
                if zjhm in seen_zjhms:
                    continue

                seen_zjhms.add(zjhm)
                rows.append(
                    {
                        "zjhm": zjhm,
                        "xm": xm,
                        "ssfjdm": ssfjdm,
                        "source_sheet_name": sheet_name,
                        "source_row_no": row_no,
                        "source_seq_no": source_seq_no,
                    }
                )

        return ParsedBaseImportResult(
            imported_row_count=imported_row_count,
            deduplicated_person_count=len(rows),
            rows=rows,
            all_zjhms=[str(item["zjhm"]) for item in rows],
        )
    finally:
        workbook.close()