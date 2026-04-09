from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Any, BinaryIO, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from jszahzyj.dao import jszahz_topic_dao
from jszahzyj.service.jszahz_topic_relation_service import attach_relation_counts


logger = logging.getLogger(__name__)


PERSON_TYPE_OPTIONS = [
    "不规律服药",
    "弱监护",
    "无监护",
    "既往有严重自杀或伤人行为",
]

RISK_OPTIONS = [
    "0级患者",
    "1级患者",
    "2级患者",
    "3级患者",
    "4级患者",
    "5级患者",
    "无数据",
]

PERSON_TYPE_RULES = {
    "服药情况": {
        "不规律服药": "不规律服药",
    },
    "监护情况": {
        "弱监护": "弱监护",
        "无监护": "无监护",
    },
    "既往有自杀或严重伤人": {
        "是": "既往有严重自杀或伤人行为",
    },
}


@dataclass(frozen=True)
class ParsedImportResult:
    imported_row_count: int
    generated_tag_count: int
    tagged_person_count: int
    rows: List[Dict[str, Any]]


def default_time_range() -> Tuple[str, str]:
    now = datetime.now()
    end_dt = datetime(now.year, now.month, now.day, 0, 0, 0)
    start_dt = datetime(2025, 1, 1, 0, 0, 0)
    return (
        start_dt.strftime("%Y-%m-%d %H:%M:%S"),
        end_dt.strftime("%Y-%m-%d %H:%M:%S"),
    )


def _normalize_branch_options(raw_rows: List[Dict[str, Any]]) -> List[Dict[str, str]]:
    seen = set()
    out: List[Dict[str, str]] = []
    for row in raw_rows or []:
        value = str(row.get("value") or "").strip()
        label = str(row.get("label") or "").strip()
        if not value or not label or value in seen:
            continue
        seen.add(value)
        out.append({"value": value, "label": label})
    return out


def _normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def _normalize_id_card(value: Any) -> str:
    text = _normalize_text(value).replace(" ", "")
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text.upper()


def parse_person_type_workbook(file_obj: BinaryIO) -> ParsedImportResult:
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    if sheet.title != "汇总":
        raise ValueError("Excel 第一个 sheet 名称必须为“汇总”")

    imported_row_count = 0
    rows: List[Dict[str, Any]] = []
    seen = set()
    tagged_people = set()

    for row_no in range(4, sheet.max_row + 1):
        zjhm = _normalize_id_card(sheet.cell(row=row_no, column=5).value)
        if not zjhm:
            continue
        imported_row_count += 1
        matched_labels: List[str] = []
        medicine_value = _normalize_text(sheet.cell(row=row_no, column=7).value)
        guardian_value = _normalize_text(sheet.cell(row=row_no, column=8).value)
        history_value = _normalize_text(sheet.cell(row=row_no, column=9).value)

        if medicine_value in PERSON_TYPE_RULES["服药情况"]:
            matched_labels.append(PERSON_TYPE_RULES["服药情况"][medicine_value])
        if guardian_value in PERSON_TYPE_RULES["监护情况"]:
            matched_labels.append(PERSON_TYPE_RULES["监护情况"][guardian_value])
        if history_value in PERSON_TYPE_RULES["既往有自杀或严重伤人"]:
            matched_labels.append(PERSON_TYPE_RULES["既往有自杀或严重伤人"][history_value])

        if not matched_labels:
            continue
        for label in matched_labels:
            key = (zjhm, label)
            if key in seen:
                continue
            seen.add(key)
            tagged_people.add(zjhm)
            rows.append(
                {
                    "zjhm": zjhm,
                    "person_type": label,
                    "source_row_no": row_no,
                }
            )

    return ParsedImportResult(
        imported_row_count=imported_row_count,
        generated_tag_count=len(rows),
        tagged_person_count=len(tagged_people),
        rows=rows,
    )


def _serialize_batch(batch: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if not batch:
        return None
    data = dict(batch)
    for key in ("created_at", "activated_at"):
        value = data.get(key)
        if isinstance(value, datetime):
            data[key] = value.strftime("%Y-%m-%d %H:%M:%S")
    return data


def defaults_payload() -> Dict[str, Any]:
    start_time, end_time = default_time_range()
    return {
        "success": True,
        "start_time": start_time,
        "end_time": end_time,
        "branch_options": _normalize_branch_options(jszahz_topic_dao.query_branch_options()),
        "person_type_options": [{"value": item, "label": item} for item in PERSON_TYPE_OPTIONS],
        "risk_options": [{"value": item, "label": item} for item in RISK_OPTIONS],
        "active_batch": _serialize_batch(jszahz_topic_dao.get_active_batch()),
    }


def import_jszahz_topic_excel(*, file_obj: BinaryIO, filename: str, created_by: str) -> Dict[str, Any]:
    started_at = datetime.now()
    batch_id = jszahz_topic_dao.create_pending_batch(
        source_file_name=filename or "upload.xlsx",
        sheet_name="汇总",
        created_by=created_by or "",
    )

    try:
        parsed = parse_person_type_workbook(file_obj)
        logger.info(
            "jszahz topic import parsed: batch_id=%s file=%s imported_rows=%s generated_tags=%s",
            batch_id,
            filename,
            parsed.imported_row_count,
            parsed.generated_tag_count,
        )
        matched_person_count = jszahz_topic_dao.save_batch_data_and_activate(
            batch_id=batch_id,
            imported_row_count=parsed.imported_row_count,
            person_type_rows=parsed.rows,
        )
    except Exception as exc:
        parsed = locals().get("parsed")
        imported_count = parsed.imported_row_count if parsed else 0
        tag_count = parsed.generated_tag_count if parsed else 0
        jszahz_topic_dao.mark_batch_failed(
            batch_id=batch_id,
            imported_row_count=imported_count,
            generated_tag_count=tag_count,
            error_message=str(exc),
        )
        logger.exception("jszahz topic import failed: batch_id=%s file=%s", batch_id, filename)
        raise

    active_batch = jszahz_topic_dao.get_active_batch()
    elapsed_seconds = (datetime.now() - started_at).total_seconds()
    logger.info(
        "jszahz topic import completed: batch_id=%s file=%s matched_person_count=%s elapsed_seconds=%.3f",
        batch_id,
        filename,
        matched_person_count,
        elapsed_seconds,
    )
    return {
        "success": True,
        "batch_id": batch_id,
        "imported_row_count": parsed.imported_row_count,
        "generated_tag_count": parsed.generated_tag_count,
        "tagged_person_count": parsed.tagged_person_count,
        "matched_person_count": matched_person_count,
        "active_batch": _serialize_batch(active_batch),
    }


def _normalize_filters(
    *,
    start_time: str,
    end_time: str,
    branch_codes: Iterable[str] | None,
    person_types: Iterable[str] | None,
    risk_labels: Iterable[str] | None,
) -> Dict[str, Any]:
    default_start, default_end = default_time_range()
    start_text = jszahz_topic_dao.normalize_datetime_text(start_time or default_start)
    end_text = jszahz_topic_dao.normalize_datetime_text(end_time or default_end)
    if start_text > end_text:
        raise ValueError("开始时间不能大于结束时间")

    return {
        "start_time": start_text,
        "end_time": end_text,
        "branch_codes": [x.strip() for x in (branch_codes or []) if x and x.strip()],
        "person_types": [x.strip() for x in (person_types or []) if x and x.strip()],
        "risk_labels": [x.strip() for x in (risk_labels or []) if x and x.strip()],
    }


def query_summary_payload(
    *,
    start_time: str,
    end_time: str,
    branch_codes: Iterable[str] | None,
    person_types: Iterable[str] | None,
    risk_labels: Iterable[str] | None,
) -> Dict[str, Any]:
    filters = _normalize_filters(
        start_time=start_time,
        end_time=end_time,
        branch_codes=branch_codes,
        person_types=person_types,
        risk_labels=risk_labels,
    )
    active_batch = jszahz_topic_dao.get_active_batch()
    if not active_batch:
        return {
            "success": True,
            "records": [],
            "count": 0,
            "message": "暂无生效批次，请先上传人员类型 Excel",
            "filters": filters,
            "active_batch": None,
        }

    records = jszahz_topic_dao.query_summary_rows(
        batch_id=int(active_batch["id"]),
        start_time=filters["start_time"],
        end_time=filters["end_time"],
        branch_codes=filters["branch_codes"],
        person_types=filters["person_types"],
        risk_labels=filters["risk_labels"],
    )
    total_count = sum(int(row.get("去重患者数") or 0) for row in records)
    records_with_total = list(records)
    records_with_total.append(
        {
            "分局代码": "__ALL__",
            "分局名称": "汇总",
            "去重患者数": total_count,
        }
    )
    return {
        "success": True,
        "records": records_with_total,
        "count": total_count,
        "message": "" if records else "当前筛选条件下暂无数据",
        "filters": filters,
        "active_batch": _serialize_batch(active_batch),
    }


def query_detail_payload(
    *,
    branch_code: Optional[str],
    start_time: str,
    end_time: str,
    person_types: Iterable[str] | None,
    risk_labels: Iterable[str] | None,
) -> Dict[str, Any]:
    filters = _normalize_filters(
        start_time=start_time,
        end_time=end_time,
        branch_codes=None,
        person_types=person_types,
        risk_labels=risk_labels,
    )
    active_batch = jszahz_topic_dao.get_active_batch()
    if not active_batch:
        return {
            "success": True,
            "records": [],
            "count": 0,
            "message": "暂无生效批次，请先上传人员类型 Excel",
            "filters": filters,
            "active_batch": None,
        }

    records = jszahz_topic_dao.query_detail_rows(
        batch_id=int(active_batch["id"]),
        start_time=filters["start_time"],
        end_time=filters["end_time"],
        branch_code=(branch_code or "").strip() or None,
        person_types=filters["person_types"],
        risk_labels=filters["risk_labels"],
    )
    detail_records = attach_relation_counts(records)
    return {
        "success": True,
        "records": detail_records,
        "count": len(detail_records),
        "message": "" if detail_records else "当前筛选条件下暂无明细数据",
        "filters": filters,
        "active_batch": _serialize_batch(active_batch),
    }


def _sanitize_filename_text(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "-", (text or "").strip())


def _build_filter_segment(values: Iterable[str] | None, fallback: str) -> str:
    items = [str(item).strip() for item in (values or []) if str(item).strip()]
    if not items:
        return fallback
    return "_".join(_sanitize_filename_text(item) for item in items)


def _build_export_filename(
    *,
    start_time: str,
    end_time: str,
    person_types: Iterable[str] | None,
    risk_labels: Iterable[str] | None,
    is_detail: bool,
) -> str:
    start_dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
    end_dt = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")
    date_range = f"{start_dt.strftime('%Y%m%d')}-{end_dt.strftime('%Y%m%d')}"
    suffix = "精神障碍患者_详细数据" if is_detail else "精神障碍患者"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return (
        f"{date_range}_"
        f"{_build_filter_segment(person_types, '全部')}_"
        f"{_build_filter_segment(risk_labels, '全部')}_"
        f"{suffix}_{timestamp}.xlsx"
    )


def _records_to_xlsx_bytes(records: List[Dict[str, Any]], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title
    if records:
        headers = list(records[0].keys())
        ws.append(headers)
        for row in records:
            ws.append([row.get(header, "") for header in headers])
    else:
        ws.append(["暂无数据"])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_summary_xlsx(
    *,
    start_time: str,
    end_time: str,
    branch_codes: Iterable[str] | None,
    person_types: Iterable[str] | None,
    risk_labels: Iterable[str] | None,
) -> Tuple[bytes, str]:
    payload = query_summary_payload(
        start_time=start_time,
        end_time=end_time,
        branch_codes=branch_codes,
        person_types=person_types,
        risk_labels=risk_labels,
    )
    data = _records_to_xlsx_bytes(payload["records"], "精神患者主题库汇总")
    filename = _build_export_filename(
        start_time=payload["filters"]["start_time"],
        end_time=payload["filters"]["end_time"],
        person_types=payload["filters"]["person_types"],
        risk_labels=payload["filters"]["risk_labels"],
        is_detail=False,
    )
    return data, filename


def export_detail_xlsx(
    *,
    branch_code: Optional[str],
    start_time: str,
    end_time: str,
    person_types: Iterable[str] | None,
    risk_labels: Iterable[str] | None,
) -> Tuple[bytes, str]:
    payload = query_detail_payload(
        branch_code=branch_code,
        start_time=start_time,
        end_time=end_time,
        person_types=person_types,
        risk_labels=risk_labels,
    )
    data = _records_to_xlsx_bytes(payload["records"], "精神患者主题库明细")
    filename = _build_export_filename(
        start_time=payload["filters"]["start_time"],
        end_time=payload["filters"]["end_time"],
        person_types=payload["filters"]["person_types"],
        risk_labels=payload["filters"]["risk_labels"],
        is_detail=True,
    )
    return data, filename
