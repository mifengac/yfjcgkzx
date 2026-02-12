from __future__ import annotations

import csv
import io
import os
import re
from typing import Any, Dict, List, Sequence, Tuple

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from gonggong.config.database import get_database_connection
from weichengnianren.dao.wcnr_9lbq_dao import query_9lbq_rows


MAX_UPLOAD_SIZE = 20 * 1024 * 1024
IDCARD_PATTERN = re.compile(r"^\d{17}[\dXx]$")
SUPPORTED_EXTS = {".xlsx", ".csv"}


def _normalize_headers(headers: Sequence[Any]) -> List[str]:
    return [str(h).strip() if h is not None else "" for h in headers]


def _find_column_index(headers: Sequence[str], column_name: str) -> int:
    target = str(column_name or "").strip()
    if not target:
        raise ValueError("请输入身份证号码列名")

    for idx, name in enumerate(headers):
        if name == target:
            return idx
    raise ValueError(f'文件中未找到列名 "{target}"，可用列: {", ".join(headers)}')


def _extract_from_csv(data: bytes, column_name: str) -> List[str]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "gb18030"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 文件编码无法识别，请使用 UTF-8 或 GBK")

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("上传文件为空")

    headers = _normalize_headers(rows[0])
    idx = _find_column_index(headers, column_name)
    out: List[str] = []
    for row in rows[1:]:
        value = ""
        if idx < len(row):
            value = str(row[idx] or "").strip()
        out.append(value)
    return out


def _extract_from_xlsx(data: bytes, column_name: str) -> List[str]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if header_row is None:
            raise ValueError("上传文件为空")

        headers = _normalize_headers(header_row)
        idx = _find_column_index(headers, column_name)

        out: List[str] = []
        for row in rows:
            value = ""
            if row is not None and idx < len(row):
                value = str(row[idx] or "").strip()
            out.append(value)
        return out
    finally:
        wb.close()


def _validate_and_dedup(raw_values: Sequence[str]) -> Tuple[List[str], Dict[str, int]]:
    non_empty = [str(v or "").strip() for v in raw_values if str(v or "").strip()]
    valid = [v.upper() for v in non_empty if IDCARD_PATTERN.match(v)]
    unique_valid = list(dict.fromkeys(valid))
    info = {
        "源数据行数": len(raw_values),
        "非空身份证数": len(non_empty),
        "有效18位身份证数": len(valid),
        "无效身份证数": len(non_empty) - len(valid),
        "去重后身份证数": len(unique_valid),
    }
    return unique_valid, info


def extract_id_cards(upload: FileStorage, column_name: str) -> Tuple[List[str], Dict[str, int]]:
    if upload is None:
        raise ValueError("请上传文件")

    filename = str(upload.filename or "").strip()
    if not filename:
        raise ValueError("文件名为空")

    ext = os.path.splitext(filename)[1].lower()
    if ext not in SUPPORTED_EXTS:
        raise ValueError("仅支持 .xlsx 或 .csv 文件")

    data = upload.read()
    if not data:
        raise ValueError("上传文件为空")
    if len(data) > MAX_UPLOAD_SIZE:
        raise ValueError("文件过大，请控制在 20MB 以内")

    if ext == ".csv":
        raw_values = _extract_from_csv(data, column_name)
    else:
        raw_values = _extract_from_xlsx(data, column_name)

    id_cards, info = _validate_and_dedup(raw_values)
    if not id_cards:
        raise ValueError("未提取到有效的18位身份证号码")
    return id_cards, info


def query_by_upload(upload: FileStorage, column_name: str) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    id_cards, extract_info = extract_id_cards(upload, column_name)
    conn = get_database_connection()
    try:
        rows = query_9lbq_rows(conn, id_cards)
    finally:
        try:
            conn.close()
        except Exception:
            pass
    return rows, extract_info

