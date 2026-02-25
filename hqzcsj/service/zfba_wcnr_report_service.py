from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

from openpyxl import load_workbook

from gonggong.config.database import get_database_connection
from hqzcsj.dao import jzqk_tongji_dao
from hqzcsj.service import wcnr_1393zhibiao_service
from hqzcsj.service import zfba_jq_aj_service
from hqzcsj.service import zfba_wcnr_jqaj_service


REPORT_ROWS: List[Tuple[str, str, int]] = [
    ("__ALL__", "全市", 6),
    ("445302", "云城区", 7),
    ("445303", "云安区", 8),
    ("445381", "罗定市", 9),
    ("445321", "新兴县", 10),
    ("445322", "郁南县", 11),
]

REPORT_REGION_CODES = {code for code, _name, _row in REPORT_ROWS if code != "__ALL__"}

CSV_COLUMNS = [
    "地区",
    "警情",
    "同比警情比例",
    "转案率",
    "转案率-同比转案率",
    "刑事",
    "刑事占比",
    "场所案件",
    "同比场所案件比例",
    "违法犯罪人数",
    "矫治覆盖率",
    "责令加强监护率",
    "涉刑人员送矫率",
    "符合送生数/已送人数",
    "结业后再犯数(违法犯罪)",
    "结业后再犯数(犯罪)",
]


def _as_int(value: Any) -> int:
    try:
        if value is None:
            return 0
        return int(value)
    except Exception:
        return 0


def _is_yes(value: Any) -> bool:
    return str(value or "").strip() == "是"


def _is_no(value: Any) -> bool:
    return str(value or "").strip() == "否"


def _parse_percent(value: Any) -> float:
    s = str(value or "").strip()
    if s.endswith("%"):
        s = s[:-1]
    try:
        return float(s)
    except Exception:
        return 0.0


def _fmt_percent(numerator: int, denominator: int) -> str:
    if denominator <= 0:
        return "0.00%"
    return f"{(numerator / denominator) * 100:.2f}%"


def _fmt_percent_diff(current_percent: Any, compare_percent: Any) -> str:
    diff = _parse_percent(current_percent) - _parse_percent(compare_percent)
    return f"{diff:+.2f}%"


def _rows_by_code(rows: Sequence[Dict[str, Any]], *, code_keys: Sequence[str], total_code: str) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for row in rows or []:
        code = ""
        for key in code_keys:
            v = str(row.get(key) or "").strip()
            if v:
                code = v
                break
        if not code and str(row.get("地区") or "").strip() == "全市":
            code = total_code
        if code:
            out[code] = row
    return out


def _empty_jzqk_stats() -> Dict[str, int]:
    return {
        "wfzf": 0,
        "jiaozhi": 0,
        "jianhu": 0,
        "sx_songxue": 0,
        "sx_fanzui": 0,
        "fuhe_songsheng": 0,
        "yisong": 0,
    }


def _accumulate_jzqk(stats: Dict[str, int], row: Dict[str, Any]) -> None:
    stats["wfzf"] += 1

    is_jiaozhi = (
        _is_yes(row.get("是否开具矫治文书"))
        or _is_yes(row.get("是否刑拘"))
        or (_is_yes(row.get("治拘大于4天")) and _is_no(row.get("是否治拘不送")))
        or _is_yes(row.get("是否送校"))
    )
    if is_jiaozhi:
        stats["jiaozhi"] += 1

    if _is_yes(row.get("是否开具家庭教育指导书")):
        stats["jianhu"] += 1

    is_xingshi_not_xingju = str(row.get("案件类型") or "").strip() == "刑事" and _is_no(row.get("是否刑拘"))
    if is_xingshi_not_xingju:
        stats["sx_fanzui"] += 1
        if _is_yes(row.get("是否送校")):
            stats["sx_songxue"] += 1

    if _is_yes(row.get("是否符合送生")):
        stats["fuhe_songsheng"] += 1

    is_yisong = (
        _is_yes(row.get("是否刑拘"))
        or (_is_yes(row.get("治拘大于4天")) and _is_no(row.get("是否治拘不送")))
        or _is_yes(row.get("是否送校"))
    )
    if is_yisong:
        stats["yisong"] += 1


def _build_jzqk_stats_by_code(detail_rows: Sequence[Dict[str, Any]]) -> Dict[str, Dict[str, int]]:
    out: Dict[str, Dict[str, int]] = {"__ALL__": _empty_jzqk_stats()}
    for code in REPORT_REGION_CODES:
        out[code] = _empty_jzqk_stats()

    for row in detail_rows or []:
        code = str(row.get("地区") or "").strip()
        _accumulate_jzqk(out["__ALL__"], row)
        if code in REPORT_REGION_CODES:
            _accumulate_jzqk(out[code], row)
    return out


def _build_report_rows(*, start_time: str, end_time: str, leixing_list: Sequence[str]) -> List[Dict[str, Any]]:
    conn = get_database_connection()
    try:
        _meta_wcnr, wcnr_rows_raw = zfba_wcnr_jqaj_service.build_summary(
            start_time=start_time,
            end_time=end_time,
            hb_start_time=None,
            hb_end_time=None,
            leixing_list=leixing_list,
            za_types=[],
        )
        wcnr_rows = zfba_wcnr_jqaj_service.append_ratio_columns(wcnr_rows_raw)

        _meta_jq, jq_rows = zfba_jq_aj_service.build_summary(
            start_time=start_time,
            end_time=end_time,
            hb_start_time=None,
            hb_end_time=None,
            leixing_list=leixing_list,
            za_types=[],
        )

        _meta_1393, zhibiao_rows = wcnr_1393zhibiao_service.build_summary(
            start_time=start_time,
            end_time=end_time,
            leixing_list=leixing_list,
        )

        jzqk_detail_rows = jzqk_tongji_dao.fetch_jzqk_data(
            conn, start_time=start_time, end_time=end_time, leixing_list=leixing_list
        )
    finally:
        try:
            conn.close()
        except Exception:
            pass

    wcnr_by_code = _rows_by_code(wcnr_rows, code_keys=["地区代码"], total_code="__ALL__")
    jq_by_code = _rows_by_code(jq_rows, code_keys=["地区代码"], total_code="__ALL__")
    zhibiao_by_code = _rows_by_code(zhibiao_rows, code_keys=["__diqu_code"], total_code="ALL")
    jzqk_by_code = _build_jzqk_stats_by_code(jzqk_detail_rows)

    out: List[Dict[str, Any]] = []
    for code, region_name, row_idx in REPORT_ROWS:
        row_wcnr = wcnr_by_code.get(code) or {}
        row_jq = jq_by_code.get(code) or {}
        row_1393 = zhibiao_by_code.get("ALL" if code == "__ALL__" else code) or {}
        row_jzqk = jzqk_by_code.get(code) or _empty_jzqk_stats()

        value_f = _as_int(row_wcnr.get("刑事"))
        value_j = int(row_jzqk["wfzf"])
        value_k_num = int(row_jzqk["jiaozhi"])
        value_l_num = int(row_jzqk["jianhu"])
        value_m_num = int(row_jzqk["sx_songxue"])
        value_m_den = int(row_jzqk["sx_fanzui"])
        value_n_num = int(row_jzqk["fuhe_songsheng"])
        value_n_den = int(row_jzqk["yisong"])

        out.append(
            {
                "row_index": row_idx,
                "region_code": code,
                "地区": region_name,
                "B": _as_int(row_wcnr.get("警情")),
                "C": str(row_wcnr.get("同比警情比例") or "0.00%"),
                "D": str(row_wcnr.get("转案率") or "0.00%"),
                "E": _fmt_percent_diff(row_wcnr.get("转案率"), row_wcnr.get("同比转案率")),
                "F": value_f,
                "G": _fmt_percent(value_f, _as_int(row_jq.get("刑事"))),
                "H": _as_int(row_wcnr.get("场所案件")),
                "I": str(row_wcnr.get("同比场所案件比例") or "0.00%"),
                "J": value_j,
                "K": _fmt_percent(value_k_num, value_j),
                "L": _fmt_percent(value_l_num, value_j),
                "M": _fmt_percent(value_m_num, value_m_den),
                "N": f"{value_n_num}/{value_n_den}",
                "O": _as_int(row_1393.get(wcnr_1393zhibiao_service.LABEL_JYH_WFZF)),
                "P": _as_int(row_1393.get(wcnr_1393zhibiao_service.LABEL_JYH_CF)),
            }
        )
    return out


def _render_report_xlsx(rows: Sequence[Dict[str, Any]]) -> bytes:
    template_path = Path(__file__).resolve().parents[1] / "templates" / "wcnr_tjb.xlsx"
    if not template_path.exists():
        raise FileNotFoundError(f"找不到模板文件: {template_path}")

    workbook = load_workbook(template_path)
    sheet = workbook.active
    for row in rows:
        r = int(row["row_index"])
        sheet.cell(r, 2, row["B"])
        sheet.cell(r, 3, row["C"])
        sheet.cell(r, 4, row["D"])
        sheet.cell(r, 5, row["E"])
        sheet.cell(r, 6, row["F"])
        sheet.cell(r, 7, row["G"])
        sheet.cell(r, 8, row["H"])
        sheet.cell(r, 9, row["I"])
        sheet.cell(r, 10, row["J"])
        sheet.cell(r, 11, row["K"])
        sheet.cell(r, 12, row["L"])
        sheet.cell(r, 13, row["M"])
        sheet.cell(r, 14, row["N"])
        sheet.cell(r, 15, row["O"])
        sheet.cell(r, 16, row["P"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _render_report_csv(rows: Sequence[Dict[str, Any]]) -> bytes:
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=CSV_COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow(
            {
                "地区": row["地区"],
                "警情": row["B"],
                "同比警情比例": row["C"],
                "转案率": row["D"],
                "转案率-同比转案率": row["E"],
                "刑事": row["F"],
                "刑事占比": row["G"],
                "场所案件": row["H"],
                "同比场所案件比例": row["I"],
                "违法犯罪人数": row["J"],
                "矫治覆盖率": row["K"],
                "责令加强监护率": row["L"],
                "涉刑人员送矫率": row["M"],
                "符合送生数/已送人数": row["N"],
                "结业后再犯数(违法犯罪)": row["O"],
                "结业后再犯数(犯罪)": row["P"],
            }
        )
    return output.getvalue().encode("utf-8-sig")


def build_report_file(
    *, fmt: str, start_time: str, end_time: str, leixing_list: Sequence[str]
) -> Tuple[bytes, str, str]:
    fmt = str(fmt or "xlsx").strip().lower()
    if fmt not in ("xlsx", "csv"):
        raise ValueError("fmt 仅支持 xlsx 或 csv")

    rows = _build_report_rows(
        start_time=start_time,
        end_time=end_time,
        leixing_list=[str(x).strip() for x in (leixing_list or []) if str(x).strip()],
    )
    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    if fmt == "csv":
        return _render_report_csv(rows), f"未成年人治理统计报表_{ts}.csv", "text/csv; charset=utf-8"

    return (
        _render_report_xlsx(rows),
        f"未成年人治理统计报表_{ts}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
