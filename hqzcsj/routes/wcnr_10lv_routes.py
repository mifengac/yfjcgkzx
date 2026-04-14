from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
import logging
import zipfile
from typing import Any, Dict, List

from flask import Blueprint, Response, abort, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook

from gonggong.config.database import get_database_connection
from hqzcsj.dao.wcnr_10lv_dao import fetch_leixing_list
from hqzcsj.routes.route_helpers import parse_bool_arg, parse_list_arg, user_has_module_access
from hqzcsj.service import wcnr_10lv_service


wcnr_10lv_bp = Blueprint("wcnr_10lv", __name__, template_folder="../templates")


@wcnr_10lv_bp.before_request
def _check_access() -> None:
    if not session.get("username"):
        return redirect(url_for("login"))
    try:
        if not user_has_module_access(get_database_connection, username=session["username"]):
            abort(403)
    except Exception:
        abort(500)


def _parse_list_args(name: str) -> List[str]:
    return parse_list_arg(name)


def _parse_bool_arg(name: str) -> bool:
    return parse_bool_arg(name)


def _safe_sheet_name(name: str) -> str:
    illegal = set(r"[]:*?/\\")
    cleaned = "".join("_" if ch in illegal else ch for ch in str(name or "sheet"))
    cleaned = cleaned.strip() or "sheet"
    return cleaned[:31]


def _download_csv(rows: List[Dict[str, Any]], filename: str) -> Response:
    output = StringIO()
    if rows:
        headers = list(rows[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: (row.get(k) if row.get(k) is not None else "") for k in headers})
    else:
        output.write("无数据\n")

    buffer = BytesIO(output.getvalue().encode("utf-8-sig"))
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv; charset=utf-8",
    )


def _download_excel(rows: List[Dict[str, Any]], filename: str, *, sheet_name: str = "数据") -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_name(sheet_name)

    if rows:
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([(row.get(k) if row.get(k) is not None else "") for k in headers])
    else:
        sheet.append(["无数据"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@wcnr_10lv_bp.route("/wcnr_10lv/api/leixing")
@wcnr_10lv_bp.route("/wcnr10lv/api/leixing")
def api_leixing() -> Any:
    try:
        conn = get_database_connection()
        try:
            items = fetch_leixing_list(conn)
        finally:
            conn.close()
        return jsonify({"success": True, "data": items})
    except Exception as exc:
        logging.exception("wcnr_10lv api_leixing failed")
        return jsonify({"success": False, "message": str(exc)}), 500


@wcnr_10lv_bp.route("/wcnr_10lv/api/summary")
@wcnr_10lv_bp.route("/wcnr10lv/api/summary")
def api_summary() -> Any:
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    hb_start_time = (request.args.get("hb_start_time") or "").strip()
    hb_end_time = (request.args.get("hb_end_time") or "").strip()
    show_hb = _parse_bool_arg("show_hb")
    show_ratio = _parse_bool_arg("show_ratio")
    profile = _parse_bool_arg("profile")
    leixing_list = _parse_list_args("leixing")

    if not start_time or not end_time:
        start_time, end_time, hb_start_default, hb_end_default = wcnr_10lv_service.default_time_range_for_page()
        if not hb_start_time:
            hb_start_time = hb_start_default
        if not hb_end_time:
            hb_end_time = hb_end_default

    try:
        meta, rows = wcnr_10lv_service.build_summary(
            start_time=start_time,
            end_time=end_time,
            hb_start_time=hb_start_time or None,
            hb_end_time=hb_end_time or None,
            leixing_list=leixing_list,
            include_hb=show_hb,
            include_perf=profile,
        )
        columns = wcnr_10lv_service.get_display_columns(show_hb=show_hb, show_ratio=show_ratio)
        return jsonify({"success": True, "meta": meta, "columns": columns, "rows": rows})
    except Exception as exc:
        logging.exception(
            "wcnr_10lv api_summary failed: start=%s end=%s hb_start=%s hb_end=%s leixing=%s",
            start_time,
            end_time,
            hb_start_time,
            hb_end_time,
            leixing_list,
        )
        return jsonify({"success": False, "message": str(exc)}), 400


@wcnr_10lv_bp.route("/wcnr_10lv/detail")
@wcnr_10lv_bp.route("/wcnr10lv/detail")
def detail_page() -> Any:
    metric = (request.args.get("metric") or "").strip()
    part = (request.args.get("part") or "").strip() or "value"
    period = (request.args.get("period") or "").strip() or "current"
    diqu = (request.args.get("diqu") or "").strip() or "__ALL__"

    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    hb_start_time = (request.args.get("hb_start_time") or "").strip()
    hb_end_time = (request.args.get("hb_end_time") or "").strip()
    leixing_list = _parse_list_args("leixing")

    if not start_time or not end_time:
        start_time, end_time, hb_start_default, hb_end_default = wcnr_10lv_service.default_time_range_for_page()
        if not hb_start_time:
            hb_start_time = hb_start_default
        if not hb_end_time:
            hb_end_time = hb_end_default

    return render_template(
        "wcnr_10lv_detail.html",
        metric=metric,
        metric_label=wcnr_10lv_service.metric_display_name(metric, part),
        part=part,
        period=period,
        diqu=diqu,
        start_time=start_time,
        end_time=end_time,
        hb_start_time=hb_start_time,
        hb_end_time=hb_end_time,
        leixing_list=leixing_list,
    )


@wcnr_10lv_bp.route("/wcnr_10lv/api/detail")
@wcnr_10lv_bp.route("/wcnr10lv/api/detail")
def api_detail() -> Any:
    metric = (request.args.get("metric") or "").strip()
    part = (request.args.get("part") or "").strip() or "value"
    period = (request.args.get("period") or "").strip() or "current"
    diqu = (request.args.get("diqu") or "").strip() or "__ALL__"

    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    hb_start_time = (request.args.get("hb_start_time") or "").strip()
    hb_end_time = (request.args.get("hb_end_time") or "").strip()
    leixing_list = _parse_list_args("leixing")

    if not start_time or not end_time:
        start_time, end_time, hb_start_default, hb_end_default = wcnr_10lv_service.default_time_range_for_page()
        if not hb_start_time:
            hb_start_time = hb_start_default
        if not hb_end_time:
            hb_end_time = hb_end_default

    try:
        rows = wcnr_10lv_service.fetch_detail(
            metric=metric,
            part=part,
            period=period,
            diqu=diqu,
            start_time=start_time,
            end_time=end_time,
            hb_start_time=hb_start_time or None,
            hb_end_time=hb_end_time or None,
            leixing_list=leixing_list,
        )
        return jsonify({"success": True, "rows": rows})
    except Exception as exc:
        logging.exception("wcnr_10lv api_detail failed")
        return jsonify({"success": False, "message": str(exc)}), 400


@wcnr_10lv_bp.route("/wcnr_10lv/detail/export")
@wcnr_10lv_bp.route("/wcnr10lv/detail/export")
def export_detail_single() -> Response:
    fmt = (request.args.get("fmt") or "xlsx").strip().lower()
    metric = (request.args.get("metric") or "").strip()
    part = (request.args.get("part") or "").strip() or "value"
    period = (request.args.get("period") or "").strip() or "current"
    diqu = (request.args.get("diqu") or "").strip() or "__ALL__"

    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    hb_start_time = (request.args.get("hb_start_time") or "").strip()
    hb_end_time = (request.args.get("hb_end_time") or "").strip()
    leixing_list = _parse_list_args("leixing")

    if not start_time or not end_time:
        start_time, end_time, hb_start_default, hb_end_default = wcnr_10lv_service.default_time_range_for_page()
        if not hb_start_time:
            hb_start_time = hb_start_default
        if not hb_end_time:
            hb_end_time = hb_end_default

    rows = wcnr_10lv_service.fetch_detail(
        metric=metric,
        part=part,
        period=period,
        diqu=diqu,
        start_time=start_time,
        end_time=end_time,
        hb_start_time=hb_start_time or None,
        hb_end_time=hb_end_time or None,
        leixing_list=leixing_list,
    )

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"未成年人10个率明细{ts}.{fmt}"
    if fmt == "csv":
        return _download_csv(rows, filename)
    return _download_excel(rows, filename, sheet_name=wcnr_10lv_service.metric_display_name(metric, part))


@wcnr_10lv_bp.route("/wcnr_10lv/export")
@wcnr_10lv_bp.route("/wcnr10lv/export")
def export_summary() -> Response:
    fmt = (request.args.get("fmt") or "xlsx").strip().lower()
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    hb_start_time = (request.args.get("hb_start_time") or "").strip()
    hb_end_time = (request.args.get("hb_end_time") or "").strip()
    show_hb = _parse_bool_arg("show_hb")
    show_ratio = _parse_bool_arg("show_ratio")
    leixing_list = _parse_list_args("leixing")

    if not start_time or not end_time:
        start_time, end_time, hb_start_default, hb_end_default = wcnr_10lv_service.default_time_range_for_page()
        if not hb_start_time:
            hb_start_time = hb_start_default
        if not hb_end_time:
            hb_end_time = hb_end_default

    meta, rows = wcnr_10lv_service.build_summary(
        start_time=start_time,
        end_time=end_time,
        hb_start_time=hb_start_time or None,
        hb_end_time=hb_end_time or None,
        leixing_list=leixing_list,
        include_hb=show_hb,
    )
    _ = meta
    visible_rows = wcnr_10lv_service.trim_rows_for_display(rows=rows, show_hb=show_hb, show_ratio=show_ratio)

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"未成年人10个率报表统计{ts}.{fmt}"
    if fmt == "csv":
        return _download_csv(visible_rows, filename)
    return _download_excel(visible_rows, filename, sheet_name="未成年人10个率")


@wcnr_10lv_bp.route("/wcnr_10lv/export_detail")
@wcnr_10lv_bp.route("/wcnr10lv/export_detail")
def export_detail_all() -> Response:
    fmt = (request.args.get("fmt") or "xlsx").strip().lower()
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    hb_start_time = (request.args.get("hb_start_time") or "").strip()
    hb_end_time = (request.args.get("hb_end_time") or "").strip()
    show_hb = _parse_bool_arg("show_hb")
    leixing_list = _parse_list_args("leixing")

    if not start_time or not end_time:
        start_time, end_time, hb_start_default, hb_end_default = wcnr_10lv_service.default_time_range_for_page()
        if not hb_start_time:
            hb_start_time = hb_start_default
        if not hb_end_time:
            hb_end_time = hb_end_default

    sheets = wcnr_10lv_service.build_detail_export_sheets(
        start_time=start_time,
        end_time=end_time,
        hb_start_time=hb_start_time or None,
        hb_end_time=hb_end_time or None,
        leixing_list=leixing_list,
        show_hb=show_hb,
    )

    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    if fmt == "csv":
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for sheet in sheets:
                out = StringIO()
                rows = sheet.get("rows") or []
                if rows:
                    headers = list(rows[0].keys())
                    writer = csv.DictWriter(out, fieldnames=headers)
                    writer.writeheader()
                    for row in rows:
                        writer.writerow({k: (row.get(k) if row.get(k) is not None else "") for k in headers})
                else:
                    out.write("无数据\n")
                csv_name = f"{_safe_sheet_name(sheet.get('name') or '数据')}.csv"
                zf.writestr(csv_name, out.getvalue().encode("utf-8-sig"))

        zip_buffer.seek(0)
        filename = f"未成年人10个率数据{ts}.zip"
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/zip",
        )

    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet in sheets:
        ws = workbook.create_sheet(title=_safe_sheet_name(sheet.get("name") or "数据"))
        rows = sheet.get("rows") or []
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for row in rows:
                ws.append([(row.get(k) if row.get(k) is not None else "") for k in headers])
        else:
            ws.append(["无数据"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"未成年人10个率数据{ts}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
