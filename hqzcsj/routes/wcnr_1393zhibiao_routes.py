from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO, StringIO
import logging
import zipfile
from typing import Any, Dict, List

from flask import Blueprint, Response, abort, jsonify, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook

from gonggong.config.database import get_database_connection
from hqzcsj.dao.wcnr_1393zhibiao_dao import fetch_leixing_list
from hqzcsj.service.wcnr_1393zhibiao_service import (
    METRICS,
    build_summary,
    default_time_range_for_page,
    fetch_all_details,
    fetch_detail,
    map_diqu_name,
    metric_label,
)


wcnr_1393zhibiao_bp = Blueprint("wcnr_1393zhibiao", __name__, template_folder="../templates")


@wcnr_1393zhibiao_bp.before_request
def _check_access() -> None:
    if not session.get("username"):
        return redirect(url_for("login"))
    try:
        conn = get_database_connection()
        with conn.cursor() as cur:
            cur.execute(
                'SELECT 1 FROM "ywdata"."jcgkzx_permission" WHERE username=%s AND module=%s',
                (session["username"], "获取综查数据"),
            )
            row = cur.fetchone()
        conn.close()
        if not row:
            abort(403)
    except Exception:
        abort(500)


def _parse_list_args(name: str) -> List[str]:
    vals = request.args.getlist(name)
    out: List[str] = []
    for v in vals:
        s = (v or "").strip()
        if s:
            out.append(s)
    return out


_ILLEGAL_SHEET_CHARS = set(r"[]:*?/\\")


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if ch in _ILLEGAL_SHEET_CHARS else ch for ch in (name or "sheet"))
    cleaned = cleaned.strip() or "sheet"
    return cleaned[:31]


def _safe_filename_part(val: str) -> str:
    s = str(val or "").strip()
    return (
        s.replace(":", "-")
        .replace("/", "-")
        .replace("\\", "-")
        .replace(" ", "_")
        .replace("\t", "_")
    )


def _download_csv(
    rows: List[Dict[str, Any]], filename: str, *, headers: List[str] | None = None
) -> Response:
    output = StringIO()
    if rows:
        headers2 = headers or list(rows[0].keys())
        writer = csv.DictWriter(output, fieldnames=headers2)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: (row.get(k) if row.get(k) is not None else "") for k in headers2})
    else:
        output.write("无数据\n")

    buffer = BytesIO(output.getvalue().encode("utf-8-sig"))
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv; charset=utf-8",
    )


def _download_excel(
    rows: List[Dict[str, Any]],
    filename: str,
    *,
    sheet_name: str = "数据",
    headers: List[str] | None = None,
) -> Response:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _safe_sheet_name(sheet_name)

    if rows:
        headers2 = headers or list(rows[0].keys())
        sheet.append(headers2)
        for row in rows:
            sheet.append([(row.get(k) if row.get(k) is not None else "") for k in headers2])
    else:
        sheet.append(["无数据"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@wcnr_1393zhibiao_bp.route("/wcnr_1393zhibiao/api/leixing")
def api_leixing() -> Any:
    try:
        conn = get_database_connection()
        try:
            items = fetch_leixing_list(conn)
        finally:
            conn.close()
        return jsonify({"success": True, "data": items})
    except Exception as exc:
        logging.exception("wcnr_1393zhibiao api_leixing failed")
        return jsonify({"success": False, "message": str(exc)}), 500


@wcnr_1393zhibiao_bp.route("/wcnr_1393zhibiao/api/summary")
def api_summary() -> Any:
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    if not start_time or not end_time:
        start_time, end_time = default_time_range_for_page()
    leixing_list = _parse_list_args("leixing")

    try:
        meta, rows = build_summary(start_time=start_time, end_time=end_time, leixing_list=leixing_list)
        return jsonify({"success": True, "meta": meta, "rows": rows})
    except Exception as exc:
        logging.exception(
            "wcnr_1393zhibiao api_summary failed: start_time=%s end_time=%s leixing_list=%s",
            start_time,
            end_time,
            leixing_list,
        )
        return jsonify({"success": False, "message": str(exc)}), 400


@wcnr_1393zhibiao_bp.route("/wcnr_1393zhibiao/detail")
def detail_page() -> Any:
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    if not start_time or not end_time:
        start_time, end_time = default_time_range_for_page()
    leixing_list = _parse_list_args("leixing")
    metric = (request.args.get("metric") or "").strip()
    diqu = (request.args.get("diqu") or "").strip() or "ALL"
    if diqu.upper() == "ALL":
        diqu_display = "全市"
    else:
        dn = map_diqu_name(diqu)
        diqu_display = f"{dn}({diqu})" if dn != diqu else diqu

    return render_template(
        "wcnr_1393zhibiao_detail.html",
        start_time=start_time,
        end_time=end_time,
        leixing_list=leixing_list,
        metric=metric,
        metric_label=metric_label(metric),
        diqu=diqu,
        diqu_display=diqu_display,
    )


@wcnr_1393zhibiao_bp.route("/wcnr_1393zhibiao/api/detail")
def api_detail() -> Any:
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    if not start_time or not end_time:
        start_time, end_time = default_time_range_for_page()
    leixing_list = _parse_list_args("leixing")
    metric = (request.args.get("metric") or "").strip()
    diqu = (request.args.get("diqu") or "").strip() or "ALL"

    try:
        rows = fetch_detail(metric=metric, start_time=start_time, end_time=end_time, leixing_list=leixing_list, diqu=diqu)
        return jsonify({"success": True, "rows": rows})
    except Exception as exc:
        logging.exception("wcnr_1393zhibiao api_detail failed")
        return jsonify({"success": False, "message": str(exc)}), 400


@wcnr_1393zhibiao_bp.route("/wcnr_1393zhibiao/export")
def export_summary() -> Response:
    fmt = (request.args.get("fmt") or "xlsx").lower()
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    if not start_time or not end_time:
        start_time, end_time = default_time_range_for_page()
    leixing_list = _parse_list_args("leixing")

    meta, rows = build_summary(start_time=start_time, end_time=end_time, leixing_list=leixing_list)
    _ = meta

    headers = ["地区"] + [m.label for m in METRICS]
    visible_rows = [{k: r.get(k) for k in headers} for r in rows]

    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"未成年人1393指标{ts}.{fmt}"

    if fmt == "csv":
        return _download_csv(visible_rows, filename, headers=headers)
    return _download_excel(visible_rows, filename, headers=headers, sheet_name="未成年人1393指标")


@wcnr_1393zhibiao_bp.route("/wcnr_1393zhibiao/export_metric_detail")
def export_metric_detail() -> Response:
    fmt = (request.args.get("fmt") or "xlsx").lower()
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    if not start_time or not end_time:
        start_time, end_time = default_time_range_for_page()
    leixing_list = _parse_list_args("leixing")
    metric = (request.args.get("metric") or "").strip()
    diqu = (request.args.get("diqu") or "").strip() or "ALL"

    rows = fetch_detail(metric=metric, start_time=start_time, end_time=end_time, leixing_list=leixing_list, diqu=diqu)
    ts = datetime.now().strftime("%Y%m%d%H%M%S")
    metric_name = metric_label(metric)
    filename = f"未成年人1393指标_{_safe_filename_part(metric_name)}_{_safe_filename_part(diqu)}_{ts}.{fmt}"

    if fmt == "csv":
        return _download_csv(rows, filename)
    return _download_excel(rows, filename, sheet_name=metric_name)


@wcnr_1393zhibiao_bp.route("/wcnr_1393zhibiao/export_detail")
def export_detail() -> Response:
    fmt = (request.args.get("fmt") or "xlsx").lower()
    start_time = (request.args.get("start_time") or "").strip()
    end_time = (request.args.get("end_time") or "").strip()
    if not start_time or not end_time:
        start_time, end_time = default_time_range_for_page()
    leixing_list = _parse_list_args("leixing")

    ts = datetime.now().strftime("%Y%m%d%H%M%S")

    detail_map = fetch_all_details(start_time=start_time, end_time=end_time, leixing_list=leixing_list)

    if fmt == "csv":
        # csv 以 zip 形式下载：每个数据源一个 csv
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for m in METRICS:
                rows = detail_map.get(m.metric) or []
                out = StringIO()
                if rows:
                    headers = list(rows[0].keys())
                    writer = csv.DictWriter(out, fieldnames=headers)
                    writer.writeheader()
                    for r in rows:
                        writer.writerow({k: (r.get(k) if r.get(k) is not None else "") for k in headers})
                else:
                    out.write("无数据\n")

                csv_bytes = out.getvalue().encode("utf-8-sig")
                zf.writestr(f"{_safe_filename_part(m.label)}.csv", csv_bytes)

        zip_buffer.seek(0)
        filename = f"未成年人1393指标详情{ts}.zip"
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=filename,
            mimetype="application/zip",
        )

    # xlsx：每个数据源一个 sheet
    workbook = Workbook()
    # 删除默认 Sheet
    workbook.remove(workbook.active)
    for m in METRICS:
        ws = workbook.create_sheet(_safe_sheet_name(m.label))
        rows = detail_map.get(m.metric) or []
        if rows:
            headers = list(rows[0].keys())
            ws.append(headers)
            for r in rows:
                ws.append([(r.get(k) if r.get(k) is not None else "") for k in headers])
        else:
            ws.append(["无数据"])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"未成年人1393指标详情{ts}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
