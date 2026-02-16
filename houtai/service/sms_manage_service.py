from __future__ import annotations

import re
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook

from houtai.dao.sms_manage_dao import SmsManageDAO


HEADER_FIELDS: List[str] = ["xh", "xq", "xqdm", "sspcs", "sspcsdm", "xm", "zw", "lxdh", "bz"]
REQUIRED_IMPORT_FIELDS: List[str] = ["xq", "xqdm", "sspcs", "sspcsdm", "xm", "zw", "lxdh", "bz"]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_xh(value: Any) -> int | None:
    text = normalize_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except Exception as exc:  # noqa: BLE001
        raise ValueError("字段 xh 必须为整数") from exc


def normalize_lxdh(raw: Any) -> str:
    text = normalize_text(raw)
    if not text:
        return ""

    normalized = re.sub(r"[，；;、\s]+", ",", text)
    parts = [p.strip() for p in normalized.split(",") if p and p.strip()]

    dedup: List[str] = []
    seen = set()
    for p in parts:
        if p in seen:
            continue
        seen.add(p)
        dedup.append(p)

    return ",".join(dedup)


def _normalize_item(item: Dict[str, Any]) -> Dict[str, Any]:
    out = {
        "xh": normalize_xh(item.get("xh")),
        "xq": normalize_text(item.get("xq")),
        "xqdm": normalize_text(item.get("xqdm")),
        "sspcs": normalize_text(item.get("sspcs")),
        "sspcsdm": normalize_text(item.get("sspcsdm")),
        "xm": normalize_text(item.get("xm")),
        "zw": normalize_text(item.get("zw")),
        "lxdh": normalize_lxdh(item.get("lxdh")),
        "bz": normalize_text(item.get("bz")),
    }

    if not out["sspcsdm"] or not out["xm"]:
        raise ValueError("字段 sspcsdm、xm 不能为空")

    return out


def _validate_import_item(item: Dict[str, Any]) -> None:
    # xh 为展示字段，不作为导入必填
    for field in REQUIRED_IMPORT_FIELDS:
        if not normalize_text(item.get(field)):
            raise ValueError(f"字段 {field} 不能为空")


def parse_rows_from_excel(ws: Any) -> Tuple[List[Dict[str, Any]], List[str]]:
    headers = [normalize_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]

    missing = [x for x in REQUIRED_IMPORT_FIELDS if x not in headers]
    if missing:
        raise ValueError(f"缺少必要列: {', '.join(missing)}")

    idx = {name: headers.index(name) for name in headers}

    valid_rows: List[Dict[str, Any]] = []
    errors: List[str] = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        item = {field: row[idx[field]].value if field in idx else None for field in HEADER_FIELDS}
        # 空行跳过
        if all(not normalize_text(v) for v in item.values()):
            continue

        try:
            _validate_import_item(item)
            valid_rows.append(_normalize_item(item))
        except Exception as exc:  # noqa: BLE001
            errors.append(f"第{row_num}行: {exc}")

    return valid_rows, errors


def build_template_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "短信发送管理"
    ws.append(HEADER_FIELDS)
    ws.append(["", "云城区", "445302", "云城分局某派出所", "445302001", "张三", "民警", "13800000000,13900000000", "示例备注"])
    return wb


class SmsManageService:
    @staticmethod
    def list_rows(*, keyword: str, page: int, page_size: int) -> Tuple[List[Dict[str, Any]], int]:
        return SmsManageDAO.list_rows(keyword=keyword, page=page, page_size=page_size)

    @staticmethod
    def save_one(item: Dict[str, Any]) -> int:
        normalized = _normalize_item(item)
        return SmsManageDAO.upsert_rows([normalized])

    @staticmethod
    def save_many(items: Iterable[Dict[str, Any]]) -> int:
        normalized_items = [_normalize_item(x) for x in items]
        return SmsManageDAO.upsert_rows(normalized_items)

    @staticmethod
    def delete_by_keys(keys: Sequence[Tuple[str, str]]) -> int:
        return SmsManageDAO.delete_by_keys(keys)
