from __future__ import annotations

"""
后台模块（houtai）：
- 用户管理：增删改查、权限管理
- Excel批量导入用户与权限
"""

import io
import os
from typing import List, Tuple

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from openpyxl import Workbook, load_workbook

from gonggong.config.database import get_database_connection


houtai_bp = Blueprint("houtai", __name__, template_folder="../../templates")


# ─── 用户管理页面 ───

@houtai_bp.route("/")
def import_page():
    """用户管理页面"""
    if not session.get("username"):
        return redirect(url_for("login"))
    return render_template("houtai_user_manage.html")


# ─── 用户管理 API ───

@houtai_bp.route("/api/users")
def list_users():
    """获取用户列表（含权限）"""
    if not session.get("username"):
        return jsonify({"success": False, "message": "请先登录"}), 401

    try:
        conn = get_database_connection()
        try:
            with conn.cursor() as cur:
                # 获取所有用户
                cur.execute('SELECT id, username FROM "ywdata"."jcgkzx_user" ORDER BY id')
                rows = cur.fetchall()
                users = []
                for row in rows:
                    user_id, username = row
                    # 获取该用户的权限
                    cur.execute(
                        'SELECT module FROM "ywdata"."jcgkzx_permission" WHERE username = %s',
                        (username,)
                    )
                    perms = [r[0] for r in cur.fetchall()]
                    users.append({
                        "id": user_id,
                        "username": username,
                        "perms": perms
                    })
                return jsonify({"success": True, "data": users})
        finally:
            conn.close()
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 500


@houtai_bp.route("/api/users/save", methods=["POST"])
def save_user():
    """新增或编辑用户"""
    if not session.get("username"):
        return jsonify({"success": False, "message": "请先登录"}), 401

    data = request.get_json(silent=True) or {}
    original_username = (data.get("original_username") or "").strip()
    username = (data.get("username") or "").strip()
    password = (data.get("password") or "").strip()
    permissions = data.get("permissions") or []

    if not username:
        return jsonify({"success": False, "message": "用户名不能为空"}), 400

    try:
        conn = get_database_connection()
        try:
            with conn:
                with conn.cursor() as cur:
                    # 确保表存在
                    cur.execute('''
                        CREATE TABLE IF NOT EXISTS "ywdata"."jcgkzx_user" (
                            id SERIAL PRIMARY KEY,
                            username VARCHAR(50) UNIQUE NOT NULL,
                            password VARCHAR(255) NOT NULL
                        )
                    ''')
                    cur.execute('''
                        CREATE TABLE IF NOT EXISTS "ywdata"."jcgkzx_permission" (
                            id SERIAL PRIMARY KEY,
                            username VARCHAR(50) NOT NULL,
                            module VARCHAR(50) NOT NULL,
                            UNIQUE (username, module)
                        )
                    ''')

                    if original_username:
                        # 编辑模式
                        if password:
                            hashed = generate_password_hash(password)
                            cur.execute(
                                'UPDATE "ywdata"."jcgkzx_user" SET password = %s WHERE username = %s',
                                (hashed, original_username)
                            )
                        # 更新权限：先删后增
                        cur.execute(
                            'DELETE FROM "ywdata"."jcgkzx_permission" WHERE username = %s',
                            (original_username,)
                        )
                        for perm in permissions:
                            if perm.strip():
                                cur.execute(
                                    'INSERT INTO "ywdata"."jcgkzx_permission" (username, module) VALUES (%s, %s) ON CONFLICT DO NOTHING',
                                    (original_username, perm.strip())
                                )
                        # 如果用户名也改了
                        if username != original_username:
                            cur.execute(
                                'UPDATE "ywdata"."jcgkzx_user" SET username = %s WHERE username = %s',
                                (username, original_username)
                            )
                            cur.execute(
                                'UPDATE "ywdata"."jcgkzx_permission" SET username = %s WHERE username = %s',
                                (username, original_username)
                            )
                    else:
                        # 新增模式
                        if not password:
                            return jsonify({"success": False, "message": "密码不能为空"}), 400
                        # 检查用户名是否已存在
                        cur.execute(
                            'SELECT 1 FROM "ywdata"."jcgkzx_user" WHERE username = %s',
                            (username,)
                        )
                        if cur.fetchone():
                            return jsonify({"success": False, "message": "用户名已存在"}), 400

                        hashed = generate_password_hash(password)
                        cur.execute(
                            'INSERT INTO "ywdata"."jcgkzx_user" (username, password) VALUES (%s, %s)',
                            (username, hashed)
                        )
                        for perm in permissions:
                            if perm.strip():
                                cur.execute(
                                    'INSERT INTO "ywdata"."jcgkzx_permission" (username, module) VALUES (%s, %s) ON CONFLICT DO NOTHING',
                                    (username, perm.strip())
                                )

            return jsonify({"success": True, "message": "保存成功"})
        finally:
            conn.close()
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 500


@houtai_bp.route("/api/users/delete", methods=["POST"])
def delete_user():
    """删除用户"""
    if not session.get("username"):
        return jsonify({"success": False, "message": "请先登录"}), 401

    data = request.get_json(silent=True) or {}
    username = (data.get("username") or "").strip()

    if not username:
        return jsonify({"success": False, "message": "用户名不能为空"}), 400

    # 不允许删除自己
    if username == session.get("username"):
        return jsonify({"success": False, "message": "不能删除当前登录用户"}), 400

    try:
        conn = get_database_connection()
        try:
            with conn:
                with conn.cursor() as cur:
                    # 删除权限
                    cur.execute(
                        'DELETE FROM "ywdata"."jcgkzx_permission" WHERE username = %s',
                        (username,)
                    )
                    # 删除用户
                    cur.execute(
                        'DELETE FROM "ywdata"."jcgkzx_user" WHERE username = %s',
                        (username,)
                    )
            return jsonify({"success": True, "message": "删除成功"})
        finally:
            conn.close()
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 500


# ─── Excel批量导入 ───

@houtai_bp.route("/download_template")
def download_template():
    """下载Excel导入模板"""
    if not session.get("username"):
        return redirect(url_for("login"))

    wb = Workbook()
    ws = wb.active
    ws.title = "用户权限导入"
    ws.append(["username", "password", "module"])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name="用户与权限导入模板.xlsx")


@houtai_bp.route("/upload", methods=["POST"])
def upload_import():
    """处理Excel上传并导入用户与权限"""
    if not session.get("username"):
        return redirect(url_for("login"))

    file = request.files.get("file")
    if not file:
        return jsonify({"success": False, "message": "请选择Excel文件"}), 400

    filename = secure_filename(file.filename)
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"success": False, "message": "文件格式错误，请上传xlsx/xlsm格式"}), 400

    try:
        wb = load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        required = ["username", "password", "module"]
        for col in required:
            if col not in headers:
                return jsonify({"success": False, "message": f"缺少必要列：{col}"}), 400
        idx = {name: headers.index(name) for name in required}

        to_users: List[Tuple[str, str]] = []
        to_perms: List[Tuple[str, str]] = []
        for r in ws.iter_rows(min_row=2):
            u = str(r[idx["username"]].value).strip() if r[idx["username"]].value is not None else ""
            p = str(r[idx["password"]].value).strip() if r[idx["password"]].value is not None else ""
            m = str(r[idx["module"]].value).strip() if r[idx["module"]].value is not None else ""
            if not u or not p or not m:
                continue
            to_users.append((u, generate_password_hash(p)))
            to_perms.append((u, m))

        if not to_users:
            return jsonify({"success": False, "message": "文件中没有有效数据"}), 400

        conn = get_database_connection()
        try:
            with conn:
                with conn.cursor() as cur:
                    cur.execute('''
                        CREATE TABLE IF NOT EXISTS "ywdata"."jcgkzx_user" (
                            id SERIAL PRIMARY KEY,
                            username VARCHAR(50) UNIQUE NOT NULL,
                            password VARCHAR(255) NOT NULL
                        )
                    ''')
                    cur.execute('''
                        CREATE TABLE IF NOT EXISTS "ywdata"."jcgkzx_permission" (
                            id SERIAL PRIMARY KEY,
                            username VARCHAR(50) NOT NULL,
                            module VARCHAR(50) NOT NULL,
                            UNIQUE (username, module)
                        )
                    ''')

                    for u, hp in to_users:
                        cur.execute(
                            'INSERT INTO "ywdata"."jcgkzx_user" (username, password) VALUES (%s, %s) '
                            'ON CONFLICT (username) DO UPDATE SET password = EXCLUDED.password',
                            (u, hp),
                        )

                    for u, m in to_perms:
                        cur.execute(
                            'INSERT INTO "ywdata"."jcgkzx_permission" (username, module) VALUES (%s, %s) '
                            'ON CONFLICT (username, module) DO NOTHING',
                            (u, m),
                        )

            return jsonify({"success": True, "message": f"导入完成，用户{len(to_users)}条，权限{len(to_perms)}条"})
        finally:
            conn.close()
    except Exception as exc:
        return jsonify({"success": False, "message": str(exc)}), 500
