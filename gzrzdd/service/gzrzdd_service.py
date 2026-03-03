from __future__ import annotations

import csv
import io
import math
import re
import time
import uuid
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from gonggong.utils.error_handler import log_info
from gzrzdd.dao.gzrzdd_dao import DEFAULT_GZRZ_SQL, find_col, query_to_dataframe


COL_ID = "证件号码"
COL_TEXT = "工作日志工作情况说明"
COL_BRANCH = "分局名称"
COL_STATION = "所属派出所"
COL_NAME = "姓名"
COL_SORT = "列管时间"
COL_WORK_TIME = "工作日志开展工作时间"


def _now_ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _norm(s: Any) -> str:
    return ("" if s is None else str(s)).strip()


def parse_threshold_percent(v: Any) -> float:
    s = _norm(v)
    if not s:
        return 0.8
    try:
        f = float(s)
    except Exception:
        return 0.8
    if f > 1.0:
        f = f / 100.0
    if f < 0:
        f = 0.0
    if f > 1:
        f = 1.0
    return float(f)


_RE_DATE_PREFIX = re.compile(
    r"^\s*"
    r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"
    r"(?:\s*(?:上午|下午|中午|晚上|早上|凌晨|傍晚|夜间|夜晚|[上中下]午))?"
    r"(?:\s*[，,、]\s*|\s+)?"
)
_RE_DATE_PREFIX2 = re.compile(
    r"^\s*"
    r"(20\d{2})[-/\.](\d{1,2})[-/\.](\d{1,2})"
    r"(?:\s*(?:\d{1,2}:\d{2}(?::\d{2})?)?)?"
    r"(?:\s*[，,、]\s*|\s+)?"
)


def clean_text(text: str, *, station: str, name: str) -> str:
    t = _norm(text)
    if not t:
        return ""
    t = _RE_DATE_PREFIX.sub("", t)
    t = _RE_DATE_PREFIX2.sub("", t)
    if station and len(station) > 1:
        t = t.replace(station, "")
    if name and len(name) > 1:
        t = t.replace(name, "")
    t = re.sub(r"\s+", " ", t).strip()
    t = re.sub(r"^[，,、:：;；\-\s]+", "", t)
    return t.strip()


_RE_TOKEN = re.compile(r"[A-Za-z0-9]+|[\u4e00-\u9fff]+")


def tokenize(text: str) -> List[str]:
    text = _norm(text)
    if not text:
        return []
    tokens: List[str] = []
    for m in _RE_TOKEN.finditer(text):
        w = m.group(0)
        if not w:
            continue
        if re.fullmatch(r"[A-Za-z0-9]+", w):
            tokens.append(w.lower())
            continue
        if len(w) <= 2:
            tokens.extend(list(w))
        else:
            for i in range(len(w) - 1):
                tokens.append(w[i : i + 2])
    return tokens


def build_tfidf_vectors(texts: List[str]) -> List[Dict[str, float]]:
    docs = [tokenize(t) for t in texts]
    n = len(docs)
    if n == 0:
        return []
    df: Dict[str, int] = {}
    for toks in docs:
        for tok in set(toks):
            df[tok] = df.get(tok, 0) + 1
    idf: Dict[str, float] = {}
    for tok, dfi in df.items():
        idf[tok] = math.log((n + 1.0) / (dfi + 1.0)) + 1.0
    vecs: List[Dict[str, float]] = []
    for toks in docs:
        tf: Dict[str, int] = {}
        for tok in toks:
            tf[tok] = tf.get(tok, 0) + 1
        vec: Dict[str, float] = {}
        for tok, c in tf.items():
            w = (1.0 + math.log(c)) * idf.get(tok, 0.0)
            if w:
                vec[tok] = w
        norm = math.sqrt(sum(v * v for v in vec.values())) or 1.0
        for tok in list(vec.keys()):
            vec[tok] = vec[tok] / norm
        vecs.append(vec)
    return vecs


def cosine_sparse(a: Dict[str, float], b: Dict[str, float]) -> float:
    if not a or not b:
        return 0.0
    if len(a) > len(b):
        a, b = b, a
    s = 0.0
    for k, va in a.items():
        vb = b.get(k)
        if vb is not None:
            s += va * vb
    return float(s)


class UnionFind:
    def __init__(self, n: int) -> None:
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, x: int) -> int:
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a: int, b: int) -> None:
        ra = self.find(a)
        rb = self.find(b)
        if ra == rb:
            return
        if self.rank[ra] < self.rank[rb]:
            self.parent[ra] = rb
            return
        if self.rank[ra] > self.rank[rb]:
            self.parent[rb] = ra
            return
        self.parent[rb] = ra
        self.rank[ra] += 1


def components_by_similarity(texts: List[str], threshold: float) -> Tuple[List[List[int]], List[float]]:
    n = len(texts)
    if n < 2:
        return [], []
    vecs = build_tfidf_vectors(texts)
    uf = UnionFind(n)
    max_edge = [0.0] * n
    for i in range(n):
        for j in range(i + 1, n):
            sim = cosine_sparse(vecs[i], vecs[j])
            if sim >= threshold:
                uf.union(i, j)
                if sim > max_edge[i]:
                    max_edge[i] = sim
                if sim > max_edge[j]:
                    max_edge[j] = sim
    comps: Dict[int, List[int]] = {}
    for i in range(n):
        comps.setdefault(uf.find(i), []).append(i)
    out_comps: List[List[int]] = []
    out_scores: List[float] = []
    for members in comps.values():
        if len(members) < 2:
            continue
        score = max(max_edge[m] for m in members)
        if score >= threshold:
            out_comps.append(sorted(members))
            out_scores.append(float(score))
    paired = list(zip(out_comps, out_scores))
    paired.sort(key=lambda x: (x[1], len(x[0])), reverse=True)
    return [p[0] for p in paired], [p[1] for p in paired]


@dataclass
class CachedResult:
    created_at: float
    count: int
    threshold: float
    pivot: pd.DataFrame
    dup_person_df: pd.DataFrame


CACHE: Dict[str, CachedResult] = {}
CACHE_TTL_SECONDS = 2 * 60 * 60


def cache_gc() -> None:
    now = time.time()
    stale = [k for k, v in CACHE.items() if now - v.created_at > CACHE_TTL_SECONDS]
    for k in stale:
        CACHE.pop(k, None)


def _required_columns(df: pd.DataFrame) -> Dict[str, str]:
    c_id = find_col(df, COL_ID)
    c_text = find_col(df, COL_TEXT)
    c_branch = find_col(df, COL_BRANCH)
    c_station = find_col(df, COL_STATION)
    c_name = find_col(df, COL_NAME)  # optional
    c_sort = find_col(df, COL_SORT)
    c_work_time = find_col(df, COL_WORK_TIME)
    missing = [
        n
        for n, c in [
            (COL_ID, c_id),
            (COL_TEXT, c_text),
            (COL_BRANCH, c_branch),
            (COL_STATION, c_station),
            (COL_SORT, c_sort),
            (COL_WORK_TIME, c_work_time),
        ]
        if not c
    ]
    if missing:
        raise ValueError(f"SQL 结果缺少必要字段：{', '.join(missing)}（可通过 SQL AS 对齐字段名）")
    return {
        "id": c_id or COL_ID,
        "text": c_text or COL_TEXT,
        "branch": c_branch or COL_BRANCH,
        "station": c_station or COL_STATION,
        "name": c_name or "",
        "sort": c_sort or COL_SORT,
        "work_time": c_work_time or COL_WORK_TIME,
    }

def _format_dt_any(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (pd.Timestamp, datetime)):
        if pd.isna(v):
            return ""
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return str(v)
    return str(v).strip()


def _join_times(values: Iterable[Any]) -> str:
    seen = set()
    out: List[str] = []
    for v in values:
        s = _format_dt_any(v)
        if not s:
            continue
        if s in seen:
            continue
        seen.add(s)
        out.append(s)
    lines: List[str] = []
    for i, s in enumerate(out, start=1):
        lines.append(f"{i}. {s}")
    return "\n".join(lines)

def _join_texts(values: Iterable[Any]) -> str:
    out: List[str] = []
    for v in values:
        s = "" if v is None else str(v).strip()
        s = re.sub(r"\s+", " ", s)
        if not s:
            continue
        out.append(s)
    lines: List[str] = []
    for i, s in enumerate(out, start=1):
        lines.append(f"{i}. {s}")
    return "\n".join(lines)


def compute_stats(*, count: int, threshold_percent: float) -> Tuple[str, Dict[str, Any]]:
    cache_gc()
    df = query_to_dataframe(DEFAULT_GZRZ_SQL)
    if df.empty:
        rid = uuid.uuid4().hex
        CACHE[rid] = CachedResult(time.time(), count, parse_threshold_percent(threshold_percent), pd.DataFrame(), df.copy())
        return rid, {"rows": [], "cols": [], "data": []}

    cols = _required_columns(df)
    c_id = cols["id"]
    c_text = cols["text"]
    c_branch = cols["branch"]
    c_station = cols["station"]
    c_name = cols["name"]
    c_sort = cols["sort"]
    c_work_time = cols["work_time"]

    thr = parse_threshold_percent(threshold_percent)

    work = df.copy()
    work[c_sort] = pd.to_datetime(work[c_sort], errors="coerce")
    # 取“最近N条”时，按“开展工作时间”倒序优先（同一人内）
    work["__work_dt_sel"] = pd.to_datetime(work[c_work_time], errors="coerce")
    work["__work_dt_sel_filled"] = work["__work_dt_sel"].fillna(pd.Timestamp.min)
    work = work.sort_values(
        by=[c_id, "__work_dt_sel_filled", c_sort],
        ascending=[True, False, False],
        kind="mergesort",
    )
    latest = work.groupby(c_id, sort=False).head(int(count)).copy()
    latest = latest.reset_index(drop=True)
    latest = latest.drop(columns=["__work_dt_sel", "__work_dt_sel_filled"], errors="ignore")

    station_vals = latest[c_station].astype(str).fillna("")
    name_vals = latest[c_name].astype(str).fillna("") if c_name else [""] * len(latest)
    latest["__clean_text"] = [
        clean_text(str(t), station=str(st), name=str(nm))
        for t, st, nm in zip(latest[c_text].astype(str).fillna(""), station_vals, name_vals)
    ]

    rows_out: List[pd.DataFrame] = []
    serial = 1
    for _, g in latest.groupby(c_id, sort=False):
        texts = g["__clean_text"].tolist()
        comps, scores = components_by_similarity(texts, thr)
        if not comps:
            continue
        for comp, score in zip(comps, scores):
            sub = g.iloc[comp].copy()
            sub["序号"] = serial
            sub["重复度"] = f"{score * 100:.2f}%"
            rows_out.append(sub)
            serial += 1

    if rows_out:
        dup_df = pd.concat(rows_out, ignore_index=True)
    else:
        dup_df = latest.iloc[0:0].copy()

    # 关键改动：重复度计算完成后，先按“证件号码”聚合，并拼接“工作日志开展工作时间”
    if dup_df.empty:
        dup_person_df = dup_df.copy()
        pivot = pd.DataFrame()
    else:
        tmp = dup_df.copy()
        tmp["__rep_score"] = (
            tmp.get("重复度", "").astype(str).str.replace("%", "", regex=False).astype(float).fillna(0.0)
        )
        # 按开展工作时间排序后再聚合，保证“开展工作时间”和“工作情况说明”拼接顺序一致
        tmp["__work_dt"] = pd.to_datetime(tmp[c_work_time], errors="coerce")
        tmp = tmp.sort_values(by=[c_id, "__work_dt"], ascending=[True, True], kind="mergesort")

        def _agg_one(g: pd.DataFrame) -> pd.Series:
            return pd.Series(
                {
                    c_branch: g[c_branch].iloc[0],
                    c_station: g[c_station].iloc[0],
                    (c_name if c_name else c_id): (g[c_name].iloc[0] if c_name else g[c_id].iloc[0]),
                    c_sort: g[c_sort].max(),
                    c_work_time: _join_times(g[c_work_time].tolist()),
                    c_text: _join_texts(g[c_text].tolist()),
                    "__rep_score": g["__rep_score"].max(),
                }
            )

        agg = tmp.groupby(c_id, as_index=False).apply(_agg_one).reset_index(drop=True).rename(
            columns={"__rep_score": "重复度(%)"}
        )
        agg["重复度"] = agg["重复度(%)"].map(lambda x: f"{float(x):.2f}%")
        agg.insert(0, "序号", range(1, len(agg) + 1))
        dup_person_df = agg.drop(columns=["重复度(%)"])

        # 交叉表：按“分局名称(列)”汇总（页面不再按“所属派出所”分组）
        series = dup_person_df.groupby(c_branch, dropna=False)[c_id].count()
        pivot = pd.DataFrame([series])
        pivot.index = ["合计"]
        pivot = pivot.fillna(0).astype(int)

    rid = uuid.uuid4().hex
    CACHE[rid] = CachedResult(time.time(), count, thr, pivot, dup_person_df)

    payload = pivot_to_payload(pivot)
    log_info(f"gzrzdd stats computed: result_id={rid}, person_rows={len(dup_person_df)}")
    return rid, payload


def pivot_to_payload(pivot: pd.DataFrame) -> Dict[str, Any]:
    if pivot is None or pivot.empty:
        return {"rows": [], "cols": [], "data": []}
    rows = [str(x) for x in pivot.index.tolist()]
    cols = [str(x) for x in pivot.columns.tolist()]
    data = [[int(pivot.loc[r, c]) for c in cols] for r in rows]
    return {"rows": rows, "cols": cols, "data": data}


def get_detail_records(result_id: str, *, branch: str, station: str, limit: int = 5000) -> List[Dict[str, Any]]:
    cache_gc()
    res = CACHE.get(result_id)
    if not res:
        raise ValueError("result_id 不存在或已过期，请重新统计")
    df = res.dup_person_df
    if df is None or df.empty:
        return []

    c_branch = find_col(df, COL_BRANCH) or COL_BRANCH
    c_station = find_col(df, COL_STATION) or COL_STATION
    if station and station != "__ALL__":
        sub = df[(df[c_branch].astype(str) == branch) & (df[c_station].astype(str) == station)].copy()
    else:
        sub = df[(df[c_branch].astype(str) == branch)].copy()

    front = [c for c in ["序号", "重复度"] if c in sub.columns]
    rest = [c for c in sub.columns if c not in front and c not in ("__clean_text",)]
    sub = sub[front + rest]

    if len(sub) > int(limit):
        sub = sub.head(int(limit)).copy()

    out: List[Dict[str, Any]] = []
    for r in sub.to_dict(orient="records"):
        clean: Dict[str, Any] = {}
        for k, v in r.items():
            if k == "__clean_text":
                continue
            if isinstance(v, (pd.Timestamp,)):
                clean[k] = "" if pd.isna(v) else v.strftime("%Y-%m-%d %H:%M:%S")
            else:
                clean[k] = "" if pd.isna(v) else v
        out.append(clean)
    return out


def export_summary(result_id: str, *, fmt: str, count: int) -> Tuple[bytes, str, str]:
    cache_gc()
    res = CACHE.get(result_id)
    if not res:
        raise ValueError("result_id 不存在或已过期，请重新统计")
    pivot = res.pivot
    if pivot is None or pivot.empty:
        df_out = pd.DataFrame(columns=[COL_STATION])
    else:
        df_out = pivot.reset_index()

    ts = _now_ts()
    fmt = (fmt or "xlsx").lower()
    filename = f"矛盾纠纷风险人员工作日志重复度统计_{count}_{ts}.{fmt}"
    if fmt == "csv":
        return _df_to_csv_bytes(df_out), "text/csv; charset=utf-8", filename
    return _df_to_xlsx_bytes(df_out, sheet="汇总"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename


def export_detail(result_id: str, *, branch: str, station: str, fmt: str, count: int) -> Tuple[bytes, str, str]:
    cache_gc()
    res = CACHE.get(result_id)
    if not res:
        raise ValueError("result_id 不存在或已过期，请重新统计")
    df = res.dup_person_df
    if df is None or df.empty:
        df_out = pd.DataFrame()
    else:
        c_branch = find_col(df, COL_BRANCH) or COL_BRANCH
        c_station = find_col(df, COL_STATION) or COL_STATION
        if station and station != "__ALL__":
            df_out = df[(df[c_branch].astype(str) == branch) & (df[c_station].astype(str) == station)].copy()
        else:
            df_out = df[(df[c_branch].astype(str) == branch)].copy()
        front = [c for c in ["序号", "重复度"] if c in df_out.columns]
        rest = [c for c in df_out.columns if c not in front and c not in ("__clean_text",)]
        df_out = df_out[front + rest]

    ts = _now_ts()
    fmt = (fmt or "xlsx").lower()
    filename = f"矛盾纠纷风险人员工作日志重复度统计明细_{count}_{ts}.{fmt}"
    if fmt == "csv":
        return _df_to_csv_bytes(df_out), "text/csv; charset=utf-8", filename
    return _df_to_xlsx_bytes(df_out, sheet="明细"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename


def _df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(list(df.columns))
    for _, row in df.iterrows():
        w.writerow(["" if pd.isna(x) else x for x in row.tolist()])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _df_to_xlsx_bytes(df: pd.DataFrame, *, sheet: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append(["" if pd.isna(x) else x for x in row.tolist()])

    # 让“工作日志开展工作时间”能完整显示类似“1. 2025-01-01”，避免列太窄导致同一行内折行
    for col_idx, col_name in enumerate(df.columns, start=1):
        name = "" if col_name is None else str(col_name).strip()
        if name != COL_WORK_TIME and "开展工作时间" not in name:
            continue
        max_len = len(name)
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            s = "" if v is None else str(v)
            for line in s.splitlines() or [""]:
                if len(line) > max_len:
                    max_len = len(line)
            ws.cell(row=r, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, min(80, max_len + 2))
        break
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

