import io
import unittest
from datetime import datetime as real_datetime
from unittest.mock import patch

from flask import Flask
from openpyxl import load_workbook

from jingqing_fenxi.routes.jingqing_fenxi_routes import jingqing_fenxi_bp
from jingqing_fenxi.service import gambling_analysis_export_service as export_service


class TestGamblingAnalysisExportService(unittest.TestCase):
    def test_generate_analysis_data_export_uses_case_list_and_sanitized_sheets(self) -> None:
        params = {
            "beginDate": "2026-04-01 00:00:00",
            "endDate": "2026-04-08 00:00:00",
        }
        rows = [
            {
                "caseNo": "JQ001",
                "callTime": "2026-04-03 01:00:00",
                "cmdId": "445300000000",
                "dutyDeptNo": "445302010000",
                "dutyDeptName": "测试派出所",
                "newCaseSourceNo": "01",
                "newCaseSource": "电话报警",
                "newCharaSubclassNo": "0301",
                "newOriCharaSubclassNo": "0301",
                "callerPhone": "13800138000",
                "occurAddress": "测试地址",
                "caseContents": "有人打麻将扰民",
                "replies": "到场未发现赌博",
                "lngOfCriterion": "112.1",
                "latOfCriterion": "22.1",
            }
        ]
        with patch.object(export_service, "resolve_gambling_topic_tags", return_value=("0301", "赌博")), patch.object(
            export_service,
            "fetch_all_case_list",
            return_value=rows,
        ) as mock_fetch, patch.object(export_service, "_build_database_sheets", return_value=[]):
            buffer = export_service.generate_gambling_analysis_export(params)

        workbook = load_workbook(io.BytesIO(buffer.getvalue()))
        self.assertIn("01_警情脱敏明细", workbook.sheetnames)
        self.assertIn("05_派出所有效性", workbook.sheetnames)
        self.assertIn("16_连续上升周", workbook.sheetnames)
        self.assertIn("17_处警结果细分", workbook.sheetnames)
        self.assertIn("18_赌资赌具证据", workbook.sheetnames)
        self.assertIn("19_线索质量风险", workbook.sheetnames)
        headers = [cell.value for cell in workbook["01_警情脱敏明细"][1]]
        self.assertIn("duty_dept_no", headers)
        self.assertIn("disposal_result", headers)
        self.assertIn("evidence_signal", headers)
        self.assertIn("report_risk_signal", headers)
        self.assertIn("consistency_signal", headers)
        self.assertNotIn("dutyDeptName", headers)
        values = [cell.value for row in workbook["01_警情脱敏明细"].iter_rows() for cell in row]
        self.assertNotIn("测试派出所", values)
        self.assertIn("举报后未抓现行", values)
        current_payload = mock_fetch.call_args_list[0].args[0]
        self.assertEqual(current_payload["newCharaSubclassNo"], "")
        self.assertEqual(current_payload["newOriCharaSubclassNo"], "0301")
        self.assertEqual(mock_fetch.call_args.kwargs["max_page_size"], export_service.GAMBLING_TOPIC_UPSTREAM_PAGE_SIZE)

    def test_generate_analysis_data_export_can_include_plain_detail_attachment(self) -> None:
        params = {
            "beginDate": "2026-04-01 00:00:00",
            "endDate": "2026-04-08 00:00:00",
            "desensitized": "0",
        }
        rows = [
            {
                "caseNo": "JQ001",
                "callTime": "2026-04-03 01:00:00",
                "occurTime": "2026-04-03 00:50:00",
                "cmdId": "445300000000",
                "cmdName": "测试地区",
                "dutyDeptNo": "445302010000",
                "dutyDeptName": "测试派出所",
                "newRecvTypeNo": "01",
                "newRecvTypeName": "电话报警",
                "newCharaSubclassNo": "0301",
                "newCharaSubclassName": "赌博",
                "newOriCharaSubclassNo": "0301",
                "newOriCharaSubclassName": "举报赌博",
                "callerName": "张三",
                "callerPhone": "13800138000",
                "occurAddress": "测试地址",
                "caseContents": "有人打麻将扰民",
                "replies": "到场未发现赌博",
                "lngOfCriterion": "112.1",
                "latOfCriterion": "22.1",
            }
        ]
        with patch.object(export_service, "resolve_gambling_topic_tags", return_value=("0301", "赌博")), patch.object(
            export_service,
            "fetch_all_case_list",
            return_value=rows,
        ), patch.object(export_service, "_build_database_sheets", return_value=[]):
            buffer = export_service.generate_gambling_analysis_export(params)

        workbook = load_workbook(io.BytesIO(buffer.getvalue()))
        self.assertIn("01_警情明细", workbook.sheetnames)
        self.assertNotIn("01_警情脱敏明细", workbook.sheetnames)
        headers = [cell.value for cell in workbook["01_警情明细"][1]]
        self.assertIn("报警电话", headers)
        self.assertIn("报警内容", headers)
        self.assertIn("处警情况", headers)
        values = [cell.value for row in workbook["01_警情明细"].iter_rows() for cell in row]
        self.assertIn("JQ001", values)
        self.assertIn("测试派出所", values)
        self.assertIn("13800138000", values)
        self.assertIn("测试地址", values)
        self.assertIn("有人打麻将扰民", values)
        self.assertIn("未脱敏", [cell.value for row in workbook["00_导出说明"].iter_rows() for cell in row])

    def test_generate_analysis_data_export_adds_yoy_sheets(self) -> None:
        params = {
            "beginDate": "2026-01-01 00:00:00",
            "endDate": "2026-03-01 00:00:00",
        }
        current_rows = [
            {
                "caseNo": "JQ001",
                "callTime": "2026-01-03 01:00:00",
                "cmdId": "445300000000",
                "dutyDeptNo": "445302010000",
                "newCharaSubclassNo": "0301",
                "newOriCharaSubclassNo": "0301",
                "caseContents": "群众举报有人打麻将赌博",
                "replies": "到场查获并受理行政案件",
            },
            {
                "caseNo": "JQ002",
                "callTime": "2026-02-03 01:00:00",
                "cmdId": "445300000000",
                "dutyDeptNo": "445302010000",
                "newCharaSubclassNo": "0301",
                "newOriCharaSubclassNo": "0301",
                "caseContents": "群众举报有人打麻将赌博",
                "replies": "到场未发现",
            },
        ]
        previous_rows = [
            {
                "caseNo": "JQ101",
                "callTime": "2025-01-03 01:00:00",
                "cmdId": "445300000000",
                "dutyDeptNo": "445302010000",
                "newCharaSubclassNo": "0301",
                "newOriCharaSubclassNo": "0301",
                "caseContents": "群众举报有人打麻将赌博",
                "replies": "到场查获并受理行政案件",
            }
        ]
        with patch.object(export_service, "resolve_gambling_topic_tags", return_value=("0301", "赌博")), patch.object(
            export_service,
            "fetch_all_case_list",
            side_effect=[current_rows, previous_rows],
        ) as mock_fetch, patch.object(export_service, "_build_database_sheets", return_value=[]):
            buffer = export_service.generate_gambling_analysis_export(params)

        workbook = load_workbook(io.BytesIO(buffer.getvalue()))
        self.assertIn("25_警情同比汇总", workbook.sheetnames)
        self.assertIn("26_月度同比_代码", workbook.sheetnames)
        self.assertIn("27_派出所同比_代码", workbook.sheetnames)
        self.assertEqual(mock_fetch.call_args_list[0].args[0]["beginDate"], "2026-01-01 00:00:00")
        self.assertEqual(mock_fetch.call_args_list[1].args[0]["beginDate"], "2025-01-01 00:00:00")
        self.assertEqual(mock_fetch.call_args_list[0].args[0]["newCharaSubclassNo"], "")
        self.assertEqual(mock_fetch.call_args_list[0].args[0]["newOriCharaSubclassNo"], "0301")
        self.assertEqual(mock_fetch.call_args_list[1].args[0]["newCharaSubclassNo"], "")
        self.assertEqual(mock_fetch.call_args_list[1].args[0]["newOriCharaSubclassNo"], "0301")

        summary = workbook["25_警情同比汇总"]
        headers = [cell.value for cell in summary[1]]
        total_row = {headers[idx]: summary[2][idx].value for idx in range(len(headers))}
        self.assertEqual(total_row["indicator"], "警情总数")
        self.assertEqual(total_row["current_count"], 2)
        self.assertEqual(total_row["previous_count"], 1)
        self.assertEqual(total_row["yoy_rate_pct"], 100)

    def test_build_analysis_export_filename_marks_plain_attachment(self) -> None:
        filename = export_service.build_gambling_analysis_export_filename(
            "2026-04-01 00:00:00",
            "2026-04-02 00:00:00",
            now=real_datetime(2026, 4, 20, 8, 9, 10),
            desensitized=False,
        )
        self.assertEqual(filename, "2026-04-01-2026-04-02赌博分析报告附件20260420080910.xlsx")


class TestGamblingAnalysisExportRoute(unittest.TestCase):
    def setUp(self) -> None:
        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.register_blueprint(jingqing_fenxi_bp, url_prefix="/jingqing_fenxi")
        self.client = app.test_client()

    def test_analysis_data_download_route_returns_workbook(self) -> None:
        with patch(
            "jingqing_fenxi.routes.gambling_topic_routes.generate_gambling_analysis_export",
            return_value=io.BytesIO(b"analysis-export"),
        ) as mock_generate, patch(
            "jingqing_fenxi.routes.gambling_topic_routes.build_gambling_analysis_export_filename",
            return_value="analysis.xlsx",
        ) as mock_filename:
            response = self.client.get(
                "/jingqing_fenxi/download/gambling-topic/analysis-data"
                "?beginDate=2026-04-01%2000:00:00&endDate=2026-04-02%2000:00:00&desensitized=0"
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b"analysis-export")
        self.assertIn("analysis.xlsx", response.headers["Content-Disposition"])
        self.assertEqual(mock_generate.call_args.args[0]["desensitized"], "0")
        self.assertFalse(mock_filename.call_args.kwargs["desensitized"])


if __name__ == "__main__":
    unittest.main()
