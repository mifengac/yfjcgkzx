import io
import unittest
from datetime import datetime
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from jszahzyj.service import jszahz_topic_service


def _build_tag_workbook_bytes() -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "汇总"

    sheet["A2"] = "所属县(市)区"
    sheet["B2"] = "派出所"
    sheet["C2"] = "姓名"
    sheet["D2"] = "人员状态"
    sheet["E2"] = "身份证号码"
    sheet["F2"] = "手机号"
    sheet["G2"] = "服药情况"
    sheet["H2"] = "监护情况"
    sheet["I2"] = "既往有自杀或严重伤人"
    sheet["J2"] = "列为重点关注人员"

    sheet["E4"] = "440123199001011111"
    sheet["G4"] = "不规律服药"
    sheet["H4"] = "弱监护"
    sheet["I4"] = "是"

    sheet["E5"] = "440123199001011111"
    sheet["G5"] = "不规律服药"

    sheet["E6"] = "440123199001011112"
    sheet["H6"] = "无监护"
    sheet["J6"] = "是"

    data = io.BytesIO()
    workbook.save(data)
    data.seek(0)
    return data


def _build_group_header_workbook() -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "汇总"

    sheet["A2"] = "所属县(市)区"
    sheet["B2"] = "派出所"
    sheet["C2"] = "姓名"
    sheet["D2"] = "人员状态"
    sheet["E2"] = "身份证号码"
    sheet["F2"] = "手机号"
    sheet.merge_cells("G2:J2")
    sheet["G2"] = "人员标签"
    sheet["G3"] = "服药情况"
    sheet["H3"] = "监护情况"
    sheet["I3"] = "既往有自杀或严重伤人"
    sheet["J3"] = "列为重点关注人员"

    sheet["E4"] = "440123199001011111"
    sheet["G4"] = "不规律服药"
    sheet["H4"] = "弱监护"
    sheet["I4"] = "是"
    sheet["J4"] = "是"

    data = io.BytesIO()
    workbook.save(data)
    data.seek(0)
    return data


def _build_base_workbook_bytes() -> io.BytesIO:
    workbook = Workbook()
    first = workbook.active
    first.title = "云城"
    sheets = [first]
    for name in ["云安", "罗定", "新兴", "郁南"]:
        sheets.append(workbook.create_sheet(title=name))

    for sheet in sheets:
        sheet["A1"] = "序号"
        sheet["B1"] = "姓名"
        sheet["C1"] = "证件号码"

    sheets[0]["A2"] = "1"
    sheets[0]["B2"] = "张三"
    sheets[0]["C2"] = "440123199001011111"

    sheets[1]["A2"] = "2"
    sheets[1]["B2"] = "李四"
    sheets[1]["C2"] = "440123199001011112"

    sheets[2]["A2"] = "3"
    sheets[2]["B2"] = "重复张三"
    sheets[2]["C2"] = "440123199001011111"

    data = io.BytesIO()
    workbook.save(data)
    data.seek(0)
    return data


class TestJszahzTopicService(unittest.TestCase):
    def test_parse_person_type_workbook_extracts_exact_labels(self) -> None:
        payload = jszahz_topic_service.parse_person_type_workbook(_build_tag_workbook_bytes())

        labels = {(row["zjhm"], row["person_type"]) for row in payload.rows}
        self.assertEqual(payload.imported_row_count, 3)
        self.assertEqual(payload.generated_tag_count, 5)
        self.assertEqual(payload.tagged_person_count, 2)
        self.assertEqual(sorted(payload.all_zjhms), ["440123199001011111", "440123199001011112"])
        self.assertIn(("440123199001011111", "不规律服药"), labels)
        self.assertIn(("440123199001011111", "弱监护"), labels)
        self.assertIn(("440123199001011111", "既往有严重自杀或伤人行为"), labels)
        self.assertIn(("440123199001011112", "无监护"), labels)
        self.assertIn(("440123199001011112", "列为重点关注人员"), labels)

    def test_parse_person_type_workbook_ignores_header_merges(self) -> None:
        payload = jszahz_topic_service.parse_person_type_workbook(_build_group_header_workbook())

        labels = {(row["zjhm"], row["person_type"]) for row in payload.rows}
        self.assertEqual(payload.imported_row_count, 1)
        self.assertEqual(payload.generated_tag_count, 4)
        self.assertEqual(payload.all_zjhms, ["440123199001011111"])
        self.assertIn(("440123199001011111", "不规律服药"), labels)
        self.assertIn(("440123199001011111", "弱监护"), labels)
        self.assertIn(("440123199001011111", "既往有严重自杀或伤人行为"), labels)
        self.assertIn(("440123199001011111", "列为重点关注人员"), labels)

    def test_parse_base_person_workbook_deduplicates_by_id_card(self) -> None:
        payload = jszahz_topic_service.parse_base_person_workbook(_build_base_workbook_bytes())

        self.assertEqual(payload.imported_row_count, 3)
        self.assertEqual(payload.deduplicated_person_count, 2)
        self.assertEqual([row["ssfjdm"] for row in payload.rows], ["445302000000", "445303000000"])
        self.assertEqual(payload.all_zjhms, ["440123199001011111", "440123199001011112"])

    def test_default_time_range_uses_midnight_window(self) -> None:
        start_time, end_time = jszahz_topic_service.default_time_range()
        start_dt = datetime.strptime(start_time, "%Y-%m-%d %H:%M:%S")
        end_dt = datetime.strptime(end_time, "%Y-%m-%d %H:%M:%S")

        self.assertEqual(start_dt, datetime(2025, 1, 1, 0, 0, 0))
        self.assertEqual(end_dt.hour, 0)
        self.assertEqual(end_dt.minute, 0)
        self.assertEqual(end_dt.second, 0)

    def test_query_summary_payload_merges_sources_and_appends_total_row(self) -> None:
        with patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "get_active_batches",
            return_value={"base_batch": {"id": 7}, "tag_batch": {"id": 8}},
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_base_person_rows",
            return_value=[
                {"zjhm": "4401", "xm": "张三", "ssfjdm": "445302000000", "ssfj": "云城分局"},
            ],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_live_person_rows",
            return_value=[
                {"zjhm": "4402", "xm": "李四", "lgsj": "2026-04-01 08:00:00", "lgdw": "云安派出所", "fxdj_label": "1级患者", "ssfjdm": "445303000000", "ssfj": "云安分局"},
            ],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_tag_rows",
            return_value=[
                {"zjhm": "4401", "person_type": "弱监护"},
                {"zjhm": "4403", "person_type": "列为重点关注人员"},
            ],
        ):
            payload = jszahz_topic_service.query_summary_payload(
                branch_codes=[],
                person_types=[],
                risk_labels=[],
                managed_only=True,
            )

        self.assertTrue(payload["success"])
        self.assertEqual(payload["count"], 3)
        self.assertEqual(payload["records"][-1]["分局名称"], "汇总")
        self.assertEqual(payload["records"][-1]["去重患者数"], 3)

    def test_query_summary_payload_returns_live_data_without_uploaded_batches(self) -> None:
        with patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "get_active_batches",
            return_value={"base_batch": None, "tag_batch": None},
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_base_person_rows",
            return_value=[],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_live_person_rows",
            return_value=[
                {"zjhm": "4402", "xm": "李四", "lgsj": "2026-04-01 08:00:00", "lgdw": "云安派出所", "fxdj_label": "1级患者", "ssfjdm": "445303000000", "ssfj": "云安分局"},
            ],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_tag_rows",
            return_value=[],
        ):
            payload = jszahz_topic_service.query_summary_payload(
                branch_codes=[],
                person_types=[],
                risk_labels=[],
                managed_only=True,
            )

        self.assertTrue(payload["success"])
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["records"][0]["分局名称"], "云安分局")

    def test_query_detail_payload_attaches_relation_counts(self) -> None:
        enhanced_records = [
            {
                "姓名": "张三",
                "身份证号": "4401",
                "列管时间": "2026-04-01 08:00:00",
                "列管单位": "云城派出所",
                "分局": "云城分局",
                "人员风险": "无数据",
                "人员类型": "弱监护",
                "关联案件": 0,
                "关联警情": 1,
                "关联机动车": 2,
                "关联视频云": 3,
                "关联门诊": 4,
                "关联飙车炸街": 5,
            }
        ]
        with patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "get_active_batches",
            return_value={"base_batch": {"id": 7}, "tag_batch": {"id": 8}},
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_base_person_rows",
            return_value=[
                {"zjhm": "4401", "xm": "张三", "ssfjdm": "445302000000", "ssfj": "云城分局"},
            ],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_live_person_rows",
            return_value=[],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_tag_rows",
            return_value=[{"zjhm": "4401", "person_type": "弱监护"}],
        ), patch(
            "jszahzyj.service.jszahz_topic_service.attach_relation_counts",
            return_value=enhanced_records,
        ) as mock_attach:
            payload = jszahz_topic_service.query_detail_payload(
                branch_code="445302000000",
                person_types=[],
                risk_labels=[],
                managed_only=True,
            )

        self.assertTrue(payload["success"])
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["records"][0]["关联飙车炸街"], 5)
        mock_attach.assert_called_once()

    def test_query_detail_payload_can_skip_relation_counts(self) -> None:
        initialized_records = [
            {
                "姓名": "张三",
                "身份证号": "4401",
                "列管时间": "",
                "列管单位": "",
                "分局": "云城分局",
                "人员风险": "无数据",
                "人员类型": "弱监护",
                "关联案件": None,
                "关联警情": None,
                "关联机动车": None,
                "关联视频云": None,
                "关联门诊": None,
                "关联飙车炸街": None,
            }
        ]
        with patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "get_active_batches",
            return_value={"base_batch": {"id": 7}, "tag_batch": {"id": 8}},
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_base_person_rows",
            return_value=[
                {"zjhm": "4401", "xm": "张三", "ssfjdm": "445302000000", "ssfj": "云城分局"},
            ],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_live_person_rows",
            return_value=[],
        ), patch.object(
            jszahz_topic_service.jszahz_topic_dao,
            "query_tag_rows",
            return_value=[{"zjhm": "4401", "person_type": "弱监护"}],
        ), patch(
            "jszahzyj.service.jszahz_topic_service.initialize_relation_placeholders",
            return_value=initialized_records,
        ) as mock_initialize, patch(
            "jszahzyj.service.jszahz_topic_service.attach_relation_counts",
        ) as mock_attach:
            payload = jszahz_topic_service.query_detail_payload(
                branch_code="445302000000",
                person_types=[],
                risk_labels=[],
                managed_only=True,
                include_relation_counts=False,
            )

        self.assertTrue(payload["success"])
        self.assertIsNone(payload["records"][0]["关联警情"])
        mock_initialize.assert_called_once()
        mock_attach.assert_not_called()

    def test_query_detail_page_payload_paginates_and_keeps_total_count(self) -> None:
        fake_payload = {
            "success": True,
            "records": [
                {"姓名": f"人员{i}", "身份证号": f"4401{i:02d}", "关联案件": None, "关联警情": None}
                for i in range(21)
            ],
            "count": 21,
            "message": "",
            "filters": {
                "branch_codes": ["445302000000"],
                "person_types": [],
                "risk_labels": [],
                "managed_only": True,
            },
            "active_batches": {"base_batch": None, "tag_batch": None},
        }
        with patch(
            "jszahzyj.service.jszahz_topic_service.query_detail_payload",
            return_value=fake_payload,
        ):
            payload = jszahz_topic_service.query_detail_page_payload(
                branch_code="445302000000",
                person_types=[],
                risk_labels=[],
                managed_only=True,
                relation_types=["case", "clinic"],
                page=2,
                page_size="20",
            )

        self.assertTrue(payload["success"])
        self.assertEqual(payload["count"], 21)
        self.assertEqual(payload["page"], 2)
        self.assertEqual(payload["page_size_value"], "20")
        self.assertEqual(payload["page_record_count"], 1)
        self.assertEqual(payload["selected_relation_types"], ["case", "clinic"])
        self.assertEqual(payload["visible_relation_columns"], ["关联案件", "关联门诊"])

    def test_query_detail_page_payload_supports_all_page_size(self) -> None:
        fake_payload = {
            "success": True,
            "records": [
                {"姓名": "张三", "身份证号": "4401", "关联案件": None},
                {"姓名": "李四", "身份证号": "4402", "关联案件": None},
            ],
            "count": 2,
            "message": "",
            "filters": {
                "branch_codes": [],
                "person_types": [],
                "risk_labels": [],
                "managed_only": False,
            },
            "active_batches": {"base_batch": None, "tag_batch": None},
        }
        with patch(
            "jszahzyj.service.jszahz_topic_service.query_detail_payload",
            return_value=fake_payload,
        ):
            payload = jszahz_topic_service.query_detail_page_payload(
                branch_code="__ALL__",
                person_types=[],
                risk_labels=[],
                managed_only=False,
                relation_types=["case"],
                page=5,
                page_size="all",
            )

        self.assertEqual(payload["page"], 1)
        self.assertEqual(payload["total_pages"], 1)
        self.assertEqual(payload["page_size"], None)
        self.assertEqual(payload["page_record_count"], 2)

    def test_export_detail_xlsx_builds_multi_sheet_workbook(self) -> None:
        fake_payload = {
            "success": True,
            "records": [
                {
                    "姓名": "张三",
                    "身份证号": "4401",
                    "列管时间": "2026-04-01 08:00:00",
                    "列管单位": "云城派出所",
                    "分局": "云城分局",
                    "人员风险": "1级患者",
                    "人员类型": "弱监护",
                    "关联案件": None,
                    "关联门诊": None,
                }
            ],
            "count": 1,
            "message": "",
            "filters": {
                "branch_codes": ["445302000000"],
                "person_types": ["弱监护"],
                "risk_labels": ["1级患者"],
                "managed_only": True,
            },
            "active_batches": {"base_batch": None, "tag_batch": None},
        }
        with patch(
            "jszahzyj.service.jszahz_topic_service.query_detail_payload",
            return_value=fake_payload,
        ), patch(
            "jszahzyj.service.jszahz_topic_detail_page_service.jszahz_topic_relation_dao.query_relation_detail_rows_batch",
            side_effect=[
                [{"身份证号": "4401", "案件编号": "AJ001", "案件类型": "刑事"}],
                [],
            ],
        ):
            data, filename = jszahz_topic_service.export_detail_xlsx(
                branch_code="445302000000",
                branch_name="云城分局",
                person_types=["弱监护"],
                risk_labels=["1级患者"],
                managed_only=True,
                relation_types=["case", "clinic"],
            )

        workbook = load_workbook(io.BytesIO(data))
        self.assertEqual(workbook.sheetnames, ["患者明细", "关联案件", "关联门诊"])
        self.assertEqual(workbook["关联案件"]["A1"].value, "患者姓名")
        self.assertEqual(workbook["关联案件"]["B2"].value, "4401")
        self.assertEqual(workbook["关联门诊"]["A1"].value, "暂无数据")
        self.assertIn("精神患者主题库详情_云城分局_已列管_弱监护_1级患者_案件_门诊_", filename)


if __name__ == "__main__":
    unittest.main()
