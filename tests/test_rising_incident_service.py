import io
import unittest
from datetime import datetime as real_datetime
from unittest.mock import patch

from flask import Flask
from openpyxl import load_workbook

from jingqing_fenxi.routes.jingqing_fenxi_routes import jingqing_fenxi_bp
from jingqing_fenxi.service import rising_incident_service as service


def _case_list_counts(counts):
    period_days = ["2026-04-03", "2026-04-10", "2026-04-17"]
    rows = []
    for idx, count in enumerate(counts):
        for row_idx in range(count):
            rows.append(
                {
                    "caseNo": f"JQ{idx}{row_idx}",
                    "callTime": f"{period_days[idx]} 01:{row_idx:02d}:00",
                    "dutyDeptNo": "445302000000",
                    "dutyDeptName": "测试派出所",
                }
            )
    return {"code": 0, "total": len(rows), "rows": rows}


class TestRisingIncidentService(unittest.TestCase):
    def test_business_week_period_uses_friday_to_thursday(self) -> None:
        periods = service.build_periods(
            real_datetime(2026, 4, 17, 0, 0, 0),
            real_datetime(2026, 4, 24, 0, 0, 0),
            "business_week",
        )

        self.assertEqual(len(periods), 1)
        self.assertEqual(periods[0].label, "2026-04-17至2026-04-23")
        self.assertEqual(periods[0].start, real_datetime(2026, 4, 17, 0, 0, 0))
        self.assertEqual(periods[0].end, real_datetime(2026, 4, 24, 0, 0, 0))

    def test_month_periods_use_natural_months(self) -> None:
        periods = service.build_periods(
            real_datetime(2026, 2, 1, 0, 0, 0),
            real_datetime(2026, 5, 1, 0, 0, 0),
            "month",
        )

        self.assertEqual([period.label for period in periods], ["2026-02", "2026-03", "2026-04"])

    def test_rising_sequence_returns_station(self) -> None:
        with patch.object(service.api_client, "get_case_list", return_value=_case_list_counts([1, 2, 3])):
            result = service.run_rising_incident_analysis(
                {
                    "beginDate": "2026-04-03 00:00:00",
                    "endDate": "2026-04-24 00:00:00",
                    "periodType": "business_week",
                    "minPeriods": "3",
                    "currentOnly": "1",
                }
            )

        self.assertEqual(len(result["rows"]), 1)
        row = result["rows"][0]
        self.assertEqual(row["派出所名称"], "测试派出所")
        self.assertEqual(row["风险等级"], "低风险")
        self.assertEqual(row["当前连续上升周期数"], 3)
        self.assertEqual(row["当前连续上升次数"], 2)
        self.assertEqual(row["趋势序列"], "1 -> 2 -> 3")

    def test_risk_level_marks_fast_rising_station_as_high(self) -> None:
        with patch.object(service.api_client, "get_case_list", return_value=_case_list_counts([1, 8, 20])):
            result = service.run_rising_incident_analysis(
                {
                    "beginDate": "2026-04-03 00:00:00",
                    "endDate": "2026-04-24 00:00:00",
                    "periodType": "business_week",
                    "minPeriods": "3",
                    "currentOnly": "1",
                }
            )

        self.assertEqual(result["rows"][0]["风险等级"], "高风险")
        self.assertEqual(result["rows"][0]["增量"], 12)

    def test_case_type_filter_is_forwarded_to_case_payload(self) -> None:
        with patch.object(
            service.api_client,
            "get_case_list",
            return_value=_case_list_counts([1, 2, 3]),
        ) as mock_get:
            result = service.run_rising_incident_analysis(
                {
                    "beginDate": "2026-04-03 00:00:00",
                    "endDate": "2026-04-24 00:00:00",
                    "periodType": "business_week",
                    "minPeriods": "3",
                    "currentOnly": "1",
                    "caseTypeSource": "nature",
                    "newOriCharaSubclassNo": "0101,0102",
                    "newOriCharaSubclass": "赌博,聚众赌博",
                }
            )

        payload = mock_get.call_args.args[0]
        self.assertEqual(payload["newOriCharaSubclassNo"], "0101,0102")
        self.assertEqual(payload["newOriCharaSubclass"], "赌博,聚众赌博")
        self.assertEqual(payload["beginDate"], "2026-04-03 00:00:00")
        self.assertEqual(payload["endDate"], "2026-04-23 23:59:59")
        self.assertEqual(payload["pageNum"], "1")
        self.assertEqual(result["meta"]["chara"], "赌博,聚众赌博")

    def test_current_only_filters_when_latest_period_falls(self) -> None:
        with patch.object(service.api_client, "get_case_list", return_value=_case_list_counts([1, 3, 2])):
            result = service.run_rising_incident_analysis(
                {
                    "beginDate": "2026-04-03 00:00:00",
                    "endDate": "2026-04-24 00:00:00",
                    "periodType": "business_week",
                    "minPeriods": "3",
                    "currentOnly": "1",
                }
            )

        self.assertEqual(result["rows"], [])

    def test_missing_period_counts_as_zero(self) -> None:
        with patch.object(service.api_client, "get_case_list", return_value=_case_list_counts([0, 1, 2])):
            result = service.run_rising_incident_analysis(
                {
                    "beginDate": "2026-04-03 00:00:00",
                    "endDate": "2026-04-24 00:00:00",
                    "periodType": "business_week",
                    "minPeriods": "3",
                    "currentOnly": "1",
                }
            )

        self.assertEqual(result["rows"][0]["趋势序列"], "0 -> 1 -> 2")
        self.assertEqual(len(result["periodDetails"]), 3)

    def test_generate_excel_has_summary_and_period_detail_sheets(self) -> None:
        result = {
            "rows": [
                {
                    "派出所名称": "测试派出所",
                    "派出所代码": "445302000000",
                    "风险等级": "低风险",
                    "周期类型": "业务周",
                    "最新周期": "2026-04-17至2026-04-23",
                    "上期数量": 2,
                    "最新数量": 3,
                    "增量": 1,
                    "当前连续上升周期数": 3,
                    "当前连续上升次数": 2,
                    "趋势序列": "1 -> 2 -> 3",
                    "涉及周期范围": "2026-04-03至2026-04-09 至 2026-04-17至2026-04-23",
                }
            ],
            "periodDetails": [{"派出所名称": "测试派出所", "周期": "2026-04-17至2026-04-23", "警情数": 3}],
        }

        buffer = service.generate_rising_incident_excel(result)
        workbook = load_workbook(io.BytesIO(buffer.getvalue()))

        self.assertEqual(workbook.sheetnames, ["升势预警", "周期明细"])
        self.assertEqual(workbook["升势预警"]["C1"].value, "风险等级")
        self.assertEqual(workbook["升势预警"]["A2"].value, "测试派出所")
        self.assertEqual(workbook["升势预警"]["C2"].value, "低风险")
        self.assertEqual(workbook["周期明细"]["A2"].value, "测试派出所")


class TestRisingIncidentRoutes(unittest.TestCase):
    def setUp(self) -> None:
        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.register_blueprint(jingqing_fenxi_bp, url_prefix="/jingqing_fenxi")
        self.client = app.test_client()

    def test_analyze_route_returns_service_result(self) -> None:
        with patch(
            "jingqing_fenxi.routes.rising_incident_routes.run_rising_incident_analysis",
            return_value={"rows": [{"派出所名称": "测试派出所"}], "meta": {}, "periods": []},
        ) as mock_run:
            response = self.client.post(
                "/jingqing_fenxi/api/rising-incident/analyze",
                data={
                    "beginDate": "2026-04-03 00:00:00",
                    "endDate": "2026-04-24 00:00:00",
                    "caseTypeSource": "nature",
                    "caseTypeIds[]": ["1", "2"],
                    "newOriCharaSubclassNo": "0101,0102",
                    "newOriCharaSubclass": "赌博,聚众赌博",
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["code"], 0)
        self.assertEqual(mock_run.call_args.args[0]["periodType"], "business_week")
        self.assertEqual(mock_run.call_args.args[0]["caseTypeIds"], ["1", "2"])
        self.assertEqual(mock_run.call_args.args[0]["newOriCharaSubclassNo"], "0101,0102")

    def test_analyze_route_returns_readable_error(self) -> None:
        with patch(
            "jingqing_fenxi.routes.rising_incident_routes.run_rising_incident_analysis",
            side_effect=ValueError("统计周期数量少于连续上升阈值"),
        ):
            response = self.client.post("/jingqing_fenxi/api/rising-incident/analyze")

        self.assertEqual(response.status_code, 400)
        self.assertIn("统计周期数量", response.get_json()["message"])

    def test_download_route_returns_workbook(self) -> None:
        with patch(
            "jingqing_fenxi.routes.rising_incident_routes.run_rising_incident_analysis",
            return_value={"rows": [], "periodDetails": [], "meta": {"beginDate": "2026-04-03", "endDate": "2026-04-24"}},
        ), patch(
            "jingqing_fenxi.routes.rising_incident_routes.generate_rising_incident_excel",
            return_value=io.BytesIO(b"test-export"),
        ), patch(
            "jingqing_fenxi.routes.rising_incident_routes.build_export_filename",
            return_value="rising.xlsx",
        ):
            response = self.client.get(
                "/jingqing_fenxi/download/rising-incident?beginDate=2026-04-03 00:00:00&endDate=2026-04-24 00:00:00"
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b"test-export")
        self.assertIn("rising.xlsx", response.headers["Content-Disposition"])


if __name__ == "__main__":
    unittest.main()
