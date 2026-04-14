import io
import unittest
from unittest.mock import patch

from flask import Flask
from openpyxl import load_workbook

from jszahzyj.service import jszahzyj_service


class TestJszahzyjService(unittest.TestCase):
    def setUp(self) -> None:
        self.app = Flask(__name__)

    def test_get_jszahzyj_data_passes_through(self) -> None:
        expected_rows = [{"姓名": "张三"}]
        with patch.object(
            jszahzyj_service,
            "query_jszahzyj_data",
            return_value=(expected_rows, 3),
        ) as mock_query:
            rows, total = jszahzyj_service.get_jszahzyj_data(page=2, page_size=50)

        self.assertEqual(rows, expected_rows)
        self.assertEqual(total, 3)
        self.assertEqual(mock_query.call_args.kwargs["page"], 2)
        self.assertEqual(mock_query.call_args.kwargs["page_size"], 50)

    def test_export_to_csv_returns_bom_content(self) -> None:
        rows = [{"姓名": "张三", "证件号码": "440123199001011111"}]
        with self.app.test_request_context(), patch.object(
            jszahzyj_service,
            "get_all_jszahzyj_data",
            return_value=rows,
        ):
            response = jszahzyj_service.export_to_csv()
            response.direct_passthrough = False
            body = response.get_data()

        self.assertTrue(body.startswith(b"\xef\xbb\xbf"))
        text = body.decode("utf-8-sig")
        self.assertIn("姓名,证件号码", text)
        self.assertIn("张三,440123199001011111", text)
        self.assertIn(".csv", response.headers.get("Content-Disposition", ""))

    def test_export_to_xlsx_serializes_complex_values(self) -> None:
        rows = [{"姓名": "张三", "扩展": {"风险": 1}, "标签": ["弱监护", "无监护"]}]
        with self.app.test_request_context(), patch.object(
            jszahzyj_service,
            "get_all_jszahzyj_data",
            return_value=rows,
        ):
            response = jszahzyj_service.export_to_xlsx()
            response.direct_passthrough = False
            body = response.get_data()

        workbook = load_workbook(io.BytesIO(body))
        sheet = workbook.active
        self.assertEqual(sheet.title, "精神障碍患者预警")
        self.assertEqual(sheet.cell(2, 1).value, "张三")
        self.assertEqual(sheet.cell(2, 2).value, '{"风险": 1}')
        self.assertEqual(sheet.cell(2, 3).value, '["弱监护", "无监护"]')
        self.assertIn(".xlsx", response.headers.get("Content-Disposition", ""))


if __name__ == "__main__":
    unittest.main()