import unittest
from datetime import datetime as real_datetime
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from jingqing_fenxi.service import fight_topic_service as service


class _FrozenDateTime(real_datetime):
    @classmethod
    def now(cls, tz=None):  # type: ignore[override]
        return cls(2026, 4, 3, 9, 30, 0)


class TestFightTopicService(unittest.TestCase):
    def test_default_time_range_uses_recent_7_days_midnight(self) -> None:
        with patch.object(service, "datetime", _FrozenDateTime):
            start_time, end_time = service.default_time_range()

        self.assertEqual(start_time, "2026-03-27 00:00:00")
        self.assertEqual(end_time, "2026-04-03 00:00:00")

    def test_resolve_fight_topic_tags_only_uses_target_parent_and_dedupes(self) -> None:
        tree_nodes = [
            {"id": "1", "pId": "OTHER", "name": "其他", "tag": "999"},
            {"id": "2", "pId": service.FIGHT_TOPIC_PARENT_ID, "name": "结伙斗殴", "tag": "02010801"},
            {"id": "3", "pId": service.FIGHT_TOPIC_PARENT_ID, "name": "结伙斗殴", "tag": "02010801"},
            {"id": "4", "pId": service.FIGHT_TOPIC_PARENT_ID, "name": "故意伤害", "tag": "01030300"},
            {"id": "5", "pId": service.FIGHT_TOPIC_PARENT_ID, "name": "空标签", "tag": ""},
        ]

        tag_csv, name_csv = service.resolve_fight_topic_tags(tree_nodes)

        self.assertEqual(tag_csv, "02010801,01030300")
        self.assertEqual(name_csv, "结伙斗殴,故意伤害")

    def test_classify_reason_uses_first_matching_rule(self) -> None:
        label, keyword = service.classify_reason("酒后因夫妻感情问题发生争吵后打架")
        self.assertEqual(label, "酒后冲突")
        self.assertEqual(keyword, "酒后")

        other_label, other_keyword = service.classify_reason("现场多人扭打，暂无法判断原因")
        self.assertEqual(other_label, "其他原因")
        self.assertEqual(other_keyword, "")

    def test_build_case_payload_uses_new_chara_subclass_no_and_default_name(self) -> None:
        payload = service._build_case_payload(
            "2026-03-27 00:00:00",
            "2026-04-03 00:00:00",
            "02010801,01030300",
        )

        self.assertEqual(payload["newCharaSubclassNo"], "02010801,01030300")
        self.assertEqual(payload["newCharaSubclass"], "全部")
        self.assertEqual(payload["newOriCharaSubclassNo"], "")

    def test_run_analysis_only_returns_selected_dimensions(self) -> None:
        rows = [
            {
                "callTime": "2026-04-01 01:00:00",
                "caseContents": "酒后发生冲突",
                "occurAddress": "测试地址A",
                "callerPhone": "13800138000",
                "dutyDeptName": "测试派出所",
            }
        ]
        params = {
            "beginDate": "2026-03-27 00:00:00",
            "endDate": "2026-04-03 00:00:00",
            "m2mStartTime": "2026-03-20 00:00:00",
            "m2mEndTime": "2026-03-27 00:00:00",
        }

        with patch.object(service, "resolve_fight_topic_tags", return_value=("02010801,01030300", "结伙斗殴,故意伤害")), patch.object(
            service,
            "fetch_all_case_list",
            return_value=rows,
        ) as mock_fetch:
            results, analysis_base, all_data, _options, _meta = service.run_fight_topic_analysis(
                params,
                ["time", "reason"],
            )

        self.assertIn("time", results)
        self.assertIn("reason", results)
        self.assertNotIn("dept", results)
        self.assertNotIn("srr", results)
        self.assertIn("timeHourly", analysis_base)
        self.assertEqual(all_data[0]["fightReasonLabel"], "酒后冲突")
        self.assertEqual(mock_fetch.call_args.kwargs["max_page_size"], service.FIGHT_TOPIC_UPSTREAM_PAGE_SIZE)

    def test_address_model_error_only_affects_address_dimension(self) -> None:
        rows = [
            {
                "callTime": "2026-04-01 01:00:00",
                "caseContents": "酒后发生冲突",
                "occurAddress": "测试地址A",
                "callerPhone": "13800138000",
                "dutyDeptName": "测试派出所",
            }
        ]
        params = {
            "beginDate": "2026-03-27 00:00:00",
            "endDate": "2026-04-03 00:00:00",
            "m2mStartTime": "2026-03-20 00:00:00",
            "m2mEndTime": "2026-03-27 00:00:00",
        }

        with patch.object(service, "resolve_fight_topic_tags", return_value=("02010801", "结伙斗殴")), patch.object(
            service,
            "fetch_all_case_list",
            return_value=rows,
        ), patch.object(
            service,
            "_predict_address_labels",
            side_effect=RuntimeError("model down"),
        ):
            results, _analysis_base, all_data, _options, _meta = service.run_fight_topic_analysis(
                params,
                ["time", "addr"],
            )

        self.assertIn("time", results)
        self.assertEqual(results["addr"], [])
        self.assertEqual(results["addr_error"]["message"], "model down")
        self.assertEqual(all_data[0]["fightAddrLabel"], "")
        self.assertEqual(all_data[0]["fightAddrProb"], "0.00000")

    def test_generate_excel_uses_single_sheet_and_appends_detail_columns(self) -> None:
        analysis_results = {
            "time": [("0-3时", 5)],
            "addr": [("街面与公共区域", 3)],
            "reason": [("酒后冲突", 2)],
        }
        all_data = [
            {
                "caseNo": "A001",
                "callTime": "2026-04-01 01:00:00",
                "cmdId": "445300000000",
                "dutyDeptName": "测试派出所",
                "caseLevelName": "一级",
                "occurAddress": "测试地址A",
                "callerName": "张三",
                "callerPhone": "13800138000",
                "caseContents": "酒后发生冲突",
                "replies": "已处理",
                "fightAddrLabel": "街面与公共区域",
                "fightAddrProb": "0.99000",
                "fightReasonLabel": "酒后冲突",
                "fightReasonKeyword": "酒后",
            }
        ]

        buffer = service.generate_fight_topic_excel(
            analysis_results,
            all_data,
            ["time", "addr", "reason"],
            begin_date="2026-03-27 00:00:00",
            end_date="2026-04-03 00:00:00",
            analysis_options={
                "timeBucketHours": 3,
                "repeatPhoneMinCount": 2,
                "repeatAddrRadiusMeters": 50,
            },
        )
        workbook = load_workbook(BytesIO(buffer.getvalue()))

        self.assertEqual(len(workbook.sheetnames), 1)
        sheet = workbook.active
        values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None]
        self.assertIn("打架斗殴警情分析", values)
        self.assertIn("详细数据", values)
        self.assertIn("地址分类结果", values)
        self.assertIn("打架原因分类", values)
        self.assertIn("酒后", values)

    def test_build_export_filename_matches_expected_format(self) -> None:
        filename = service.build_export_filename(
            "2026-03-27 00:00:00",
            "2026-04-03 00:00:00",
            now=real_datetime(2026, 4, 3, 10, 11, 12),
        )
        self.assertEqual(filename, "2026-03-27-2026-04-03打架斗殴警情分析20260403101112.xlsx")


if __name__ == "__main__":
    unittest.main()
