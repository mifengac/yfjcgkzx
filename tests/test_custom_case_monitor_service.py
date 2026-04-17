import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from jingqing_fenxi.service import custom_case_monitor_service as monitor_service
from jingqing_fenxi.service import special_case_tab_service as special_service


class TestCustomCaseMonitorService(unittest.TestCase):
    def tearDown(self) -> None:
        with monitor_service._STATUS_LOCK:
            monitor_service._QUERY_JOB_STATUS.clear()

    def test_validate_rules_accepts_multiline_values(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "流浪\n乞讨",
                }
            ]
        )

        self.assertEqual(rules[0]["rule_values"], ["流浪", "乞讨"])
        self.assertTrue(rules[0]["is_enabled"])

    def test_validate_rules_accepts_regex_values_with_quantifier_commas(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "regex_any",
                    "rule_values": "(?:持|拿).{0,6}(?:刀|械)\n(?:手持|携带).{0,4}(?:钢管|铁棍)",
                }
            ]
        )

        self.assertEqual(
            rules[0]["rule_values"],
            ["(?:持|拿).{0,6}(?:刀|械)", "(?:手持|携带).{0,4}(?:钢管|铁棍)"],
        )

    def test_validate_rules_rejects_empty_values(self) -> None:
        with self.assertRaises(ValueError):
            special_service.validate_scheme_rules(
                [
                    {
                        "field_name": "combined_text",
                        "operator": "contains_any",
                        "rule_values": "",
                    }
                ]
            )

    def test_validate_rules_rejects_invalid_regex(self) -> None:
        with self.assertRaisesRegex(ValueError, "正则表达式不合法"):
            special_service.validate_scheme_rules(
                [
                    {
                        "field_name": "combined_text",
                        "operator": "regex_any",
                        "rule_values": "(持刀",
                    }
                ]
            )

    def test_validate_rules_preserves_group_no(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "knife",
                    "group_no": "2",
                }
            ]
        )

        self.assertEqual(rules[0]["group_no"], 2)

    def test_filter_rows_by_rules_uses_or_between_groups(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "knife",
                    "group_no": 1,
                },
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "conflict,harm",
                    "group_no": 1,
                },
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "weapon",
                    "group_no": 2,
                },
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "theft,steal",
                    "group_no": 2,
                },
            ]
        )
        rows = [
            {"caseNo": "A001", "caseContents": "knife conflict", "replies": ""},
            {"caseNo": "A002", "caseContents": "weapon theft", "replies": ""},
            {"caseNo": "A003", "caseContents": "knife patrol", "replies": ""},
            {"caseNo": "A004", "caseContents": "weapon conflict", "replies": ""},
        ]

        matched = special_service.filter_rows_by_rules(rows, rules)

        self.assertEqual([row["caseNo"] for row in matched], ["A001", "A002"])

    def test_filter_rows_by_regex_rules_supports_variable_middle_text(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "regex_any",
                    "rule_values": "(?:持|拿|携带|手持).{0,6}(?:刀|匕首|砍刀|械|钢管)",
                    "group_no": 1,
                }
            ]
        )
        rows = [
            {"caseNo": "A001", "caseContents": "现场有人持砍刀伤人", "replies": ""},
            {"caseNo": "A002", "caseContents": "对方拿着菜刀威胁", "replies": ""},
            {"caseNo": "A003", "caseContents": "", "replies": "反馈称手持钢管追赶他人"},
            {"caseNo": "A004", "caseContents": "普通纠纷", "replies": ""},
        ]

        matched = special_service.filter_rows_by_rules(rows, rules)

        self.assertEqual([row["caseNo"] for row in matched], ["A001", "A002", "A003"])

    def test_filter_rows_by_rules_reports_progress(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "持刀,持械",
                    "group_no": 1,
                }
            ]
        )
        events = []

        special_service.filter_rows_by_rules(
            [
                {"caseNo": "A001", "caseContents": "持刀警情", "replies": ""},
                {"caseNo": "A002", "caseContents": "普通警情", "replies": ""},
                {"caseNo": "A003", "caseContents": "", "replies": "反馈提到持械"},
            ],
            rules,
            progress_callback=events.append,
            progress_step=2,
        )

        self.assertGreaterEqual(len(events), 2)
        self.assertEqual(events[-1]["stage"], "rule_filtering")
        self.assertEqual(events[-1]["stats"]["rule_scanned_count"], 3)
        self.assertEqual(events[-1]["stats"]["rule_match_count"], 2)

    def test_collect_rule_hit_keywords_only_from_matched_groups(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "knife",
                    "group_no": 1,
                },
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "conflict",
                    "group_no": 1,
                },
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "weapon",
                    "group_no": 2,
                },
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "theft",
                    "group_no": 2,
                },
            ]
        )

        keywords = special_service.collect_rule_hit_keywords(
            {"caseContents": "knife conflict", "replies": "weapon seen nearby"},
            rules,
        )

        self.assertEqual(keywords, ["knife", "conflict"])

    def test_collect_rule_hit_keyword_details_marks_actual_source_field(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": "持刀,持械",
                    "group_no": 1,
                }
            ]
        )

        details = special_service.collect_rule_hit_keyword_details(
            {"caseContents": "报警内容提到持刀", "replies": "反馈内容提到持械"},
            rules,
        )

        self.assertEqual(details, ["报警内容→持刀", "反馈内容→持械"])

    def test_collect_rule_hit_keyword_details_uses_actual_regex_match_fragment(self) -> None:
        rules = special_service.validate_scheme_rules(
            [
                {
                    "field_name": "combined_text",
                    "operator": "regex_all",
                    "rule_values": "(?:持|拿).{0,6}(?:刀|械)\n(?:伤人|打架|威胁)",
                    "group_no": 1,
                }
            ]
        )

        details = special_service.collect_rule_hit_keyword_details(
            {"caseContents": "报警称有人持砍刀伤人", "replies": ""},
            rules,
        )

        self.assertEqual(details, ["报警内容→持砍刀", "报警内容→伤人"])

    def test_combined_text_matches_when_case_contents_only_hits(self) -> None:
        rule = {
            "field_name": "combined_text",
            "operator": "contains_any",
            "rule_values": ["持刀"],
            "is_enabled": True,
        }

        matched = special_service.filter_rows_by_rules(
            [
                {"caseNo": "A001", "caseContents": "报警人称现场有人持刀", "replies": ""},
                {"caseNo": "A002", "caseContents": "普通纠纷", "replies": ""},
            ],
            [rule],
        )

        self.assertEqual([row["caseNo"] for row in matched], ["A001"])

    def test_query_custom_case_monitor_records_matches_seeded_logic(self) -> None:
        rows = [
            {
                "caseNo": "A001",
                "cmdId": "445300000000",
                "caseContents": "发现流浪人员求助",
                "replies": "",
                "callTime": "2026-04-05 08:00:00",
                "dutyDeptName": "测试派出所",
            },
            {
                "caseNo": "A002",
                "cmdId": "445300000000",
                "caseContents": "普通警情",
                "replies": "现场涉及乞讨人员聚集",
                "callTime": "2026-04-05 09:00:00",
                "dutyDeptName": "测试派出所",
            },
            {
                "caseNo": "A003",
                "cmdId": "445300000000",
                "caseContents": "普通纠纷",
                "replies": "",
                "callTime": "2026-04-05 10:00:00",
                "dutyDeptName": "测试派出所",
            },
        ]
        scheme = {
            "id": 1,
            "scheme_name": "流浪/乞讨警情",
            "is_enabled": True,
            "rules": [
                {
                    "field_name": "combined_text",
                    "operator": "contains_any",
                    "rule_values": ["流浪", "乞讨"],
                    "is_enabled": True,
                }
            ],
        }

        with patch.object(monitor_service.dao, "get_scheme_by_id", return_value=scheme), patch.object(
            special_service,
            "fetch_all_case_list",
            return_value=rows,
        ):
            result = monitor_service.query_custom_case_monitor_records(
                scheme_id=1,
                start_time="2026-04-05 00:00:00",
                end_time="2026-04-05 23:59:59",
                branches=[],
                page_num=1,
                page_size=15,
            )

        self.assertTrue(result["success"])
        self.assertEqual(result["total"], 2)
        self.assertEqual(result["scheme_name"], "流浪/乞讨警情")
        self.assertEqual(result["rows"][0]["hitKeywordDetails"], "报警内容→流浪")
        self.assertEqual(result["rows"][1]["hitKeywordDetails"], "反馈内容→乞讨")

    def test_query_rejects_disabled_scheme(self) -> None:
        with patch.object(monitor_service.dao, "get_scheme_by_id", return_value={"id": 1, "is_enabled": False}):
            with self.assertRaises(ValueError):
                monitor_service.query_custom_case_monitor_records(
                    scheme_id=1,
                    start_time="2026-04-05 00:00:00",
                    end_time="2026-04-05 23:59:59",
                    branches=[],
                    page_num=1,
                    page_size=15,
                )

    def test_start_query_job_stores_success_result(self) -> None:
        scheme = {"id": 1, "scheme_name": "涉刀警情", "is_enabled": True, "rules": []}
        final_result = {
            "success": True,
            "scheme_name": "涉刀警情",
            "page_num": 1,
            "page_size": 15,
            "total": 2,
            "rows": [{"caseNo": "A001"}, {"caseNo": "A002"}],
        }

        class _ImmediateThread:
            def __init__(self, *, target, kwargs=None, daemon=None):
                self._target = target
                self._kwargs = kwargs or {}

            def start(self):
                self._target(**self._kwargs)

        def _fake_query(**kwargs):
            kwargs["progress_callback"](
                {
                    "stage": "rule_filtering",
                    "message": "规则过滤中：已扫描 5 条，命中 3 条",
                    "stats": {
                        "upstream_row_count": 5,
                        "rule_scanned_count": 5,
                        "rule_match_count": 3,
                        "branch_scanned_count": 0,
                        "branch_filtered_count": 0,
                    },
                }
            )
            kwargs["progress_callback"](
                {
                    "stage": "branch_filtering",
                    "message": "分局过滤中：已扫描 3 条，保留 2 条",
                    "stats": {
                        "upstream_row_count": 5,
                        "rule_scanned_count": 5,
                        "rule_match_count": 3,
                        "branch_scanned_count": 3,
                        "branch_filtered_count": 2,
                    },
                }
            )
            return final_result

        with patch.object(monitor_service, "_load_enabled_scheme_or_raise", return_value=scheme), patch.object(
            monitor_service,
            "_query_records_with_scheme",
            side_effect=_fake_query,
        ), patch.object(monitor_service.threading, "Thread", _ImmediateThread):
            job_id = monitor_service.start_query_custom_case_monitor_job(
                username="tester",
                scheme_id=1,
                start_time="2026-04-05 00:00:00",
                end_time="2026-04-05 23:59:59",
                branches=["云城"],
                page_num=1,
                page_size=15,
            )

        status = monitor_service.get_query_custom_case_monitor_job_status(username="tester", job_id=job_id)
        self.assertIsNotNone(status)
        self.assertEqual(status["state"], "success")
        self.assertEqual(status["stage"], "done")
        self.assertEqual(status["stats"]["branch_filtered_count"], 2)
        self.assertEqual(status["result"]["total"], 2)

    def test_start_query_job_marks_failure_and_scopes_username(self) -> None:
        scheme = {"id": 1, "scheme_name": "涉刀警情", "is_enabled": True, "rules": []}

        class _ImmediateThread:
            def __init__(self, *, target, kwargs=None, daemon=None):
                self._target = target
                self._kwargs = kwargs or {}

            def start(self):
                self._target(**self._kwargs)

        with patch.object(monitor_service, "_load_enabled_scheme_or_raise", return_value=scheme), patch.object(
            monitor_service,
            "_query_records_with_scheme",
            side_effect=RuntimeError("上游查询失败"),
        ), patch.object(monitor_service.threading, "Thread", _ImmediateThread):
            job_id = monitor_service.start_query_custom_case_monitor_job(
                username="tester",
                scheme_id=1,
                start_time="2026-04-05 00:00:00",
                end_time="2026-04-05 23:59:59",
                branches=[],
                page_num=1,
                page_size=15,
            )

        self.assertIsNone(monitor_service.get_query_custom_case_monitor_job_status(username="other", job_id=job_id))
        status = monitor_service.get_query_custom_case_monitor_job_status(username="tester", job_id=job_id)
        self.assertEqual(status["state"], "failed")
        self.assertEqual(status["message"], "上游查询失败")

    def test_export_xlsx_adds_hit_keywords_and_safe_filename(self) -> None:
        rules = [
            {
                "field_name": "combined_text",
                "operator": "contains_any",
                "rule_values": ["坟地", "林地", "纠纷"],
                "is_enabled": True,
            }
        ]
        rows = [
            {
                "caseNo": "A001",
                "callTime": "2026-04-05 08:00:00",
                "cmdId": "445300000000",
                "dutyDeptName": "测试派出所",
                "caseLevelName": "一级",
                "occurAddress": "测试地址",
                "callerName": "张三",
                "callerPhone": "13800138000",
                "caseContents": "坟地纠纷",
                "replies": "林地争执",
            }
        ]

        with patch.object(monitor_service.dao, "get_scheme_by_id", return_value={
            "id": 1,
            "scheme_name": "流浪/乞讨警情",
            "is_enabled": True,
            "rules": rules,
        }), patch.object(special_service, "fetch_all_case_list", return_value=rows):
            buffer, mimetype, filename = monitor_service.export_custom_case_monitor_records(
                export_format="xlsx",
                scheme_id=1,
                start_time="2026-04-05 00:00:00",
                end_time="2026-04-05 23:59:59",
                branches=[],
            )

        workbook = load_workbook(BytesIO(buffer.getvalue()))
        worksheet = workbook.active
        headers = [cell.value for cell in worksheet[1]]

        self.assertEqual(mimetype, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertEqual(filename, "流浪_乞讨警情_2026-04-05_2026-04-05.xlsx")
        self.assertEqual(worksheet.title, "流浪_乞讨警情")
        self.assertIn("命中关键字", headers)
        self.assertEqual(worksheet.cell(row=2, column=len(headers)).value, "坟地、林地、纠纷")

    def test_export_csv_keeps_columns_and_uses_safe_filename(self) -> None:
        rules = [
            {
                "field_name": "combined_text",
                "operator": "contains_any",
                "rule_values": ["坟地", "林地", "纠纷"],
                "is_enabled": True,
            }
        ]
        rows = [
            {
                "caseNo": "A001",
                "callTime": "2026-04-05 08:00:00",
                "cmdId": "445300000000",
                "dutyDeptName": "测试派出所",
                "caseLevelName": "一级",
                "occurAddress": "测试地址",
                "callerName": "张三",
                "callerPhone": "13800138000",
                "caseContents": "坟地纠纷",
                "replies": "林地争执",
            }
        ]

        with patch.object(monitor_service.dao, "get_scheme_by_id", return_value={
            "id": 1,
            "scheme_name": "流浪/乞讨警情",
            "is_enabled": True,
            "rules": rules,
        }), patch.object(special_service, "fetch_all_case_list", return_value=rows):
            buffer, mimetype, filename = monitor_service.export_custom_case_monitor_records(
                export_format="csv",
                scheme_id=1,
                start_time="2026-04-05 00:00:00",
                end_time="2026-04-05 23:59:59",
                branches=[],
            )

        csv_text = buffer.getvalue().decode("utf-8-sig")
        header_line = csv_text.splitlines()[0]

        self.assertEqual(mimetype, "text/csv; charset=utf-8")
        self.assertEqual(filename, "流浪_乞讨警情_2026-04-05_2026-04-05.csv")
        self.assertNotIn("命中关键字", header_line)


if __name__ == "__main__":
    unittest.main()
