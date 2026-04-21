import io
import unittest
from unittest.mock import DEFAULT, patch

from flask import Flask
from openpyxl import load_workbook

from mdjfxsyj.service import mdjfxsyj_mdjfjqfx_export as export_service
from mdjfxsyj.service import mdjfxsyj_mdjfjqfx_service as service


RAW_ROWS = [
    {
        "caseNo": "J1",
        "callTime": "2026-04-01 08:00:00",
        "callerPhone": "138-0000-0001",
        "cmdId": "B1",
        "cmdName": "上游分局1",
        "dutyDeptNo": "S1",
        "dutyDeptName": "上游所1",
        "newOriCharaSubclass": "0801",
        "newOriCharaSubclassName": "家庭纠纷",
        "newCharaSubclass": "0802",
        "newCharaSubclassName": "邻里纠纷",
        "occurAddress": "地址1",
        "caseContents": "内容1",
        "replies": "处警1",
    },
    {
        "caseno": "J2",
        "calltime": "2026-04-01 09:00:00",
        "callerphone": "13800000001",
        "cmdid": "B1",
        "dutydeptno": "S1",
        "neworicharasubclass": "0801",
        "newcharasubclass": "0901",
    },
    {
        "caseNo": "J3",
        "callTime": "2026-04-01 10:00:00",
        "callerPhone": "gffdsagsa",
        "cmdId": "B2",
        "dutyDeptNo": "S2",
        "newOriCharaSubclass": "0803",
        "newCharaSubclass": "0803",
    },
]


class _FakeNatureTreeResponse:
    status_code = 200

    def json(self):
        return [{"id": "08"}, {"id": "0801", "pId": "08"}]


class _LegacyApiClient:
    def __init__(self):
        self.calls = []

    def request_with_retry(self, *args, **kwargs):
        self.calls.append((args, kwargs))
        return _FakeNatureTreeResponse()


class TestMdjfjqfxService(unittest.TestCase):
    def _patch_sources(self):
        return patch.multiple(
            service,
            fetch_raw_mdj_cases=DEFAULT,
            query_org_mappings=DEFAULT,
            query_case_conversion_map=DEFAULT,
        )

    def test_get_mdj_category_code_csv_includes_08_and_direct_children(self):
        csv_text = service.get_mdj_category_code_csv(
            [
                {"id": "08"},
                {"id": "0801", "pId": "08"},
                {"id": "080101", "pId": "0801"},
                {"id": "09", "pId": ""},
            ]
        )

        self.assertEqual(csv_text, "08,0801")

    def test_get_mdj_category_code_csv_falls_back_to_documented_endpoint(self):
        legacy_client = _LegacyApiClient()

        with patch.object(service, "api_client", legacy_client):
            csv_text = service.get_mdj_category_code_csv()

        self.assertEqual(csv_text, "08,0801")
        self.assertEqual(
            legacy_client.calls,
            [(("GET", "/dsjfx/nature/treeNewViewData"), {"timeout": 15})],
        )

    def test_clean_phone_filters_obvious_invalid_values(self):
        self.assertEqual(service.clean_phone("138-0000-0001"), "13800000001")
        self.assertEqual(service.clean_phone("00000"), "")
        self.assertEqual(service.clean_phone("gffdsagsa"), "")
        self.assertEqual(service.clean_phone("11111111111"), "")

    def test_summary_counts_conversion_repeat_and_fine_categories(self):
        with self._patch_sources() as mocks:
            mocks["fetch_raw_mdj_cases"].return_value = RAW_ROWS
            mocks["query_org_mappings"].return_value = (
                {"B1": "一分局", "B2": "二分局"},
                {"S1": "一所", "S2": "二所"},
            )
            mocks["query_case_conversion_map"].return_value = {"J1": ["A1"]}

            payload = service.get_summary_payload(
                start_time="2026-04-01 00:00:00",
                end_time="2026-04-02 00:00:00",
                ssfjdm_list=[],
                group_by="fenju",
                repeat_min=2,
            )

        branch_one = next(row for row in payload["overall"] if row["group_code"] == "B1")
        self.assertEqual(branch_one["原始警情数"], 2)
        self.assertEqual(branch_one["转案数"], 1)
        self.assertEqual(branch_one["转案率"], "50.00%")
        self.assertEqual(branch_one["原始确认均纠纷性质"], 1)
        self.assertEqual(branch_one["重复警情数"], 2)
        self.assertEqual(branch_one["重复警情转案数"], 1)
        self.assertEqual(branch_one["重复警情转案率"], "50.00%")

        fine = next(row for row in payload["fine"] if row["group_code"] == "B1")
        self.assertEqual(fine["原始细类"], "家庭纠纷")
        self.assertEqual(fine["确认细类"], "邻里纠纷")
        self.assertEqual(fine["警情数"], 1)

        repeat = next(row for row in payload["repeat"] if row["group_code"] == "B1")
        self.assertEqual(repeat["重复电话数"], 1)
        self.assertEqual(repeat["重复警情数"], 2)

    def test_branch_filter_and_detail_dimension(self):
        with self._patch_sources() as mocks:
            mocks["fetch_raw_mdj_cases"].return_value = RAW_ROWS
            mocks["query_org_mappings"].return_value = (
                {"B1": "一分局", "B2": "二分局"},
                {"S1": "一所", "S2": "二所"},
            )
            mocks["query_case_conversion_map"].return_value = {"J1": ["A1"]}

            payload = service.get_detail_payload(
                start_time="2026-04-01 00:00:00",
                end_time="2026-04-02 00:00:00",
                ssfjdm_list=["B1"],
                group_by="fenju",
                repeat_min=2,
                dimension="converted",
                group_code="B1",
                page=1,
                page_size=20,
            )

        self.assertEqual(payload["total"], 1)
        self.assertEqual(payload["rows"][0]["警情编号"], "J1")
        self.assertEqual(payload["rows"][0]["是否转案"], "是")
        self.assertEqual(payload["rows"][0]["案件编号"], "A1")

    def test_summary_export_builds_three_sheets(self):
        fake_payload = {
            "start_time": "2026-04-01 00:00:00",
            "end_time": "2026-04-02 00:00:00",
            "overall": [{"分组": "总计", "原始警情数": 1}],
            "fine": [{"分组": "一分局", "原始细类": "家庭纠纷", "警情数": 1}],
            "repeat": [{"分组": "一分局", "重复警情数": 1}],
        }

        app = Flask(__name__)
        with app.test_request_context("/"), patch.object(export_service, "get_summary_payload", return_value=fake_payload):
            response = export_service.build_summary_export(
                start_time="2026-04-01 00:00:00",
                end_time="2026-04-02 00:00:00",
                ssfjdm_list=[],
                group_by="fenju",
                repeat_min=2,
            )
            response.direct_passthrough = False
            workbook = load_workbook(io.BytesIO(response.get_data()))

        self.assertEqual(workbook.sheetnames, ["总体统计", "细类统计", "重复统计"])
        self.assertEqual(workbook["总体统计"]["A2"].value, "总计")


if __name__ == "__main__":
    unittest.main()
