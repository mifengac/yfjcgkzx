import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from xunfang.service import jiemiansanlei_service as service


class TestJiemiansanleiService(unittest.TestCase):
    def test_query_classified_filters_street_rows_before_paging(self) -> None:
        rows = [
            {
                "leixing": "盗窃",
                "source": "原始",
                "bureau": "云城分局",
                "station_no": "001",
                "station_name": "A所",
                "call_time": "2026-03-01 10:00:00",
                "address": "地址1",
                "lng": "113.1",
                "lat": "22.1",
                "case_type_name": "盗窃",
                "pred_label": "街面与公共区域",
                "pred_prob": 0.9,
            },
            {
                "leixing": "盗窃",
                "source": "原始",
                "bureau": "云城分局",
                "station_no": "002",
                "station_name": "B所",
                "call_time": "2026-03-01 11:00:00",
                "address": "地址2",
                "lng": "113.2",
                "lat": "22.2",
                "case_type_name": "盗窃",
                "pred_label": "住宅小区",
                "pred_prob": 0.8,
            },
            {
                "leixing": "盗窃",
                "source": "原始",
                "bureau": "云城分局",
                "station_no": "003",
                "station_name": "C所",
                "call_time": "2026-03-01 12:00:00",
                "address": "地址3",
                "lng": "113.3",
                "lat": "22.3",
                "case_type_name": "盗窃",
                "pred_label": "街面与公共区域",
                "pred_prob": 0.7,
            },
        ]

        with patch.object(service, "_fetch_rows_for_filters", return_value=rows), patch.object(
            service, "_append_predictions", side_effect=lambda _rows: None
        ):
            result = service.query_classified(
                start_time="2026-03-01 00:00:00",
                end_time="2026-03-02 00:00:00",
                leixing_list=["盗窃"],
                source_list=["原始"],
                page=2,
                page_size=1,
                street_only=True,
                minor_only=False,
            )

        self.assertEqual(result["total"], 2)
        self.assertEqual(result["page"], 2)
        self.assertEqual(len(result["rows"]), 1)
        self.assertEqual(result["rows"][0]["分类结果"], "街面与公共区域")

    def test_build_case_payload_uses_minor_case_mark_and_original_codes(self) -> None:
        payload = service._build_case_payload(  # noqa: SLF001
            start_time="2026-03-01 00:00:00",
            end_time="2026-03-02 00:00:00",
            source="原始",
            code_csv="0101,0102",
            minor_only=True,
            page_num=1,
            page_size=200,
        )

        self.assertEqual(payload["newOriCharaSubclassNo"], "0101,0102")
        self.assertEqual(payload["newCharaSubclassNo"], "")
        self.assertEqual(payload["caseMarkNo"], service.MINOR_CASE_MARK_NO)

    def test_fetch_source_rows_uses_confirmed_codes_and_matches_leixing(self) -> None:
        with patch.object(
            service.api_client,
            "get_case_list",
            return_value={
                "code": 0,
                "total": 1,
                "rows": [
                    {
                        "caseNo": "A001",
                        "newCharaSubclassNo": "0201,0202",
                        "cmdName": "云安分局",
                        "dutyDeptNo": "1001",
                        "dutyDeptName": "巡逻队",
                        "callTime": "2026-03-01 12:00:00",
                        "occurAddress": "测试地址",
                    }
                ],
            },
        ) as mock_case_list:
            rows = service._fetch_source_rows(  # noqa: SLF001
                start_time="2026-03-01 00:00:00",
                end_time="2026-03-02 00:00:00",
                source="确认",
                leixing="侵财类",
                code_list=["0201", "0301"],
                minor_only=True,
            )

        payload = mock_case_list.call_args.args[0]
        self.assertEqual(payload["newCharaSubclassNo"], "0201,0301")
        self.assertEqual(payload["newOriCharaSubclassNo"], "")
        self.assertEqual(payload["caseMarkNo"], service.MINOR_CASE_MARK_NO)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["leixing"], "侵财类")
        self.assertEqual(rows[0]["bureau"], "云安分局")

    def test_export_report_only_uses_time_filters_for_db_query(self) -> None:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        for name in list(service.REPORT_LEIXING_LIST) + ["三类合计"]:
            sheet = workbook.create_sheet(title=name)
            sheet["A6"] = ""

        db_row = {
            "caseno": "A001",
            "leixing": "人身伤害类",
            "yuanshiqueren": "原始",
            "分局": "云城分局",
            "派出所编号": "001",
            "派出所名称": "测试所",
            "报警时间": "2026-03-01 10:00:00",
            "警情地址": "测试地址",
            "报警内容": "测试内容",
            "处警情况": "已处置",
            "经度": "113.1",
            "纬度": "22.1",
            "jq_type": "测试类型",
        }

        def fake_append_predictions(rows):
            for row in rows:
                row["pred_label"] = service.STREET_LABEL
                row["pred_prob"] = 0.99

        with patch.object(service.os.path, "exists", return_value=True), patch(
            "openpyxl.load_workbook",
            return_value=workbook,
        ), patch.object(
            service,
            "fetch_db_jingqings",
            return_value=[db_row],
        ) as mock_fetch, patch.object(service, "_append_predictions", side_effect=fake_append_predictions):
            file_bytes, _mimetype, _filename = service.export_report(
                start_time="2026-03-01 00:00:00",
                end_time="2026-03-02 00:00:00",
                hb_start_time="2026-02-22 00:00:00",
                hb_end_time="2026-02-23 00:00:00",
            )

        self.assertEqual(mock_fetch.call_count, 2)
        for call in mock_fetch.call_args_list:
            query = call.args[0]
            self.assertFalse(query.minor_only)
            self.assertEqual(list(query.leixing_list), list(service.REPORT_LEIXING_LIST))
            self.assertEqual(list(query.source_list), ["原始", "确认"])

        exported = load_workbook(BytesIO(file_bytes))
        self.assertEqual(exported["人身伤害类"]["C6"].value, 1)
        self.assertEqual(exported["三类合计"]["C6"].value, 1)


if __name__ == "__main__":
    unittest.main()
