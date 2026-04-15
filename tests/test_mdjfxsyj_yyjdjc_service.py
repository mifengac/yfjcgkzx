import io
import unittest
from datetime import datetime
from unittest.mock import patch

from flask import Flask
from openpyxl import load_workbook

from mdjfxsyj.service import mdjfxsyj_yyjdjc_service as service
from mdjfxsyj.service import mdjfxsyj_yyjdjc_workorder_support as workorder_support


class _FixedDatetime(service.datetime):
    @classmethod
    def now(cls, tz=None):
        return cls(2026, 4, 3, 15, 30, 0)


class TestMdjfxsyjYyjdjcService(unittest.TestCase):
    def test_find_matched_keywords_keeps_config_order(self) -> None:
        matched = service.find_matched_keywords(
            ["\u6709\u4eba\u626c\u8a00\u62a5\u590d\u793e\u4f1a\uff0c\u8fd8\u8bf4\u8981\u653e\u706b\u5e76\u4e14\u518d\u6b21\u626c\u8a00"],
            ["\u626c\u8a00", "\u653e\u706b", "\u62a5\u590d\u793e\u4f1a", "\u626c\u8a00"],
        )

        self.assertEqual(matched, ["\u626c\u8a00", "\u653e\u706b", "\u62a5\u590d\u793e\u4f1a"])

    def test_default_range_uses_today_midnight_and_minus_eight_days(self) -> None:
        with patch.object(service, "datetime", _FixedDatetime):
            start_dt, end_dt = service.default_range()

        self.assertEqual(start_dt.strftime("%Y-%m-%d %H:%M:%S"), "2026-03-26 00:00:00")
        self.assertEqual(end_dt.strftime("%Y-%m-%d %H:%M:%S"), "2026-04-03 00:00:00")

    def test_fetch_workorder_source_rows_maps_new_source_fields(self) -> None:
        raw_rows = [
            {
                "orderid": "GD-1",
                "name": "\u5f20\u4e09",
                "mobile": "13800000000",
                "ordertitle": "\u5de5\u5355\u6807\u9898",
                "ordercont": "\u91cd\u590d\u626c\u8a00\u81ea\u6740\u5e76\u62a5\u590d\u793e\u4f1a",
                "keyword": "\u77db\u76fe\u7ea0\u7eb7",
                "registertime": "20260403112233",
                "caseaddr": '{"coordinate":[121.4737,31.2304],"address":"\u95f5\u884c\u533a\u67d0\u8def"}',
                "objectname": "\u67d0\u5355\u4f4d",
                "orderstatuscd": "01",
            }
        ]
        keywords = ["\u81ea\u6740", "\u62a5\u590d\u793e\u4f1a"]

        with patch.object(service, "query_workorder_rows", return_value=raw_rows) as mock_query:
            rows = service.fetch_workorder_source_rows(
                start_dt=datetime(2026, 4, 1, 0, 0, 0),
                end_dt=datetime(2026, 4, 3, 23, 59, 59),
                keywords=keywords,
            )

        mock_query.assert_called_once_with(
            start_time="20260401000000",
            end_time="20260403235959",
            keywords=keywords,
        )

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[0]], "GD-1")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[1]], "\u5f20\u4e09")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[2]], "13800000000")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[3]], "\u5de5\u5355\u6807\u9898")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[4]], "\u91cd\u590d\u626c\u8a00\u81ea\u6740\u5e76\u62a5\u590d\u793e\u4f1a")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[5]], "\u77db\u76fe\u7ea0\u7eb7")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[6]], "2026-04-03 11:22:33")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[7]], "\u95f5\u884c\u533a\u67d0\u8def")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[8]], "\u67d0\u5355\u4f4d")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[9]], "01")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[10]], "\u81ea\u6740\u3001\u62a5\u590d\u793e\u4f1a")

    def test_fetch_workorder_source_rows_ignores_title_only_keyword_matches(self) -> None:
        raw_rows = [
            {
                "orderid": "GD-2",
                "name": "\u674e\u56db",
                "mobile": "13900000000",
                "ordertitle": "\u5de5\u5355\u6807\u9898\u63d0\u5230\u81ea\u6740",
                "ordercont": "\u666e\u901a\u6295\u8bc9\u5185\u5bb9",
                "keyword": "",
                "registertime": "20260403112233",
                "caseaddr": "",
                "objectname": "",
                "orderstatuscd": "",
            }
        ]

        with patch.object(service, "query_workorder_rows", return_value=raw_rows):
            rows = service.fetch_workorder_source_rows(
                start_dt=datetime(2026, 4, 1, 0, 0, 0),
                end_dt=datetime(2026, 4, 3, 23, 59, 59),
                keywords=["\u81ea\u6740"],
            )

        self.assertEqual(rows, [])

    def test_fetch_workorder_source_rows_handles_missing_contact_and_bad_address(self) -> None:
        raw_rows = [
            {
                "orderid": "GD-3",
                "name": None,
                "mobile": None,
                "ordertitle": "\u5de5\u5355\u6807\u9898",
                "ordercont": "\u53cd\u590d\u626c\u8a00\u62a5\u590d",
                "keyword": "",
                "registertime": "bad-time",
                "caseaddr": "bad-json",
                "objectname": "",
                "orderstatuscd": "",
            }
        ]

        with patch.object(service, "query_workorder_rows", return_value=raw_rows):
            rows = service.fetch_workorder_source_rows(
                start_dt=datetime(2026, 4, 1, 0, 0, 0),
                end_dt=datetime(2026, 4, 3, 23, 59, 59),
                keywords=["\u62a5\u590d"],
            )

        self.assertEqual(len(rows), 1)
        row = rows[0]
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[1]], "")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[2]], "")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[6]], "bad-time")
        self.assertEqual(row[workorder_support.WORKORDER_COLUMNS[7]], "bad-json")

    def test_get_monitor_data_collects_source_errors_individually(self) -> None:
        with (
            patch.object(service, "fetch_police_rows", side_effect=RuntimeError("\u8b66\u60c5\u63a5\u53e3\u4e0d\u53ef\u7528")),
            patch.object(
                service,
                "fetch_workorder_source_rows",
                return_value=[
                    {
                        workorder_support.WORKORDER_COLUMNS[0]: "GD-1",
                        workorder_support.WORKORDER_COLUMNS[10]: "\u626c\u8a00",
                    }
                ],
            ),
            patch.object(service, "fetch_dispute_source_rows", return_value=[]),
        ):
            payload = service.get_monitor_data(
                start_time="2026-04-01 00:00:00",
                end_time="2026-04-03 00:00:00",
            )

        self.assertEqual(payload["sources"]["police"]["count"], 0)
        self.assertIn("\u8b66\u60c5\u63a5\u53e3\u4e0d\u53ef\u7528", payload["sources"]["police"]["error"])
        self.assertEqual(payload["sources"]["workorder"]["count"], 1)
        self.assertEqual(payload["sources"]["dispute"]["rows"], [])

    def test_build_all_sources_export_has_three_fixed_sheets(self) -> None:
        fake_payload = {
            "start_time": "2026-04-01 00:00:00",
            "end_time": "2026-04-03 00:00:00",
            "keywords": list(service.DEFAULT_KEYWORDS),
            "sources": {
                "police": {
                    "rows": [{"\u8b66\u60c5\u7f16\u53f7": "J1", "\u547d\u4e2d\u5173\u952e\u8bcd": "\u626c\u8a00"}],
                    "columns": service.SOURCE_SPECS["police"]["columns"],
                    "count": 1,
                    "error": "",
                },
                "workorder": {
                    "rows": [{workorder_support.WORKORDER_COLUMNS[0]: "GD-1"}],
                    "columns": service.SOURCE_SPECS["workorder"]["columns"],
                    "count": 1,
                    "error": "",
                },
                "dispute": {
                    "rows": [],
                    "columns": service.SOURCE_SPECS["dispute"]["columns"],
                    "count": 0,
                    "error": "\u6d4b\u8bd5\u9519\u8bef",
                },
            },
        }

        app = Flask(__name__)
        with app.test_request_context("/"), patch.object(service, "get_monitor_data", return_value=fake_payload):
            response = service.build_all_sources_export(
                start_time="2026-04-01 00:00:00",
                end_time="2026-04-03 00:00:00",
            )
            response.direct_passthrough = False
            workbook = load_workbook(io.BytesIO(response.get_data()))

        self.assertEqual(
            workbook.sheetnames,
            [
                service.SOURCE_SPECS["police"]["sheet_name"],
                service.SOURCE_SPECS["workorder"]["sheet_name"],
                service.SOURCE_SPECS["dispute"]["sheet_name"],
            ],
        )
        self.assertEqual(workbook[service.SOURCE_SPECS["police"]["sheet_name"]]["A2"].value, "J1")
        self.assertEqual(
            workbook[service.SOURCE_SPECS["workorder"]["sheet_name"]]["A1"].value,
            workorder_support.WORKORDER_COLUMNS[0],
        )
        self.assertEqual(
            workbook[service.SOURCE_SPECS["workorder"]["sheet_name"]]["A2"].value,
            "GD-1",
        )
        self.assertEqual(
            workbook[service.SOURCE_SPECS["dispute"]["sheet_name"]]["A1"].value,
            "\u9519\u8bef\u4fe1\u606f",
        )


if __name__ == "__main__":
    unittest.main()
