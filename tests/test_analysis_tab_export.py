import io
import unittest
from unittest.mock import patch

from flask import Flask
from openpyxl import load_workbook

from jingqing_fenxi.routes.jingqing_fenxi_routes import jingqing_fenxi_bp
from jingqing_fenxi.service.jingqing_fenxi_service import generate_excel_report


def _find_row(sheet, cell_value):
    for row_idx in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_idx, column=1).value == cell_value:
            return row_idx
    return None


class TestGenerateExcelReport(unittest.TestCase):
    def test_generate_excel_report_uses_single_sheet_and_single_detail_section(self):
        workbook_stream = generate_excel_report(
            analysis_results={
                "srr": [
                    {
                        "name": "云城分局",
                        "presentCycle": 6,
                        "upperY2yCycle": 5,
                        "y2yProportion": "20%",
                        "upperM2mCycle": 4,
                        "m2mProportion": "50%",
                    }
                ],
                "time": [("0-3时", 7)],
                "dept": [("城中所", 3)],
            },
            all_data=[
                {
                    "caseNo": "JQ001",
                    "callTime": "2026-04-10 08:00:00",
                    "caseLevelName": "一级",
                    "occurAddress": "云城区建设路",
                    "callerPhone": "13800138000",
                    "dutyDeptName": "城中所",
                    "caseState": "已处置",
                    "caseContents": "持刀纠纷",
                }
            ],
            dimensions_selected=["srr", "time", "dept"],
            analysis_options={"timeBucketHours": 3},
            begin_date="2026-04-01 00:00:00",
            end_date="2026-04-10 23:59:59",
        )

        workbook = load_workbook(workbook_stream)
        sheet = workbook.active

        self.assertEqual(workbook.sheetnames, ["警情分析报表"])
        self.assertEqual(sheet["A1"].value, "警情分析报表")
        self.assertEqual(sheet["A2"].value, "开始时间")
        self.assertEqual(sheet["B2"].value, "2026-04-01 00:00:00")
        self.assertEqual(sheet["A3"].value, "结束时间")
        self.assertEqual(sheet["B3"].value, "2026-04-10 23:59:59")

        srr_row = _find_row(sheet, "各地同环比统计")
        time_row = _find_row(sheet, "时段(每3小时)统计")
        dept_row = _find_row(sheet, "派出所统计")
        detail_row = _find_row(sheet, "分析源数据明细")

        self.assertIsNotNone(srr_row)
        self.assertIsNotNone(time_row)
        self.assertIsNotNone(dept_row)
        self.assertIsNotNone(detail_row)
        self.assertLess(srr_row, time_row)
        self.assertLess(time_row, dept_row)
        self.assertLess(dept_row, detail_row)
        self.assertEqual(
            sum(1 for row_idx in range(1, sheet.max_row + 1) if sheet.cell(row=row_idx, column=1).value == "分析源数据明细"),
            1,
        )
        self.assertEqual(sheet.cell(row=detail_row + 1, column=1).value, "接警号")
        self.assertEqual(sheet.cell(row=detail_row + 2, column=1).value, "JQ001")

    def test_generate_excel_report_without_dimensions_still_writes_detail_section(self):
        workbook_stream = generate_excel_report(
            analysis_results={},
            all_data=[
                {
                    "caseNo": "JQ002",
                    "callTime": "2026-04-11 09:00:00",
                    "caseLevelName": "二级",
                    "occurAddress": "罗定市人民路",
                    "callerPhone": "13900139000",
                    "dutyDeptName": "人民所",
                    "caseState": "已反馈",
                    "caseContents": "警情说明",
                }
            ],
            dimensions_selected=[],
            begin_date="2026-04-11 00:00:00",
            end_date="2026-04-11 23:59:59",
        )

        workbook = load_workbook(workbook_stream)
        sheet = workbook.active
        detail_row = _find_row(sheet, "分析源数据明细")

        self.assertEqual(workbook.sheetnames, ["警情分析报表"])
        self.assertEqual(sheet["B2"].value, "2026-04-11 00:00:00")
        self.assertEqual(sheet["B3"].value, "2026-04-11 23:59:59")
        self.assertEqual(detail_row, 5)
        self.assertEqual(sheet.cell(row=detail_row + 2, column=1).value, "JQ002")
        self.assertIsNone(_find_row(sheet, "无数据"))


class TestAnalysisExportRoute(unittest.TestCase):
    def setUp(self):
        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.register_blueprint(jingqing_fenxi_bp, url_prefix="/jingqing_fenxi")
        self.client = app.test_client()

    def test_export_route_passes_begin_and_end_time_to_excel_report(self):
        with patch(
            "jingqing_fenxi.routes.analysis_tab_routes.run_analysis",
            return_value=({"time": [("0-3时", 1)]}, {}, [], {"timeBucketHours": 3}),
        ), patch(
            "jingqing_fenxi.routes.analysis_tab_routes.generate_excel_report",
            return_value=io.BytesIO(b"test-export"),
        ) as mock_generate:
            response = self.client.post(
                "/jingqing_fenxi/export",
                data={
                    "beginDate": "2026-04-01 00:00:00",
                    "endDate": "2026-04-10 23:59:59",
                    "dimensions[]": ["time", "dept"],
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(mock_generate.call_args.kwargs["begin_date"], "2026-04-01 00:00:00")
        self.assertEqual(mock_generate.call_args.kwargs["end_date"], "2026-04-10 23:59:59")
        self.assertEqual(response.data, b"test-export")


class TestAnalysisTreeRoutes(unittest.TestCase):
    def setUp(self):
        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.register_blueprint(jingqing_fenxi_bp, url_prefix="/jingqing_fenxi")
        self.client = app.test_client()

    def test_tree_routes_return_matching_sources(self):
        with patch(
            "jingqing_fenxi.routes.analysis_tab_routes.get_tree_view_data",
            return_value=[{"id": "plan-root", "name": "预案"}],
        ), patch(
            "jingqing_fenxi.routes.analysis_tab_routes.get_nature_tree_new_view_data",
            return_value=[{"id": "01", "name": "警情性质"}],
        ):
            legacy_response = self.client.get("/jingqing_fenxi/treeData")
            plan_response = self.client.get("/jingqing_fenxi/planTreeData")
            nature_response = self.client.get("/jingqing_fenxi/natureTreeData")

        self.assertEqual(legacy_response.get_json(), [{"id": "plan-root", "name": "预案"}])
        self.assertEqual(plan_response.get_json(), [{"id": "plan-root", "name": "预案"}])
        self.assertEqual(nature_response.get_json(), [{"id": "01", "name": "警情性质"}])


if __name__ == "__main__":
    unittest.main()
