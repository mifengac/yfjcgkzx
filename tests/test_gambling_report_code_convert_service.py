import io
import unittest
from datetime import datetime as real_datetime
from unittest.mock import patch

from docx import Document
from flask import Flask
from openpyxl import Workbook, load_workbook

from jingqing_fenxi.routes.jingqing_fenxi_routes import jingqing_fenxi_bp
from jingqing_fenxi.service import gambling_report_code_convert_service as service


class TestGamblingReportCodeConvertService(unittest.TestCase):
    def test_extract_station_codes_returns_unique_12_digit_codes(self) -> None:
        text = "445321510000、445321510000、445381650000，2026年1024起"

        self.assertEqual(service.extract_station_codes(text), ["445321510000", "445381650000"])

    def test_convert_markdown_station_codes_to_docx_replaces_known_codes(self) -> None:
        markdown = "# 赌博分析报告\n\n一、数据分析\n\n445321510000的89起，445381650000的57起。"
        with patch.object(
            service,
            "fetch_station_name_map",
            return_value={"445321510000": "测试分局测试派出所"},
        ) as mock_fetch:
            output = service.convert_markdown_station_codes_to_docx(markdown.encode("utf-8"), "report.md")

        mock_fetch.assert_called_once_with(["445321510000", "445381650000"])
        document = Document(io.BytesIO(output.getvalue()))
        text = "\n".join(paragraph.text for paragraph in document.paragraphs)
        self.assertIn("测试分局测试派出所的89起", text)
        self.assertIn("445381650000", text)
        self.assertIn("未转换派出所代码", text)

    def test_fetch_station_name_map_prefixes_branch_name(self) -> None:
        class FakeCursor:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return None

            def execute(self, sql, params):
                self.sql = sql
                self.params = params

            def fetchall(self):
                return [
                    {
                        "code": "445321510000",
                        "branch_name": "测试分局",
                        "station_name": "测试派出所",
                    }
                ]

        class FakeConnection:
            def __init__(self):
                self.cursor_obj = FakeCursor()
                self.closed = False

            def cursor(self, *_, **__):
                return self.cursor_obj

            def close(self):
                self.closed = True

        connection = FakeConnection()
        with patch.object(service, "get_database_connection", return_value=connection):
            station_map = service.fetch_station_name_map(["445321510000"])

        self.assertEqual(station_map, {"445321510000": "测试分局测试派出所"})
        self.assertIn("ssfj", connection.cursor_obj.sql)
        self.assertEqual(connection.cursor_obj.params, ["445321510000"])
        self.assertTrue(connection.closed)

    def test_build_code_convert_filename_uses_original_stem(self) -> None:
        filename = service.build_code_convert_filename(
            "赌博警情分析报告.md",
            now=real_datetime(2026, 4, 28, 18, 1, 2),
        )

        self.assertEqual(filename, "赌博警情分析报告_派出所名称转换20260428180102.docx")

    def test_build_code_convert_filename_keeps_xlsx_extension(self) -> None:
        filename = service.build_code_convert_filename(
            "赌博分析附件.xlsx",
            now=real_datetime(2026, 4, 29, 9, 10, 11),
        )

        self.assertEqual(filename, "赌博分析附件_派出所名称转换20260429091011.xlsx")

    def test_convert_xlsx_station_codes_keeps_workbook_format(self) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "明细"
        worksheet["A1"] = "派出所"
        worksheet["A2"] = "445321510000的89起，445381650000的57起"
        worksheet["B2"] = 445321510000
        worksheet["C2"] = "=A2"
        source = io.BytesIO()
        workbook.save(source)

        with patch.object(
            service,
            "fetch_station_name_map",
            return_value={"445321510000": "测试分局测试派出所"},
        ) as mock_fetch:
            output = service.convert_xlsx_station_codes(source.getvalue(), "report.xlsx")

        mock_fetch.assert_called_once_with(["445321510000", "445381650000"])
        converted = load_workbook(io.BytesIO(output.getvalue()), data_only=False)
        sheet = converted["明细"]
        self.assertEqual(sheet["A2"].value, "测试分局测试派出所的89起，445381650000的57起")
        self.assertEqual(sheet["B2"].value, "测试分局测试派出所")
        self.assertEqual(sheet["C2"].value, "=A2")


class TestGamblingReportCodeConvertRoute(unittest.TestCase):
    def setUp(self) -> None:
        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.register_blueprint(jingqing_fenxi_bp, url_prefix="/jingqing_fenxi")
        self.client = app.test_client()

    def test_code_convert_route_downloads_docx(self) -> None:
        with patch(
            "jingqing_fenxi.routes.gambling_topic_routes.convert_markdown_station_codes_to_docx",
            return_value=io.BytesIO(b"docx-data"),
        ) as mock_convert, patch(
            "jingqing_fenxi.routes.gambling_topic_routes.build_code_convert_filename",
            return_value="converted.docx",
        ):
            response = self.client.post(
                "/jingqing_fenxi/download/gambling-topic/code-convert",
                data={"file": (io.BytesIO("# 报告\n445321510000".encode("utf-8")), "report.md")},
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b"docx-data")
        self.assertIn("converted.docx", response.headers["Content-Disposition"])
        self.assertEqual(mock_convert.call_args.args[1], "report.md")

    def test_code_convert_route_downloads_xlsx(self) -> None:
        with patch(
            "jingqing_fenxi.routes.gambling_topic_routes.convert_xlsx_station_codes",
            return_value=io.BytesIO(b"xlsx-data"),
        ) as mock_convert, patch(
            "jingqing_fenxi.routes.gambling_topic_routes.build_code_convert_filename",
            return_value="converted.xlsx",
        ):
            response = self.client.post(
                "/jingqing_fenxi/download/gambling-topic/code-convert",
                data={"file": (io.BytesIO(b"xlsx-content"), "report.xlsx")},
                content_type="multipart/form-data",
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b"xlsx-data")
        self.assertIn("converted.xlsx", response.headers["Content-Disposition"])
        self.assertEqual(
            response.headers["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(mock_convert.call_args.args[1], "report.xlsx")

    def test_code_convert_route_rejects_unsupported_file(self) -> None:
        response = self.client.post(
            "/jingqing_fenxi/download/gambling-topic/code-convert",
            data={"file": (io.BytesIO(b"content"), "report.txt")},
            content_type="multipart/form-data",
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn("只支持上传 .md 或 .xlsx", response.get_json()["message"])


if __name__ == "__main__":
    unittest.main()
