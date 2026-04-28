import io
import unittest
from datetime import datetime as real_datetime
from unittest.mock import patch

from flask import Flask
from openpyxl import load_workbook

from jingqing_fenxi.routes.jingqing_fenxi_routes import jingqing_fenxi_bp
from jingqing_fenxi.service import gambling_topic_service as service


class TestGamblingTopicService(unittest.TestCase):
    def test_resolve_gambling_topic_tags_uses_target_parent(self) -> None:
        tree_nodes = [
            {"id": "1", "pId": "OTHER", "name": "其他", "tag": "999"},
            {"id": "2", "pId": service.GAMBLING_TOPIC_PARENT_ID, "name": "赌博", "tag": "0301"},
            {"id": "3", "pId": service.GAMBLING_TOPIC_PARENT_ID, "name": "赌博", "tag": "0301"},
            {"id": "4", "pId": service.GAMBLING_TOPIC_PARENT_ID, "name": "开设赌场", "tag": "0302"},
        ]

        tag_csv, name_csv = service.resolve_gambling_topic_tags(tree_nodes)

        self.assertEqual(tag_csv, "0301,0302")
        self.assertEqual(name_csv, "赌博,开设赌场")

    def test_build_gambling_case_payload_uses_original_chara(self) -> None:
        payload = service._build_gambling_case_payload(
            "2026-01-01 00:00:00",
            "2026-04-28 00:00:00",
            "09010200,02052001",
            "举报聚众赌博,赌博",
        )

        self.assertEqual(payload["newCharaSubclassNo"], "")
        self.assertEqual(payload["newCharaSubclass"], "全部")
        self.assertEqual(payload["newOriCharaSubclassNo"], "09010200,02052001")
        self.assertEqual(payload["newOriCharaSubclass"], "举报聚众赌博,赌博")
        self.assertEqual(payload["beginDate"], "2026-01-01 00:00:00")
        self.assertEqual(payload["endDate"], "2026-04-28 00:00:00")

    def test_summarize_gambling_way_by_region_counts_keyword_categories(self) -> None:
        rows = [
            {"cmdName": "云城", "caseContents": "有人聚众打麻将赌博", "replies": ""},
            {"cmdName": "云城", "caseContents": "现场有人玩牛牛和炸金花", "replies": ""},
            {"cmdName": "罗定", "caseContents": "群众举报网络赌博平台", "replies": ""},
        ]

        result = service.summarize_gambling_way_by_region(rows)

        self.assertEqual(result["rows"][0]["cmdName"], "云城")
        self.assertEqual(result["rows"][0]["counts"]["麻将"], 1)
        self.assertEqual(result["rows"][0]["counts"]["斗牛"], 1)
        self.assertEqual(result["rows"][0]["counts"]["扑克"], 1)
        self.assertEqual(result["rows"][0]["total"], 3)
        self.assertEqual(rows[1]["gamblingWayLabels"], "斗牛、扑克")
        self.assertIn("炸金花", rows[1]["gamblingWayKeywords"])

    def test_summarize_wilderness_by_region_uses_reply_keywords(self) -> None:
        rows = [
            {"cmdName": "云城", "caseContents": "赌博", "replies": "民警到山腰树林处查处"},
            {"cmdName": "罗定", "caseContents": "赌博", "replies": "室内麻将档"},
            {"cmdName": "云城", "caseContents": "赌博", "replies": "山脚野外有人聚赌"},
        ]

        result = service.summarize_wilderness_by_region(rows)

        self.assertEqual(result["rows"], [{"cmdName": "云城", "total": 2}])
        self.assertEqual(len(result["details"]), 2)
        self.assertIn("山腰", rows[0]["gamblingWildernessKeywords"])
        self.assertEqual(rows[1]["gamblingWildernessKeywords"], "")

    def test_summarize_venue_by_cmd_id_uses_content_reply_and_address(self) -> None:
        rows = [
            {"cmdId": "445302000000", "cmdName": "云城", "caseContents": "棋牌室有人赌博", "replies": "", "occurAddress": ""},
            {"cmdId": "445302000000", "cmdName": "云城", "caseContents": "", "replies": "现场为麻将馆", "occurAddress": ""},
            {"cmdId": "445381000000", "cmdName": "罗定", "caseContents": "", "replies": "", "occurAddress": "某小卖部"},
            {"cmdId": "445381000000", "cmdName": "罗定", "caseContents": "住宅内赌博", "replies": "", "occurAddress": ""},
        ]

        result = service.summarize_venue_by_cmd_id(rows)

        self.assertEqual(result["rows"][0]["cmdId"], "445302000000")
        self.assertEqual(result["rows"][0]["total"], 2)
        self.assertEqual(result["rows"][1]["cmdId"], "445381000000")
        self.assertEqual(result["rows"][1]["total"], 1)
        self.assertEqual(rows[0]["gamblingVenueFields"], "报警内容")
        self.assertEqual(rows[1]["gamblingVenueFields"], "处警情况")
        self.assertEqual(rows[2]["gamblingVenueFields"], "警情地址")
        self.assertEqual(rows[3]["gamblingVenueKeywords"], "")

    def test_run_analysis_returns_selected_custom_dimensions(self) -> None:
        params = {
            "beginDate": "2026-04-01 00:00:00",
            "endDate": "2026-04-02 00:00:00",
            "m2mStartTime": "2026-03-31 00:00:00",
            "m2mEndTime": "2026-04-01 00:00:00",
        }
        rows = [
            {
                "cmdName": "云城",
                "cmdId": "445302000000",
                "caseContents": "出租屋有人打麻将赌博",
                "replies": "处警发现山脚树林有人聚赌",
                "occurAddress": "某棋牌室",
                "callTime": "2026-04-01 01:00:00",
            }
        ]

        with patch.object(service, "resolve_gambling_topic_tags", return_value=("0301", "赌博")), patch.object(
            service,
            "fetch_all_case_list",
            return_value=rows,
        ) as mock_fetch:
            results, _base, all_data, _options, _meta = service.run_gambling_topic_analysis(
                params,
                ["gambling_way", "wilderness", "venue"],
            )

        self.assertIn("gambling_way", results)
        self.assertIn("wilderness", results)
        self.assertIn("venue", results)
        self.assertEqual(results["gambling_way"]["rows"][0]["counts"]["麻将"], 1)
        self.assertEqual(results["wilderness"]["rows"][0]["total"], 1)
        self.assertEqual(results["venue"]["rows"][0]["cmdId"], "445302000000")
        self.assertEqual(results["venue"]["rows"][0]["total"], 1)
        self.assertEqual(all_data[0]["gamblingWayLabels"], "麻将")
        self.assertEqual(all_data[0]["gamblingVenueKeywords"], "棋牌室")
        payload = mock_fetch.call_args.args[0]
        self.assertEqual(payload["newCharaSubclassNo"], "")
        self.assertEqual(payload["newOriCharaSubclassNo"], "0301")
        self.assertEqual(mock_fetch.call_args.kwargs["max_page_size"], service.GAMBLING_TOPIC_UPSTREAM_PAGE_SIZE)

    def test_generate_excel_adds_custom_dimension_sheets(self) -> None:
        all_data = [
            {
                "caseNo": "JQ001",
                "callTime": "2026-04-01 01:00:00",
                "cmdName": "云城",
                "cmdId": "445302000000",
                "dutyDeptName": "测试所",
                "caseContents": "有人打麻将赌博",
                "replies": "处警发现山腰树林有人聚赌",
                "occurAddress": "某棋牌室",
                "fightAddrLabel": "商业场所",
                "gamblingWayLabels": "麻将",
                "gamblingWayKeywords": "麻将",
                "gamblingWildernessKeywords": "山腰、树林",
                "gamblingVenueFields": "警情地址",
                "gamblingVenueKeywords": "棋牌室",
            }
        ]
        analysis_results = {
            "addr": [("商业场所", 1)],
            "gambling_way": {
                "columns": ["麻将"],
                "rows": [{"cmdName": "云城", "counts": {"麻将": 1}, "total": 1}],
                "details": all_data,
            },
            "wilderness": {
                "rows": [{"cmdName": "云城", "total": 1}],
                "details": all_data,
            },
            "venue": {
                "rows": [{"cmdId": "445302000000", "cmdName": "云城", "total": 1}],
                "details": all_data,
            },
        }

        buffer = service.generate_gambling_topic_excel(
            analysis_results,
            all_data,
            ["addr", "gambling_way", "wilderness", "venue"],
            begin_date="2026-04-01 00:00:00",
            end_date="2026-04-02 00:00:00",
        )
        workbook = load_workbook(io.BytesIO(buffer.getvalue()))

        self.assertEqual(workbook.sheetnames, ["赌博专题", "赌博方式", "涉山林野外赌博", "棋牌麻将小卖部"])
        main_values = [cell.value for row in workbook["赌博专题"].iter_rows() for cell in row]
        self.assertIn("警情地址统计", main_values)
        self.assertIn("商业场所", main_values)
        way_sheet = workbook["赌博方式"]
        wild_sheet = workbook["涉山林野外赌博"]
        venue_sheet = workbook["棋牌麻将小卖部"]
        self.assertEqual(way_sheet["A1"].value, "赌博方式")
        self.assertIn("命中关键词", [cell.value for row in way_sheet.iter_rows() for cell in row])
        self.assertEqual(wild_sheet["A1"].value, "涉山林野外赌博")
        self.assertIn("山腰、树林", [cell.value for row in wild_sheet.iter_rows() for cell in row])
        self.assertEqual(venue_sheet["A1"].value, "棋牌室/麻将馆/小卖部")
        self.assertIn("棋牌室", [cell.value for row in venue_sheet.iter_rows() for cell in row])

    def test_build_export_filename_matches_expected_format(self) -> None:
        filename = service.build_export_filename(
            "2026-04-01 00:00:00",
            "2026-04-02 00:00:00",
            now=real_datetime(2026, 4, 20, 8, 9, 10),
        )
        self.assertEqual(filename, "2026-04-01-2026-04-02赌博专题警情分析20260420080910.xlsx")

class TestGamblingTopicRoutes(unittest.TestCase):
    def setUp(self) -> None:
        app = Flask(__name__)
        app.secret_key = "test-secret"
        app.register_blueprint(jingqing_fenxi_bp, url_prefix="/jingqing_fenxi")
        self.client = app.test_client()

    def test_analyze_route_returns_service_result(self) -> None:
        with patch(
            "jingqing_fenxi.routes.gambling_topic_routes.run_gambling_topic_analysis",
            return_value=({"gambling_way": {"rows": []}}, {}, [], {}, {}),
        ) as mock_run:
            response = self.client.post(
                "/jingqing_fenxi/api/gambling-topic/analyze",
                data={"dimensions[]": ["gambling_way"]},
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["code"], 0)
        self.assertEqual(mock_run.call_args.args[1], ["gambling_way"])

    def test_download_route_returns_workbook(self) -> None:
        with patch(
            "jingqing_fenxi.routes.gambling_topic_routes.run_gambling_topic_analysis",
            return_value=({}, {}, [], {}, {"beginDate": "2026-04-01 00:00:00", "endDate": "2026-04-02 00:00:00"}),
        ), patch(
            "jingqing_fenxi.routes.gambling_topic_routes.generate_gambling_topic_excel",
            return_value=io.BytesIO(b"test-export"),
        ), patch(
            "jingqing_fenxi.routes.gambling_topic_routes.build_export_filename",
            return_value="gambling.xlsx",
        ):
            response = self.client.get("/jingqing_fenxi/download/gambling-topic?dimensions[]=gambling_way")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b"test-export")
        self.assertIn("gambling.xlsx", response.headers["Content-Disposition"])

if __name__ == "__main__":
    unittest.main()
