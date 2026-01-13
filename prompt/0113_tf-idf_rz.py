#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
使用 TF-IDF + 余弦相似度判断工作日志重复度：
- 按 “证件号码” 分组
- 在每组内计算 “工作日志工作情况说明” 的相似度
- 导出重复度 >= 阈值（默认 0.8）的记录到当前目录 CSV：
  gongzuorizhi_chongfudu_{时间戳}.csv

输出要求：
- CSV 增加一列：序号（同一重复簇序号一样，按 1,2,3...）
- 仅导出命中重复簇的记录

输入：
- 支持 xlsx（依赖 openpyxl）或 csv（标准库）
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple


DEFAULT_ID_COL = "证件号码"
DEFAULT_TEXT_COL = "工作日志工作情况说明"
DEFAULT_STATION_COL = "所属派出所"
DEFAULT_NAME_COL = "姓名"


def log(msg: str, *, quiet: bool = False) -> None:
    if quiet:
        return
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{ts}] {msg}")


def _norm(s: Any) -> str:
    return ("" if s is None else str(s)).strip()


def _norm_key(s: Any) -> str:
    return _norm(s).lower()


def _first(items: Iterable[str], default: str = "") -> str:
    for it in items:
        if it:
            return it
    return default


def resolve_col_name(row: Dict[str, Any], want: str) -> str:
    w = _norm_key(want)
    for k in row.keys():
        if _norm_key(k) == w:
            return k
    return want


def read_rows_from_csv(path: str) -> List[Dict[str, Any]]:
    # 简单编码兜底：utf-8-sig / gb18030
    encodings = ["utf-8-sig", "utf-8", "gb18030"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.DictReader(f)
                return [dict(r) for r in reader]
        except Exception as e:
            last_err = e
    raise RuntimeError(f"读取CSV失败：{path} ({last_err})")


def read_rows_from_xlsx(path: str, *, sheet: str = "") -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(f"缺少依赖 openpyxl，无法读取 xlsx：{e}")

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header = next(it, None)
    if not header:
        return []
    keys = [str(x).strip() if x is not None else "" for x in header]
    out: List[Dict[str, Any]] = []
    for r in it:
        if r is None:
            continue
        d: Dict[str, Any] = {}
        for i, k in enumerate(keys):
            if not k:
                continue
            d[k] = r[i] if i < len(r) else None
        if d:
            out.append(d)
    return out


def load_rows(path: str, *, sheet: str = "") -> List[Dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return read_rows_from_xlsx(path, sheet=sheet)
    if ext in (".csv", ".txt"):
        return read_rows_from_csv(path)
    raise RuntimeError(f"不支持的输入格式：{path}（仅支持 xlsx/csv）")


_RE_TOKEN = re.compile(r"[A-Za-z0-9]+|[\u4e00-\u9fff]+")


def tokenize(text: str) -> List[str]:
    """
    轻量 tokenizer：
    - 英文数字按连续串
    - 中文按2-gram（更适配短句相似度），很短则按单字
    """
    text = _norm(text)
    if not text:
        return []

    tokens: List[str] = []
    for m in _RE_TOKEN.finditer(text):
        w = m.group(0)
        if not w:
            continue
        if re.fullmatch(r"[A-Za-z0-9]+", w):
            tokens.append(w.lower())
            continue
        # 中文串
        if len(w) <= 2:
            tokens.extend(list(w))
        else:
            for i in range(len(w) - 1):
                tokens.append(w[i : i + 2])
    return tokens


_RE_DATE_PREFIX = re.compile(
    r"^\s*"
    r"(20\d{2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日"
    r"(?:\s*(?:上午|下午|中午|晚上|早上|凌晨|傍晚|夜间|夜晚|[上中下]午))?"
    r"(?:\s*[，,、]\s*|\s+)?"
)
_RE_DATE_PREFIX2 = re.compile(
    r"^\s*"
    r"(20\d{2})[-/\.](\d{1,2})[-/\.](\d{1,2})"
    r"(?:\s*(?:\d{1,2}:\d{2}(?::\d{2})?)?)?"
    r"(?:\s*[，,、]\s*|\s+)?"
)


def clean_log_text(
    row: Dict[str, Any],
    *,
    text_col: str,
    station_col: Optional[str],
    name_col: Optional[str],
) -> str:
    """
    清洗规则（按你的要求）：
    1) 删除前缀日期时间，如“2025年1月2日上午，”或“2024年2月2日”
    2) 读取“所属派出所”“姓名”列的值，在日志中匹配后替换为空
    """
    text = _norm(row.get(text_col))
    if not text:
        return ""

    # 1) 去掉开头日期时间
    text = _RE_DATE_PREFIX.sub("", text)
    text = _RE_DATE_PREFIX2.sub("", text)

    # 2) 去掉姓名/派出所信息
    replace_items: List[str] = []
    if station_col:
        replace_items.append(_norm(row.get(station_col)))
    if name_col:
        replace_items.append(_norm(row.get(name_col)))

    for it in replace_items:
        if not it:
            continue
        # 避免极短字符串导致误删
        if len(it) <= 1:
            continue
        text = text.replace(it, "")

    # 收尾清理：去掉多余空白与开头分隔符
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"^[，,、:：;；\-\s]+", "", text)
    return text.strip()


def build_tfidf_vectors(texts: Sequence[str]) -> List[Dict[str, float]]:
    """
    返回每条文本的 TF-IDF 稀疏向量（已做 L2 归一化）。
    """
    docs_tokens: List[List[str]] = [tokenize(t) for t in texts]
    n = len(docs_tokens)
    if n == 0:
        return []

    df: Dict[str, int] = {}
    for toks in docs_tokens:
        seen = set(toks)
        for tok in seen:
            df[tok] = df.get(tok, 0) + 1

    idf: Dict[str, float] = {}
    for tok, dfi in df.items():
        # 平滑 idf
        idf[tok] = math.log((n + 1.0) / (dfi + 1.0)) + 1.0

    vectors: List[Dict[str, float]] = []
    for toks in docs_tokens:
        tf: Dict[str, int] = {}
        for tok in toks:
            tf[tok] = tf.get(tok, 0) + 1
        vec: Dict[str, float] = {}
        for tok, c in tf.items():
            # sublinear tf
            w = (1.0 + math.log(c)) * idf.get(tok, 0.0)
            if w:
                vec[tok] = w
        # L2 norm
        norm = math.sqrt(sum(v * v for v in vec.values())) or 1.0
        for tok in list(vec.keys()):
            vec[tok] = vec[tok] / norm
        vectors.append(vec)
    return vectors


def cosine_sparse(a: Dict[str, float], b: Dict[str, float]) -> float:
    if not a or not b:
        return 0.0
    # 遍历更小的字典
    if len(a) > len(b):
        a, b = b, a
    s = 0.0
    for k, va in a.items():
        vb = b.get(k)
        if vb is not None:
            s += va * vb
    return float(s)


class UnionFind:
    def __init__(self, n: int) -> None:
        self.parent = list(range(n))
        self.rank = [0] * n

    def find(self, x: int) -> int:
        while self.parent[x] != x:
            self.parent[x] = self.parent[self.parent[x]]
            x = self.parent[x]
        return x

    def union(self, a: int, b: int) -> None:
        ra = self.find(a)
        rb = self.find(b)
        if ra == rb:
            return
        if self.rank[ra] < self.rank[rb]:
            self.parent[ra] = rb
            return
        if self.rank[ra] > self.rank[rb]:
            self.parent[rb] = ra
            return
        self.parent[rb] = ra
        self.rank[ra] += 1


def _parse_threshold(v: str) -> float:
    s = _norm(v)
    if not s:
        return 0.8
    try:
        f = float(s)
    except Exception:
        raise ValueError(f"无法解析重复度阈值：{v}")
    # 兼容 80 表示 80%
    if f > 1.0:
        f = f / 100.0
    if f < 0:
        f = 0.0
    if f > 1:
        f = 1.0
    return float(f)


def build_candidate_buckets(vectors: List[Dict[str, float]], *, topk: int = 8, keyk: int = 3) -> Dict[Tuple[str, ...], List[int]]:
    buckets: Dict[Tuple[str, ...], List[int]] = {}
    for i, vec in enumerate(vectors):
        if not vec:
            continue
        top = sorted(vec.items(), key=lambda x: x[1], reverse=True)[: max(1, int(topk))]
        key = tuple(k for k, _ in top[: max(1, int(keyk))])
        if not key:
            continue
        buckets.setdefault(key, []).append(i)
    return buckets


def find_duplicate_components(
    texts: Sequence[str],
    *,
    threshold: float,
    group_full_scan_max: int = 120,
) -> Tuple[List[List[int]], List[float]]:
    """
    返回：
    - components: 每个重复簇包含的文档下标（簇大小>=2 且簇内存在>=threshold边）
    - comp_score: 每个簇的“重复度”（簇内最大相似度）
    """
    n = len(texts)
    if n < 2:
        return [], []

    vectors = build_tfidf_vectors(texts)
    uf = UnionFind(n)
    max_edge = [0.0] * n  # 每个点参与的最大相似度（用于后续簇分数）

    # 候选对生成：小组全量扫描，大组用 bucket 降维
    candidate_pairs: List[Tuple[int, int]] = []
    if n <= int(group_full_scan_max):
        for i in range(n):
            for j in range(i + 1, n):
                candidate_pairs.append((i, j))
    else:
        buckets = build_candidate_buckets(vectors, topk=8, keyk=3)
        for idxs in buckets.values():
            if len(idxs) < 2:
                continue
            # bucket 内做全量对比
            for a in range(len(idxs)):
                for b in range(a + 1, len(idxs)):
                    i = idxs[a]
                    j = idxs[b]
                    candidate_pairs.append((i, j))

    # 去重候选对（大组时不同 bucket 可能重复）
    if len(candidate_pairs) > 1:
        seen = set()
        uniq: List[Tuple[int, int]] = []
        for i, j in candidate_pairs:
            key = (i, j) if i < j else (j, i)
            if key in seen:
                continue
            seen.add(key)
            uniq.append(key)
        candidate_pairs = uniq

    for i, j in candidate_pairs:
        sim = cosine_sparse(vectors[i], vectors[j])
        if sim >= float(threshold):
            uf.union(i, j)
            if sim > max_edge[i]:
                max_edge[i] = sim
            if sim > max_edge[j]:
                max_edge[j] = sim

    comps: Dict[int, List[int]] = {}
    for i in range(n):
        r = uf.find(i)
        comps.setdefault(r, []).append(i)

    out_comps: List[List[int]] = []
    out_scores: List[float] = []
    for _, members in comps.items():
        if len(members) < 2:
            continue
        # 簇分数：簇内最大相似度
        score = 0.0
        for m in members:
            if max_edge[m] > score:
                score = max_edge[m]
        if score >= float(threshold):
            out_comps.append(sorted(members))
            out_scores.append(float(score))

    # 按重复度降序、簇大小降序
    paired = list(zip(out_comps, out_scores))
    paired.sort(key=lambda x: (x[1], len(x[0])), reverse=True)
    return [p[0] for p in paired], [p[1] for p in paired]


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="输入文件（xlsx/csv）")
    parser.add_argument("--sheet", default="", help="xlsx sheet（默认第一个）")
    parser.add_argument("--id-col", default=DEFAULT_ID_COL)
    parser.add_argument("--text-col", default=DEFAULT_TEXT_COL)
    parser.add_argument("--station-col", default=DEFAULT_STATION_COL)
    parser.add_argument("--name-col", default=DEFAULT_NAME_COL)
    parser.add_argument("--threshold", default="0.8", help="重复度阈值：0-1 或 0-100(百分比)")
    parser.add_argument("--group-full-scan-max", type=int, default=120, help="组内全量两两对比的最大组大小")
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args()
    quiet = bool(args.quiet)

    threshold = _parse_threshold(args.threshold)

    log("步骤：读取数据", quiet=quiet)
    rows = load_rows(args.input, sheet=args.sheet)
    if not rows:
        raise SystemExit("输入数据为空")

    id_col = resolve_col_name(rows[0], args.id_col)
    text_col = resolve_col_name(rows[0], args.text_col)
    station_col = resolve_col_name(rows[0], args.station_col)
    name_col = resolve_col_name(rows[0], args.name_col)
    # 若列不存在则跳过替换（保持兼容）
    if station_col not in rows[0]:
        station_col = None
    if name_col not in rows[0]:
        name_col = None

    log(
        f"字段：证件号码={id_col}，日志={text_col}，所属派出所={station_col or '未找到'}，姓名={name_col or '未找到'}",
        quiet=quiet,
    )

    # 分组（保留原顺序）
    groups: Dict[str, List[int]] = {}
    for i, r in enumerate(rows):
        gid = _norm(r.get(id_col))
        if not gid:
            continue
        groups.setdefault(gid, []).append(i)

    log(f"步骤：分组完成（组数={len(groups)}）", quiet=quiet)

    # 输出：仅命中重复簇的行
    out_rows: List[Dict[str, Any]] = []
    serial = 1

    for gid, idxs in groups.items():
        if len(idxs) < 2:
            continue
        texts = [
            clean_log_text(rows[i], text_col=text_col, station_col=station_col, name_col=name_col) for i in idxs
        ]
        comps, scores = find_duplicate_components(
            texts, threshold=threshold, group_full_scan_max=int(args.group_full_scan_max)
        )
        if not comps:
            continue

        for comp, score in zip(comps, scores):
            for local_i in comp:
                src_idx = idxs[local_i]
                r = dict(rows[src_idx])
                r["序号"] = serial
                r["重复度"] = f"{score * 100:.2f}%"
                out_rows.append(r)
            serial += 1

    if not out_rows:
        raise SystemExit("未发现超过阈值的重复日志")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = os.path.join(os.getcwd(), f"gongzuorizhi_chongfudu_{ts}.csv")

    # 写 CSV：序号、重复度放前面，其余列按出现顺序
    keys: List[str] = ["序号", "重复度"]
    seen = set(keys)
    for r in out_rows:
        for k in r.keys():
            if k in seen:
                continue
            seen.add(k)
            keys.append(k)

    log(f"步骤：导出CSV -> {out_path}（行数={len(out_rows)}）", quiet=quiet)
    with open(out_path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        for r in out_rows:
            w.writerow(r)

    print(f"已导出：{out_path}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("Interrupted", file=sys.stderr)
        raise
