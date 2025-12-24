"""巡防统计模块路由。负责处理巡防统计相关的 HTTP 请求，并在请求前增加 IP 访问控制。"""
from flask import Blueprint, render_template, request, jsonify, send_file, abort
from xunfang.service.xunfang_service import calculate_xunfang_for_date_range, export_online_rate_for_date_range, export_quadrant_chart_for_date_range, export_police_force_for_date_range
from gonggong.utils.error_handler import handle_errors, log_info, log_error, log_warning
from flask import session as flask_session, redirect, url_for
from gonggong.config.database import get_database_connection

from io import BytesIO
import requests
import json
import csv
from openpyxl import Workbook
from datetime import datetime
import tempfile
import os
import zipfile
from urllib.parse import unquote_plus, parse_qs

# 创建蓝图
xunfang_bp = Blueprint('xunfang', __name__, template_folder='../templates')


@xunfang_bp.before_request
def _ensure_access() -> None:
    """请求前拦截，基于 IP 做巡防模块的访问控制。"""
    if not flask_session.get('username'):
        return redirect(url_for('login'))
    try:
        conn = get_database_connection()
        with conn.cursor() as cur:
            cur.execute('SELECT 1 FROM "ywdata"."jcgkzx_permission" WHERE username=%s AND module=%s', (flask_session['username'], '巡防'))
            row = cur.fetchone()
        conn.close()
        if not row:
            abort(403)
    except Exception:
        abort(500)

@xunfang_bp.route('/')
def xunfang():
    return render_template('xunfang.html')

@xunfang_bp.route('/calculate', methods=['POST'])
def calculate_xunfang():
    data = request.json
    start_time = data.get('startTime')
    end_time = data.get('endTime')
    
    if not start_time or not end_time:
        return jsonify({
            'success': False,
            'message': '寮€濮嬫椂闂村拰缁撴潫鏃堕棿涓嶈兘涓虹┖'
        })
    
    try:
        result = calculate_xunfang_for_date_range(start_time, end_time)
        
        return jsonify({
            'success': True,
            'message': f'璁＄畻瀹屾垚锛?,
            'success_count': result['success_count'],
            'total_tasks': result['total_tasks']
        })
        
    except Exception as e:
        print(f"宸￠槻缁熻璁＄畻杩囩▼涓彂鐢熼敊璇? {e}")
        return jsonify({
            'success': False,
            'message': f'璁＄畻杩囩▼涓彂鐢熼敊璇? {str(e)}'
        })

@xunfang_bp.route('/export_online_rate', methods=['POST'])
def export_online_rate():
    data = request.json
    start_time = data.get('startTime')
    end_time = data.get('endTime')
    
    if not start_time or not end_time:
        return jsonify({
            'success': False,
            'message': '寮€濮嬫椂闂村拰缁撴潫鏃堕棿涓嶈兘涓虹┖'
        }), 400
    
    try:
        wb = export_online_rate_for_date_range(start_time, end_time)
        
        # 灏咵xcel鏂囦欢淇濆瓨鍒板唴瀛樹腑骞惰繑鍥?        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 杩斿洖Excel鏂囦欢
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name=f'{start_time}鑷硔end_time}鍏ㄥ競鍦ㄥ矖鍦ㄧ嚎鐜?xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"瀵煎嚭鍦ㄧ嚎鍦ㄥ矖鏁版嵁杩囩▼涓彂鐢熼敊璇? {e}")
        return jsonify({
            'success': False,
            'message': f'瀵煎嚭杩囩▼涓彂鐢熼敊璇? {str(e)}'
        }), 500

@xunfang_bp.route('/export_quadrant_chart', methods=['POST'])
def export_quadrant_chart():
    data = request.json
    start_time = data.get('startTime')
    end_time = data.get('endTime')

    if not start_time or not end_time:
        return jsonify({
            'success': False,
            'message': '寮€濮嬫椂闂村拰缁撴潫鏃堕棿涓嶈兘涓虹┖'
        }), 400

    try:
        # 浣跨敤matplotlib瀵煎嚭璞￠檺鍥?        image_buffer, wb = export_quadrant_chart_for_date_range(start_time, end_time)

        # 鍒涘缓ZIP鏂囦欢锛屽寘鍚玃NG鍥剧墖鍜孍xcel鏂囦欢
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 娣诲姞PNG鍥剧墖鍒癦IP
            if image_buffer:
                image_buffer.seek(0)
                zip_file.writestr(f'{start_time}鑷硔end_time}娲惧嚭鎵€宸￠槻璀︽儏璞￠檺鍥?png', image_buffer.read())

            # 娣诲姞Excel鏂囦欢鍒癦IP
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            zip_file.writestr(f'{start_time}鑷硔end_time}娲惧嚭鎵€宸￠槻璀︽儏璞￠檺鍥?xlsx', excel_buffer.read())

        zip_buffer.seek(0)

        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f'{start_time}鑷硔end_time}娲惧嚭鎵€宸￠槻璀︽儏璞￠檺鍥?zip',
            mimetype='application/zip'
        )

    except Exception as e:
        print(f"瀵煎嚭璞￠檺鍥捐繃绋嬩腑鍙戠敓閿欒: {e}")
        return jsonify({
            'success': False,
            'message': f'瀵煎嚭杩囩▼涓彂鐢熼敊璇? {str(e)}'
        }), 500

@xunfang_bp.route('/export_police_force', methods=['POST'])
def export_police_force():
    data = request.json
    start_time = data.get('startTime')
    end_time = data.get('endTime')

    if not start_time or not end_time:
        return jsonify({
            'success': False,
            'message': '寮€濮嬫椂闂村拰缁撴潫鏃堕棿涓嶈兘涓虹┖'
        }), 400

    try:
        wb = export_police_force_for_date_range(start_time, end_time)

        # 灏咵xcel鏂囦欢淇濆瓨鍒板唴瀛樹腑骞惰繑鍥?        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 杩斿洖Excel鏂囦欢
        return send_file(
            output,
            as_attachment=True,
            download_name=f'{start_time}鑷硔end_time}宸￠槻璀﹀姏琛?xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        print(f"瀵煎嚭璀﹀姏琛ㄨ繃绋嬩腑鍙戠敓閿欒: {e}")
        return jsonify({
            'success': False,
            'message': f'瀵煎嚭杩囩▼涓彂鐢熼敊璇? {str(e)}'
        }), 500


@xunfang_bp.route('/request_data', methods=['POST'])
@handle_errors("璇锋眰鏁版嵁")
def request_data():
    data = request.json
    method = data.get('method', 'GET')
    url = data.get('url')
    params = data.get('params', {})
    export_format = data.get('exportFormat', 'csv')

    if not url:
        return jsonify({
            'success': False,
            'message': '璇锋眰閾炬帴涓嶈兘涓虹┖'
        }), 400

    log_info(f"寮€濮嬪鐞唟method.upper()}璇锋眰: {url}")

    # 瀵煎叆session_manager
    from gonggong.service.session_manager import session_manager

    # 璁剧疆榛樿璇锋眰澶?    headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.6261.95 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'Accept-Language': 'zh-CN,zh;q=0.9',
            'Connection': 'keep-alive'
        }

        # 鏍规嵁璇锋眰鏂规硶澶勭悊鍙傛暟鍜岃姹?    try:
        log_info(f"寮€濮嬪鐞唟method.upper()}璇锋眰")
        log_info(f"璇锋眰URL: {url}")
        log_info(f"璇锋眰鍙傛暟绫诲瀷: {type(params)}")
        log_info(f"璇锋眰鍙傛暟鍐呭: {params}")

        # 妫€鏌ュ弬鏁版槸鍚︿负绌哄瓧绗︿覆
        if isinstance(params, str) and params.strip() == '':
            params = {}
            log_info("妫€娴嬪埌绌哄瓧绗︿覆鍙傛暟锛岃浆鎹负绌哄瓧鍏?)

        if method.upper() == 'GET':
            # 瀵逛簬GET璇锋眰锛屽鏋淯RL涓凡缁忓寘鍚弬鏁帮紝闇€瑕佽В鏋愬嚭鏉?            from urllib.parse import urlparse, parse_qs

            parsed_url = urlparse(url)
            base_url = f"{parsed_url.scheme}://{parsed_url.netloc}{parsed_url.path}"

            # 瑙ｆ瀽URL涓殑鏌ヨ鍙傛暟
            url_params = {}
            if parsed_url.query:
                url_params = {k: v[0] if v and len(v) == 1 else v for k, v in parse_qs(parsed_url.query).items()}

            # 鍚堝苟URL鍙傛暟鍜岀敤鎴蜂紶鍏ョ殑鍙傛暟
            if isinstance(params, dict):
                merged_params = {**url_params, **params}
            else:
                merged_params = url_params

            log_info(f"GET璇锋眰URL: {base_url}")
            log_info(f"GET璇锋眰鍙傛暟: {merged_params}")

            # 浣跨敤session_manager鍙戦€佽姹傦紝璁剧疆杈冪煭鐨勮秴鏃舵椂闂?            response = session_manager.make_request('GET', base_url, headers=headers, params=merged_params, timeout=30)

        elif method.upper() == 'POST':
            log_info(f"澶勭悊POST璇锋眰锛屽弬鏁扮被鍨? {type(params)}")

            # 鐜板湪params搴旇濮嬬粓鏄瓧鍏告牸寮?            if isinstance(params, dict):
                log_info(f"鎺ユ敹鍒板瓧鍏告牸寮忓弬鏁? {params}")

                # 鏍规嵁鍙傛暟鍐呭鍒ゆ柇鏄惁闇€瑕佽浆鎹负琛ㄥ崟鏁版嵁
                should_send_as_form = False
                try:
                    # 妫€鏌ユ墍鏈夊€兼槸鍚﹂兘鏄畝鍗曞瓧绗︿覆锛屽鏋滄槸鍒欏彲浠ヨ€冭檻鍙戦€佽〃鍗曟暟鎹?                    all_simple_values = all(isinstance(v, str) for v in params.values())
                    if all_simple_values and len(params) > 0:
                        should_send_as_form = True
                except:
                    should_send_as_form = False

                if should_send_as_form:
                    # 鍙戦€佽〃鍗曟暟鎹牸寮?                    headers['Content-Type'] = 'application/x-www-form-urlencoded; charset=UTF-8'
                    log_info(f"鍙戦€佽〃鍗曟暟鎹牸寮廝OST璇锋眰锛孶RL: {url}")
                    log_info(f"璇锋眰鍙傛暟: {params}")
                    response = session_manager.make_request('POST', url, headers=headers, data=params, timeout=30)
                else:
                    # 鍙戦€丣SON鏍煎紡
                    headers['Content-Type'] = 'application/json'
                    log_info(f"鍙戦€丣SON鏍煎紡POST璇锋眰锛孶RL: {url}")
                    log_info(f"璇锋眰鍙傛暟: {params}")
                    response = session_manager.make_request('POST', url, headers=headers, json=params, timeout=30)
            else:
                # 濡傛灉params涓嶆槸瀛楀吀锛岃褰曢敊璇?                log_error(f"POST璇锋眰鍙傛暟鏍煎紡閿欒锛屾湡鏈涘瓧鍏告牸寮忥紝瀹為檯鏀跺埌: {type(params)} - {params}")
                return jsonify({
                    'success': False,
                    'message': f'POST璇锋眰鍙傛暟鏍煎紡閿欒锛屾湡鏈涘璞℃牸寮忥紝瀹為檯鏀跺埌: {type(params).__name__}'
                }), 400
        else:
            return jsonify({
                'success': False,
                'message': f'涓嶆敮鎸佺殑璇锋眰鏂规硶: {method}'
            }), 400

        log_info(f"鍝嶅簲鐘舵€佺爜: {response.status_code}")

        if response.status_code != 200:
            return jsonify({
                'success': False,
                'message': f'璇锋眰澶辫触锛岀姸鎬佺爜: {response.status_code}'
            }), 400

        # 瑙ｆ瀽鍝嶅簲鏁版嵁
        try:
            json_data = response.json()
            log_info(f"鎴愬姛瑙ｆ瀽JSON鏁版嵁锛屾暟鎹被鍨? {type(json_data)}")
        except ValueError as e:
            log_info(f"JSON瑙ｆ瀽澶辫触: {e}")
            log_info(f"鍝嶅簲鍐呭: {response.text[:500]}...")  # 鎵撳嵃鍓?00涓瓧绗?            return jsonify({
                'success': False,
                'message': '鍝嶅簲鏁版嵁涓嶆槸鏈夋晥鐨凧SON鏍煎紡'
            }), 400

        # 澶勭悊宓屽JSON鏁版嵁锛屾壘鍒板疄闄呯殑鏁版嵁閮ㄥ垎
        flattened_data = extract_data_from_nested_json(json_data)

        if not flattened_data:
            return jsonify({
                'success': False,
                'message': '鏈壘鍒版湁鏁堢殑鏁版嵁'
            }), 400

        # 纭繚鏁版嵁鏄垪琛ㄦ牸寮?        if not isinstance(flattened_data, list):
            if isinstance(flattened_data, dict):
                flattened_data = [flattened_data]  # 杞崲涓哄垪琛?            else:
                return jsonify({
                    'success': False,
                    'message': '鏁版嵁鏍煎紡涓嶆纭紝搴斾负瀵硅薄鎴栧璞℃暟缁?
                }), 400

        log_info(f"澶勭悊瀹屾垚锛屾暟鎹潯鏁? {len(flattened_data)}")

        return jsonify({
            'success': True,
            'data': flattened_data,
            'count': len(flattened_data),
            'message': '璇锋眰鎴愬姛'
        })

    except Exception as e:
        log_error(f"璇锋眰杩囩▼涓彂鐢熼敊璇? {e}")
        return jsonify({
            'success': False,
            'message': f'璇锋眰杩囩▼涓彂鐢熼敊璇? {str(e)}'
        }), 500


def parse_form_data_string(form_string):
    """
    瑙ｆ瀽琛ㄥ崟鏁版嵁瀛楃涓蹭负瀛楀吀
    鏀寔URL缂栫爜鐨勮〃鍗曟暟鎹紝濡傦細key1=value1&key2=value2
    鏀寔宓屽鍙傛暟锛屽锛歱arams[beginTime]=2025-10-02 00:00:00
    """
    try:
        # 瑙ｆ瀽URL缂栫爜鐨勮〃鍗曟暟鎹?        parsed_data = {}

        print(f"鍘熷琛ㄥ崟鏁版嵁: {form_string}")

        # 濡傛灉瀛楃涓插寘鍚玌RL缂栫爜瀛楃锛屽厛瑙ｇ爜
        decoded_string = unquote_plus(form_string)
        print(f"瑙ｇ爜鍚庢暟鎹? {decoded_string}")

        # 鎸夌収鍙傛暟鍒嗗壊
        params = decoded_string.split('&')

        for param in params:
            if '=' in param:
                key, value = param.split('=', 1)
                value = value or ''  # 纭繚value涓嶄负None

                # 澶勭悊宓屽鍙傛暟锛屽 params[beginTime]
                if '[' in key and key.endswith(']'):
                    # 澶勭悊宓屽鍙傛暟锛屽 params[beginTime]
                    base_key = key.split('[')[0]
                    nested_key = key.split('[')[1].rstrip(']')

                    if base_key not in parsed_data:
                        parsed_data[base_key] = {}

                    parsed_data[base_key][nested_key] = value
                    print(f"澶勭悊宓屽鍙傛暟: {base_key}[{nested_key}] = {value}")
                else:
                    # 鏅€氬弬鏁?                    # 濡傛灉value涓虹┖瀛楃涓诧紝淇濇寔涓虹┖瀛楃涓?                    parsed_data[key] = value
                    print(f"澶勭悊鏅€氬弬鏁? {key} = {value}")

        print(f"瑙ｆ瀽琛ㄥ崟鏁版嵁鎴愬姛: {parsed_data}")
        return parsed_data

    except Exception as e:
        print(f"瑙ｆ瀽琛ㄥ崟鏁版嵁澶辫触: {e}")
        import traceback
        print(f"閿欒璇︽儏: {traceback.format_exc()}")
        # 濡傛灉瑙ｆ瀽澶辫触锛岃繑鍥炶В鐮佸悗鐨勫師濮嬪瓧绗︿覆
        try:
            return unquote_plus(form_string)
        except:
            return form_string


def extract_data_from_nested_json(data, max_depth=5):
    """
    鎻愬彇宓屽JSON涓殑瀹為檯鏁版嵁
    閫掑綊鏌ユ壘鍊间负鍒楄〃鎴栧瓧鍏哥殑瀛楁锛岀洿鍒版壘鍒扮湡姝ｇ殑鏁版嵁
    """
    if max_depth <= 0:
        return data

    # 濡傛灉鏁版嵁鏈韩灏辨槸鍒楄〃锛岀洿鎺ヨ繑鍥?    if isinstance(data, list):
        print(f"鍙戠幇鍒楄〃鏁版嵁锛屽寘鍚?{len(data)} 涓厓绱?)
        return data

    if isinstance(data, dict):
        print(f"澶勭悊瀛楀吀鏁版嵁锛屽寘鍚敭: {list(data.keys())}")

        # 浼樺厛妫€鏌ヤ竴浜涘父瑙佺殑鏁版嵁瀛楁锛岀壒鍒槸閽堝鍒嗛〉鏁版嵁缁撴瀯
        priority_keys = ['rows', 'data', 'list', 'items', 'result', 'results']
        for key in priority_keys:
            if key in data:
                value = data[key]
                if isinstance(value, list):
                    if len(value) > 0:
                        print(f"鍦ㄩ敭 '{key}' 涓壘鍒板垪琛ㄦ暟鎹紝鍖呭惈 {len(value)} 涓厓绱?)
                        return value
                    else:
                        print(f"閿?'{key}' 鏄┖鍒楄〃锛岀户缁煡鎵惧叾浠栨暟鎹?)
                elif isinstance(value, dict):
                    nested_result = extract_data_from_nested_json(value, max_depth - 1)
                    if isinstance(nested_result, list) and len(nested_result) > 0:
                        return nested_result

        # 妫€鏌ユ墍鏈夊瓧娈碉紝瀵绘壘鍒楄〃鏁版嵁
        for key, value in data.items():
            if isinstance(value, list):
                if len(value) > 0:
                    print(f"鍦ㄩ敭 '{key}' 涓壘鍒板垪琛ㄦ暟鎹紝鍖呭惈 {len(value)} 涓厓绱?)
                    return value
                else:
                    print(f"閿?'{key}' 鏄┖鍒楄〃锛岀户缁煡鎵?)
            elif isinstance(value, dict):
                # 濡傛灉鍊兼槸瀛楀吀锛岄€掑綊妫€鏌?                nested_result = extract_data_from_nested_json(value, max_depth - 1)
                if isinstance(nested_result, list) and len(nested_result) > 0:
                    return nested_result
                elif nested_result != value:  # 濡傛灉閫掑綊瀵艰嚧浜嗘暟鎹敼鍙橈紝杩斿洖缁撴灉
                    return nested_result

        # 濡傛灉娌℃湁鎵惧埌鍒楄〃锛屼絾瀛楀吀鏈韩鍖呭惈鏈夋剰涔夌殑鏁版嵁锛岃繑鍥炲瓧鍏告湰韬?        if len(data) > 0:
            print(f"娌℃湁鎵惧埌鍒楄〃鏁版嵁锛岃繑鍥炲瓧鍏告湰韬紝鍖呭惈 {len(data)} 涓瓧娈?)
            return data

        # 濡傛灉娌℃湁鎵惧埌鍒楄〃锛岃繑鍥炲師濮嬫暟鎹?        return data
    else:
        # 鍏朵粬鎯呭喌杩斿洖鍘熷鏁版嵁
        print(f"鏁版嵁绫诲瀷涓嶆槸瀛楀吀鎴栧垪琛紝杩斿洖鍘熷鏁版嵁: {type(data)}")
        return data


@xunfang_bp.route('/download_result', methods=['POST'])
def download_result():
    try:
        data = request.form.get('data')
        if not data:
            return jsonify({
                'success': False,
                'message': '娌℃湁鎺ユ敹鍒版暟鎹?
            }), 400
        
        data_dict = json.loads(data)
        result_data = data_dict.get('data', [])
        export_format = data_dict.get('exportFormat', 'csv')
        filename = data_dict.get('filename', 'data')
        
        if export_format.lower() == 'csv':
            return download_csv(result_data, filename)
        elif export_format.lower() == 'xlsx':
            return download_xlsx(result_data, filename)
        else:
            return jsonify({
                'success': False,
                'message': f'涓嶆敮鎸佺殑瀵煎嚭鏍煎紡: {export_format}'
            }), 400
    
    except Exception as e:
        print(f"涓嬭浇缁撴灉杩囩▼涓彂鐢熼敊璇? {e}")
        return jsonify({
            'success': False,
            'message': f'涓嬭浇缁撴灉杩囩▼涓彂鐢熼敊璇? {str(e)}'
        }), 500


def download_csv(data, filename):
    """涓嬭浇CSV鏂囦欢"""
    try:
        # 鍒涘缓涓存椂鏂囦欢
        temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.csv', encoding='utf-8-sig')
        
        if data:
            # 鑾峰彇鎵€鏈夐敭浣滀负CSV鏍囬
            all_keys = set()
            for item in data:
                if isinstance(item, dict):
                    all_keys.update(item.keys())
            
            # 鍐欏叆CSV
            writer = csv.DictWriter(temp_file, fieldnames=sorted(list(all_keys)))
            writer.writeheader()
            for item in data:
                if isinstance(item, dict):
                    writer.writerow(item)
        
        temp_file.close()
        
        # 杩斿洖鏂囦欢
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'{filename}.csv',
            mimetype='text/csv'
        )
        
    except Exception as e:
        print(f"鐢熸垚CSV鏂囦欢鏃跺彂鐢熼敊璇? {e}")
        raise e
    finally:
        # 娓呯悊涓存椂鏂囦欢
        try:
            os.unlink(temp_file.name)
        except:
            pass


def download_xlsx(data, filename):
    """涓嬭浇XLSX鏂囦欢"""
    try:
        # 鍒涘缓宸ヤ綔绨垮拰宸ヤ綔琛?        wb = Workbook()
        ws = wb.active
        ws.title = "璇锋眰鏁版嵁"
        
        if data:
            # 鑾峰彇鎵€鏈夐敭浣滀负鍒楁爣棰?            all_keys = set()
            for item in data:
                if isinstance(item, dict):
                    all_keys.update(item.keys())
            
            # 鍐欏叆鍒楁爣棰?            headers = sorted(list(all_keys))
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)
            
            # 鍐欏叆鏁版嵁琛?            for row_num, item in enumerate(data, 2):
                if isinstance(item, dict):
                    for col_num, header in enumerate(headers, 1):
                        value = item.get(header, "")
                        ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
        
        # 淇濆瓨鍒板唴瀛樹腑
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 杩斿洖鏂囦欢
        return send_file(
            output,
            as_attachment=True,
            download_name=f'{filename}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    except Exception as e:
        print(f"鐢熸垚XLSX鏂囦欢鏃跺彂鐢熼敊璇? {e}")
        raise e


