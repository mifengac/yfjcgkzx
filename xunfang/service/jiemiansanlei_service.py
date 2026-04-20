from __future__ import annotations

import json
import os
import threading
from dataclasses import dataclass
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Literal, Optional, Sequence, Tuple

from openpyxl import Workbook

from gonggong.service.upstream_jingqing_client import api_client
from xunfang.dao.jiemiansanlei_dao import (
    JiemianSanleiDbQuery,
    fetch_db_jingqings,
    get_case_type_code_map,
    list_case_types,
)


SourceType = Literal["原始", "确认"]
ExportFormat = Literal["xlsx", "xls"]
ReportBureau = Literal["云城分局", "云安分局", "罗定市公安局", "新兴县公安局", "郁南县公安局", "ALL"]
StreetFilterMode = Literal[
    "none",
    "model",
    "content_road",
    "content_public",
    "reply_road",
    "reply_public",
    "text_any",
]

REPORT_LEIXING_LIST = ["人身伤害类", "侵犯财产类", "扰乱秩序类"]
MINOR_CASE_MARK_NO = "01020201,0102020101,0102020102,0102020103"
STREET_LABEL = "街面与公共区域"
STREET_FILTER_MODE_DEFAULT: StreetFilterMode = "model"
STREET_FILTER_MODES = {
    "none",
    "model",
    "content_road",
    "content_public",
    "reply_road",
    "reply_public",
    "text_any",
}
_ROAD_KEYWORDS = (
    "街面",
    "路面",
    "路边",
    "路口",
    "路段",
    "道路",
    "马路",
    "街道",
    "人行道",
    "斑马线",
    "公路",
    "大道",
    "桥",
)
_PUBLIC_KEYWORDS = (
    "广场",
    "公园",
    "市场",
    "商场",
    "超市",
    "车站",
    "公交站",
    "公共场所",
    "门口",
    "现场",
)
_STREET_FILTER_LABELS = {
    "none": "不限街面",
    "model": "街面(模型)",
    "content_road": "街面(报警内容-路面)",
    "content_public": "街面(报警内容-公共)",
    "reply_road": "街面(处警-路面)",
    "reply_public": "街面(处警-公共)",
    "text_any": "街面(综合关键字)",
}
_STREET_FIELD_LABELS = {
    "case_contents": "报警内容",
    "replies": "处警情况",
}
_STREET_KEYWORD_RULES: Dict[str, Tuple[Tuple[str, ...], Tuple[str, ...]]] = {
    "content_road": (("case_contents",), _ROAD_KEYWORDS),
    "content_public": (("case_contents",), _PUBLIC_KEYWORDS),
    "reply_road": (("replies",), _ROAD_KEYWORDS),
    "reply_public": (("replies",), _PUBLIC_KEYWORDS),
    "text_any": (("case_contents", "replies"), _ROAD_KEYWORDS + _PUBLIC_KEYWORDS),
}
_PAGE_SIZE = 5000

_MODEL_LOCK = threading.Lock()
_MODEL_BUNDLE: Optional["ModelBundle"] = None


@dataclass(frozen=True)
class ModelBundle:
    tokenizer: Any
    model: Any
    id2label: Dict[int, str]
    device: str


def get_case_types() -> List[str]:
    return list_case_types()


def query_classified(
    *,
    start_time: str,
    end_time: str,
    leixing_list: Sequence[str],
    source_list: Sequence[SourceType],
    page: int,
    page_size: Optional[int],
    street_only: bool = False,
    street_filter_mode: StreetFilterMode = STREET_FILTER_MODE_DEFAULT,
    minor_only: bool = False,
) -> Dict[str, Any]:
    rows = _fetch_rows_for_filters(
        start_time=start_time,
        end_time=end_time,
        leixing_list=leixing_list,
        source_list=source_list,
        minor_only=minor_only,
    )
    _append_predictions(rows)

    effective_street_mode = _resolve_street_filter_mode(street_filter_mode, street_only=street_only)
    rows = _filter_street_rows(rows, effective_street_mode)

    total = len(rows)
    current_page = 1 if page_size is None else max(1, int(page or 1))
    page_rows = _paginate_rows(rows, page=current_page, page_size=page_size)
    return {
        "total": total,
        "page": current_page,
        "page_size": page_size,
        "street_filter": get_street_filter_description(effective_street_mode),
        "rows": _serialize_rows(page_rows),
    }


def get_street_filter_description(mode: Any, *, street_only: bool = True) -> Dict[str, Any]:
    effective_mode = _resolve_street_filter_mode(mode, street_only=street_only)
    label = _STREET_FILTER_LABELS.get(effective_mode, _STREET_FILTER_LABELS[STREET_FILTER_MODE_DEFAULT])

    if effective_mode == "none":
        return {
            "mode": effective_mode,
            "label": label,
            "fields": [],
            "keywords": [],
            "description": "当前未启用街面过滤，统计结果包含所选警情性质和口径下的全部数据。",
        }

    if effective_mode == "model":
        return {
            "mode": effective_mode,
            "label": label,
            "fields": ["警情地址"],
            "keywords": [STREET_LABEL],
            "description": f"当前按警情地址模型分类结果过滤，分类结果为“{STREET_LABEL}”。",
        }

    fields, keywords = _STREET_KEYWORD_RULES.get(effective_mode, ((), ()))
    field_labels = [_STREET_FIELD_LABELS.get(field, field) for field in fields]
    keyword_list = list(keywords)
    return {
        "mode": effective_mode,
        "label": label,
        "fields": field_labels,
        "keywords": keyword_list,
        "description": f"当前按{'、'.join(field_labels)}字段过滤，关键字：{'、'.join(keyword_list)}。",
    }


def export_classified(
    *,
    start_time: str,
    end_time: str,
    leixing_list: Sequence[str],
    source_list: Sequence[SourceType],
    fmt: ExportFormat,
    street_only: bool = False,
    street_filter_mode: StreetFilterMode = STREET_FILTER_MODE_DEFAULT,
    minor_only: bool = False,
) -> Tuple[bytes, str, str]:
    rows = _fetch_rows_for_filters(
        start_time=start_time,
        end_time=end_time,
        leixing_list=leixing_list,
        source_list=source_list,
        minor_only=minor_only,
    )
    _append_predictions(rows)
    effective_street_mode = _resolve_street_filter_mode(street_filter_mode, street_only=street_only)
    rows = _filter_street_rows(rows, effective_street_mode)

    grouped_rows: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
    for row in rows:
        key = (str(row.get("source") or ""), str(row.get("leixing") or ""))
        grouped_rows.setdefault(key, []).append(row)

    combos: List[Tuple[str, str]] = []
    for source in _normalize_source_list(source_list):
        for leixing in _normalize_leixing_list(leixing_list):
            combos.append((source, leixing))

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"街面三类警情地址分类{timestamp}.{fmt}"

    if fmt == "xlsx":
        bio = BytesIO()
        wb = _build_xlsx_workbook(combos, grouped_rows)
        wb.save(bio)
        bio.seek(0)
        return (
            bio.read(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename,
        )

    xls_bytes = _build_xls_bytes(combos, grouped_rows)
    return xls_bytes, "application/vnd.ms-excel", filename


def export_report(
    *,
    start_time: str,
    end_time: str,
    hb_start_time: str,
    hb_end_time: str,
    street_filter_mode: StreetFilterMode = STREET_FILTER_MODE_DEFAULT,
) -> Tuple[bytes, str, str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"缺少依赖 openpyxl，无法导出报表：{exc}") from exc

    current_start = _parse_dt(start_time)
    current_end = _parse_dt(end_time)
    hb_start = _parse_dt(hb_start_time)
    hb_end = _parse_dt(hb_end_time)
    effective_street_mode = _resolve_street_filter_mode(street_filter_mode, street_only=True)

    if current_start >= current_end:
        raise ValueError("开始时间必须早于结束时间")
    if hb_start >= hb_end:
        raise ValueError("环比开始必须早于环比结束")

    yoy_current_start = _shift_year(current_start, -1)
    yoy_current_end = _shift_year(current_end, -1)

    ytd_start = datetime(current_end.year, 1, 1, 0, 0, 0)
    x_days = (current_end - ytd_start).days
    hb_ytd_start = ytd_start - timedelta(days=x_days)
    hb_ytd_end = ytd_start
    yoy_ytd_start = _shift_year(ytd_start, -1)
    yoy_ytd_end = _shift_year(current_end, -1)

    year_window_start = min(current_start, hb_start, ytd_start, hb_ytd_start)
    year_window_end = current_end
    last_year_window_start = yoy_ytd_start
    last_year_window_end = yoy_ytd_end

    rows_year = [
        _normalize_db_report_row(row)
        for row in fetch_db_jingqings(
            JiemianSanleiDbQuery(
                start_time=_format_dt(year_window_start),
                end_time=_format_dt(year_window_end),
                leixing_list=REPORT_LEIXING_LIST,
                source_list=["原始", "确认"],
                minor_only=False,
                limit=None,
                offset=0,
            )
        )
    ]
    if effective_street_mode == "model":
        _append_predictions(rows_year)

    rows_last_year = [
        _normalize_db_report_row(row)
        for row in fetch_db_jingqings(
            JiemianSanleiDbQuery(
                start_time=_format_dt(last_year_window_start),
                end_time=_format_dt(last_year_window_end),
                leixing_list=REPORT_LEIXING_LIST,
                source_list=["原始", "确认"],
                minor_only=False,
                limit=None,
                offset=0,
            )
        )
    ]
    if effective_street_mode == "model":
        _append_predictions(rows_last_year)

    counts = _build_report_counts(
        rows_year=rows_year,
        rows_last_year=rows_last_year,
        street_filter_mode=effective_street_mode,
        segments_year=[
            ("current", current_start, current_end, "C", "D"),
            ("hb", hb_start, hb_end, "K", "N"),
            ("ytd", ytd_start, current_end, "Q", "R"),
            ("hb_ytd", hb_ytd_start, hb_ytd_end, "Y", "AB"),
        ],
        segments_last_year=[
            ("yoy_current", yoy_current_start, yoy_current_end, "E", "F"),
            ("yoy_ytd", yoy_ytd_start, yoy_ytd_end, "S", "T"),
        ],
    )

    template_path = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "templates", "jiemiansanleijingqing_template.xlsx")
    )
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"未找到报表模板文件：{template_path}")

    wb = load_workbook(template_path)
    time_range_text = f"{_format_zh_date(current_start)}-{_format_zh_date(current_end)}"
    for sheet_name in wb.sheetnames:
        try:
            wb[sheet_name]["A6"].value = time_range_text
        except Exception:
            continue

    expected_sheets = REPORT_LEIXING_LIST + ["三类合计"]
    missing_sheets = [sheet_name for sheet_name in expected_sheets if sheet_name not in wb.sheetnames]
    if missing_sheets:
        raise RuntimeError(f"模板缺少 sheet：{'、'.join(missing_sheets)}")

    for sheet_name in expected_sheets:
        ws = wb[sheet_name]
        for bureau, row_idx in _REPORT_BUREAU_ROW.items():
            for col in _REPORT_COLS:
                ws[f"{col}{row_idx}"].value = int(counts.get((sheet_name, bureau, col), 0))

    filename = f"{_safe_filename_part(start_time)}-{_safe_filename_part(end_time)}_街面三类警情统计表.xlsx"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return (
        bio.read(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename,
    )


def _fetch_rows_for_filters(
    *,
    start_time: str,
    end_time: str,
    leixing_list: Sequence[str],
    source_list: Sequence[SourceType],
    minor_only: bool,
) -> List[Dict[str, Any]]:
    leixing_values = _normalize_leixing_list(leixing_list)
    if not leixing_values:
        return []

    source_values = _normalize_source_list(source_list)
    if not source_values:
        return []

    case_type_code_map = get_case_type_code_map(leixing_values)
    rows: List[Dict[str, Any]] = []
    for source in source_values:
        for leixing in leixing_values:
            codes = case_type_code_map.get(leixing, [])
            combo_rows = _fetch_source_rows(
                start_time=start_time,
                end_time=end_time,
                source=source,
                leixing=leixing,
                code_list=codes,
                minor_only=minor_only,
            )
            rows.extend(combo_rows)
    return rows


def _fetch_source_rows(
    *,
    start_time: str,
    end_time: str,
    source: SourceType,
    leixing: str,
    code_list: Sequence[str],
    minor_only: bool,
) -> List[Dict[str, Any]]:
    code_csv = _build_union_code_csv([code_list])
    if not code_csv:
        return []

    all_rows: List[Dict[str, Any]] = []
    page_num = 1
    total: Optional[int] = None

    while True:
        payload = _build_case_payload(
            start_time=start_time,
            end_time=end_time,
            source=source,
            code_csv=code_csv,
            minor_only=minor_only,
            page_num=page_num,
            page_size=_PAGE_SIZE,
        )
        result = api_client.get_case_list(payload)
        if not isinstance(result, dict):
            raise RuntimeError("case/list 响应格式异常")

        code = result.get("code")
        if code == -1:
            raise RuntimeError("111警情系统登录或取数超时，请检查网络连通性和上游系统状态")
        if code not in (None, 0):
            raise RuntimeError(f"case/list 返回异常，code={code}，msg={result.get('msg', '')}")

        raw_rows = result.get("rows") or []
        if not isinstance(raw_rows, list):
            raise RuntimeError("case/list rows 不是数组")

        if total is None:
            try:
                total = int(result.get("total", 0) or 0)
            except Exception:
                total = 0
        for raw_row in raw_rows:
            if not isinstance(raw_row, dict):
                continue
            all_rows.append(_standardize_case_row(raw_row, source=source, leixing=leixing))

        if not raw_rows:
            break
        if len(raw_rows) < _PAGE_SIZE:
            break
        if total is not None and page_num * _PAGE_SIZE >= total:
            break
        page_num += 1

    return all_rows


def _build_case_payload(
    *,
    start_time: str,
    end_time: str,
    source: SourceType,
    code_csv: str,
    minor_only: bool,
    page_num: int,
    page_size: int,
) -> Dict[str, str]:
    payload = {
        "params[colArray]": "",
        "beginDate": start_time,
        "endDate": end_time,
        "newCaseSourceNo": "",
        "newCaseSource": "全部",
        "dutyDeptNo": "",
        "dutyDeptName": "全部",
        "newCharaSubclassNo": "",
        "newCharaSubclass": "全部",
        "newOriCharaSubclassNo": "",
        "newOriCharaSubclass": "全部",
        "caseNo": "",
        "callerName": "",
        "callerPhone": "",
        "phoneAddress": "",
        "callerIdentity": "",
        "operatorNo": "",
        "operatorName": "",
        "params[isInvalidCase]": "",
        "occurAddress": "",
        "caseMarkNo": MINOR_CASE_MARK_NO if minor_only else "",
        "caseMark": "全部",
        "params[repetitionCase]": "",
        "params[originalDuplicateCase]": "",
        "params[startTimePeriod]": "",
        "params[endTimePeriod]": "",
        "caseContents": "",
        "replies": "",
        "params[sinceRecord]": "",
        "dossierResult": "",
        "params[isVideo]": "",
        "params[isConversation]": "",
        "pageSize": str(page_size),
        "pageNum": str(page_num),
        "orderByColumn": "callTime",
        "isAsc": "desc",
    }

    if source == "原始":
        payload["newOriCharaSubclassNo"] = code_csv
        payload["newOriCharaSubclass"] = ""
    else:
        payload["newCharaSubclassNo"] = code_csv
        payload["newCharaSubclass"] = ""
    return payload


def _standardize_case_row(row: Dict[str, Any], *, source: SourceType, leixing: str) -> Dict[str, Any]:
    return {
        "case_no": _first_non_empty(row, "caseNo", "caseno"),
        "leixing": leixing,
        "source": source,
        "bureau": _normalize_bureau_name(_first_non_empty(row, "cmdName", "cmdname")),
        "station_no": _first_non_empty(row, "dutyDeptNo", "dutydeptno"),
        "station_name": _first_non_empty(row, "dutyDeptName", "dutydeptname"),
        "call_time": _first_non_empty(row, "callTime", "calltime"),
        "address": _first_non_empty(row, "occurAddress", "occuraddress", "address", "caseAddress", "caseaddress"),
        "lng": _first_non_empty(row, "lngOfCriterion", "lngofcriterion"),
        "lat": _first_non_empty(row, "latOfCriterion", "latofcriterion"),
        "case_contents": _first_non_empty(row, "caseContents", "casecontents"),
        "replies": _first_non_empty(row, "replies"),
        "case_type_name": _extract_case_type_name(row, source=source, fallback=leixing),
        "pred_label": "",
        "pred_prob": 0.0,
    }


def _normalize_db_report_row(row: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "case_no": _first_non_empty(row, "case_no", "caseno", "caseNo"),
        "leixing": _first_non_empty(row, "leixing"),
        "source": _first_non_empty(row, "source", "yuanshiqueren"),
        "bureau": _normalize_bureau_name(_first_non_empty(row, "bureau", "分局", "cmdName", "cmdname")),
        "station_no": _first_non_empty(row, "station_no", "派出所编号", "dutyDeptNo", "dutydeptno"),
        "station_name": _first_non_empty(row, "station_name", "派出所名称", "dutyDeptName", "dutydeptname"),
        "call_time": _first_non_empty(row, "call_time", "报警时间", "callTime", "calltime"),
        "address": _first_non_empty(row, "address", "警情地址", "occurAddress", "occuraddress"),
        "lng": _first_non_empty(row, "lng", "经度", "lngOfCriterion", "lngofcriterion"),
        "lat": _first_non_empty(row, "lat", "纬度", "latOfCriterion", "latofcriterion"),
        "case_contents": _first_non_empty(row, "case_contents", "报警内容", "caseContents", "casecontents"),
        "replies": _first_non_empty(row, "replies", "处警情况"),
        "case_type_name": _first_non_empty(
            row,
            "case_type_name",
            "jq_type",
            "newOriCharaSubclass",
            "newCharaSubclass",
        ),
        "pred_label": str(row.get("pred_label") or "").strip(),
        "pred_prob": row.get("pred_prob") or 0.0,
    }


def _extract_case_type_name(row: Dict[str, Any], *, source: SourceType, fallback: str) -> str:
    if source == "原始":
        value = _first_non_empty(
            row,
            "newOriCharaSubclass",
            "newOriCharaSubclassName",
            "newOriCharaSubcategoryName",
        )
    else:
        value = _first_non_empty(
            row,
            "newCharaSubclass",
            "newCharaSubclassName",
            "newCharaSubcategoryName",
        )
    return value or fallback


def _build_union_code_csv(code_groups: Sequence[Sequence[str]]) -> str:
    codes: List[str] = []
    seen = set()
    for group in code_groups:
        for code in group:
            normalized = str(code or "").strip()
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            codes.append(normalized)
    return ",".join(codes)


def _parse_csv_codes(value: Any) -> List[str]:
    if value is None:
        return []
    if isinstance(value, (list, tuple, set)):
        raw_items = value
    else:
        text = str(value).strip()
        if text.startswith("{") and text.endswith("}"):
            text = text[1:-1]
        raw_items = text.split(",") if text else []

    codes: List[str] = []
    seen = set()
    for item in raw_items:
        code = str(item or "").strip().strip('"').strip("'")
        if not code or code in seen:
            continue
        seen.add(code)
        codes.append(code)
    return codes


def _normalize_leixing_list(leixing_list: Sequence[str]) -> List[str]:
    values: List[str] = []
    seen = set()
    for item in leixing_list or []:
        value = str(item or "").strip()
        if not value or value in seen:
            continue
        seen.add(value)
        values.append(value)
    return values


def _normalize_source_list(source_list: Sequence[SourceType]) -> List[SourceType]:
    values: List[SourceType] = []
    seen = set()
    for item in source_list or []:
        value = str(item or "").strip()
        if value not in {"原始", "确认"} or value in seen:
            continue
        seen.add(value)
        values.append(value)  # type: ignore[arg-type]
    return values


def _first_non_empty(row: Dict[str, Any], *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value is None:
            continue
        text = str(value).strip()
        if text:
            return text
    return ""


def _normalize_bureau_name(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    if "云城" in text:
        return "云城分局"
    if "云安" in text:
        return "云安分局"
    if "罗定" in text:
        return "罗定市公安局"
    if "新兴" in text:
        return "新兴县公安局"
    if "郁南" in text:
        return "郁南县公安局"
    return text


def _build_xlsx_workbook(
    combos: Sequence[Tuple[str, str]],
    grouped_rows: Dict[Tuple[str, str], List[Dict[str, Any]]],
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for source, leixing in combos:
        ws = wb.create_sheet(title=_safe_sheet_name(f"{source}{leixing}地址分类"))
        _write_table_xlsx(ws, grouped_rows.get((source, leixing), []))
    return wb


def _build_xls_bytes(
    combos: Sequence[Tuple[str, str]],
    grouped_rows: Dict[Tuple[str, str], List[Dict[str, Any]]],
) -> bytes:
    try:
        import xlwt  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"缺少依赖 xlwt，无法导出 xls：{exc}") from exc

    wb = xlwt.Workbook(encoding="utf-8")
    for source, leixing in combos:
        ws = wb.add_sheet(_safe_sheet_name(f"{source}{leixing}地址分类"))
        _write_table_xls(ws, grouped_rows.get((source, leixing), []))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def _resolve_street_filter_mode(mode: Any, *, street_only: bool) -> StreetFilterMode:
    if not street_only:
        return "none"
    value = str(mode or STREET_FILTER_MODE_DEFAULT).strip()
    if value in STREET_FILTER_MODES:
        return value  # type: ignore[return-value]
    return STREET_FILTER_MODE_DEFAULT


def _filter_street_rows(rows: Sequence[Dict[str, Any]], mode: StreetFilterMode) -> List[Dict[str, Any]]:
    if mode == "none":
        return list(rows)
    return [row for row in rows if _row_matches_street_filter(row, mode)]


def _row_matches_street_filter(row: Dict[str, Any], mode: StreetFilterMode) -> bool:
    if mode == "none":
        return True
    if mode == "model":
        return str(row.get("pred_label") or "").strip() == STREET_LABEL

    fields, keywords = _STREET_KEYWORD_RULES.get(mode, ((), ()))
    for field in fields:
        text = str(row.get(field) or "")
        if any(keyword in text for keyword in keywords):
            return True
    return False


def _paginate_rows(
    rows: Sequence[Dict[str, Any]],
    *,
    page: int,
    page_size: Optional[int],
) -> List[Dict[str, Any]]:
    if page_size is None:
        return list(rows)
    size = max(1, int(page_size))
    current_page = max(1, int(page or 1))
    offset = (current_page - 1) * size
    return list(rows[offset : offset + size])


def _serialize_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [
        {
            "警情性质": row.get("leixing") or "",
            "警情性质口径": row.get("source") or "",
            "分局": row.get("bureau") or "",
            "派出所编号": row.get("station_no") or "",
            "派出所名称": row.get("station_name") or "",
            "报警时间": _format_dt(row.get("call_time")),
            "警情地址": row.get("address") or "",
            "经度": _format_coord(row.get("lng")),
            "纬度": _format_coord(row.get("lat")),
            "警情类型": row.get("case_type_name") or "",
            "分类结果": row.get("pred_label") or "",
            "置信度": _format_prob(row.get("pred_prob")),
        }
        for row in rows
    ]


def _write_table_xlsx(ws: Any, rows: Sequence[Dict[str, Any]]) -> None:
    headers = [
        "分局",
        "派出所编号",
        "派出所名称",
        "报警时间",
        "警情地址",
        "经度",
        "纬度",
        "报警内容",
        "处警情况",
        "警情类型",
        "分类结果",
        "置信度",
    ]
    ws.append(headers)
    for row in rows:
        ws.append(
            [
                row.get("bureau") or "",
                row.get("station_no") or "",
                row.get("station_name") or "",
                _format_dt(row.get("call_time")),
                row.get("address") or "",
                _excel_number_or_blank(row.get("lng")),
                _excel_number_or_blank(row.get("lat")),
                row.get("case_contents") or "",
                row.get("replies") or "",
                row.get("case_type_name") or "",
                row.get("pred_label") or "",
                _format_prob(row.get("pred_prob")),
            ]
        )


def _write_table_xls(ws: Any, rows: Sequence[Dict[str, Any]]) -> None:
    headers = [
        "分局",
        "派出所编号",
        "派出所名称",
        "报警时间",
        "警情地址",
        "经度",
        "纬度",
        "报警内容",
        "处警情况",
        "警情类型",
        "分类结果",
        "置信度",
    ]
    for col, header in enumerate(headers):
        ws.write(0, col, header)

    for row_index, row in enumerate(rows, start=1):
        ws.write(row_index, 0, row.get("bureau") or "")
        ws.write(row_index, 1, row.get("station_no") or "")
        ws.write(row_index, 2, row.get("station_name") or "")
        ws.write(row_index, 3, _format_dt(row.get("call_time")))
        ws.write(row_index, 4, row.get("address") or "")
        ws.write(row_index, 5, _excel_number_or_blank(row.get("lng")))
        ws.write(row_index, 6, _excel_number_or_blank(row.get("lat")))
        ws.write(row_index, 7, row.get("case_contents") or "")
        ws.write(row_index, 8, row.get("replies") or "")
        ws.write(row_index, 9, row.get("case_type_name") or "")
        ws.write(row_index, 10, row.get("pred_label") or "")
        ws.write(row_index, 11, _format_prob(row.get("pred_prob")))


def _append_predictions(rows: List[Dict[str, Any]]) -> None:
    texts = [str(row.get("address") or "").strip() for row in rows]
    predictions = predict_addresses(texts)
    for row, (label, prob) in zip(rows, predictions):
        row["pred_label"] = label
        row["pred_prob"] = prob


def predict_addresses(texts: Sequence[str]) -> List[Tuple[str, float]]:
    bundle = _get_model_bundle()
    if not texts:
        return []

    import torch  # type: ignore
    import torch.nn.functional as F  # type: ignore

    results: List[Tuple[str, float]] = []
    batch_size = 64
    for start in range(0, len(texts), batch_size):
        batch = list(texts[start : start + batch_size])
        encoded = bundle.tokenizer(
            batch,
            padding=True,
            truncation=True,
            max_length=128,
            return_tensors="pt",
        )
        encoded = {key: value.to(bundle.device) for key, value in encoded.items()}
        with torch.no_grad():
            outputs = bundle.model(**encoded)
            probs = F.softmax(outputs.logits, dim=-1)
            best_prob, best_idx = torch.max(probs, dim=-1)

        for text, prob, index in zip(batch, best_prob.tolist(), best_idx.tolist()):
            if not str(text).strip():
                results.append(("", 0.0))
                continue
            label = bundle.id2label.get(int(index), str(index))
            results.append((label, float(prob)))
    return results


def _get_model_bundle() -> ModelBundle:
    global _MODEL_BUNDLE  # noqa: PLW0603
    if _MODEL_BUNDLE is not None:
        return _MODEL_BUNDLE

    with _MODEL_LOCK:
        if _MODEL_BUNDLE is not None:
            return _MODEL_BUNDLE

        model_root = os.path.abspath(
            os.path.join(os.path.dirname(__file__), "..", "..", "gonggong", "5lei_dizhi_model")
        )
        model_dir = os.path.join(model_root, "best_model")
        id2label_path = os.path.join(model_root, "id2label.json")

        try:
            import torch  # type: ignore
            from transformers import AutoModelForSequenceClassification, AutoTokenizer  # type: ignore
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"缺少依赖 torch/transformers，无法加载地址分类模型：{exc}") from exc

        with open(id2label_path, "r", encoding="utf-8") as file:
            raw = json.load(file)
        id2label = {int(key): str(value) for key, value in raw.items()}

        tokenizer = AutoTokenizer.from_pretrained(model_dir)
        model = AutoModelForSequenceClassification.from_pretrained(model_dir)
        model.to("cpu")
        model.eval()

        _MODEL_BUNDLE = ModelBundle(tokenizer=tokenizer, model=model, id2label=id2label, device="cpu")
        return _MODEL_BUNDLE


def _format_dt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _format_coord(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        return f"{float(value):.6f}"
    except Exception:
        return text


def _excel_number_or_blank(value: Any) -> Any:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        return float(value)
    except Exception:
        return text


def _format_prob(value: Any) -> str:
    try:
        return f"{float(value):.5f}"
    except Exception:
        return "0.00000"


_ILLEGAL_SHEET_CHARS = set(r"[]:*?/\\")


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if ch in _ILLEGAL_SHEET_CHARS else ch for ch in (name or "sheet"))
    cleaned = cleaned.strip() or "sheet"
    return cleaned[:31]


def _safe_filename_part(value: str) -> str:
    return (
        str(value or "")
        .strip()
        .replace(":", "-")
        .replace("/", "-")
        .replace("\\", "-")
        .replace(" ", "_")
        .replace("\t", "_")
    )


def _format_zh_date(dt: datetime) -> str:
    return f"{dt.year}年{dt.month}月{dt.day}日"


_REPORT_BUREAU_ROW: Dict[ReportBureau, int] = {
    "云城分局": 6,
    "云安分局": 7,
    "罗定市公安局": 8,
    "新兴县公安局": 9,
    "郁南县公安局": 10,
    "ALL": 11,
}

_REPORT_COLS = ["C", "D", "E", "F", "K", "N", "Q", "R", "S", "T", "Y", "AB"]


def _parse_dt(value: str) -> datetime:
    return datetime.strptime(str(value).strip(), "%Y-%m-%d %H:%M:%S")


def _shift_year(dt: datetime, years: int) -> datetime:
    try:
        return dt.replace(year=dt.year + years)
    except ValueError:
        base = dt.replace(day=1, month=dt.month, year=dt.year + years)
        next_month = base.replace(day=28) + timedelta(days=4)
        last_day = next_month - timedelta(days=next_month.day)
        return dt.replace(year=dt.year + years, day=last_day.day)


def _as_dt(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except Exception:
            continue
    return None


def _build_report_counts(
    *,
    rows_year: Sequence[Dict[str, Any]],
    rows_last_year: Sequence[Dict[str, Any]],
    street_filter_mode: StreetFilterMode,
    segments_year: Sequence[Tuple[str, datetime, datetime, str, str]],
    segments_last_year: Sequence[Tuple[str, datetime, datetime, str, str]],
) -> Dict[Tuple[str, ReportBureau, str], int]:
    target_sources = {"原始", "确认"}
    bureaus = {"云城分局", "云安分局", "罗定市公安局", "新兴县公安局", "郁南县公安局"}
    counts: Dict[Tuple[str, ReportBureau, str], int] = {}
    seen: Dict[Tuple[str, ReportBureau, str], set] = {}

    def increase(sheet_name: str, bureau: ReportBureau, col: str, case_key: str) -> None:
        key = (sheet_name, bureau, col)
        bucket = seen.setdefault(key, set())
        if case_key in bucket:
            return
        bucket.add(case_key)
        counts[key] = counts.get(key, 0) + 1

    def process(rows: Sequence[Dict[str, Any]], segments: Sequence[Tuple[str, datetime, datetime, str, str]]) -> None:
        for index, row in enumerate(rows):
            leixing = str(row.get("leixing") or "").strip()
            source = str(row.get("source") or "").strip()
            if leixing not in REPORT_LEIXING_LIST or source not in target_sources:
                continue
            if not _row_matches_street_filter(row, street_filter_mode):
                continue

            call_time = _as_dt(row.get("call_time"))
            if call_time is None:
                continue

            bureau_raw = str(row.get("bureau") or "").strip()
            bureau_keys: List[ReportBureau] = ["ALL"]
            if bureau_raw in bureaus:
                bureau_keys = [bureau_raw, "ALL"]  # type: ignore[list-item]

            case_key = str(row.get("case_no") or "").strip() or f"__row__{index}"
            for _, start, end, col_orig, col_confirm in segments:
                if call_time < start or call_time >= end:
                    continue
                col = col_orig if source == "原始" else col_confirm
                for sheet_name in (leixing, "三类合计"):
                    for bureau_key in bureau_keys:
                        increase(sheet_name, bureau_key, col, case_key)

    process(rows_year, segments_year)
    process(rows_last_year, segments_last_year)
    return counts
