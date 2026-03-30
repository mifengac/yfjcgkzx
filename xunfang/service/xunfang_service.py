"""
巡防统计业务逻辑层
处理巡防统计相关的业务逻辑
"""
import requests
from typing import Dict, Any, List
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import logging
from xunfang.dao.xunfang_dao import get_cross_day_data
from gonggong.config.upstream_zhksh import build_zhksh_url
from gonggong.service.session_manager import session_manager
from gonggong.config.database import DB_CONFIG, execute_query
import psycopg2
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.font_manager as fm
import numpy as np
from io import BytesIO
import base64
import os

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)


def _to_str(val) -> str:
    return "" if val is None else str(val)


def _to_float(val, default: float = 0.0) -> float:
    try:
        if val is None:
            return default
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        if s == "":
            return default
        return float(s)
    except Exception:
        return default


def get_diqu_mapping():
    """
    获取地区映射字典
    """
    return {
        '445302': '云城',
        '445303': '云安',
        '445321': '新兴',
        '445322': '郁南',
        '445381': '罗定'
    }


def get_region_name_by_diqu(diqu_code):
    """
    根据地区代码获取地区名称
    """
    diqu_mapping = get_diqu_mapping()
    return diqu_mapping.get(diqu_code, f'未知地区({diqu_code})')


def setup_chinese_font():
    """
    设置matplotlib中文字体
    """
    try:
        # 尝试设置系统中文字体
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False  # 解决负号显示问题

        # 测试中文字体是否可用
        fig, ax = plt.subplots(figsize=(1, 1))
        ax.text(0.5, 0.5, '测试', fontsize=12)
        plt.close(fig)

        logging.info("中文字体设置成功")
        return True
    except Exception as e:
        logging.warning(f"中文字体设置失败: {e}")
        # 使用默认字体，可能会显示方块
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
        return False


def create_quadrant_chart_with_matplotlib(final_chart_data, start_time, end_time):
    """
    使用matplotlib创建象限图，包含智能标签位置调整
    """
    # 设置中文字体
    setup_chinese_font()

    if not final_chart_data:
        raise ValueError("没有数据可用于创建象限图")

    # 提取数据
    patrol_values = [data['巡防投入指数'] for data in final_chart_data]
    risk_values = [data['警情风险指数'] for data in final_chart_data]
    station_names = [data['派出所名称'] for data in final_chart_data]

    # 计算数据范围和中心点
    min_patrol = min(patrol_values) if patrol_values else 0
    max_patrol = max(patrol_values) if patrol_values else 1
    min_risk = min(risk_values) if risk_values else 0
    max_risk = max(risk_values) if risk_values else 1

    patrol_center = (min_patrol + max_patrol) / 2
    risk_center = (min_risk + max_risk) / 2

    # 添加边距（10%）
    patrol_range = max_patrol - min_patrol if max_patrol != min_patrol else 1
    risk_range = max_risk - min_risk if max_risk != min_risk else 1

    patrol_margin = patrol_range * 0.1
    risk_margin = risk_range * 0.1

    # 创建图表，调整尺寸以适应底部说明文字
    plt.figure(figsize=(14, 12))
    ax = plt.gca()

    # 设置坐标轴范围
    ax.set_xlim(min_patrol - patrol_margin, max_patrol + patrol_margin)
    ax.set_ylim(min_risk - risk_margin, max_risk + risk_margin)

    # 为不同象限的数据点分配颜色
    colors = []
    for x, y in zip(patrol_values, risk_values):
        if x >= patrol_center and y >= risk_center:
            colors.append('red')      # 右上：高投入/高风险 - 红色
        elif x < patrol_center and y >= risk_center:
            colors.append('orange')   # 左上：低投入/高风险 - 橙色
        elif x < patrol_center and y < risk_center:
            colors.append('green')    # 左下：低投入/低风险 - 绿色
        else:
            colors.append('blue')     # 右下：高投入/低风险 - 蓝色

    # 绘制散点
    scatter = ax.scatter(patrol_values, risk_values,
                      c=colors, s=100, alpha=0.7, edgecolors='black', linewidth=1)

    # 智能标签位置调整算法
    def calculate_label_positions(x_coords, y_coords, labels, patrol_center, risk_center):
        """
        计算避免重叠的标签位置
        """
        positions = []
        used_positions = []  # 记录已使用的位置

        for i, (x, y, label) in enumerate(zip(x_coords, y_coords, labels)):
            # 基础偏移量
            base_offset = 8

            # 根据象限确定初始偏移方向
            if x >= patrol_center and y >= risk_center:  # 右上
                candidates = [
                    (base_offset, base_offset),
                    (base_offset + 5, base_offset),
                    (base_offset, base_offset + 5),
                    (base_offset - 5, base_offset),
                    (base_offset, base_offset - 5),
                    (base_offset + 10, base_offset),
                    (base_offset, base_offset + 10),
                ]
            elif x < patrol_center and y >= risk_center:  # 左上
                candidates = [
                    (-base_offset - len(label) * 3, base_offset),
                    (-base_offset - len(label) * 3 - 5, base_offset),
                    (-base_offset - len(label) * 3, base_offset + 5),
                    (-base_offset - len(label) * 3 + 5, base_offset),
                    (-base_offset - len(label) * 3, base_offset - 5),
                    (-base_offset - len(label) * 3 - 10, base_offset),
                    (-base_offset - len(label) * 3, base_offset + 10),
                ]
            elif x < patrol_center and y < risk_center:  # 左下
                candidates = [
                    (-base_offset - len(label) * 3, -base_offset - 15),
                    (-base_offset - len(label) * 3 - 5, -base_offset - 15),
                    (-base_offset - len(label) * 3, -base_offset - 10),
                    (-base_offset - len(label) * 3 + 5, -base_offset - 15),
                    (-base_offset - len(label) * 3, -base_offset - 20),
                    (-base_offset - len(label) * 3 - 10, -base_offset - 15),
                    (-base_offset - len(label) * 3, -base_offset - 25),
                ]
            else:  # 右下
                candidates = [
                    (base_offset, -base_offset - 15),
                    (base_offset + 5, -base_offset - 15),
                    (base_offset, -base_offset - 10),
                    (base_offset - 5, -base_offset - 15),
                    (base_offset, -base_offset - 20),
                    (base_offset + 10, -base_offset - 15),
                    (base_offset, -base_offset - 25),
                ]

            # 检查哪个位置不会与其他标签重叠
            best_position = candidates[0]
            min_distance = float('inf')

            for candidate in candidates:
                # 计算标签的实际位置
                label_x = x + candidate[0] / 72  # 转换points到数据坐标（近似）
                label_y = y + candidate[1] / 72

                # 计算与已使用位置的最小距离
                min_candidate_distance = float('inf')
                for used_x, used_y in used_positions:
                    distance = ((label_x - used_x) ** 2 + (label_y - used_y) ** 2) ** 0.5
                    min_candidate_distance = min(min_candidate_distance, distance)

                # 如果没有重叠或距离足够大，选择这个位置
                if min_candidate_distance > 0.5 or len(used_positions) == 0:  # 最小距离阈值
                    best_position = candidate
                    break

                # 否则选择距离最大的位置
                if min_candidate_distance > min_distance:
                    min_distance = min_candidate_distance
                    best_position = candidate

            # 记录选中的位置
            label_x = x + best_position[0] / 72
            label_y = y + best_position[1] / 72
            used_positions.append((label_x, label_y))

            positions.append(best_position)

        return positions

    # 计算智能标签位置
    label_positions = calculate_label_positions(patrol_values, risk_values, station_names, patrol_center, risk_center)

    # 添加数据标签（使用智能调整后的位置）
    for i, (x, y, name, color, position) in enumerate(zip(patrol_values, risk_values, station_names, colors, label_positions)):
        ax.annotate(name, (x, y),
                   xytext=position, textcoords='offset points',
                   fontsize=8, alpha=0.9,
                   bbox=dict(boxstyle='round,pad=0.3', facecolor=color, alpha=0.2),
                   arrowprops=dict(arrowstyle='-', color=color, alpha=0.5, lw=0.5))

    # 添加象限分割线
    ax.axvline(x=patrol_center, color='gray', linestyle='--', alpha=0.8, linewidth=2)
    ax.axhline(y=risk_center, color='gray', linestyle='--', alpha=0.8, linewidth=2)

    # 设置坐标轴标签和标题
    ax.set_xlabel('巡防投入指数', fontsize=14, fontweight='bold')
    ax.set_ylabel('警情风险指数', fontsize=14, fontweight='bold')
    ax.set_title(f'{start_time}至{end_time}派出所巡防警情象限图',
                fontsize=16, fontweight='bold', pad=20)

    # 添加象限标签（修正位置和内容）
    ax.text(0.98, 0.98, '高投入/高风险', transform=ax.transAxes,
           fontsize=12, fontweight='bold', va='top', ha='right',
           bbox=dict(boxstyle='round,pad=0.5', facecolor='red', alpha=0.3))
    ax.text(0.02, 0.98, '低投入/高风险', transform=ax.transAxes,
           fontsize=12, fontweight='bold', va='top', ha='left',
           bbox=dict(boxstyle='round,pad=0.5', facecolor='orange', alpha=0.3))
    ax.text(0.02, 0.02, '低投入/低风险', transform=ax.transAxes,
           fontsize=12, fontweight='bold', va='bottom', ha='left',
           bbox=dict(boxstyle='round,pad=0.5', facecolor='green', alpha=0.3))
    ax.text(0.98, 0.02, '高投入/低风险', transform=ax.transAxes,
           fontsize=12, fontweight='bold', va='bottom', ha='right',
           bbox=dict(boxstyle='round,pad=0.5', facecolor='blue', alpha=0.3))

    # 优化坐标轴刻度显示
    def format_tick_label(value, is_percentage=False):
        """格式化刻度标签"""
        if is_percentage:
            return f'{value:.1f}%'
        else:
            return f'{value:.1f}'

    # 设置X轴刻度
    ax.tick_params(axis='x', labelsize=10, rotation=45)
    ax.xaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format_tick_label(x, False)))

    # 设置Y轴刻度
    ax.tick_params(axis='y', labelsize=10)
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format_tick_label(x, False)))

    # 在分割线位置添加特殊标注
    ax.text(patrol_center, min_risk - risk_margin * 0.8,
           f'平均值: {patrol_center:.2f}',
           ha='center', va='top', fontsize=9, color='gray')
    ax.text(min_patrol - patrol_margin * 0.8, risk_center,
           f'平均值: {risk_center:.2f}',
           ha='right', va='center', fontsize=9, color='gray',
           rotation=90)

    # 添加网格线
    ax.grid(True, alpha=0.3, linestyle='-', linewidth=0.5)

    # 设置背景色
    ax.set_facecolor('#f8f9fa')

    # 添加象限水印文字
    def add_quadrant_watermark(ax, patrol_center, risk_center, min_patrol, max_patrol, min_risk, max_risk):
        """
        在象限内添加水印文字
        """
        # 计算每个象限的中心位置
        q1_x = patrol_center + (max_patrol - patrol_center) * 0.5
        q1_y = risk_center + (max_risk - risk_center) * 0.5
        q2_x = patrol_center - (patrol_center - min_patrol) * 0.5
        q2_y = risk_center + (max_risk - risk_center) * 0.5
        q3_x = patrol_center - (patrol_center - min_patrol) * 0.5
        q3_y = risk_center - (risk_center - min_risk) * 0.5
        q4_x = patrol_center + (max_patrol - patrol_center) * 0.5
        q4_y = risk_center - (risk_center - min_risk) * 0.5

        # 添加水印文字，设置透明度和字体大小
        watermark_props = {
            'fontsize': 16,
            'fontweight': 'bold',
            'alpha': 0.15,  # 透明度，不影响正常显示
            'ha': 'center',
            'va': 'center',
            'color': 'gray'
        }

        # 第一象限（右上）
        ax.text(q1_x, q1_y, '第一象限', **watermark_props)

        # 第二象限（左上）
        ax.text(q2_x, q2_y, '第二象限', **watermark_props)

        # 第三象限（左下）
        ax.text(q3_x, q3_y, '第三象限', **watermark_props)

        # 第四象限（右下）
        ax.text(q4_x, q4_y, '第四象限', **watermark_props)

    # 调用水印添加函数
    add_quadrant_watermark(ax, patrol_center, risk_center, min_patrol, max_patrol, min_risk, max_risk)

    # 在图片底部添加指数说明
    def add_index_explanation(ax):
        """
        在图片底部添加巡防投入指数和警情风险指数的详细说明
        """
        explanation_text = (
            "巡防投入指数说明：\n"
            "• 车巡 (权重1.3): 巡逻时长(小时) + 巡逻里程(公里) × 1.3\n"
            "• 步巡 (权重1.0): 巡逻时长(小时) + 巡逻里程(公里) × 1.0\n"
            "• 摩托车巡 (权重1.3): 巡逻时长(小时) + 巡逻里程(公里) × 1.3\n"
            "• 车巡+步巡 (权重1.5): 巡逻时长(小时) + 巡逻里程(公里) × 1.5\n\n"
            "警情风险指数说明：\n"
            "• 人身伤害类警情数 × 1.5 (权重1.5)\n"
            "• 侵犯财产类警情数 × 1.2 (权重1.2)\n"
            "• 扰乱秩序类警情数 × 1.0 (权重1.0)"
        )

        # 在底部添加说明文字，字体比坐标轴标题小一些
        ax.text(0.5, -0.25, explanation_text,
               transform=ax.transAxes,
               fontsize=9,  # 比坐标轴标题(14)小
               ha='center', va='top',
               bbox=dict(boxstyle='round,pad=0.5', facecolor='lightgray', alpha=0.8))

    # 调用说明添加函数
    add_index_explanation(ax)

    # 调整布局，为底部文字留出空间
    plt.subplots_adjust(bottom=0.3)  # 增加底部边距

    # 保存到内存
    buffer = BytesIO()
    plt.savefig(buffer, format='png', dpi=300, bbox_inches='tight')
    buffer.seek(0)

    logging.info(f"matplotlib象限图创建成功，数据点数量: {len(final_chart_data)}")
    return buffer


def calculate_xunfang_for_date_range(start_time: str, end_time: str) -> Dict[str, Any]:
    """
    根据日期范围计算巡防统计数据
    """
    success_count = 0
    success_count_lock = threading.Lock()

    def update_duty_schedule_safe(url: str, data: Dict[str, Any]) -> bool:
        """
        使用全局会话管理器发起更新请求
        """
        try:
            response = session_manager.make_request('POST', url, data=data,
                headers={
                    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
                    'X-Requested-With': 'XMLHttpRequest'
                })
            return response.status_code == 200
        except Exception as e:
            logging.error(f"更新勤务数据请求失败: {e}")
            return False

    def update_duty_schedule_threaded(args):
        """
        在单独的线程中更新勤务安排
        """
        nonlocal success_count
        update_url, update_data, index, total = args

        try:
            # 调用更新接口
            success = update_duty_schedule_safe(update_url, update_data)
            if success:
                with success_count_lock:
                    success_count += 1
                logging.info(f'更新第 {index}/{total} 条记录成功')
                return True, index
            else:
                logging.warning(f'更新第 {index}/{total} 条记录失败')
                return False, index
        except Exception as e:
            logging.error(f'更新第 {index}/{total} 条记录时出错: {e}')
            return False, index

    # 解析日期时间
    start_date = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
    end_date = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
    
    # 验证日期范围
    if start_date > end_date:
        raise ValueError('开始时间不能晚于结束时间')
    
    # 跨天数据接口URL
    cross_day_url = build_zhksh_url("/zhksh/dutySchedule/crossDayList")
    # 更新接口URL
    update_url = build_zhksh_url("/zhksh/dutySchedule/updateDutyScheduleInfo")

    # 一次性获取整个时间范围的数据
    try:
        # 直接使用用户输入的时间范围
        logging.info(f"获取 {start_time} 至 {end_time} 的跨天数据...")
        rows_data = get_cross_day_data(cross_day_url, start_time, end_time)
        logging.info(f"获取到 {len(rows_data)} 条数据")

        # 准备更新任务
        all_tasks = []
        task_index = 1

        for row in rows_data:
            # 构建请求参数，使用row中的数据替换默认值
            # 根据xunfang2.txt，只包含form表单中的字段
            update_data = {
                'createBy': row.get('createBy', '匿名'),
                'createTime': row.get('createTime', '2025-09-24 09:46:06'),
                'updateBy': row.get('updateBy', '匿名'),
                'updateTime': row.get('updateTime', '2025-09-26 04:55:32'),
                'deptId': row.get('deptId', '445300210805'),
                'deptName': row.get('deptName', '云浮市局交通警察支队市区大队二中队'),
                'scheduleId': row.get('scheduleId', ''),
                'schemeId': row.get('schemeId', '2DF029677D5F0B9BE063CF01FC446B74'),
                'shiftId': row.get('shiftId', '2DF041755D2151CBE063CF01FC44ACDD'),
                'userNo': row.get('userNo', ''),
                'userName': row.get('userName', ''),
                'shiftName': row.get('shiftName', '夜班'),
                'startTime': row.get('startTime', start_time),
                'endTime': row.get('endTime', end_time),
                'scheduleDate': row.get('scheduleDate', start_time[:10]),
                'patrolMileage': row.get('patrolMileage', 0),
                'onlineDuration': row.get('onlineDuration', 0),
                'onDutyOnlineDuration': row.get('onDutyOnlineDuration', 0),
                'dutyTypeName': row.get('dutyTypeName', '巡逻宣防岗'),
                'deploymentTypeName': row.get('deploymentTypeName', '机构部门'),
                'deploymentName': row.get('deploymentName', '云浮市局交通警察支队市区大队二中队'),
                'bindPdtNo': row.get('bindPdtNo', '76626224'),
                'leaderName': row.get('leaderName', '林树生'),
                'userPhone': row.get('userPhone', '18023799748'),
                'dutyPhone': row.get('dutyPhone', '18023799748'),
                'dutyTypeCode': row.get('dutyTypeCode', 'jj_xlxfg'),
                'dutyLevelCode': row.get('dutyLevelCode', 0),
                'dutyLevelName': row.get('dutyLevelName', '日常勤务'),
                'duty': row.get('duty', 0),
                'leaderId': row.get('leaderId', '270320'),
                'userId': row.get('userId', '8A40EB9F54B8425D84257A5FC43D97FD'),
                'userCategory': row.get('userCategory', 0),
                'deploymentType': row.get('deploymentType', 15),
                'deploymentId': row.get('deploymentId', '445300210805'),
                'serviceStyle': row.get('serviceStyle', 0),
                'reviewState': row.get('reviewState', 0),
                'reportType': row.get('reportType', 0),
                'reportState': row.get('reportState', 0),
                'lng': row.get('lng', 0),
                'lat': row.get('lat', 0),
                'scheduleOrganId': row.get('scheduleOrganId', '445300210805'),
                'status': row.get('status', 'Y'),
                'orderNum': row.get('orderNum', 0),
                'inTaskNum': row.get('inTaskNum', 0),
                'countNum': row.get('countNum', 0),
                'repairDuration': row.get('repairDuration', 0),
                'dutyViolationDetail': row.get('dutyViolationDetail', '无违规情况')
            }

            all_tasks.append((update_url, update_data, task_index, len(rows_data)))
            task_index += 1

    except Exception as e:
        logging.error(f"获取跨天数据失败: {e}")
        all_tasks = []
    
    # 使用线程池处理所有任务
    max_workers = min(10, len(all_tasks))  # 限制最大线程数，避免服务器压力过大

    if all_tasks:
        logging.info(f"开始处理 {len(all_tasks)} 个任务，使用 {max_workers} 个线程")
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有任务
            future_to_index = {executor.submit(update_duty_schedule_threaded, task): task[3] for task in all_tasks}

            # 等待所有任务完成
            for future in as_completed(future_to_index):
                result, index = future.result()
    
    # 返回结果
    return {
        'success_count': success_count,
        'total_tasks': len(all_tasks)
    }


def export_online_rate_for_date_range(start_time: str, end_time: str):
    """
    根据日期范围导出在线在岗率数据
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.utils import get_column_letter
    import json
    from datetime import datetime, timedelta

    # 解析日期时间
    start_date = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
    end_date = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')
    
    # 验证日期范围
    if start_date > end_date:
        raise ValueError('开始时间不能晚于结束时间')
    
    # 跨天数据接口URL
    cross_day_url = build_zhksh_url("/zhksh/dutySchedule/crossDayList")

    # 一次性获取整个时间范围的数据
    all_data = []
    try:
        logging.info(f"获取 {start_time} 至 {end_time} 的跨天数据用于导出...")
        rows_data = get_cross_day_data(cross_day_url, start_time, end_time)
        logging.info(f"获取到 {len(rows_data)} 条数据")

        for row in rows_data:
            # 添加排班时长字段（分钟）
            start_time_dt = datetime.strptime(row['startTime'], '%Y-%m-%d %H:%M:%S')
            end_time_dt = datetime.strptime(row['endTime'], '%Y-%m-%d %H:%M:%S')
            shift_duration = (end_time_dt - start_time_dt).total_seconds() / 60
            row['排班时长'] = shift_duration

            # 添加地区字段，截取deptId前6位
            dept_id = row.get('deptId', '')
            row['diqu'] = dept_id[:6] if len(dept_id) >= 6 else dept_id

        all_data = rows_data
    except Exception as e:
        logging.error(f"获取跨天数据失败: {e}")
        all_data = []
    
    # 定义各分局的organ_id映射
    organ_ids = {
        '云城分局': '445302000000',
        '云安分局': '445303000000',
        '新兴县局': '445321000000',
        '郁南县局': '445322000000',
        '罗定市局': '445381000000'
    }

    # 创建Excel工作簿
    wb = Workbook()

    # 存储各地区数据的汇总信息，用于计算全市数据
    all_region_summary = {
        'onlineDuration': 0,
        'onDutyOnlineDuration': 0,
        '排班时长': 0
    }

    # 添加数据统计日志
    logging.info(f"开始处理导出在线在岗数据，总数据条数: {len(all_data)}")

    # 统计所有数据中的organ_id分布
    all_organ_ids = {}
    for item in all_data:
        organ_id = item.get('scheduleOrganId', '未知')
        if organ_id not in all_organ_ids:
            all_organ_ids[organ_id] = 0
        all_organ_ids[organ_id] += 1

    logging.info(f"所有数据的organ_id分布: {all_organ_ids}")

    # 统计在定义的5个地区中的数据条数
    included_data_count = 0
    for organ_id in organ_ids.values():
        included_data_count += all_organ_ids.get(organ_id, 0)

    excluded_data_count = len(all_data) - included_data_count
    logging.info(f"包含在5个地区中的数据条数: {included_data_count}")
    logging.info(f"不在5个地区中的数据条数: {excluded_data_count}")

    if excluded_data_count > 0:
        logging.warning(f"有 {excluded_data_count} 条数据不在5个定义的地区中，这些数据将不被包含在导出中")

    # 先创建各个地区的sheet
    region_sheets_data = {}
    for sheet_name, organ_id in organ_ids.items():
        ws = wb.create_sheet(title=sheet_name)

        # 过滤对应organ_id的数据
        filtered_data = [item for item in all_data if item.get('scheduleOrganId') == organ_id]
        logging.info(f"{sheet_name} (organ_id: {organ_id}) 过滤后数据条数: {len(filtered_data)}")

        # 按deptId分组，但存储deptName用于显示
        dept_grouped = {}
        for item in filtered_data:
            dept_id = item.get('deptId', '未知')
            dept_name = item.get('deptName', '未知')

            if dept_id not in dept_grouped:
                dept_grouped[dept_id] = {
                    'dept_name': dept_name,  # 存储显示名称
                    'onlineDuration': 0,
                    'onDutyOnlineDuration': 0,
                    '排班时长': 0
                }
            dept_grouped[dept_id]['onlineDuration'] += item.get('onlineDuration', 0)
            dept_grouped[dept_id]['onDutyOnlineDuration'] += item.get('onDutyOnlineDuration', 0)
            dept_grouped[dept_id]['排班时长'] += item.get('排班时长', 0)

        logging.info(f"{sheet_name} 按部门ID分组后，共有 {len(dept_grouped)} 个部门")

        # 在工作表中写入数据
        headers = ['部门', '在线率', '在岗在线率']
        ws.append(headers)

        for dept_id, values in dept_grouped.items():
            display_name = values['dept_name']  # 使用存储的部门名称显示
            if values['排班时长'] > 0:
                online_rate = round((values['onlineDuration'] / values['排班时长']), 4)  # 不再乘以100
                onduty_rate = round((values['onDutyOnlineDuration'] / values['排班时长']), 4)  # 不再乘以100
                ws.append([display_name, online_rate, onduty_rate])
            else:
                ws.append([display_name, 0.00, 0.00])

        # 为该sheet创建柱状图
        if len(dept_grouped) > 0:
            # 创建柱状图
            chart = BarChart()
            chart.type = "col"
            chart.title = f"{start_time}至{end_time}{sheet_name}在岗在线率"
            chart.style = 12
            chart.y_axis.title = '百分比(%)'
            chart.x_axis.title = '部门'

            # 定义数据范围
            data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=len(dept_grouped) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(dept_grouped) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

            # 显示图表数据标签
            try:
                from openpyxl.chart.label import DataLabelList
                chart.dLbls = DataLabelList()
                chart.dLbls.showVal = True  # 显示数值
                chart.dLbls.showCatName = False  # 不显示类别名称
                chart.dLbls.showSerName = False  # 不显示系列名称
                chart.dLbls.showPercent = False  # 不显示百分比（因为我们使用的是已经转换为百分比格式的数据）
                logging.info(f"{sheet_name}图表数据标签设置成功")
            except ImportError:
                logging.warning(f"无法导入DataLabelList，{sheet_name}图表将不显示数据标签")

            # 添加图表到工作表
            ws.add_chart(chart, "F2")

            # 设置百分比格式，Excel会自动将小数转换为百分比显示
            for row in range(2, len(dept_grouped) + 2):
                ws.cell(row=row, column=2).number_format = '0.00%'
                ws.cell(row=row, column=3).number_format = '0.00%'

        # 保存该地区的数据汇总，用于计算全市数据
        region_summary = {
            'onlineDuration': 0,
            'onDutyOnlineDuration': 0,
            '排班时长': 0
        }
        for values in dept_grouped.values():
            region_summary['onlineDuration'] += values['onlineDuration']
            region_summary['onDutyOnlineDuration'] += values['onDutyOnlineDuration']
            region_summary['排班时长'] += values['排班时长']

        # 记录地区汇总数据
        logging.info(f"{sheet_name} 地区汇总数据:")
        logging.info(f"  在线时长: {region_summary['onlineDuration']} 分钟")
        logging.info(f"  在岗在线时长: {region_summary['onDutyOnlineDuration']} 分钟")
        logging.info(f"  排班时长: {region_summary['排班时长']} 分钟")
        if region_summary['排班时长'] > 0:
            region_online_rate = round((region_summary['onlineDuration'] / region_summary['排班时长']) * 100, 2)
            region_onduty_rate = round((region_summary['onDutyOnlineDuration'] / region_summary['排班时长']) * 100, 2)
            logging.info(f"  地区在线率: {region_online_rate}%")
            logging.info(f"  地区在岗在线率: {region_onduty_rate}%")

        region_sheets_data[sheet_name] = {
            'dept_grouped': dept_grouped,
            'summary': region_summary
        }

        # 累加到全市汇总数据
        all_region_summary['onlineDuration'] += region_summary['onlineDuration']
        all_region_summary['onDutyOnlineDuration'] += region_summary['onDutyOnlineDuration']
        all_region_summary['排班时长'] += region_summary['排班时长']

    # 记录全市汇总数据
    logging.info(f"全市汇总数据:")
    logging.info(f"  总在线时长: {all_region_summary['onlineDuration']} 分钟")
    logging.info(f"  总在岗在线时长: {all_region_summary['onDutyOnlineDuration']} 分钟")
    logging.info(f"  总排班时长: {all_region_summary['排班时长']} 分钟")

    # 最后创建全市sheet，基于diqu字段进行统计
    ws1 = wb.create_sheet(title="全市", index=0)  # 插入到第一个位置

    # 在全市sheet中写入各地区数据
    headers = ['地区', '在线率', '在岗在线率']
    ws1.append(headers)

    # 使用diqu字段进行全市统计
    diqu_mapping = get_diqu_mapping()
    diqu_stats = {}

    # 统计各地区数据
    for item in all_data:
        diqu = item.get('diqu', '')
        if diqu in diqu_mapping:  # 只统计定义的地区
            region_name = diqu_mapping[diqu]
            if region_name not in diqu_stats:
                diqu_stats[region_name] = {
                    'onlineDuration': 0,
                    'onDutyOnlineDuration': 0,
                    '排班时长': 0
                }

            diqu_stats[region_name]['onlineDuration'] += item.get('onlineDuration', 0)
            diqu_stats[region_name]['onDutyOnlineDuration'] += item.get('onDutyOnlineDuration', 0)
            diqu_stats[region_name]['排班时长'] += item.get('排班时长', 0)

    logging.info("全市统计 - 按diqu字段统计结果:")
    for region_name, stats in diqu_stats.items():
        if stats['排班时长'] > 0:
            online_rate = round((stats['onlineDuration'] / stats['排班时长']), 4)
            onduty_rate = round((stats['onDutyOnlineDuration'] / stats['排班时长']), 4)

            # 记录全市计算过程
            actual_online_rate = round(online_rate * 100, 2)
            actual_onduty_rate = round(onduty_rate * 100, 2)
            logging.info(f"全市计算 - {region_name}:")
            logging.info(f"  在线时长: {stats['onlineDuration']} 分钟")
            logging.info(f"  在岗在线时长: {stats['onDutyOnlineDuration']} 分钟")
            logging.info(f"  排班时长: {stats['排班时长']} 分钟")
            logging.info(f"  在线率: {actual_online_rate}% (小数: {online_rate})")
            logging.info(f"  在岗在线率: {actual_onduty_rate}% (小数: {onduty_rate})")

            ws1.append([region_name, online_rate, onduty_rate])
        else:
            ws1.append([region_name, 0.00, 0.00])
            logging.warning(f"全市计算 - {region_name}: 排班时长为0，比率设为0%")

    # 为全市sheet创建柱状图
    if len(diqu_stats) > 0:
        # 创建柱状图
        chart = BarChart()
        chart.type = "col"
        chart.title = f"{start_time}至{end_time}全市在岗在线率"
        chart.style = 12
        chart.y_axis.title = '百分比(%)'
        chart.x_axis.title = '地区'

        # 定义数据范围（现在从第2行开始，因为只有标题行）
        # 第1行：标题，第2行开始是数据
        data_start_row = 2  # 数据开始行
        data_end_row = len(diqu_stats) + data_start_row - 1  # 数据结束行

        data = Reference(ws1, min_col=2, min_row=1, max_col=3, max_row=data_end_row)
        cats = Reference(ws1, min_col=1, min_row=data_start_row, max_row=data_end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # 显示图表数据标签
        try:
            from openpyxl.chart.label import DataLabelList
            chart.dLbls = DataLabelList()
            chart.dLbls.showVal = True  # 显示数值
            chart.dLbls.showCatName = False  # 不显示类别名称
            chart.dLbls.showSerName = False  # 不显示系列名称
            chart.dLbls.showPercent = False  # 不显示百分比（因为我们使用的是已经转换为百分比格式的数据）
            logging.info("全市图表数据标签设置成功")
        except ImportError:
            logging.warning("无法导入DataLabelList，图表将不显示数据标签")

        # 添加图表到工作表
        ws1.add_chart(chart, "F2")

        # 设置百分比格式，Excel会自动将小数转换为百分比显示
        for row in range(data_start_row, data_end_row + 1):
            ws1.cell(row=row, column=2).number_format = '0.00%'
            ws1.cell(row=row, column=3).number_format = '0.00%'

    return wb


def get_police_stations():
    """
    获取全市派出所列表
    """
    try:
        # 获取全市派出所编码
        query = '''SELECT dq."parent_code", dq."code", dq."name"
                   FROM "ywdata"."sys_dq_pcs" dq
                   WHERE dq."level"=2 and dq."name" ~~ '%派出所%' '''

        results = execute_query(query)

        police_stations = []
        for row in results:
            police_stations.append({
                'parent_code': row['parent_code'],
                'code': row['code'],
                'name': row['name']
            })

        logging.info(f"获取到 {len(police_stations)} 个派出所")
        return police_stations

    except Exception as e:
        logging.error(f"获取派出所列表失败: {e}")
        return []


def get_case_statistics_by_type(start_time: str, end_time: str, case_type: str):
    """
    根据案件类型获取警情统计
    """
    try:
        # 根据案件类型构建查询 - 修正SQL语法
        query = '''SELECT jq."dutydeptno", jq."dutydeptname", count(jq."caseno") as 警情数量
                    FROM "ywdata"."zq_kshddpt_dsjfx_jq" jq
                    WHERE jq."calltime" BETWEEN %s AND %s
                    AND EXISTS (
                        SELECT 1 FROM "ywdata"."case_type_config" ct
                        WHERE ct."leixing" = %s
                        AND ct."newcharasubclass_list" @> ARRAY[jq."neworicharasubclass"]
                    )
                    GROUP BY jq."dutydeptno", jq."dutydeptname" '''

        results = execute_query(query, (start_time + ' 00:00:00', end_time + ' 23:59:59', case_type))

        case_stats = {}
        for row in results:
            case_stats[row['dutydeptno']] = {
                'dutydeptno': row['dutydeptno'],
                'dutydeptname': row['dutydeptname'],
                'count': row['警情数量']
            }

        logging.info(f"获取{case_type}警情统计 {len(case_stats)} 条记录")
        return case_stats

    except Exception as e:
        logging.error(f"获取{case_type}警情统计失败: {e}")
        return {}


def calculate_patrol_input_index(qwsj_result, police_stations):
    """
    计算巡防投入指数，返回详细原始数据
    """
    # 获取派出所代码列表
    police_station_codes = {station['code'] for station in police_stations}

    # 过滤派出所数据（统一将 deptId 转为字符串匹配）
    filtered_data = [item for item in qwsj_result if _to_str(item.get('deptId')) in police_station_codes]

    # 初始化原始数据存储结构
    patrol_raw_data = {}
    patrol_weights = {}

    # 初始化每个派出所的原始数据
    for station in police_stations:
        dept_code = station['code']
        patrol_raw_data[dept_code] = {
            '派出所名称': station['name'],
            '派出所代码': dept_code,
            '所属地区代码': station['parent_code'],
            '车巡时长(小时)': 0,
            '车巡里程(公里)': 0,
            '步巡时长(小时)': 0,
            '步巡里程(公里)': 0,
            '摩托车巡时长(小时)': 0,
            '摩托车巡里程(公里)': 0,
            '车巡+步巡时长(小时)': 0,
            '车巡+步巡里程(公里)': 0,
            '巡防投入指数': 0
        }

    # 车巡 (typeCode='1', 权重=1.3)
    car_patrol = [item for item in filtered_data if _to_str(item.get('typeCode')) == '1']
    car_patrol_grouped = {}
    for item in car_patrol:
        dept_id = _to_str(item.get('deptId'))
        if dept_id not in car_patrol_grouped:
            car_patrol_grouped[dept_id] = {'onlineDuration': 0, 'patrolMileage': 0}
        car_patrol_grouped[dept_id]['onlineDuration'] += _to_float(item.get('onlineDuration', 0))
        car_patrol_grouped[dept_id]['patrolMileage'] += _to_float(item.get('patrolMileage', 0))

    for dept_id, data in car_patrol_grouped.items():
        patrol_hours = round(data['onlineDuration'] / 60, 2)  # 转换为小时
        patrol_km = round(data['patrolMileage'] / 1000, 2)  # 转换为公里
        weight = round((patrol_hours + patrol_km) * 1.3, 2)

        if dept_id not in patrol_weights:
            patrol_weights[dept_id] = 0
        patrol_weights[dept_id] += weight

        # 保存原始数据
        if dept_id in patrol_raw_data:
            patrol_raw_data[dept_id]['车巡时长(小时)'] = patrol_hours
            patrol_raw_data[dept_id]['车巡里程(公里)'] = patrol_km

    # 步巡 (typeCode='2', 权重=1)
    foot_patrol = [item for item in filtered_data if _to_str(item.get('typeCode')) == '2']
    foot_patrol_grouped = {}
    for item in foot_patrol:
        dept_id = _to_str(item.get('deptId'))
        if dept_id not in foot_patrol_grouped:
            foot_patrol_grouped[dept_id] = {'onlineDuration': 0, 'patrolMileage': 0}
        foot_patrol_grouped[dept_id]['onlineDuration'] += _to_float(item.get('onlineDuration', 0))
        foot_patrol_grouped[dept_id]['patrolMileage'] += _to_float(item.get('patrolMileage', 0))

    for dept_id, data in foot_patrol_grouped.items():
        patrol_hours = round(data['onlineDuration'] / 60, 2)
        patrol_km = round(data['patrolMileage'] / 1000, 2)
        weight = round((patrol_hours + patrol_km) * 1, 2)

        if dept_id not in patrol_weights:
            patrol_weights[dept_id] = 0
        patrol_weights[dept_id] += weight

        # 保存原始数据
        if dept_id in patrol_raw_data:
            patrol_raw_data[dept_id]['步巡时长(小时)'] = patrol_hours
            patrol_raw_data[dept_id]['步巡里程(公里)'] = patrol_km

    # 摩托车巡 (typeCode='5', 权重=1.3)
    motorcycle_patrol = [item for item in filtered_data if _to_str(item.get('typeCode')) == '5']
    motorcycle_patrol_grouped = {}
    for item in motorcycle_patrol:
        dept_id = _to_str(item.get('deptId'))
        if dept_id not in motorcycle_patrol_grouped:
            motorcycle_patrol_grouped[dept_id] = {'onlineDuration': 0, 'patrolMileage': 0}
        motorcycle_patrol_grouped[dept_id]['onlineDuration'] += _to_float(item.get('onlineDuration', 0))
        motorcycle_patrol_grouped[dept_id]['patrolMileage'] += _to_float(item.get('patrolMileage', 0))

    for dept_id, data in motorcycle_patrol_grouped.items():
        patrol_hours = round(data['onlineDuration'] / 60, 2)
        patrol_km = round(data['patrolMileage'] / 1000, 2)
        weight = round((patrol_hours + patrol_km) * 1.3, 2)

        if dept_id not in patrol_weights:
            patrol_weights[dept_id] = 0
        patrol_weights[dept_id] += weight

        # 保存原始数据
        if dept_id in patrol_raw_data:
            patrol_raw_data[dept_id]['摩托车巡时长(小时)'] = patrol_hours
            patrol_raw_data[dept_id]['摩托车巡里程(公里)'] = patrol_km

    # 车巡+步巡 (typeCode='4', 权重=1.5)
    car_foot_patrol = [item for item in filtered_data if _to_str(item.get('typeCode')) == '4']
    car_foot_patrol_grouped = {}
    for item in car_foot_patrol:
        dept_id = _to_str(item.get('deptId'))
        if dept_id not in car_foot_patrol_grouped:
            car_foot_patrol_grouped[dept_id] = {'onlineDuration': 0, 'patrolMileage': 0}
        car_foot_patrol_grouped[dept_id]['onlineDuration'] += _to_float(item.get('onlineDuration', 0))
        car_foot_patrol_grouped[dept_id]['patrolMileage'] += _to_float(item.get('patrolMileage', 0))

    for dept_id, data in car_foot_patrol_grouped.items():
        patrol_hours = round(data['onlineDuration'] / 60, 2)
        patrol_km = round(data['patrolMileage'] / 1000, 2)
        weight = round((patrol_hours + patrol_km) * 1.5, 2)

        if dept_id not in patrol_weights:
            patrol_weights[dept_id] = 0
        patrol_weights[dept_id] += weight

        # 保存原始数据
        if dept_id in patrol_raw_data:
            patrol_raw_data[dept_id]['车巡+步巡时长(小时)'] = patrol_hours
            patrol_raw_data[dept_id]['车巡+步巡里程(公里)'] = patrol_km

    # 为每个派出所计算最终的巡防投入指数
    for dept_code in patrol_raw_data:
        patrol_raw_data[dept_code]['巡防投入指数'] = patrol_weights.get(dept_code, 0)

    return patrol_weights, patrol_raw_data


def calculate_risk_index(start_time: str, end_time: str, police_stations):
    """
    计算警情风险指数，返回详细原始数据
    """
    # 获取派出所代码列表
    police_station_codes = {station['code'] for station in police_stations}

    # 初始化原始数据存储结构
    risk_raw_data = {}
    risk_weights = {}

    # 初始化每个派出所的原始数据
    for station in police_stations:
        dept_code = station['code']
        risk_raw_data[dept_code] = {
            '派出所名称': station['name'],
            '派出所代码': dept_code,
            '所属地区代码': station['parent_code'],
            '人身伤害类案件数': 0,
            '侵犯财产类案件数': 0,
            '扰乱秩序类案件数': 0,
            '警情风险指数': 0
        }

    # 人身伤害类 (权重=1.5)
    rsshl_result = get_case_statistics_by_type(start_time, end_time, '人身伤害类')

    # 侵犯财产类 (权重=1.2)
    qfccl_result = get_case_statistics_by_type(start_time, end_time, '侵犯财产类')

    # 扰乱秩序类 (权重=1)
    rlzxl_result = get_case_statistics_by_type(start_time, end_time, '扰乱秩序类')

    # 合并所有派出所的deptid
    all_dept_ids = set()
    all_dept_ids.update(rsshl_result.keys())
    all_dept_ids.update(qfccl_result.keys())
    all_dept_ids.update(rlzxl_result.keys())

    for dept_id in all_dept_ids:
        # 只计算派出所的数据
        if dept_id in police_station_codes:
            risk_weight = 0

            # 人身伤害类权重
            rsshl_count = 0
            if dept_id in rsshl_result:
                rsshl_count = rsshl_result[dept_id]['count']
                risk_weight += rsshl_count * 1.5

            # 侵犯财产类权重
            qfccl_count = 0
            if dept_id in qfccl_result:
                qfccl_count = qfccl_result[dept_id]['count']
                risk_weight += qfccl_count * 1.2

            # 扰乱秩序类权重
            rlzxl_count = 0
            if dept_id in rlzxl_result:
                rlzxl_count = rlzxl_result[dept_id]['count']
                risk_weight += rlzxl_count * 1

            risk_weights[dept_id] = round(risk_weight, 2)

            # 保存原始数据
            if dept_id in risk_raw_data:
                risk_raw_data[dept_id]['人身伤害类案件数'] = rsshl_count
                risk_raw_data[dept_id]['侵犯财产类案件数'] = qfccl_count
                risk_raw_data[dept_id]['扰乱秩序类案件数'] = rlzxl_count
                risk_raw_data[dept_id]['警情风险指数'] = risk_weight

    return risk_weights, risk_raw_data


def export_quadrant_chart_for_date_range(start_time: str, end_time: str):
    """
    导出象限图数据（使用matplotlib绘制）
    返回: (图片buffer, Excel工作簿)
    """
    # 解析日期时间
    start_date = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
    end_date = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')

    # 验证日期范围
    if start_date > end_date:
        raise ValueError('开始时间不能晚于结束时间')

    # 1. 获取全市派出所列表
    logging.info("获取全市派出所列表...")
    police_stations = get_police_stations()
    logging.info(f"获取到 {len(police_stations)} 个派出所")

    # 2. 获取巡防投入数据（改为从数据库获取）
    logging.info("获取巡防投入数据（数据库）...")

    def fetch_zxzgl_rows(start_time: str, end_time: str) -> List[Dict[str, Any]]:
        """
        从数据库读取巡防部署记录：
        SELECT * FROM ywdata.zq_kshddpt_zxzgl zkz
        WHERE (zkz."startTime" BETWEEN %s AND %s)
           OR (zkz."endTime" BETWEEN %s AND %s)
        """
        sql = (
            'SELECT * FROM "ywdata"."zq_kshddpt_zxzgl" zkz '
            'WHERE (zkz."startTime" BETWEEN %s AND %s) '
            '   OR (zkz."endTime" BETWEEN %s AND %s)'
        )
        try:
            return execute_query(sql, (start_time, end_time, start_time, end_time))
        except Exception as e:
            logging.error(f"数据库查询象限图原始数据失败: {e}")
            return []

    all_qwsj_data = fetch_zxzgl_rows(start_time, end_time)

    # 3. 计算巡防投入指数
    logging.info("计算巡防投入指数...")
    patrol_weights, patrol_raw_data = calculate_patrol_input_index(all_qwsj_data, police_stations)

    # 4. 计算警情风险指数
    logging.info("计算警情风险指数...")
    risk_weights, risk_raw_data = calculate_risk_index(start_time, end_time, police_stations)

    # 5. 生成最终数据
    final_chart_data = []
    for station in police_stations:
        dept_code = station['code']
        station_name = station['name']

        # 获取巡防投入指数
        patrol_input = patrol_weights.get(dept_code, 0)

        # 获取警情风险指数
        risk_index = risk_weights.get(dept_code, 0)

        final_chart_data.append({
            '派出所名称': station_name,
            '派出所代码': dept_code,
            '所属地区代码': station['parent_code'],
            '巡防投入指数': patrol_input,
            '警情风险指数': risk_index
        })

    # 6. 使用matplotlib创建象限图
    if len(final_chart_data) > 0:
        logging.info(f"开始使用matplotlib创建象限图，数据点数量: {len(final_chart_data)}")

        # 创建matplotlib象限图
        chart_buffer = create_quadrant_chart_with_matplotlib(final_chart_data, start_time, end_time)

        # 创建详细的Excel文件，包含原始数据
        wb = create_detailed_excel_file(patrol_raw_data, risk_raw_data, final_chart_data, start_time, end_time)

        logging.info("matplotlib象限图导出完成")
        return chart_buffer, wb
    else:
        raise ValueError("没有数据可用于创建象限图")


def create_detailed_excel_file(patrol_raw_data, risk_raw_data, final_chart_data, start_time, end_time):
    """
    创建包含详细原始数据的Excel文件
    """
    wb = Workbook()

    # 删除默认工作表
    wb.remove(wb.active)

    # 创建巡防原始数据工作表
    patrol_ws = wb.create_sheet(title="巡防原始数据")

    # 写入巡防数据表头
    patrol_headers = ['派出所名称', '派出所代码', '所属地区代码',
                     '车巡时长(小时)', '车巡里程(公里)',
                     '步巡时长(小时)', '步巡里程(公里)',
                     '摩托车巡时长(小时)', '摩托车巡里程(公里)',
                     '车巡+步巡时长(小时)', '车巡+步巡里程(公里)',
                     '巡防投入指数']

    for col_num, header in enumerate(patrol_headers, 1):
        patrol_ws.cell(row=1, column=col_num, value=header)

    # 写入巡防数据
    for row_num, (dept_code, data) in enumerate(patrol_raw_data.items(), 2):
        patrol_ws.cell(row=row_num, column=1, value=data['派出所名称'])
        patrol_ws.cell(row=row_num, column=2, value=data['派出所代码'])
        patrol_ws.cell(row=row_num, column=3, value=data['所属地区代码'])
        patrol_ws.cell(row=row_num, column=4, value=data['车巡时长(小时)'])
        patrol_ws.cell(row=row_num, column=5, value=data['车巡里程(公里)'])
        patrol_ws.cell(row=row_num, column=6, value=data['步巡时长(小时)'])
        patrol_ws.cell(row=row_num, column=7, value=data['步巡里程(公里)'])
        patrol_ws.cell(row=row_num, column=8, value=data['摩托车巡时长(小时)'])
        patrol_ws.cell(row=row_num, column=9, value=data['摩托车巡里程(公里)'])
        patrol_ws.cell(row=row_num, column=10, value=data['车巡+步巡时长(小时)'])
        patrol_ws.cell(row=row_num, column=11, value=data['车巡+步巡里程(公里)'])
        patrol_ws.cell(row=row_num, column=12, value=data['巡防投入指数'])

    # 创建警情原始数据工作表
    risk_ws = wb.create_sheet(title="警情原始数据")

    # 写入警情数据表头
    risk_headers = ['派出所名称', '派出所代码', '所属地区代码',
                   '人身伤害类案件数', '侵犯财产类案件数', '扰乱秩序类案件数',
                   '警情风险指数']

    for col_num, header in enumerate(risk_headers, 1):
        risk_ws.cell(row=1, column=col_num, value=header)

    # 写入警情数据
    for row_num, (dept_code, data) in enumerate(risk_raw_data.items(), 2):
        risk_ws.cell(row=row_num, column=1, value=data['派出所名称'])
        risk_ws.cell(row=row_num, column=2, value=data['派出所代码'])
        risk_ws.cell(row=row_num, column=3, value=data['所属地区代码'])
        risk_ws.cell(row=row_num, column=4, value=data['人身伤害类案件数'])
        risk_ws.cell(row=row_num, column=5, value=data['侵犯财产类案件数'])
        risk_ws.cell(row=row_num, column=6, value=data['扰乱秩序类案件数'])
        risk_ws.cell(row=row_num, column=7, value=data['警情风险指数'])

    # 创建汇总数据工作表
    summary_ws = wb.create_sheet(title="汇总数据")

    # 写入汇总数据表头
    summary_headers = ['派出所名称', '派出所代码', '所属地区代码', '巡防投入指数', '警情风险指数']

    for col_num, header in enumerate(summary_headers, 1):
        summary_ws.cell(row=1, column=col_num, value=header)

    # 写入汇总数据
    for row_num, data in enumerate(final_chart_data, 2):
        summary_ws.cell(row=row_num, column=1, value=data['派出所名称'])
        summary_ws.cell(row=row_num, column=2, value=data['派出所代码'])
        summary_ws.cell(row=row_num, column=3, value=data['所属地区代码'])
        summary_ws.cell(row=row_num, column=4, value=data['巡防投入指数'])
        summary_ws.cell(row=row_num, column=5, value=data['警情风险指数'])

    # 添加统计信息
    summary_ws.cell(row=len(final_chart_data) + 3, column=1, value="统计信息:")
    summary_ws.cell(row=len(final_chart_data) + 4, column=1,
                   value=f"数据时间范围: {start_time} 至 {end_time}")
    summary_ws.cell(row=len(final_chart_data) + 5, column=1,
                   value=f"派出所总数: {len(final_chart_data)}")

    # 计算巡防投入指数统计
    patrol_values = [data['巡防投入指数'] for data in final_chart_data if data['巡防投入指数'] > 0]
    if patrol_values:
        summary_ws.cell(row=len(final_chart_data) + 6, column=1,
                       value=f"巡防投入指数平均值: {round(sum(patrol_values) / len(patrol_values), 2)}")
        summary_ws.cell(row=len(final_chart_data) + 7, column=1,
                       value=f"巡防投入指数最大值: {max(patrol_values)}")
        summary_ws.cell(row=len(final_chart_data) + 8, column=1,
                       value=f"巡防投入指数最小值: {min(patrol_values)}")

    # 计算警情风险指数统计
    risk_values = [data['警情风险指数'] for data in final_chart_data if data['警情风险指数'] > 0]
    if risk_values:
        summary_ws.cell(row=len(final_chart_data) + 9, column=1,
                       value=f"警情风险指数平均值: {round(sum(risk_values) / len(risk_values), 2)}")
        summary_ws.cell(row=len(final_chart_data) + 10, column=1,
                       value=f"警情风险指数最大值: {max(risk_values)}")
        summary_ws.cell(row=len(final_chart_data) + 11, column=1,
                       value=f"警情风险指数最小值: {min(risk_values)}")

    # 设置列宽以更好地显示数据
    for ws in [patrol_ws, risk_ws, summary_ws]:
        for col in range(1, len(ws[1]) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

    return wb


def export_quadrant_chart_for_date_range_excel(start_time: str, end_time: str):
    """
    导出象限图数据（使用Excel绘制）- 原版本作为备份
    """
    # 调用新版本
    return export_quadrant_chart_for_date_range(start_time, end_time)
    
    # 验证日期范围
    if start_date > end_date:
        raise ValueError('开始时间不能晚于结束时间')
    
    # 1. 获取全市派出所列表
    logging.info("获取全市派出所列表...")
    police_stations = get_police_stations()
    logging.info(f"获取到 {len(police_stations)} 个派出所")
    
    # 2. 获取巡防投入数据
    logging.info("获取巡防投入数据...")
    cross_day_url = build_zhksh_url("/zhksh/dutySchedule/crossDayList")
    
    # 计算需要处理的日期列表
    current_date = start_date
    date_list = []
    while current_date <= end_date:
        date_list.append(current_date.strftime('%Y-%m-%d'))
        current_date += timedelta(days=1)
    
    # 获取所有日期的跨天数据
    all_qwsj_data = []
    for date_str in date_list:
        try:
            logging.info(f"获取 {date_str} 的跨天数据...")
            rows_data = get_cross_day_data(cross_day_url, date_str)
            all_qwsj_data.extend(rows_data)
        except Exception as e:
            logging.error(f"获取 {date_str} 的跨天数据失败: {e}")
            continue
    
    # 3. 计算巡防投入指数
    logging.info("计算巡防投入指数...")
    patrol_weights = calculate_patrol_input_index(all_qwsj_data, police_stations)
    
    # 4. 计算警情风险指数
    logging.info("计算警情风险指数...")
    risk_weights = calculate_risk_index(start_time, end_time, police_stations)
    
    # 5. 生成最终数据
    final_chart_data = []
    for station in police_stations:
        dept_code = station['code']
        station_name = station['name']
        
        # 获取巡防投入指数
        patrol_input = patrol_weights.get(dept_code, 0)
        
        # 获取警情风险指数
        risk_index = risk_weights.get(dept_code, 0)
        
        final_chart_data.append({
            '派出所名称': station_name,
            '派出所代码': dept_code,
            '所属地区代码': station['parent_code'],
            '巡防投入指数': patrol_input,
            '警情风险指数': risk_index
        })
    
    # 6. 创建Excel文件
    wb = Workbook()
    ws = wb.active
    ws.title = "派出所巡防警情象限图"
    
    # 写入表头
    headers = ['派出所名称', '派出所代码', '所属地区代码', '巡防投入指数', '警情风险指数']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # 写入数据
    for row_num, data in enumerate(final_chart_data, 2):
        ws.cell(row=row_num, column=1, value=data['派出所名称'])
        ws.cell(row=row_num, column=2, value=data['派出所代码'])
        ws.cell(row=row_num, column=3, value=data['所属地区代码'])
        ws.cell(row=row_num, column=4, value=data['巡防投入指数'])
        ws.cell(row=row_num, column=5, value=data['警情风险指数'])
    
    # 创建散点图（象限图）
    if len(final_chart_data) > 0:
        chart = ScatterChart()
        chart.title = f"{start_time}至{end_time}派出所巡防警情象限图"
        chart.style = 13
        chart.x_axis.title = '巡防投入指数'
        chart.y_axis.title = '警情风险指数'

        # 设置数据范围
        xvalues = Reference(ws, min_col=4, min_row=2, max_row=len(final_chart_data) + 1)
        yvalues = Reference(ws, min_col=5, min_row=2, max_row=len(final_chart_data) + 1)

        series = Series(yvalues, xvalues, title="派出所")
        from openpyxl.chart.marker import Marker
        series.marker = Marker('circle')  # 设置数据点为圆形
        series.marker.size = 6  # 设置数据点大小
        chart.series.append(series)

        # 计算数据极值和中心点
        patrol_values = [data['巡防投入指数'] for data in final_chart_data]
        risk_values = [data['警情风险指数'] for data in final_chart_data]

        # 计算最大最小值
        min_patrol = min(patrol_values)
        max_patrol = max(patrol_values)
        min_risk = min(risk_values)
        max_risk = max(risk_values)

        # 计算中心点（极值中点）
        patrol_center = (min_patrol + max_patrol) / 2
        risk_center = (min_risk + max_risk) / 2

        # 计算数据范围并添加10%边距
        patrol_range = max_patrol - min_patrol
        risk_range = max_risk - min_risk

        # 设置边距（10%）
        patrol_margin = patrol_range * 0.1 if patrol_range > 0 else 1
        risk_margin = risk_range * 0.1 if risk_range > 0 else 1

        # 设置坐标轴范围（带边距）
        try:
            chart.x_axis.scaling.min = min_patrol - patrol_margin
            chart.x_axis.scaling.max = max_patrol + patrol_margin
            chart.y_axis.scaling.min = min_risk - risk_margin
            chart.y_axis.scaling.max = max_risk + risk_margin
        except AttributeError as e:
            logging.warning(f"设置坐标轴范围时发生错误，使用默认范围: {e}")
            # 如果scaling属性不存在，跳过范围设置
        except Exception as e:
            logging.warning(f"设置坐标轴范围时发生未知错误: {e}")

        # 添加象限分割线
        from openpyxl.chart.shapes import GraphicalProperties

        # 尝试隐藏网格线（如果属性存在的话）
        try:
            # 添加水平分割线（在纵坐标中心点）
            if hasattr(chart.y_axis, 'major_gridlines'):
                chart.y_axis.major_gridlines.spPr = GraphicalProperties()
                chart.y_axis.major_gridlines.spPr.ln = None
            else:
                logging.info("y_axis.major_gridlines 属性不存在，跳过网格线设置")

            # 添加垂直分割线（在横坐标中心点）
            if hasattr(chart.x_axis, 'major_gridlines'):
                chart.x_axis.major_gridlines.spPr = GraphicalProperties()
                chart.x_axis.major_gridlines.spPr.ln = None
            else:
                logging.info("x_axis.major_gridlines 属性不存在，跳过网格线设置")
        except Exception as e:
            logging.warning(f"设置网格线时发生错误，跳过此步骤: {e}")

        # 手动添加分割线（通过添加系列实现）
        try:
            # 水平分割线
            horizontal_line_data = [risk_center] * 2

            # 安全获取坐标轴范围
            x_min = getattr(chart.x_axis.scaling, 'min', min_patrol - patrol_margin) if hasattr(chart.x_axis, 'scaling') else (min_patrol - patrol_margin)
            x_max = getattr(chart.x_axis.scaling, 'max', max_patrol + patrol_margin) if hasattr(chart.x_axis, 'scaling') else (max_patrol + patrol_margin)

            horizontal_line_x = [x_min, x_max]
            horizontal_series = Series(horizontal_line_data, horizontal_line_x, title="水平分割线")
            horizontal_series.marker = Marker()
            horizontal_series.marker.symbol = "none"
            horizontal_series.graphicalProperties = GraphicalProperties()
            horizontal_series.graphicalProperties.line = GraphicalProperties()
            horizontal_series.graphicalProperties.line.solidFill = "CCCCCC"
            horizontal_series.graphicalProperties.line.width = 0.5
            chart.series.append(horizontal_series)

            # 垂直分割线
            vertical_line_x = [patrol_center] * 2

            # 安全获取坐标轴范围
            y_min = getattr(chart.y_axis.scaling, 'min', min_risk - risk_margin) if hasattr(chart.y_axis, 'scaling') else (min_risk - risk_margin)
            y_max = getattr(chart.y_axis.scaling, 'max', max_risk + risk_margin) if hasattr(chart.y_axis, 'scaling') else (max_risk + risk_margin)

            vertical_line_y = [y_min, y_max]
            vertical_series = Series(vertical_line_y, vertical_line_x, title="垂直分割线")
            vertical_series.marker = Marker()
            vertical_series.marker.symbol = "none"
            vertical_series.graphicalProperties = GraphicalProperties()
            vertical_series.graphicalProperties.line = GraphicalProperties()
            vertical_series.graphicalProperties.line.solidFill = "CCCCCC"
            vertical_series.graphicalProperties.line.width = 0.5
            chart.series.append(vertical_series)

            logging.info("象限图分割线添加成功")
        except Exception as e:
            logging.warning(f"添加象限图分割线时发生错误，跳过此步骤: {e}")

        # 添加图表到工作表
        ws.add_chart(chart, "G2")
        
          
        # 在图表下方添加统计信息
        ws.cell(row=len(final_chart_data) + 3, column=1, value="统计信息:")
        ws.cell(row=len(final_chart_data) + 4, column=1, value=f"巡防投入指数中心点: {round(patrol_center, 2)} (范围: {round(min_patrol, 2)} - {round(max_patrol, 2)})")
        ws.cell(row=len(final_chart_data) + 5, column=1, value=f"警情风险指数中心点: {round(risk_center, 2)} (范围: {round(min_risk, 2)} - {round(max_risk, 2)})")
        ws.cell(row=len(final_chart_data) + 6, column=1, value=f"派出所总数: {len(final_chart_data)}")

        # 添加象限说明
        ws.cell(row=len(final_chart_data) + 8, column=1, value="象限说明:")
        ws.cell(row=len(final_chart_data) + 9, column=1, value="第一象限(右上): 高巡防投入，高警情风险")
        ws.cell(row=len(final_chart_data) + 10, column=1, value="第二象限(左上): 低巡防投入，高警情风险")
        ws.cell(row=len(final_chart_data) + 11, column=1, value="第三象限(左下): 低巡防投入，低警情风险")
        ws.cell(row=len(final_chart_data) + 12, column=1, value="第四象限(右下): 高巡防投入，低警情风险")
    
    logging.info(f"象限图导出完成，共 {len(final_chart_data)} 个派出所数据")
    return wb


def export_police_force_for_date_range(start_time: str, end_time: str):
    """
    根据日期范围导出巡防警力表
    """
    from openpyxl import load_workbook

    # 解析日期时间
    start_date = datetime.strptime(start_time, '%Y-%m-%d %H:%M:%S')
    end_date = datetime.strptime(end_time, '%Y-%m-%d %H:%M:%S')

    # 验证日期范围
    if start_date > end_date:
        raise ValueError('开始时间不能晚于结束时间')

    # 跨天数据接口URL
    cross_day_url = build_zhksh_url("/zhksh/dutySchedule/crossDayList")

    # 一次性获取整个时间范围的数据
    all_data = []
    try:
        logging.info(f"获取 {start_time} 至 {end_time} 的跨天数据用于导出警力表...")
        rows_data = get_cross_day_data(cross_day_url, start_time, end_time)
        all_data = rows_data
        logging.info(f"获取到 {len(all_data)} 条数据")
    except Exception as e:
        logging.error(f"获取跨天数据失败: {e}")
        all_data = []

    # 加载模板文件
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'xfjlmb.xlsx')
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"模板文件不存在: {template_path}")

    wb = load_workbook(template_path)
    ws = wb.active

    # 使用用户模板，不设置时间范围，只写入指定的统计数据单元格

    # 统计数据
    statistics = {
        '云城': {'机关单位': 0, '派出所': 0, '巡逻警察': 0, '交通警察': 0, '合计': 0},
        '云安': {'机关单位': 0, '派出所': 0, '巡逻警察': 0, '交通警察': 0, '合计': 0},
        '罗定': {'机关单位': 0, '派出所': 0, '巡逻警察': 0, '交通警察': 0, '合计': 0},
        '新兴': {'机关单位': 0, '派出所': 0, '巡逻警察': 0, '交通警察': 0, '合计': 0},
        '郁南': {'机关单位': 0, '派出所': 0, '巡逻警察': 0, '交通警察': 0, '合计': 0},
        '云浮市局交通': {'合计': 0},
        '云浮市局': {'机关单位': 0, '特警': 0}
    }

    # 遍历数据进行统计
    for item in all_data:
        dept_name = item.get('deptName', '')

        # 云城区统计
        if '云城' in dept_name:
            statistics['云城']['合计'] += 1
            if '派出所' not in dept_name and '巡逻警察' not in dept_name and '交通' not in dept_name:
                statistics['云城']['机关单位'] += 1
            if '派出所' in dept_name:
                statistics['云城']['派出所'] += 1
            if '巡逻警察' in dept_name:
                statistics['云城']['巡逻警察'] += 1
            if '交通' in dept_name:
                statistics['云城']['交通警察'] += 1

        # 云安区统计
        elif '云安' in dept_name:
            statistics['云安']['合计'] += 1
            if '派出所' not in dept_name and '巡逻警察' not in dept_name and '交通' not in dept_name:
                statistics['云安']['机关单位'] += 1
            if '派出所' in dept_name:
                statistics['云安']['派出所'] += 1
            if '巡逻警察' in dept_name:
                statistics['云安']['巡逻警察'] += 1
            if '交通' in dept_name:
                statistics['云安']['交通警察'] += 1

        # 罗定市统计
        elif '罗定' in dept_name:
            statistics['罗定']['合计'] += 1
            if '派出所' not in dept_name and '巡逻警察' not in dept_name and '交通' not in dept_name:
                statistics['罗定']['机关单位'] += 1
            if '派出所' in dept_name:
                statistics['罗定']['派出所'] += 1
            if '巡逻警察' in dept_name:
                statistics['罗定']['巡逻警察'] += 1
            if '交通' in dept_name:
                statistics['罗定']['交通警察'] += 1

        # 新兴县统计
        elif '新兴' in dept_name:
            statistics['新兴']['合计'] += 1
            if '派出所' not in dept_name and '巡逻警察' not in dept_name and '交通' not in dept_name:
                statistics['新兴']['机关单位'] += 1
            if '派出所' in dept_name:
                statistics['新兴']['派出所'] += 1
            if '巡逻警察' in dept_name:
                statistics['新兴']['巡逻警察'] += 1
            if '交通' in dept_name:
                statistics['新兴']['交通警察'] += 1

        # 郁南县统计
        elif '郁南' in dept_name:
            statistics['郁南']['合计'] += 1
            if '派出所' not in dept_name and '巡逻警察' not in dept_name and '交通' not in dept_name:
                statistics['郁南']['机关单位'] += 1
            if '派出所' in dept_name:
                statistics['郁南']['派出所'] += 1
            if '巡逻警察' in dept_name:
                statistics['郁南']['巡逻警察'] += 1
            if '交通' in dept_name:
                statistics['郁南']['交通警察'] += 1

        # 云浮市局交通警察支队统计
        elif '云浮市局交通' in dept_name:
            statistics['云浮市局交通']['合计'] += 1

        # 云浮市局统计
        elif '云浮市局' in dept_name:
            if '交通' not in dept_name and '特警' not in dept_name:
                statistics['云浮市局']['机关单位'] += 1
            if '特警' in dept_name:
                statistics['云浮市局']['特警'] += 1

    # 写入统计数据到Excel
    # 第4行：云城区
    ws.cell(row=4, column=3, value=statistics['云城']['机关单位'])  # C4
    ws.cell(row=4, column=6, value=statistics['云城']['派出所'])  # F4
    ws.cell(row=4, column=9, value=statistics['云城']['巡逻警察'])  # I4
    ws.cell(row=4, column=12, value=statistics['云城']['交通警察'])  # L4
    ws.cell(row=4, column=17, value=statistics['云城']['合计'])  # Q4

    # 第5行：云安区
    ws.cell(row=5, column=3, value=statistics['云安']['机关单位'])  # C5
    ws.cell(row=5, column=6, value=statistics['云安']['派出所'])  # F5
    ws.cell(row=5, column=9, value=statistics['云安']['巡逻警察'])  # I5
    ws.cell(row=5, column=12, value=statistics['云安']['交通警察'])  # L5
    ws.cell(row=5, column=17, value=statistics['云安']['合计'])  # Q5

    # 第6行：罗定市
    ws.cell(row=6, column=3, value=statistics['罗定']['机关单位'])  # C6
    ws.cell(row=6, column=6, value=statistics['罗定']['派出所'])  # F6
    ws.cell(row=6, column=9, value=statistics['罗定']['巡逻警察'])  # I6
    ws.cell(row=6, column=12, value=statistics['罗定']['交通警察'])  # L6
    ws.cell(row=6, column=17, value=statistics['新兴']['合计'])  # Q6（按需求写入新兴县合计）

    # 第7行：新兴县
    ws.cell(row=7, column=3, value=statistics['新兴']['机关单位'])  # C7
    ws.cell(row=7, column=6, value=statistics['新兴']['派出所'])  # F7
    ws.cell(row=7, column=9, value=statistics['新兴']['巡逻警察'])  # I7
    ws.cell(row=7, column=12, value=statistics['新兴']['交通警察'])  # L7
    ws.cell(row=7, column=17, value=statistics['郁南']['合计'])  # Q7（按需求写入郁南县合计）

    # 第8行：郁南县
    ws.cell(row=8, column=3, value=statistics['郁南']['机关单位'])  # C8
    ws.cell(row=8, column=6, value=statistics['郁南']['派出所'])  # F8
    ws.cell(row=8, column=9, value=statistics['郁南']['巡逻警察'])  # I8
    ws.cell(row=8, column=12, value=statistics['郁南']['交通警察'])  # L8
    ws.cell(row=8, column=17, value=statistics['郁南']['合计'])  # Q8

    # 第9行：云浮市局交通警察支队
    ws.cell(row=9, column=12, value=statistics['云浮市局交通']['合计'])  # L9

    # 第10行：云浮市局
    ws.cell(row=10, column=3, value=statistics['云浮市局']['机关单位'])  # C10
    ws.cell(row=10, column=9, value=statistics['云浮市局']['特警'])  # I10

    logging.info(f"警力表导出完成，统计数据: {statistics}")
    return wb
