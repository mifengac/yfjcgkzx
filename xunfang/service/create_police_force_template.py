"""
创建巡防警力表模板文件
"""
import os
from openpyxl import Workbook

def create_police_force_template():
    """创建巡防警力表模板文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "巡防警力表"

    # 设置标题（第1行）
    ws.merge_cells('A1:F1')
    ws['A1'] = "巡防警力统计表"
    from openpyxl.styles import Alignment
    ws['A1'].alignment = Alignment(horizontal='center')

    # 设置时间范围提示（第2行）
    ws['A2'] = "统计时间："

    # 设置表头（第3行）
    ws['A3'] = "地区"
    ws['B3'] = "机关单位"
    ws['C3'] = "派出所"
    ws['D3'] = "巡逻警察"
    ws['E3'] = "交通警察"
    ws['F3'] = "合计"

    # 设置地区行标签（第4-10行）
    areas = ["云城区", "云安区", "罗定市", "新兴县", "郁南县", "云浮市局交通警察支队", "云浮市局", "合计"]
    for i, area in enumerate(areas, start=4):
        ws.cell(row=i, column=1, value=area)

    # 保存模板文件
    template_path = os.path.join(os.path.dirname(__file__), '..', 'templates', 'xfjlmb.xlsx')
    wb.save(template_path)
    print(f"模板文件已创建: {template_path}")
    return template_path

if __name__ == "__main__":
    create_police_force_template()