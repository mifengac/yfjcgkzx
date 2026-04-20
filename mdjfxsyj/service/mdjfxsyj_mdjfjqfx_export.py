from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Mapping, Sequence

from flask import Response, send_file
from openpyxl import Workbook

from mdjfxsyj.service.mdjfxsyj_mdjfjqfx_service import (
    _filter_detail_rows,
    _load_cases,
    _to_detail_row,
    get_detail_payload,
    get_summary_payload,
    normalize_group_by,
    normalize_range,
)


def _append_sheet(workbook: Workbook, title: str, rows: Sequence[Mapping[str, Any]]) -> None:
    sheet = workbook.create_sheet(title=title[:31])
    if not rows:
        sheet.append(["无数据"])
        return
    headers = [key for key in rows[0].keys() if key not in {"group_code", "ori_code", "confirm_code"}]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])


def _send_workbook(workbook: Workbook, filename: str) -> Response:
    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        del workbook["Sheet"]
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def build_summary_export(**kwargs: Any) -> Response:
    payload = get_summary_payload(**kwargs)
    workbook = Workbook()
    _append_sheet(workbook, "总体统计", payload["overall"])
    _append_sheet(workbook, "细类统计", payload["fine"])
    _append_sheet(workbook, "重复统计", payload["repeat"])
    filename = f"{payload['start_time'][:10]}至{payload['end_time'][:10]}矛盾纠纷警情统计{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return _send_workbook(workbook, filename)


def build_all_details_export(**kwargs: Any) -> Response:
    _, _, start_text, end_text = normalize_range(kwargs.get("start_time"), kwargs.get("end_time"))
    group_by = normalize_group_by(kwargs.get("group_by"))
    repeat_min = int(kwargs.get("repeat_min") or 2)
    base_rows = _load_cases(start_text, end_text, kwargs.get("ssfjdm_list") or [])
    workbook = Workbook()
    dimensions = [
        ("原始警情", "original_total"),
        ("转案警情", "converted"),
        ("原始确认均纠纷性质", "both_mdj"),
        ("重复警情", "repeat"),
        ("重复转案警情", "repeat_converted"),
    ]
    for sheet_name, dimension in dimensions:
        rows = _filter_detail_rows(
            base_rows,
            group_by=group_by,
            group_code="__TOTAL__",
            dimension=dimension,
            repeat_min=repeat_min,
        )
        _append_sheet(workbook, sheet_name, [_to_detail_row(row) for row in rows])
    filename = f"{start_text[:10]}至{end_text[:10]}矛盾纠纷警情统计详情{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return _send_workbook(workbook, filename)


def build_detail_export(**kwargs: Any) -> Response:
    payload = get_detail_payload(page_size=0, **kwargs)
    workbook = Workbook()
    _append_sheet(workbook, "详情", payload["rows"])
    filename = f"{payload['start_time'][:10]}至{payload['end_time'][:10]}{kwargs.get('dimension') or '详情'}统计{datetime.now():%Y%m%d%H%M%S}.xlsx"
    return _send_workbook(workbook, filename)
