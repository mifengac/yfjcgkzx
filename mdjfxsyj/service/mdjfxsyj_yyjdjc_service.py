from __future__ import annotations

import csv
from datetime import datetime, time, timedelta
from io import BytesIO, StringIO
from typing import Any, Dict, List, Optional, Sequence, Tuple

from flask import Response, send_file
from openpyxl import Workbook

from gonggong.service.upstream_jingqing_client import api_client
from mdjfxsyj.dao.mdjfxsyj_yyjdjc_dao import query_dispute_rows, query_workorder_rows


DEFAULT_KEYWORDS: Tuple[str, ...] = (
    "杀",
    "扬言",
    "极端",
    "报复",
    "自杀",
    "轻生",
    "自残",
    "寻死",
    "跳楼",
    "跳桥",
    "跳河",
    "同归于尽",
    "报复社会",
    "炸弹",
    "爆炸",
    "放火",
    "行凶",
    "砍人",
    "捅人",
    "袭击",
    "灭门",
)

SOURCE_SPECS: Dict[str, Dict[str, Any]] = {
    "police": {
        "label": "涉扬言极端警情",
        "sheet_name": "涉扬言极端警情",
        "columns": [
            "警情编号",
            "报警时间",
            "报警电话",
            "报警地址",
            "报警内容",
            "处警情况",
            "原始警情性质",
            "所属分局",
            "所属派出所",
            "命中关键词",
        ],
    },
    "workorder": {
        "label": "涉扬言极端12345工单",
        "sheet_name": "涉扬言极端12345工单",
        "columns": [
            "业务编号",
            "来电编号",
            "紧急程度",
            "来电号码",
            "诉求人",
            "诉求内容",
            "工单标题",
            "事项分类",
            "来电时间",
            "所属机构",
            "处理单位",
            "诉求范围",
            "工单状态",
            "命中关键词",
        ],
    },
    "dispute": {
        "label": "涉扬言极端矛盾纠纷事件",
        "sheet_name": "涉扬言极端矛盾纠纷事件",
        "columns": [
            "业务流水号",
            "纠纷名称",
            "纠纷类型",
            "是否公安职责范围",
            "发生时间",
            "风险等级",
            "简要情况",
            "纠纷缘由",
            "上报人联系电话",
            "调处方案",
            "发生地址",
            "所属分局",
            "所属派出所",
            "登记时间",
            "修改人姓名",
            "命中关键词",
        ],
    },
}

_CASE_PAGE_SIZE = 1000


def _parse_datetime(value: str) -> datetime:
    text = str(value or "").strip().replace("T", " ")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(text, fmt)
            if fmt == "%Y-%m-%d":
                return datetime.combine(parsed.date(), time(0, 0, 0))
            if fmt == "%Y-%m-%d %H:%M":
                return parsed.replace(second=0)
            return parsed
        except ValueError:
            continue
    raise ValueError("时间格式错误，请使用 YYYY-MM-DD HH:MM:SS")


def default_range() -> Tuple[datetime, datetime]:
    today = datetime.now().date()
    end_dt = datetime.combine(today, time(0, 0, 0))
    start_dt = end_dt - timedelta(days=8)
    return start_dt, end_dt


def normalize_range(start_time: Optional[str], end_time: Optional[str]) -> Tuple[datetime, datetime, str, str]:
    if start_time and end_time:
        start_dt = _parse_datetime(start_time)
        end_dt = _parse_datetime(end_time)
    else:
        start_dt, end_dt = default_range()

    if start_dt > end_dt:
        raise ValueError("开始时间不能晚于结束时间")

    start_text = start_dt.strftime("%Y-%m-%d %H:%M:%S")
    end_text = end_dt.strftime("%Y-%m-%d %H:%M:%S")
    return start_dt, end_dt, start_text, end_text


def _stringify_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def find_matched_keywords(texts: Sequence[Any], keywords: Sequence[str] = DEFAULT_KEYWORDS) -> List[str]:
    haystack = "\n".join(_stringify_value(text) for text in texts if text not in (None, ""))
    matched: List[str] = []
    seen = set()
    for keyword in keywords:
        if keyword and keyword not in seen and keyword in haystack:
            seen.add(keyword)
            matched.append(keyword)
    return matched


def _build_export_basename(start_text: str, end_text: str, suffix: str) -> str:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    return f"{start_text[:10]}-{end_text[:10]}{suffix}{timestamp}"


def _build_case_payload(start_text: str, end_text: str, code_csv: str, page_num: int) -> Dict[str, str]:
    return {
        "params[colArray]": "",
        "beginDate": start_text,
        "endDate": end_text,
        "newCaseSourceNo": "",
        "newCaseSource": "全部",
        "dutyDeptNo": "",
        "dutyDeptName": "全部",
        "newCharaSubclassNo": "",
        "newCharaSubclass": "全部",
        "newOriCharaSubclassNo": code_csv,
        "newOriCharaSubclass": "",
        "caseNo": "",
        "callerName": "",
        "callerPhone": "",
        "phoneAddress": "",
        "callerIdentity": "",
        "operatorNo": "",
        "operatorName": "",
        "params[isInvalidCase]": "",
        "occurAddress": "",
        "caseMarkNo": "",
        "caseMark": "全部",
        "params[repetitionCase]": "",
        "params[originalDuplicateCase]": "",
        "params[startTimePeriod]": "",
        "params[endTimePeriod]": "",
        "caseContents": "",
        "replies": "",
        "params[sinceRecord]": "",
        "dossierResult": "",
        "params[isVideo]": "",
        "params[isConversation]": "",
        "pageSize": str(_CASE_PAGE_SIZE),
        "pageNum": str(page_num),
        "orderByColumn": "callTime",
        "isAsc": "desc",
    }


def _get_yangyan_case_codes() -> str:
    response = api_client.request_with_retry("GET", "/dsjfx/nature/treeNewViewData", timeout=20)
    if response is None or response.status_code != 200:
        raise RuntimeError("警情性质树接口调用失败")

    try:
        payload = response.json()
    except Exception as exc:  # noqa: BLE001
        raise RuntimeError(f"警情性质树接口返回解析失败: {exc}") from exc

    if not isinstance(payload, list):
        raise RuntimeError("警情性质树接口返回格式异常")

    codes: List[str] = []
    seen = set()
    for item in payload:
        if not isinstance(item, dict):
            continue
        code = str(item.get("id") or "").strip()
        if not code or code.startswith("12") or code in seen:
            continue
        seen.add(code)
        codes.append(code)

    if not codes:
        raise RuntimeError("警情性质树接口未返回可用编码")
    return ",".join(codes)


def fetch_police_rows(*, start_text: str, end_text: str, keywords: Sequence[str]) -> List[Dict[str, str]]:
    code_csv = _get_yangyan_case_codes()
    all_rows: List[Dict[str, str]] = []
    page_num = 1
    total: Optional[int] = None

    while True:
        payload = _build_case_payload(start_text, end_text, code_csv, page_num)
        result = api_client.get_case_list(payload)
        if not isinstance(result, dict):
            raise RuntimeError("警情接口返回格式异常")

        code = result.get("code")
        if code == -1:
            raise RuntimeError("警情接口登录失效或请求超时")
        if code not in (None, 0):
            raise RuntimeError(f"警情接口返回异常，code={code}，msg={result.get('msg', '')}")

        raw_rows = result.get("rows") or []
        if not isinstance(raw_rows, list):
            raise RuntimeError("警情接口 rows 格式异常")

        if total is None:
            try:
                total = int(result.get("total", 0) or 0)
            except Exception:
                total = 0

        for raw_row in raw_rows:
            if not isinstance(raw_row, dict):
                continue
            matched_keywords = find_matched_keywords(
                [raw_row.get("caseContents"), raw_row.get("replies")],
                keywords,
            )
            if not matched_keywords:
                continue
            all_rows.append(
                {
                    "警情编号": _stringify_value(raw_row.get("caseNo")),
                    "报警时间": _stringify_value(raw_row.get("callTime")),
                    "报警电话": _stringify_value(raw_row.get("callerPhone")),
                    "报警地址": _stringify_value(raw_row.get("occurAddress")),
                    "报警内容": _stringify_value(raw_row.get("caseContents")),
                    "处警情况": _stringify_value(raw_row.get("replies")),
                    "原始警情性质": _stringify_value(
                        raw_row.get("newOriCharaSubclass") or raw_row.get("newOriCharaSubclassName")
                    ),
                    "所属分局": _stringify_value(raw_row.get("cmdName")),
                    "所属派出所": _stringify_value(raw_row.get("dutyDeptName")),
                    "命中关键词": "、".join(matched_keywords),
                }
            )

        if not raw_rows:
            break
        if len(raw_rows) < _CASE_PAGE_SIZE:
            break
        if total is not None and page_num * _CASE_PAGE_SIZE >= total:
            break
        page_num += 1

    return all_rows


def _normalize_query_rows(rows: Sequence[Dict[str, Any]], text_fields: Sequence[str], keywords: Sequence[str]) -> List[Dict[str, str]]:
    normalized_rows: List[Dict[str, str]] = []
    for row in rows:
        matched_keywords = find_matched_keywords([row.get(field_name) for field_name in text_fields], keywords)
        if not matched_keywords:
            continue
        normalized_row = {key: _stringify_value(value) for key, value in row.items()}
        normalized_row["命中关键词"] = "、".join(matched_keywords)
        normalized_rows.append(normalized_row)
    return normalized_rows


def fetch_workorder_source_rows(
    *,
    start_dt: datetime,
    end_dt: datetime,
    keywords: Sequence[str],
) -> List[Dict[str, str]]:
    rows = query_workorder_rows(start_time=start_dt, end_time=end_dt, keywords=keywords)
    return _normalize_query_rows(rows, ["诉求内容"], keywords)


def fetch_dispute_source_rows(
    *,
    start_dt: datetime,
    end_dt: datetime,
    keywords: Sequence[str],
) -> List[Dict[str, str]]:
    rows = query_dispute_rows(start_time=start_dt, end_time=end_dt, keywords=keywords)
    return _normalize_query_rows(rows, ["简要情况", "纠纷缘由"], keywords)


def get_monitor_data(*, start_time: Optional[str], end_time: Optional[str]) -> Dict[str, Any]:
    start_dt, end_dt, start_text, end_text = normalize_range(start_time, end_time)
    source_results: Dict[str, Dict[str, Any]] = {}

    fetchers = {
        "police": lambda: fetch_police_rows(
            start_text=start_text,
            end_text=end_text,
            keywords=DEFAULT_KEYWORDS,
        ),
        "workorder": lambda: fetch_workorder_source_rows(
            start_dt=start_dt,
            end_dt=end_dt,
            keywords=DEFAULT_KEYWORDS,
        ),
        "dispute": lambda: fetch_dispute_source_rows(
            start_dt=start_dt,
            end_dt=end_dt,
            keywords=DEFAULT_KEYWORDS,
        ),
    }

    for source_key, spec in SOURCE_SPECS.items():
        try:
            rows = fetchers[source_key]()
            source_results[source_key] = {
                "label": spec["label"],
                "columns": spec["columns"],
                "count": len(rows),
                "rows": rows,
                "error": "",
            }
        except Exception as exc:  # noqa: BLE001
            source_results[source_key] = {
                "label": spec["label"],
                "columns": spec["columns"],
                "count": 0,
                "rows": [],
                "error": str(exc),
            }

    return {
        "start_time": start_text,
        "end_time": end_text,
        "keywords": list(DEFAULT_KEYWORDS),
        "sources": source_results,
    }


def _append_rows_to_sheet(sheet: Any, columns: Sequence[str], rows: Sequence[Dict[str, Any]], error: str = "") -> None:
    if error:
        sheet.append(["错误信息"])
        sheet.append([error])
        return

    sheet.append(list(columns))
    if not rows:
        return
    for row in rows:
        sheet.append([_stringify_value(row.get(column)) for column in columns])


def _build_csv_buffer(columns: Sequence[str], rows: Sequence[Dict[str, Any]], error: str = "") -> BytesIO:
    stream = StringIO()
    writer = csv.writer(stream)
    if error:
        writer.writerow(["错误信息"])
        writer.writerow([error])
    else:
        writer.writerow(list(columns))
        for row in rows:
            writer.writerow([_stringify_value(row.get(column)) for column in columns])
    return BytesIO(stream.getvalue().encode("utf-8-sig"))


def build_all_sources_export(*, start_time: Optional[str], end_time: Optional[str]) -> Response:
    payload = get_monitor_data(start_time=start_time, end_time=end_time)
    workbook = Workbook()
    first_sheet = workbook.active
    first_sheet.title = SOURCE_SPECS["police"]["sheet_name"]
    _append_rows_to_sheet(
        first_sheet,
        SOURCE_SPECS["police"]["columns"],
        payload["sources"]["police"]["rows"],
        payload["sources"]["police"]["error"],
    )

    for source_key in ("workorder", "dispute"):
        sheet = workbook.create_sheet(title=SOURCE_SPECS[source_key]["sheet_name"])
        _append_rows_to_sheet(
            sheet,
            SOURCE_SPECS[source_key]["columns"],
            payload["sources"][source_key]["rows"],
            payload["sources"][source_key]["error"],
        )

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = _build_export_basename(payload["start_time"], payload["end_time"], "扬言极端") + ".xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _fetch_single_source_rows(source: str, start_time: Optional[str], end_time: Optional[str]) -> Tuple[str, str, List[str], List[Dict[str, str]]]:
    if source not in SOURCE_SPECS:
        raise ValueError("未知数据源")

    start_dt, end_dt, start_text, end_text = normalize_range(start_time, end_time)
    if source == "police":
        rows = fetch_police_rows(start_text=start_text, end_text=end_text, keywords=DEFAULT_KEYWORDS)
    elif source == "workorder":
        rows = fetch_workorder_source_rows(start_dt=start_dt, end_dt=end_dt, keywords=DEFAULT_KEYWORDS)
    else:
        rows = fetch_dispute_source_rows(start_dt=start_dt, end_dt=end_dt, keywords=DEFAULT_KEYWORDS)
    return start_text, end_text, SOURCE_SPECS[source]["columns"], rows


def build_source_export(*, source: str, export_format: str, start_time: Optional[str], end_time: Optional[str]) -> Response:
    start_text, end_text, columns, rows = _fetch_single_source_rows(source, start_time, end_time)
    suffix = SOURCE_SPECS[source]["label"]
    base_name = _build_export_basename(start_text, end_text, suffix)

    if export_format == "csv":
        buffer = _build_csv_buffer(columns, rows)
        buffer.seek(0)
        return send_file(
            buffer,
            as_attachment=True,
            download_name=base_name + ".csv",
            mimetype="text/csv; charset=utf-8",
        )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SOURCE_SPECS[source]["sheet_name"]
    _append_rows_to_sheet(sheet, columns, rows)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=base_name + ".xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
