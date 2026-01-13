from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Optional
import io

from docxtpl import DocxTemplate
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import pandas as pd
import re


@dataclass
class BiaochezhajieReportResult:
    content: bytes
    filename: str


_TREND_COLUMNS = {"C", "K", "L", "N", "R", "U", "V"}


def _format_date_cn(date_value: datetime) -> str:
    return date_value.strftime("%Y年%m月%d日")


def _compute_this_week_range(today: datetime) -> str:
    yesterday = today - timedelta(days=1)
    monday = today - timedelta(days=today.weekday())
    return f"{monday.strftime('%Y年%m月%d日')}-{yesterday.strftime('%m月%d日')}"


def _format_cell_value(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, (int, float)):
        number_format = (getattr(cell, "number_format", "") or "").lower()
        if "%" in number_format:
            return f"{value * 100:.2f}%"
        if isinstance(value, float) and abs(value - round(value)) < 1e-9:
            return str(int(round(value)))
        if isinstance(value, float):
            return f"{value:.2f}"
        return str(value)

    text = str(value).strip()
    if text in {"", "-"}:
        return text

    m = re.fullmatch(r"([+-]?)(\d+)(?:\.(\d+))?(%)?", text)
    if not m:
        return text

    sign, integer_part, decimal_part, percent = m.groups()
    decimal_part = decimal_part or ""
    normalized = f"{sign}{integer_part}.{decimal_part}" if decimal_part else f"{sign}{integer_part}"
    try:
        num = float(normalized)
    except ValueError:
        return text

    formatted = f"{num:.2f}"
    if percent:
        return f"{formatted}%"
    return formatted


def _parse_numeric_for_trend(text_value: str) -> Optional[float]:
    if text_value is None:
        return None
    text = str(text_value).strip()
    if text == "" or text == "-":
        return None
    is_percent = text.endswith("%")
    if is_percent:
        text = text[:-1]
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def _parse_float_any(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return _parse_numeric_for_trend(str(value))


def _apply_trend_prefix(value_text: str) -> str:
    num = _parse_numeric_for_trend(value_text)
    if num is None:
        return value_text
    if abs(num) < 1e-12:
        return "持平"
    if num > 0:
        stripped = value_text.lstrip("+").strip()
        return f"上升{stripped}"
    stripped = value_text.strip()
    if stripped.startswith("-"):
        stripped = stripped[1:]
    return f"下降{stripped}"


def _build_result_text(ws, context: Dict[str, Any]) -> Dict[str, str]:
    def cell_text(addr: str) -> str:
        return str(context.get(addr, "") or "")

    def join_segments(segments):
        segments = [s for s in segments if s]
        if not segments:
            return ""
        return ";".join(segments) + ";"

    rows = range(8, 13)

    result1_segments = []
    for r in rows:
        result1_segments.append(
            f"{cell_text(f'A{r}')}{cell_text(f'B{r}')}起,同比{cell_text(f'C{r}')}"
        )

    result2_segments = []
    for r in rows:
        result2_segments.append(
            f"{cell_text(f'A{r}')}{cell_text(f'J{r}')}起,同比{cell_text(f'K{r}')},环比{cell_text(f'L{r}')}"
        )

    result3_segments = []
    for r in rows:
        t_value = _parse_float_any(ws[f"T{r}"].value)
        if t_value is None or t_value <= 0:
            continue
        result3_segments.append(
            f"{cell_text(f'A{r}')}{cell_text(f'T{r}')}起,同比{cell_text(f'U{r}')},环比{cell_text(f'V{r}')}"
        )

    result4_segments = []
    for r in rows:
        result4_segments.append(
            f"{cell_text(f'A{r}')}{cell_text(f'H{r}')}条,查处比值{cell_text(f'I{r}')},万人查处比{cell_text(f'F{r}')},人均查处{cell_text(f'G{r}')}起"
        )

    result5_segments = []
    for r in rows:
        o_value = cell_text(f"O{r}")
        if o_value.strip() == "-":
            result5_segments.append(f"{cell_text(f'A{r}')}人员查处数据{cell_text(f'M{r}')}条")
        else:
            result5_segments.append(
                f"{cell_text(f'A{r}')}人员查处数据{cell_text(f'M{r}')}条,查处比值{o_value}"
            )

    result6_segments = []
    for r in rows:
        result6_segments.append(
            f"{cell_text(f'A{r}')}当日查处{cell_text(f'Q{r}')}条、任务完成率{cell_text(f'S{r}')}"
        )

    return {
        "result1": join_segments(result1_segments),
        "result2": join_segments(result2_segments),
        "result3": join_segments(result3_segments),
        "result4": join_segments(result4_segments),
        "result5": join_segments(result5_segments),
        "result6": join_segments(result6_segments),
    }


def _load_worksheet_from_bytes(file_bytes: bytes, filename: str):
    """
    根据文件扩展名选择合适的库读取 Excel 文件。

    支持 .xlsx (openpyxl) 和 .xls (pandas + xlrd) 格式。
    返回一个与 openpyxl worksheet 兼容的对象，至少需要以下属性:
    - iter_rows(): 遍历行
    - calculate_dimension(): 返回范围
    - 以及 cell 对象需要: value, coordinate, column_letter, number_format
    """
    is_xls = filename.lower().endswith(".xls")

    if is_xls:
        # 使用 pandas 读取 .xls 文件
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None, engine="xlrd")

        class XlsCell:
            """模拟 openpyxl 的 Cell 对象"""
            def __init__(self, value, row_idx, col_idx):
                self.value = value
                self.row = row_idx
                self.column = col_idx
                self.coordinate = f"{_col_number_to_letter(col_idx)}{row_idx}"
                self.column_letter = _col_number_to_letter(col_idx)
                self.number_format = ""

        class XlsWorksheet:
            """模拟 openpyxl 的 Worksheet 对象"""
            def __init__(self, df):
                self.df = df
                self.max_row = len(df)
                self.max_column = len(df.columns) if len(df.columns) > 0 else 1

            def calculate_dimension(self):
                return f"A1:{_col_number_to_letter(self.max_column)}{self.max_row}"

            def __getitem__(self, key):
                """支持单元格坐标访问，如 ws['A1']"""
                # 解析坐标，如 "A1" -> col=A, row=1
                match = re.match(r"([A-Z]+)(\d+)", str(key))
                if not match:
                    raise KeyError(f"Invalid cell coordinate: {key}")

                col_letters = match.group(1)
                row_idx = int(match.group(2))
                col_idx = _letter_to_col_number(col_letters)

                # 获取单元格值
                df_row_idx = row_idx - 1
                df_col_idx = col_idx - 1
                value = None
                if 0 <= df_row_idx < len(self.df) and 0 <= df_col_idx < len(self.df.columns):
                    val = self.df.iloc[df_row_idx, df_col_idx]
                    if pd.notna(val):
                        value = val

                return XlsCell(value, row_idx, col_idx)

            def iter_rows(self, min_row=None, max_row=None, min_col=None, max_col=None):
                min_row = min_row or 1
                max_row = max_row or self.max_row
                min_col = min_col or 1
                max_col = max_col or self.max_column

                for row_idx in range(min_row, max_row + 1):
                    row_cells = []
                    for col_idx in range(min_col, max_col + 1):
                        df_row_idx = row_idx - 1
                        df_col_idx = col_idx - 1
                        value = None
                        if 0 <= df_row_idx < len(self.df) and 0 <= df_col_idx < len(self.df.columns):
                            val = self.df.iloc[df_row_idx, df_col_idx]
                            # 处理 pandas 的 NaN 值
                            if pd.notna(val):
                                value = val
                        row_cells.append(XlsCell(value, row_idx, col_idx))
                    yield row_cells

        return XlsWorksheet(df)
    else:
        # 使用 openpyxl 读取 .xlsx 文件
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        return wb.worksheets[0]


def _col_number_to_letter(col_num: int) -> str:
    """将列号转换为字母（如 1 -> A, 27 -> AA）"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(col_num % 26 + ord('A')) + result
        col_num //= 26
    return result


def _letter_to_col_number(letters: str) -> int:
    """将列字母转换为列号（如 A -> 1, AA -> 27）"""
    result = 0
    for char in letters:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result


def generate_biaochezhajie_report(
    file_bytes: bytes, filename: str, template_path: Path, today: Optional[datetime] = None
) -> BiaochezhajieReportResult:
    if today is None:
        today = datetime.now()

    ws = _load_worksheet_from_bytes(file_bytes, filename)

    context: Dict[str, Any] = {}
    context["time"] = _format_date_cn(today)
    context["last_time"] = _format_date_cn(today - timedelta(days=1))
    context["this_week"] = _compute_this_week_range(today)

    min_col, min_row, max_col, max_row = range_boundaries(ws.calculate_dimension())
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            coord = getattr(cell, "coordinate", None)
            if not coord:
                continue
            raw_text = _format_cell_value(cell)
            if getattr(cell, "column_letter", "") in _TREND_COLUMNS and raw_text:
                raw_text = _apply_trend_prefix(raw_text)
            context[coord] = raw_text

    context.update(_build_result_text(ws, context))

    doc = DocxTemplate(str(template_path))
    doc.render(context)
    output = io.BytesIO()
    doc.save(output)
    output.seek(0)

    filename = f"飙车炸街日报_{today.strftime('%Y%m%d')}.docx"
    return BiaochezhajieReportResult(content=output.getvalue(), filename=filename)
