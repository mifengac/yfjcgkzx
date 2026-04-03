from __future__ import annotations

import io
import logging
import re
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Any, Dict, Iterable, List, Mapping, Sequence, Tuple

import openpyxl
from openpyxl import Workbook

from jingqing_fenxi.service.jingqing_api_client import api_client
from jingqing_fenxi.service.jingqing_fenxi_service import (
    calc_duty_dept,
    calc_repeat_address,
    calc_repeat_phone,
    calc_time_hourly_counts,
    calc_time_period,
    fetch_all_case_list,
    fetch_srr_list,
)


logger = logging.getLogger(__name__)


FIGHT_TOPIC_PARENT_ID = "79958C902AE14BBDBB3F1FD9AD6AA3FC"
FIGHT_TOPIC_UPSTREAM_PAGE_SIZE = 5000
FIGHT_TOPIC_DIMENSIONS = ["srr", "time", "dept", "phone", "cluster", "addr", "reason"]

FIGHT_REASON_RULES: List[Tuple[str, List[str]]] = [
    ("酒后冲突", ["酒后", "醉酒", "喝酒", "饮酒"]),
    ("感情/家庭纠纷", ["感情", "情感", "婚恋", "夫妻", "老婆", "老公", "前男友", "前女友", "家暴", "离婚", "吃醋"]),
    ("经济/债务纠纷", ["欠钱", "借钱", "还钱", "债务", "货款", "赔偿", "经济纠纷"]),
    ("交通/停车纠纷", ["停车", "挪车", "车位", "刮蹭", "会车", "超车", "别车", "电动车", "摩托车"]),
    ("消费/服务纠纷", ["消费", "买单", "结账", "餐费", "顾客", "商家", "店员", "店主", "外卖"]),
    ("劳资/工地纠纷", ["工地", "工人", "包工头", "劳务", "工厂"]),
    ("校园/未成年人冲突", ["学校", "学生", "同学", "校园", "宿舍", "班级", "放学"]),
    ("邻里/口角纠纷", ["邻居", "邻里", "口角", "争吵", "吵架", "谩骂"]),
    ("寻衅/报复冲突", ["寻衅", "挑衅", "报复", "堵门", "纠集", "滋事"]),
]

FIGHT_TOPIC_SRR_TITLE = "各地同比环比"
UNCLASSIFIED_ADDRESS_LABEL = "未识别/空地址"
OTHER_REASON_LABEL = "其他原因"


def default_time_range() -> Tuple[str, str]:
    now = datetime.now()
    end_dt = datetime(now.year, now.month, now.day, 0, 0, 0)
    start_dt = end_dt - timedelta(days=7)
    return _format_datetime(start_dt), _format_datetime(end_dt)


def normalize_dimensions(dimensions_selected: Sequence[str] | None) -> List[str]:
    normalized: List[str] = []
    seen = set()
    for dim in dimensions_selected or []:
        value = str(dim or "").strip()
        if not value or value not in FIGHT_TOPIC_DIMENSIONS or value in seen:
            continue
        seen.add(value)
        normalized.append(value)
    return normalized


def _normalize_datetime(value: Any) -> str:
    text = str(value or "").strip().replace("T", " ")
    if not text:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text + " 00:00:00"
    if re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}$", text):
        return text + ":00"
    if not re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$", text):
        raise ValueError("时间格式不正确")
    return text


def _parse_datetime(value: Any) -> datetime | None:
    normalized = _normalize_datetime(value)
    if not normalized:
        return None
    return datetime.strptime(normalized, "%Y-%m-%d %H:%M:%S")


def _format_datetime(value: datetime) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S")


def _shift_year_safe(value: datetime, years: int) -> datetime:
    target_year = value.year + years
    try:
        return value.replace(year=target_year)
    except ValueError:
        return value.replace(year=target_year, month=2, day=28)


def _parse_int(value: Any, default: int | None = None) -> int | None:
    try:
        return int(str(value).strip())
    except Exception:
        return default


def _clamp(value: int, minimum: int, maximum: int) -> int:
    return max(minimum, min(value, maximum))


def _build_analysis_options(params: Mapping[str, Any]) -> Dict[str, Any]:
    valid_time_buckets = [1, 2, 3, 4, 6, 8, 12]
    time_bucket_hours = _parse_int(params.get("timeBucketHours"), 3)
    if time_bucket_hours not in valid_time_buckets:
        time_bucket_hours = 3

    dept_top_n_raw = str(params.get("deptTopN", "all")).strip().lower()
    if dept_top_n_raw in ("", "all", "0"):
        dept_top_n = None
    else:
        dept_top_n = _parse_int(dept_top_n_raw)
        if not dept_top_n or dept_top_n < 1:
            dept_top_n = None

    repeat_phone_min_count = _parse_int(params.get("repeatPhoneMinCount"), 2)
    if repeat_phone_min_count is None:
        repeat_phone_min_count = 2
    repeat_phone_min_count = _clamp(repeat_phone_min_count, 2, 10)

    repeat_addr_radius_meters = _parse_int(params.get("repeatAddrRadiusMeters"), 50)
    if repeat_addr_radius_meters is None:
        repeat_addr_radius_meters = 50
    repeat_addr_radius_meters = _clamp(repeat_addr_radius_meters, 50, 500)
    repeat_addr_radius_meters = int(round(repeat_addr_radius_meters / 50.0) * 50)
    repeat_addr_radius_meters = _clamp(repeat_addr_radius_meters, 50, 500)

    return {
        "timeBucketHours": time_bucket_hours,
        "deptTopN": dept_top_n,
        "repeatPhoneMinCount": repeat_phone_min_count,
        "repeatAddrRadiusMeters": repeat_addr_radius_meters,
    }


def _resolve_main_time_range(params: Mapping[str, Any]) -> Tuple[str, str, datetime, datetime]:
    default_start, default_end = default_time_range()
    begin_date = _normalize_datetime(params.get("beginDate") or default_start)
    end_date = _normalize_datetime(params.get("endDate") or default_end)
    start_dt = _parse_datetime(begin_date)
    end_dt = _parse_datetime(end_date)
    if not start_dt or not end_dt:
        raise ValueError("开始时间或结束时间格式不正确")
    if end_dt < start_dt:
        raise ValueError("结束时间不能早于开始时间")
    return begin_date, end_date, start_dt, end_dt


def _resolve_m2m_time_range(params: Mapping[str, Any], start_dt: datetime, end_dt: datetime) -> Tuple[str, str]:
    duration = end_dt - start_dt
    default_m2m_end = start_dt
    default_m2m_start = default_m2m_end - duration

    m2m_start_time = _normalize_datetime(params.get("m2mStartTime") or _format_datetime(default_m2m_start))
    m2m_end_time = _normalize_datetime(params.get("m2mEndTime") or _format_datetime(default_m2m_end))

    m2m_start_dt = _parse_datetime(m2m_start_time)
    m2m_end_dt = _parse_datetime(m2m_end_time)
    if not m2m_start_dt or not m2m_end_dt:
        raise ValueError("环比时间格式不正确")
    if m2m_end_dt < m2m_start_dt:
        raise ValueError("环比结束时间不能早于环比开始时间")
    return m2m_start_time, m2m_end_time


def _resolve_y2y_time_range(start_dt: datetime, end_dt: datetime) -> Tuple[str, str]:
    return _format_datetime(_shift_year_safe(start_dt, -1)), _format_datetime(_shift_year_safe(end_dt, -1))


def _get_fight_topic_tree_nodes() -> List[Dict[str, Any]]:
    return list(api_client.get_tree_view_data() or [])


def resolve_fight_topic_tags(tree_nodes: Sequence[Mapping[str, Any]] | None = None) -> Tuple[str, str]:
    tags: List[str] = []
    names: List[str] = []
    tag_seen = set()
    name_seen = set()
    nodes = _get_fight_topic_tree_nodes() if tree_nodes is None else list(tree_nodes)
    for node in nodes:
        if str(node.get("pId") or "").strip() != FIGHT_TOPIC_PARENT_ID:
            continue
        tag = str(node.get("tag") or "").strip()
        name = str(node.get("name") or "").strip()
        if tag and tag not in tag_seen:
            tag_seen.add(tag)
            tags.append(tag)
        if tag and name and name not in name_seen:
            name_seen.add(name)
            names.append(name)
    if not tags:
        raise ValueError("未获取到打架斗殴专题的警情类型编码")
    return ",".join(tags), ",".join(names)


def _build_case_payload(begin_date: str, end_date: str, tag_csv: str) -> Dict[str, Any]:
    return {
        "params[colArray]": "",
        "beginDate": begin_date,
        "endDate": end_date,
        "newCaseSourceNo": "",
        "newCaseSource": "全部",
        "dutyDeptNo": "",
        "dutyDeptName": "全部",
        "newCharaSubclassNo": tag_csv,
        "newCharaSubclass": "全部",
        "newOriCharaSubclassNo": "",
        "newOriCharaSubclass": "",
        "caseNo": "",
        "callerName": "",
        "callerPhone": "",
        "phoneAddress": "",
        "callerIdentity": "",
        "operatorNo": "",
        "operatorName": "",
        "params[isInvalidCase]": "",
        "occurAddress": "",
        "caseMarkNo": "",
        "caseMark": "全部",
        "params[repetitionCase]": "",
        "params[originalDuplicateCase]": "",
        "params[startTimePeriod]": "",
        "params[endTimePeriod]": "",
        "caseContents": "",
        "replies": "",
        "params[sinceRecord]": "",
        "dossierResult": "",
        "params[isVideo]": "",
        "params[isConversation]": "",
        "pageSize": FIGHT_TOPIC_UPSTREAM_PAGE_SIZE,
        "pageNum": 1,
        "orderByColumn": "callTime",
        "isAsc": "desc",
    }


def _build_srr_payload(
    begin_date: str,
    end_date: str,
    m2m_start_time: str,
    m2m_end_time: str,
    tag_csv: str,
    chara_name_csv: str,
) -> Dict[str, Any]:
    start_dt = _parse_datetime(begin_date)
    end_dt = _parse_datetime(end_date)
    if not start_dt or not end_dt:
        raise ValueError("开始时间或结束时间格式不正确")
    y2y_start_time, y2y_end_time = _resolve_y2y_time_range(start_dt, end_dt)
    return {
        "params[startTime]": begin_date,
        "params[endTime]": end_date,
        "params[y2yStartTime]": y2y_start_time,
        "params[y2yEndTime]": y2y_end_time,
        "params[m2mStartTime]": m2m_start_time,
        "params[m2mEndTime]": m2m_end_time,
        "charaNo": tag_csv,
        "chara": chara_name_csv,
        "groupField": "duty_dept_no",
        "charaType": "chara_ori",
        "charaLevel": "1",
        "caseLevel": "",
        "dutyDeptNo": "",
        "dutyDeptName": "全部",
        "newRecvType": "",
        "newRecvTypeName": "全部",
        "newCaseSourceNo": "",
        "newCaseSource": "全部",
        "params[searchAnd]": "",
        "params[searchOr]": "",
        "params[searchNot]": "",
        "caseContents": "on",
        "replies": "on",
        "pageNum": "NaN",
        "orderByColumn": "",
        "isAsc": "asc",
    }


def _predict_address_labels(unique_texts: Sequence[str]) -> List[Tuple[str, float]]:
    from xunfang.service.jiemiansanlei_service import predict_addresses

    return predict_addresses(unique_texts)


def summarize_address_labels(rows: Sequence[Dict[str, Any]]) -> Tuple[List[Tuple[str, int]], Dict[str, Any] | None]:
    unique_texts: List[str] = []
    seen = set()
    for row in rows:
        text = str(row.get("occurAddress") or "").strip()
        if text in seen:
            continue
        seen.add(text)
        unique_texts.append(text)

    prediction_map: Dict[str, Tuple[str, float]] = {}
    if unique_texts:
        try:
            non_empty_texts = [text for text in unique_texts if text]
            preds = _predict_address_labels(non_empty_texts) if non_empty_texts else []
            pred_idx = 0
            for text in unique_texts:
                if not text:
                    prediction_map[text] = (UNCLASSIFIED_ADDRESS_LABEL, 0.0)
                    continue
                label, prob = preds[pred_idx]
                pred_idx += 1
                normalized_label = str(label or "").strip() or UNCLASSIFIED_ADDRESS_LABEL
                try:
                    normalized_prob = float(prob or 0.0)
                except Exception:
                    normalized_prob = 0.0
                prediction_map[text] = (normalized_label, normalized_prob)
        except Exception as exc:  # noqa: BLE001
            for row in rows:
                row["fightAddrLabel"] = ""
                row["fightAddrProb"] = "0.00000"
            return [], {"message": str(exc)}

    counts: Dict[str, int] = defaultdict(int)
    for row in rows:
        text = str(row.get("occurAddress") or "").strip()
        label, prob = prediction_map.get(text, (UNCLASSIFIED_ADDRESS_LABEL, 0.0))
        row["fightAddrLabel"] = label
        row["fightAddrProb"] = f"{float(prob):.5f}"
        counts[label] += 1

    pairs = sorted(counts.items(), key=lambda item: (-item[1], item[0]))
    return pairs, None


def classify_reason(text: Any) -> Tuple[str, str]:
    content = str(text or "").strip()
    for label, keywords in FIGHT_REASON_RULES:
        for keyword in keywords:
            if keyword and keyword in content:
                return label, keyword
    return OTHER_REASON_LABEL, ""


def summarize_reason_labels(rows: Sequence[Dict[str, Any]]) -> List[Tuple[str, int]]:
    counts: Dict[str, int] = defaultdict(int)
    for row in rows:
        label, keyword = classify_reason(row.get("caseContents"))
        row["fightReasonLabel"] = label
        row["fightReasonKeyword"] = keyword
        counts[label] += 1
    return sorted(counts.items(), key=lambda item: (-item[1], item[0]))


def run_fight_topic_analysis(
    params: Mapping[str, Any],
    dimensions_selected: Sequence[str],
    *,
    trace_id: str | None = None,
    include_detail_rows: bool = False,
) -> Tuple[Dict[str, Any], Dict[str, Any], List[Dict[str, Any]], Dict[str, Any], Dict[str, Any]]:
    dims = normalize_dimensions(dimensions_selected)
    if not dims:
        raise ValueError("请至少选择一个分析维度")

    begin_date, end_date, start_dt, end_dt = _resolve_main_time_range(params)
    m2m_start_time, m2m_end_time = _resolve_m2m_time_range(params, start_dt, end_dt)
    analysis_options = _build_analysis_options(params)
    tag_csv, chara_name_csv = resolve_fight_topic_tags()

    results: Dict[str, Any] = {}
    analysis_base: Dict[str, Any] = {}
    all_data: List[Dict[str, Any]] = []
    trace = trace_id or "-"

    logger.info(
        "[trace:%s][fight-topic] analyze begin=%s end=%s dims=%s tagCount=%s options=%s",
        trace,
        begin_date,
        end_date,
        dims,
        len([item for item in tag_csv.split(",") if item]),
        analysis_options,
    )

    case_dimensions = {"time", "dept", "phone", "cluster", "addr", "reason"}
    requires_case_data = include_detail_rows or any(dim in case_dimensions for dim in dims)
    if requires_case_data:
        all_data = fetch_all_case_list(
            _build_case_payload(begin_date, end_date, tag_csv),
            max_page_size=FIGHT_TOPIC_UPSTREAM_PAGE_SIZE,
        )
        if "time" in dims:
            analysis_base["timeHourly"] = calc_time_hourly_counts(all_data)
            results["time"] = calc_time_period(all_data, bucket_hours=analysis_options["timeBucketHours"])
        if "dept" in dims:
            analysis_base["deptAll"] = calc_duty_dept(all_data, top_n=None)
            results["dept"] = calc_duty_dept(all_data, top_n=analysis_options["deptTopN"])
        if "phone" in dims:
            analysis_base["phoneAll"] = calc_repeat_phone(all_data, min_count=2)
            results["phone"] = calc_repeat_phone(all_data, min_count=analysis_options["repeatPhoneMinCount"])
        if "cluster" in dims:
            results["cluster"] = calc_repeat_address(all_data, radius_meters=analysis_options["repeatAddrRadiusMeters"])
        if "addr" in dims:
            addr_pairs, addr_error = summarize_address_labels(all_data)
            results["addr"] = addr_pairs
            if addr_error:
                results["addr_error"] = addr_error
        if "reason" in dims:
            results["reason"] = summarize_reason_labels(all_data)

    if "srr" in dims:
        srr_payload = _build_srr_payload(
            begin_date,
            end_date,
            m2m_start_time,
            m2m_end_time,
            tag_csv,
            chara_name_csv,
        )
        srr_result = fetch_srr_list(srr_payload, trace_id=trace)
        if srr_result.get("code") == 0:
            results["srr"] = srr_result.get("rows", [])
        else:
            results["srr"] = []
            results["srr_error"] = {
                "upstream_code": srr_result.get("code", -1),
                "message": srr_result.get("msg") or "上游接口异常",
                "trace_id": trace,
            }

    meta = {
        "beginDate": begin_date,
        "endDate": end_date,
        "m2mStartTime": m2m_start_time,
        "m2mEndTime": m2m_end_time,
        "tagCsv": tag_csv,
        "tagNames": chara_name_csv,
        "dimensions": dims,
    }
    return results, analysis_base, all_data, analysis_options, meta


def _write_title_row(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, title: str, max_col: int) -> int:
    ws.cell(row=row_idx, column=1, value=title).font = openpyxl.styles.Font(bold=True, size=14)
    if max_col > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
    return row_idx + 1


def _write_pair_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    title: str,
    rows: Sequence[Tuple[Any, Any]],
) -> int:
    ws.cell(row=row_idx, column=1, value=title).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="统计项").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=row_idx, column=2, value="数量").font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    if not rows:
        ws.cell(row=row_idx, column=1, value="无数据")
        row_idx += 1
        return row_idx + 1
    for label, count in rows:
        ws.cell(row=row_idx, column=1, value=str(label))
        ws.cell(row=row_idx, column=2, value=count)
        row_idx += 1
    return row_idx + 1


def _write_message_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    title: str,
    message: str,
) -> int:
    ws.cell(row=row_idx, column=1, value=title).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value=message)
    return row_idx + 2


def _write_srr_block(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    title: str,
    rows: Sequence[Mapping[str, Any]],
) -> int:
    headers = ["单位名称", "本期数", "同比上期", "同比比例", "环比上期", "环比比例"]
    keys = ["name", "presentCycle", "upperY2yCycle", "y2yProportion", "upperM2mCycle", "m2mProportion"]
    ws.cell(row=row_idx, column=1, value=title).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row_idx, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    if not rows:
        ws.cell(row=row_idx, column=1, value="无数据")
        row_idx += 1
        return row_idx + 1
    for item in rows:
        for col_idx, key in enumerate(keys, 1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))
        row_idx += 1
    return row_idx + 1


def build_export_filename(begin_date: str, end_date: str, now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d%H%M%S")
    return f"{begin_date[:10]}-{end_date[:10]}打架斗殴警情分析{timestamp}.xlsx"


def generate_fight_topic_excel(
    analysis_results: Mapping[str, Any],
    all_data: Sequence[Mapping[str, Any]],
    dimensions_selected: Sequence[str],
    *,
    begin_date: str,
    end_date: str,
    analysis_options: Mapping[str, Any] | None = None,
) -> io.BytesIO:
    dims = normalize_dimensions(dimensions_selected)
    opts = dict(analysis_options or {})

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "打架斗殴专题"

    detail_headers = [
        ("caseNo", "接警号"),
        ("callTime", "报警时间"),
        ("cmdId", "分局编码"),
        ("dutyDeptName", "管辖单位"),
        ("caseLevelName", "警情级别"),
        ("occurAddress", "涉案地址"),
        ("callerName", "报警人"),
        ("callerPhone", "报警人电话"),
        ("caseContents", "简要案情"),
        ("replies", "反馈内容"),
    ]
    if "addr" in dims:
        detail_headers.extend(
            [
                ("fightAddrLabel", "地址分类结果"),
                ("fightAddrProb", "地址分类置信度"),
            ]
        )
    if "reason" in dims:
        detail_headers.extend(
            [
                ("fightReasonLabel", "打架原因分类"),
                ("fightReasonKeyword", "命中关键词"),
            ]
        )

    row_idx = 1
    row_idx = _write_title_row(worksheet, row_idx, "打架斗殴警情分析", len(detail_headers))
    worksheet.cell(row=row_idx, column=1, value=f"查询时间范围：{begin_date} 至 {end_date}")
    row_idx += 2

    titles = {
        "srr": FIGHT_TOPIC_SRR_TITLE,
        "time": f"时段报警数（每{opts.get('timeBucketHours', 3)}小时）",
        "dept": "派出所报警数",
        "phone": f"重复报警电话（>={opts.get('repeatPhoneMinCount', 2)}次）",
        "cluster": f"重复报警地址（半径{opts.get('repeatAddrRadiusMeters', 50)}米）",
        "addr": "警情地址统计",
        "reason": "打架原因分析",
    }

    for dim in dims:
        if dim == "srr":
            if analysis_results.get("srr_error"):
                row_idx = _write_message_block(
                    worksheet,
                    row_idx,
                    titles[dim],
                    f"统计失败：{analysis_results['srr_error'].get('message', '上游接口异常')}",
                )
            else:
                row_idx = _write_srr_block(worksheet, row_idx, titles[dim], analysis_results.get("srr", []))
            continue
        if dim == "addr" and analysis_results.get("addr_error"):
            row_idx = _write_message_block(
                worksheet,
                row_idx,
                titles[dim],
                f"统计失败：{analysis_results['addr_error'].get('message', '地址分类模型不可用')}",
            )
            continue
        row_idx = _write_pair_block(worksheet, row_idx, titles[dim], analysis_results.get(dim, []))

    row_idx += 1
    worksheet.cell(row=row_idx, column=1, value="详细数据").font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    for col_idx, (_, header) in enumerate(detail_headers, 1):
        worksheet.cell(row=row_idx, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
    row_idx += 1

    for raw_row in all_data:
        for col_idx, (field, _) in enumerate(detail_headers, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=raw_row.get(field, ""))
        row_idx += 1

    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 20
    worksheet.column_dimensions["D"].width = 20
    worksheet.column_dimensions["F"].width = 32
    worksheet.column_dimensions["I"].width = 48
    worksheet.column_dimensions["J"].width = 48
    if "addr" in dims:
        worksheet.column_dimensions["K"].width = 18
        worksheet.column_dimensions["L"].width = 16

    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)
    return out
