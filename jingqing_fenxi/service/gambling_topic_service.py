from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Mapping, Sequence, Tuple

import openpyxl
from openpyxl import Workbook

from jingqing_fenxi.service.fight_topic_service import (
    _build_analysis_options,
    _build_case_payload as _build_confirm_case_payload,
    _build_srr_payload,
    _resolve_main_time_range,
    _resolve_m2m_time_range,
    _write_message_block,
    _write_pair_block,
    _write_srr_block,
    calc_duty_dept,
    calc_repeat_address,
    calc_repeat_phone,
    calc_time_hourly_counts,
    calc_time_period,
    fetch_all_case_list,
    fetch_srr_list,
    summarize_address_labels,
)
from jingqing_fenxi.service.gambling_topic_keywords import (
    GAMBLING_VENUE_SHEET_TITLE,
    GAMBLING_VENUE_TITLE,
    GAMBLING_WAY_TITLE,
    GAMBLING_WILDERNESS_TITLE,
    summarize_gambling_way_by_region,
    summarize_venue_by_cmd_id,
    summarize_wilderness_by_region,
)
from jingqing_fenxi.service.jingqing_api_client import api_client


GAMBLING_TOPIC_PARENT_ID = "5BF6A1CA6C3D4ED9896244554A1BA87C"
GAMBLING_TOPIC_UPSTREAM_PAGE_SIZE = 5000
GAMBLING_TOPIC_DIMENSIONS = [
    "srr",
    "time",
    "dept",
    "phone",
    "cluster",
    "addr",
    "gambling_way",
    "wilderness",
    "venue",
]

GAMBLING_TOPIC_SRR_TITLE = "各地同比环比"
DETAIL_HEADERS = [
    ("caseNo", "接警号"),
    ("callTime", "报警时间"),
    ("cmdName", "地区"),
    ("cmdId", "地区编码"),
    ("dutyDeptName", "管辖单位"),
    ("caseLevelName", "警情级别"),
    ("occurAddress", "警情地址"),
    ("callerName", "报警人"),
    ("callerPhone", "报警人电话"),
    ("caseContents", "报警内容"),
    ("replies", "处警情况"),
]


def normalize_dimensions(dimensions_selected: Sequence[str] | None) -> List[str]:
    normalized: List[str] = []
    seen = set()
    for dim in dimensions_selected or []:
        value = str(dim or "").strip()
        if not value or value not in GAMBLING_TOPIC_DIMENSIONS or value in seen:
            continue
        seen.add(value)
        normalized.append(value)
    return normalized


def _get_gambling_topic_tree_nodes() -> List[Dict[str, Any]]:
    return list(api_client.get_tree_view_data() or [])


def resolve_gambling_topic_tags(tree_nodes: Sequence[Mapping[str, Any]] | None = None) -> Tuple[str, str]:
    tags: List[str] = []
    names: List[str] = []
    tag_seen = set()
    name_seen = set()
    nodes = _get_gambling_topic_tree_nodes() if tree_nodes is None else list(tree_nodes)
    for node in nodes:
        if str(node.get("pId") or "").strip() != GAMBLING_TOPIC_PARENT_ID:
            continue
        tag = str(node.get("tag") or "").strip()
        name = str(node.get("name") or "").strip()
        if tag and tag not in tag_seen:
            tag_seen.add(tag)
            tags.append(tag)
        if tag and name and name not in name_seen:
            name_seen.add(name)
            names.append(name)
    if not tags:
        raise ValueError("未获取到赌博专题的警情类型编码")
    return ",".join(tags), ",".join(names)


def _build_gambling_case_payload(begin_date: str, end_date: str, tag_csv: str, tag_names: str = "") -> Dict[str, Any]:
    payload = _build_confirm_case_payload(begin_date, end_date, "")
    payload["newCharaSubclassNo"] = ""
    payload["newCharaSubclass"] = "全部"
    payload["newOriCharaSubclassNo"] = tag_csv
    payload["newOriCharaSubclass"] = tag_names or "全部"
    return payload


def run_gambling_topic_analysis(
    params: Mapping[str, Any],
    dimensions_selected: Sequence[str],
    *,
    include_detail_rows: bool = False,
) -> Tuple[Dict[str, Any], Dict[str, Any], List[Dict[str, Any]], Dict[str, Any], Dict[str, Any]]:
    dims = normalize_dimensions(dimensions_selected)
    if not dims:
        raise ValueError("请至少选择一个分析维度")

    begin_date, end_date, start_dt, end_dt = _resolve_main_time_range(params)
    m2m_start_time, m2m_end_time = _resolve_m2m_time_range(params, start_dt, end_dt)
    analysis_options = _build_analysis_options(params)
    tag_csv, chara_name_csv = resolve_gambling_topic_tags()

    results: Dict[str, Any] = {}
    analysis_base: Dict[str, Any] = {}
    all_data: List[Dict[str, Any]] = []

    case_dimensions = {"time", "dept", "phone", "cluster", "addr", "gambling_way", "wilderness", "venue"}
    requires_case_data = include_detail_rows or any(dim in case_dimensions for dim in dims)
    if requires_case_data:
        all_data = fetch_all_case_list(
            _build_gambling_case_payload(begin_date, end_date, tag_csv, chara_name_csv),
            max_page_size=GAMBLING_TOPIC_UPSTREAM_PAGE_SIZE,
        )
        if "time" in dims:
            analysis_base["timeHourly"] = calc_time_hourly_counts(all_data)
            results["time"] = calc_time_period(all_data, bucket_hours=analysis_options["timeBucketHours"])
        if "dept" in dims:
            analysis_base["deptAll"] = calc_duty_dept(all_data, top_n=None)
            results["dept"] = calc_duty_dept(all_data, top_n=analysis_options["deptTopN"])
        if "phone" in dims:
            analysis_base["phoneAll"] = calc_repeat_phone(all_data, min_count=2)
            results["phone"] = calc_repeat_phone(all_data, min_count=analysis_options["repeatPhoneMinCount"])
        if "cluster" in dims:
            results["cluster"] = calc_repeat_address(all_data, radius_meters=analysis_options["repeatAddrRadiusMeters"])
        if "addr" in dims:
            addr_pairs, addr_error = summarize_address_labels(all_data)
            results["addr"] = addr_pairs
            if addr_error:
                results["addr_error"] = addr_error
        if "gambling_way" in dims:
            results["gambling_way"] = summarize_gambling_way_by_region(all_data)
        if "wilderness" in dims:
            results["wilderness"] = summarize_wilderness_by_region(all_data)
        if "venue" in dims:
            results["venue"] = summarize_venue_by_cmd_id(all_data)

    if "srr" in dims:
        srr_payload = _build_srr_payload(
            begin_date,
            end_date,
            m2m_start_time,
            m2m_end_time,
            tag_csv,
            chara_name_csv,
        )
        srr_result = fetch_srr_list(srr_payload)
        if srr_result.get("code") == 0:
            results["srr"] = srr_result.get("rows", [])
        else:
            results["srr"] = []
            results["srr_error"] = {
                "upstream_code": srr_result.get("code", -1),
                "message": srr_result.get("msg") or "上游接口异常",
            }

    meta = {
        "beginDate": begin_date,
        "endDate": end_date,
        "m2mStartTime": m2m_start_time,
        "m2mEndTime": m2m_end_time,
        "tagCsv": tag_csv,
        "tagNames": chara_name_csv,
        "dimensions": dims,
    }
    return results, analysis_base, all_data, analysis_options, meta


def _write_title_row(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, title: str, max_col: int) -> int:
    ws.cell(row=row_idx, column=1, value=title).font = openpyxl.styles.Font(bold=True, size=14)
    if max_col > 1:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
    return row_idx + 1


def _write_detail_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row_idx: int,
    rows: Sequence[Mapping[str, Any]],
    extra_headers: Sequence[Tuple[str, str]] = (),
    dimensions_selected: Sequence[str] = (),
) -> int:
    headers = list(DETAIL_HEADERS)
    if "addr" in normalize_dimensions(dimensions_selected):
        headers.append(("fightAddrLabel", "警情地址统计"))
    headers.extend(extra_headers)
    ws.cell(row=row_idx, column=1, value="详细数据").font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    for col_idx, (_field, header) in enumerate(headers, 1):
        ws.cell(row=row_idx, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    for raw_row in rows:
        for col_idx, (field, _header) in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=raw_row.get(field, ""))
        row_idx += 1
    return row_idx


def _set_detail_widths(ws: openpyxl.worksheet.worksheet.Worksheet) -> None:
    widths = {
        "A": 22,
        "B": 20,
        "C": 16,
        "E": 20,
        "G": 30,
        "J": 48,
        "K": 48,
        "L": 18,
        "M": 24,
    }
    for column_name, width in widths.items():
        ws.column_dimensions[column_name].width = width


def _write_gambling_way_sheet(workbook: Workbook, result: Mapping[str, Any], dims: Sequence[str]) -> None:
    ws = workbook.create_sheet(GAMBLING_WAY_TITLE)
    columns = list(result.get("columns") or [])
    rows = list(result.get("rows") or [])
    row_idx = _write_title_row(ws, 1, GAMBLING_WAY_TITLE, max(2 + len(columns), len(DETAIL_HEADERS)))
    headers = ["地区"] + columns + ["合计"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row_idx, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    if not rows:
        ws.cell(row=row_idx, column=1, value="无数据")
        row_idx += 2
    else:
        for item in rows:
            ws.cell(row=row_idx, column=1, value=item.get("cmdName", ""))
            counts = item.get("counts") or {}
            for col_idx, label in enumerate(columns, 2):
                ws.cell(row=row_idx, column=col_idx, value=counts.get(label, 0))
            ws.cell(row=row_idx, column=len(headers), value=item.get("total", 0))
            row_idx += 1
        row_idx += 1
    _write_detail_rows(
        ws,
        row_idx,
        result.get("details") or [],
        [("gamblingWayLabels", "赌博方式"), ("gamblingWayKeywords", "命中关键词")],
        dims,
    )
    _set_detail_widths(ws)


def _write_wilderness_sheet(workbook: Workbook, result: Mapping[str, Any], dims: Sequence[str]) -> None:
    ws = workbook.create_sheet(GAMBLING_WILDERNESS_TITLE)
    rows = list(result.get("rows") or [])
    row_idx = _write_title_row(ws, 1, GAMBLING_WILDERNESS_TITLE, len(DETAIL_HEADERS))
    ws.cell(row=row_idx, column=1, value="地区").font = openpyxl.styles.Font(bold=True)
    ws.cell(row=row_idx, column=2, value="数量").font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    if not rows:
        ws.cell(row=row_idx, column=1, value="无数据")
        row_idx += 2
    else:
        for item in rows:
            ws.cell(row=row_idx, column=1, value=item.get("cmdName", ""))
            ws.cell(row=row_idx, column=2, value=item.get("total", 0))
            row_idx += 1
        row_idx += 1
    _write_detail_rows(
        ws,
        row_idx,
        result.get("details") or [],
        [("gamblingWildernessKeywords", "命中关键词")],
        dims,
    )
    _set_detail_widths(ws)


def _write_venue_sheet(workbook: Workbook, result: Mapping[str, Any], dims: Sequence[str]) -> None:
    ws = workbook.create_sheet(GAMBLING_VENUE_SHEET_TITLE)
    rows = list(result.get("rows") or [])
    row_idx = _write_title_row(ws, 1, GAMBLING_VENUE_TITLE, len(DETAIL_HEADERS))
    headers = ["地区编码", "地区", "数量"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=row_idx, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
    row_idx += 1
    if not rows:
        ws.cell(row=row_idx, column=1, value="无数据")
        row_idx += 2
    else:
        for item in rows:
            ws.cell(row=row_idx, column=1, value=item.get("cmdId", ""))
            ws.cell(row=row_idx, column=2, value=item.get("cmdName", ""))
            ws.cell(row=row_idx, column=3, value=item.get("total", 0))
            row_idx += 1
        row_idx += 1
    _write_detail_rows(
        ws,
        row_idx,
        result.get("details") or [],
        [("gamblingVenueFields", "命中字段"), ("gamblingVenueKeywords", "命中关键词")],
        dims,
    )
    _set_detail_widths(ws)


def build_export_filename(begin_date: str, end_date: str, now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d%H%M%S")
    return f"{begin_date[:10]}-{end_date[:10]}赌博专题警情分析{timestamp}.xlsx"


def generate_gambling_topic_excel(
    analysis_results: Mapping[str, Any],
    all_data: Sequence[Mapping[str, Any]],
    dimensions_selected: Sequence[str],
    *,
    begin_date: str,
    end_date: str,
    analysis_options: Mapping[str, Any] | None = None,
) -> io.BytesIO:
    dims = normalize_dimensions(dimensions_selected)
    opts = dict(analysis_options or {})
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "赌博专题"

    row_idx = 1
    row_idx = _write_title_row(worksheet, row_idx, "赌博专题警情分析", len(DETAIL_HEADERS))
    worksheet.cell(row=row_idx, column=1, value=f"查询时间范围：{begin_date} 至 {end_date}")
    row_idx += 2

    titles = {
        "srr": GAMBLING_TOPIC_SRR_TITLE,
        "time": f"时段报警数（每{opts.get('timeBucketHours', 3)}小时）",
        "dept": "派出所报警数",
        "phone": f"重复报警电话（>={opts.get('repeatPhoneMinCount', 2)}次）",
        "cluster": f"重复报警地址（半径{opts.get('repeatAddrRadiusMeters', 50)}米）",
        "addr": "警情地址统计",
    }
    main_dims = [dim for dim in dims if dim not in {"gambling_way", "wilderness", "venue"}]
    for dim in main_dims:
        if dim == "srr":
            if analysis_results.get("srr_error"):
                row_idx = _write_message_block(
                    worksheet,
                    row_idx,
                    titles[dim],
                    f"统计失败：{analysis_results['srr_error'].get('message', '上游接口异常')}",
                )
            else:
                row_idx = _write_srr_block(worksheet, row_idx, titles[dim], analysis_results.get("srr", []))
            continue
        if dim == "addr" and analysis_results.get("addr_error"):
            row_idx = _write_message_block(
                worksheet,
                row_idx,
                titles[dim],
                f"统计失败：{analysis_results['addr_error'].get('message', '地址分类模型不可用')}",
            )
            continue
        row_idx = _write_pair_block(worksheet, row_idx, titles[dim], analysis_results.get(dim, []))

    row_idx += 1
    _write_detail_rows(worksheet, row_idx, all_data, dimensions_selected=dims)
    _set_detail_widths(worksheet)

    if "gambling_way" in dims:
        _write_gambling_way_sheet(workbook, analysis_results.get("gambling_way") or {}, dims)
    if "wilderness" in dims:
        _write_wilderness_sheet(workbook, analysis_results.get("wilderness") or {}, dims)
    if "venue" in dims:
        _write_venue_sheet(workbook, analysis_results.get("venue") or {}, dims)

    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)
    return out
