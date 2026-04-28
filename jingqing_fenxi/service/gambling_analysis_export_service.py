from __future__ import annotations

import hashlib
import io
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Mapping, Sequence, Tuple

import openpyxl
from openpyxl import Workbook
from psycopg2.extras import RealDictCursor

from gonggong.config.database import get_database_connection
from jingqing_fenxi.service.gambling_analysis_text_features import build_text_features
from jingqing_fenxi.service.fight_topic_service import (
    _resolve_main_time_range,
    _resolve_y2y_time_range,
)
from jingqing_fenxi.service.gambling_topic_service import (
    GAMBLING_TOPIC_UPSTREAM_PAGE_SIZE,
    _build_gambling_case_payload,
    fetch_all_case_list,
    resolve_gambling_topic_tags,
)


HASH_SALT = "gambling-analysis-export-2026"
FRIDAY_ANCHOR = date(2000, 1, 7)

SheetData = Tuple[str, List[str], List[Mapping[str, Any]]]
LOW_VALUE_SIGNALS = {"噪音扰民/非赌诉求", "误报/虚假/核实无", "同行竞争/纠纷类"}
YOY_SUMMARY_SHEET_TITLE = "25_警情同比汇总"
YOY_MONTH_SHEET_TITLE = "26_月度同比_代码"
YOY_DEPT_SHEET_TITLE = "27_派出所同比_代码"
PLAIN_EXPORT_HEADER_LABELS = {
    "case_hash": "警情哈希",
    "case_no": "警情编号",
    "call_time": "接警时间",
    "occur_time": "发案时间",
    "call_date": "接警日期",
    "year_month": "月份",
    "business_week": "业务周",
    "weekday": "星期",
    "hour": "小时",
    "cmd_id": "地区编码",
    "cmd_name": "地区名称",
    "duty_dept_no": "派出所编码",
    "duty_dept_name": "派出所名称",
    "pcs_code": "派出所代码",
    "new_case_source_no": "警情来源编码",
    "new_case_source": "警情来源",
    "new_recv_type_no": "接警类型编码",
    "new_recv_type_name": "接警类型",
    "case_level": "警情等级",
    "case_mark_no": "警情标识编码",
    "case_mark": "警情标识",
    "new_chara_subclass_no": "确认警情性质编码",
    "new_chara_subclass_name": "确认警情性质",
    "new_ori_chara_subclass_no": "原始警情性质编码",
    "new_ori_chara_subclass_name": "原始警情性质",
    "caller_name": "报警人",
    "caller_phone": "报警电话",
    "caller_phone_hash": "报警电话哈希",
    "address_hash": "警情地址哈希",
    "occur_address": "警情地址",
    "case_contents": "报警内容",
    "replies": "处警情况",
    "lng": "经度",
    "lat": "纬度",
    "coord_source": "坐标来源",
    "effective_class": "有效性分类",
    "problem_signal": "问题信号",
    "gambling_way": "赌博方式",
    "venue_type": "场所类型",
    "disposal_result": "处警结果",
    "disposal_evidence": "证据情况",
    "evidence_signal": "证据信号",
    "report_risk_signal": "风险信号",
    "report_quality_signal": "线索质量",
    "watchout_signal": "看风盯梢",
    "gambling_scale_signal": "规模信号",
    "profit_signal": "赌资/抽水信号",
    "consistency_signal": "查处一致性",
    "content_len": "报警内容长度",
    "reply_len": "处警情况长度",
    "incident_count": "警情数",
    "total_count": "警情总数",
    "effective_count": "有效违法警情数",
    "left_scene_count": "到场已散/未抓现行数",
    "noise_count": "噪音扰民/非赌诉求数",
    "false_count": "误报/虚假/核实无数",
    "competition_count": "同行竞争/纠纷类数",
    "online_count": "网络/手机赌博数",
    "other_count": "其他警情数",
    "effective_rate": "有效率",
    "low_value_rate": "低质率",
    "distinct_duty_dept_count": "涉及派出所数",
    "distinct_address_count": "涉及地址数",
    "distinct_phone_count": "涉及电话数",
    "low_value_count": "低质警情数",
    "first_call_time": "首次报警时间",
    "last_call_time": "末次报警时间",
    "grid_lng": "网格经度",
    "grid_lat": "网格纬度",
    "latest_week": "最新业务周",
    "previous_count": "同期数",
    "latest_count": "最新周警情数",
    "increase_count": "增减数",
    "current_rising_weeks": "当前连续上升周数",
    "current_rising_steps": "当前连续上升步数",
    "trend_series": "近8周走势",
    "case_type": "案件类型",
    "case_source": "案件来源",
    "case_status": "案件状态",
    "case_count": "案件数",
    "case_total": "案件总数",
    "incident_transfer_count": "警情转案数",
    "self_discovered_count": "主动查处数",
    "unknown_source_count": "来源不明案件数",
    "admin_case_count": "行政案件数",
    "criminal_case_count": "刑事案件数",
    "admin_penalty_count": "行政处罚人次",
    "admin_detention_count": "行政拘留人次",
    "criminal_detention_count": "刑事拘留人次",
    "arrest_count": "逮捕人数",
    "prosecution_person_count": "移诉人数",
    "person_hash": "人员标识",
    "first_month": "首次月份",
    "last_month": "末次月份",
    "indicator": "指标",
    "current_period": "本期时间",
    "previous_period": "同期时间",
    "current_count": "本期数",
    "current_effective_count": "本期有效违法警情数",
    "previous_effective_count": "同期有效违法警情数",
    "current_effective_rate": "本期有效率",
    "previous_effective_rate": "同期有效率",
    "current_low_value_count": "本期低质警情数",
    "previous_low_value_count": "同期低质警情数",
    "current_low_value_rate": "本期低质率",
    "previous_low_value_rate": "同期低质率",
    "yoy_rate_pct": "同比增幅(%)",
    "yoy_desc": "同比说明",
    "month": "月份序号",
    "current_year_month": "本期月份",
    "previous_year_month": "同期月份",
    "error": "错误信息",
}

def build_gambling_analysis_export_filename(
    begin_date: str,
    end_date: str,
    now: datetime | None = None,
    *,
    desensitized: bool = True,
) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d%H%M%S")
    label = "赌博分析数据包" if desensitized else "赌博分析报告附件"
    return f"{begin_date[:10]}-{end_date[:10]}{label}{timestamp}.xlsx"


def generate_gambling_analysis_export(params: Mapping[str, Any]) -> io.BytesIO:
    begin_date, end_date, start_dt, end_dt = _resolve_main_time_range(params)
    y2y_start_time, y2y_end_time = _resolve_y2y_time_range(start_dt, end_dt)
    desensitized = _is_desensitized_export(params)
    tag_csv, tag_names = resolve_gambling_topic_tags()
    rows = fetch_all_case_list(
        _build_gambling_case_payload(begin_date, end_date, tag_csv, tag_names),
        max_page_size=GAMBLING_TOPIC_UPSTREAM_PAGE_SIZE,
    )
    incident_rows = _normalize_incident_rows(rows, include_plain_fields=not desensitized)

    y2y_errors: List[str] = []
    y2y_incident_rows: List[Dict[str, Any]] = []
    try:
        y2y_rows = fetch_all_case_list(
            _build_gambling_case_payload(y2y_start_time, y2y_end_time, tag_csv, tag_names),
            max_page_size=GAMBLING_TOPIC_UPSTREAM_PAGE_SIZE,
        )
        y2y_incident_rows = _normalize_incident_rows(y2y_rows, include_plain_fields=not desensitized)
    except Exception as exc:  # noqa: BLE001
        y2y_errors.append(f"同比警情未导出: {exc}")

    db_errors: List[str] = []
    db_sheets: List[SheetData] = []
    try:
        db_sheets = _build_database_sheets(begin_date, end_date)
    except Exception as exc:  # noqa: BLE001
        db_errors.append(f"业务库维度未导出: {exc}")

    workbook = Workbook()
    info_ws = workbook.active
    info_ws.title = "00_导出说明"
    _write_info_sheet(
        info_ws,
        begin_date=begin_date,
        end_date=end_date,
        tag_csv=tag_csv,
        tag_names=tag_names,
        incident_count=len(incident_rows),
        y2y_start_time=y2y_start_time,
        y2y_end_time=y2y_end_time,
        y2y_incident_count=len(y2y_incident_rows),
        y2y_errors=y2y_errors,
        db_errors=db_errors,
        desensitized=desensitized,
    )

    for title, headers, sheet_rows in _build_incident_sheets(incident_rows, start_dt, end_dt, desensitized=desensitized):
        if not desensitized:
            title, headers, sheet_rows = _localize_plain_sheet(title, headers, sheet_rows)
        _write_table_sheet(workbook, title, headers, sheet_rows)
    for title, headers, sheet_rows in db_sheets:
        if not desensitized:
            title, headers, sheet_rows = _localize_plain_sheet(title, headers, sheet_rows)
        _write_table_sheet(workbook, title, headers, sheet_rows)
    for title, headers, sheet_rows in _build_yoy_sheets(
        incident_rows,
        y2y_incident_rows,
        begin_date=begin_date,
        end_date=end_date,
        y2y_start_time=y2y_start_time,
        y2y_end_time=y2y_end_time,
        y2y_errors=y2y_errors,
    ):
        if not desensitized:
            title, headers, sheet_rows = _localize_plain_sheet(title, headers, sheet_rows)
        _write_table_sheet(workbook, title, headers, sheet_rows)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _is_desensitized_export(params: Mapping[str, Any]) -> bool:
    value = params.get("desensitized", "1")
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() not in {"0", "false", "no", "off"}


def _normalize_incident_rows(rows: Sequence[Mapping[str, Any]], *, include_plain_fields: bool = False) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for row in rows:
        call_dt = _parse_datetime(_first(row, "callTime", "alarmTime", "createTime"))
        occur_dt = _parse_datetime(_first(row, "occurTime", "happenTime"))
        lng, lat, coord_source = _extract_coordinate(row)
        duty_dept_no = _clean_code(_first(row, "dutyDeptNo", "duty_dept_no"))
        case_contents = _first(row, "caseContents", "content", "alarmContent")
        replies = _first(row, "replies", "reply", "feedback")
        address_text = _first(row, "occurAddressNorm", "occurAddress", "address")
        caller_phone = _first(row, "callerPhone", "phone", "alarmPhone")
        case_no = _first(row, "caseNo", "newCaseNo", "id")
        text_features = build_text_features(row)
        item = {
            "case_hash": _hash_value(case_no),
            "call_time": _format_dt(call_dt),
            "call_date": call_dt.date().isoformat() if call_dt else "",
            "year_month": call_dt.strftime("%Y-%m") if call_dt else "",
            "business_week": _business_week_label(call_dt.date()) if call_dt else "",
            "weekday": call_dt.isoweekday() if call_dt else "",
            "hour": call_dt.hour if call_dt else "",
            "cmd_id": _clean_code(_first(row, "cmdId", "cmd_id")),
            "duty_dept_no": duty_dept_no,
            "pcs_code": _pcs_code(duty_dept_no),
            "new_case_source_no": _clean_code(_first(row, "newCaseSourceNo", "caseSourceNo")),
            "new_case_source": _first(row, "newCaseSource", "caseSource") or "",
            "new_recv_type_no": _clean_code(_first(row, "newRecvTypeNo", "recvTypeNo", "newRecvType")),
            "case_level": _first(row, "caseLevel", "caseLevelName") or "",
            "case_mark_no": _clean_code(_first(row, "caseMarkNo")),
            "new_chara_subclass_no": _clean_code(_first(row, "newCharaSubclassNo", "newCharaSubclass")),
            "new_ori_chara_subclass_no": _clean_code(_first(row, "newOriCharaSubclassNo", "newOriCharaSubclass")),
            "caller_phone_hash": _hash_value(caller_phone),
            "address_hash": _hash_value(address_text),
            "lng": lng,
            "lat": lat,
            "coord_source": coord_source,
            **text_features,
        }
        if include_plain_fields:
            item.update(
                {
                    "case_no": case_no or "",
                    "occur_time": _format_dt(occur_dt),
                    "cmd_name": _first(row, "cmdName", "cmd_name") or "",
                    "duty_dept_name": _first(row, "dutyDeptName", "duty_dept_name") or "",
                    "caller_name": _first(row, "callerName", "alarmPersonName") or "",
                    "caller_phone": caller_phone or "",
                    "occur_address": address_text or "",
                    "case_contents": case_contents or "",
                    "replies": replies or "",
                    "new_recv_type_name": _first(row, "newRecvTypeName", "recvTypeName") or "",
                    "case_mark": _first(row, "caseMark") or "",
                    "new_chara_subclass_name": _first(row, "newCharaSubclassName", "newCharaSubclassText", "newCharaSubclass") or "",
                    "new_ori_chara_subclass_name": _first(row, "newOriCharaSubclassName", "newOriCharaSubclassText", "newOriCharaSubclass") or "",
                }
            )
        normalized.append(item)
    return normalized


def _build_incident_sheets(
    rows: Sequence[Mapping[str, Any]],
    start_dt: datetime,
    end_dt: datetime,
    *,
    desensitized: bool = True,
) -> List[SheetData]:
    if not desensitized:
        return _build_plain_incident_sheets(rows, start_dt, end_dt)

    detail_headers = [
        "case_hash",
        "call_time",
        "call_date",
        "year_month",
        "business_week",
        "weekday",
        "hour",
        "cmd_id",
        "duty_dept_no",
        "pcs_code",
        "new_case_source_no",
        "new_case_source",
        "new_recv_type_no",
        "case_level",
        "case_mark_no",
        "new_chara_subclass_no",
        "new_ori_chara_subclass_no",
        "caller_phone_hash",
        "address_hash",
        "lng",
        "lat",
        "coord_source",
        "effective_class",
        "problem_signal",
        "gambling_way",
        "venue_type",
        "disposal_result",
        "disposal_evidence",
        "evidence_signal",
        "report_risk_signal",
        "report_quality_signal",
        "watchout_signal",
        "gambling_scale_signal",
        "profit_signal",
        "consistency_signal",
        "content_len",
        "reply_len",
    ]
    return [
        ("01_警情脱敏明细", detail_headers, [dict(row) for row in rows]),
        ("02_月度警情_代码", ["year_month", "duty_dept_no", "pcs_code", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "pcs_code"])),
        ("03_业务周警情_代码", ["business_week", "duty_dept_no", "pcs_code", "incident_count"], _count_by(rows, ["business_week", "duty_dept_no", "pcs_code"])),
        ("04_处警有效性", ["year_month", "effective_class", "incident_count"], _count_by(rows, ["year_month", "effective_class"])),
        ("05_派出所有效性", _dept_effectiveness_headers(), _dept_effectiveness(rows)),
        ("06_高频电话", _repeat_headers("caller_phone_hash"), _repeat_rank(rows, "caller_phone_hash")),
        ("07_高频地址", _repeat_headers("address_hash"), _repeat_rank(rows, "address_hash")),
        ("08_时段星期", ["weekday", "hour", "incident_count"], _count_by(rows, ["weekday", "hour"])),
        ("09_来源类型", ["new_case_source_no", "new_case_source", "new_recv_type_no", "incident_count"], _count_by(rows, ["new_case_source_no", "new_case_source", "new_recv_type_no"])),
        ("10_警情类别代码", ["new_chara_subclass_no", "new_ori_chara_subclass_no", "incident_count"], _count_by(rows, ["new_chara_subclass_no", "new_ori_chara_subclass_no"])),
        ("11_赌博方式", ["gambling_way", "duty_dept_no", "incident_count"], _count_by(rows, ["gambling_way", "duty_dept_no"])),
        ("12_场所类型", ["venue_type", "duty_dept_no", "incident_count"], _count_by(rows, ["venue_type", "duty_dept_no"])),
        ("13_无效低质警情", ["year_month", "duty_dept_no", "problem_signal", "incident_count"], _low_value_rows(rows)),
        ("14_经纬度点位", _geo_point_headers(), _geo_points(rows)),
        ("15_经纬度网格", _geo_grid_headers(), _geo_grid(rows)),
        ("16_连续上升周", _rising_week_headers(), _weekly_rising(rows, start_dt, end_dt)),
        ("17_处警结果细分", ["year_month", "duty_dept_no", "pcs_code", "disposal_result", "consistency_signal", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "pcs_code", "disposal_result", "consistency_signal"])),
        ("18_赌资赌具证据", ["year_month", "duty_dept_no", "disposal_evidence", "evidence_signal", "profit_signal", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "disposal_evidence", "evidence_signal", "profit_signal"])),
        ("19_线索质量风险", ["year_month", "duty_dept_no", "report_quality_signal", "report_risk_signal", "watchout_signal", "gambling_scale_signal", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "report_quality_signal", "report_risk_signal", "watchout_signal", "gambling_scale_signal"])),
    ]


def _build_plain_incident_sheets(rows: Sequence[Mapping[str, Any]], start_dt: datetime, end_dt: datetime) -> List[SheetData]:
    public_detail_headers = [
        "警情编号",
        "接警时间",
        "发案时间",
        "日期",
        "月份",
        "业务周",
        "星期",
        "小时",
        "地区编码",
        "地区名称",
        "派出所编码",
        "派出所名称",
        "接警类型编码",
        "接警类型",
        "警情等级",
        "警情标识",
        "确认警情性质编码",
        "确认警情性质",
        "原始警情性质编码",
        "原始警情性质",
        "报警人",
        "报警电话",
        "警情地址",
        "报警内容",
        "处警情况",
        "经度",
        "纬度",
        "坐标来源",
        "有效性分类",
        "问题信号",
        "赌博方式",
        "场所类型",
        "处警结果",
        "证据情况",
        "证据信号",
        "风险信号",
        "线索质量",
        "看风盯梢",
        "规模信号",
        "赌资/抽水信号",
        "查处一致性",
        "报警内容长度",
        "处警情况长度",
    ]
    return [
        ("01_警情明细", public_detail_headers, _plain_detail_rows(rows)),
        ("02_月度警情_代码", ["year_month", "duty_dept_no", "duty_dept_name", "pcs_code", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "duty_dept_name", "pcs_code"])),
        ("03_业务周警情_代码", ["business_week", "duty_dept_no", "duty_dept_name", "pcs_code", "incident_count"], _count_by(rows, ["business_week", "duty_dept_no", "duty_dept_name", "pcs_code"])),
        ("04_处警有效性", ["year_month", "effective_class", "incident_count"], _count_by(rows, ["year_month", "effective_class"])),
        ("05_派出所有效性", _dept_effectiveness_headers(plain=True), _dept_effectiveness(rows)),
        ("06_高频电话", _repeat_headers("caller_phone"), _repeat_rank(rows, "caller_phone", address_field="occur_address", phone_field="caller_phone")),
        ("07_高频地址", _repeat_headers("occur_address"), _repeat_rank(rows, "occur_address", address_field="occur_address", phone_field="caller_phone")),
        ("08_时段星期", ["weekday", "hour", "incident_count"], _count_by(rows, ["weekday", "hour"])),
        ("09_来源类型", ["new_case_source_no", "new_case_source", "new_recv_type_no", "new_recv_type_name", "incident_count"], _count_by(rows, ["new_case_source_no", "new_case_source", "new_recv_type_no", "new_recv_type_name"])),
        ("10_警情类别代码", ["new_chara_subclass_no", "new_chara_subclass_name", "new_ori_chara_subclass_no", "new_ori_chara_subclass_name", "incident_count"], _count_by(rows, ["new_chara_subclass_no", "new_chara_subclass_name", "new_ori_chara_subclass_no", "new_ori_chara_subclass_name"])),
        ("11_赌博方式", ["gambling_way", "duty_dept_no", "duty_dept_name", "incident_count"], _count_by(rows, ["gambling_way", "duty_dept_no", "duty_dept_name"])),
        ("12_场所类型", ["venue_type", "duty_dept_no", "duty_dept_name", "incident_count"], _count_by(rows, ["venue_type", "duty_dept_no", "duty_dept_name"])),
        ("13_无效低质警情", ["year_month", "duty_dept_no", "duty_dept_name", "problem_signal", "incident_count"], _low_value_rows(rows, plain=True)),
        ("14_经纬度点位", _geo_point_headers(plain=True), _geo_points(rows, plain=True)),
        ("15_经纬度网格", _geo_grid_headers(), _geo_grid(rows)),
        ("16_连续上升周", _rising_week_headers(plain=True), _weekly_rising(rows, start_dt, end_dt)),
        ("17_处警结果细分", ["year_month", "duty_dept_no", "duty_dept_name", "pcs_code", "disposal_result", "consistency_signal", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "duty_dept_name", "pcs_code", "disposal_result", "consistency_signal"])),
        ("18_赌资赌具证据", ["year_month", "duty_dept_no", "duty_dept_name", "disposal_evidence", "evidence_signal", "profit_signal", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "duty_dept_name", "disposal_evidence", "evidence_signal", "profit_signal"])),
        ("19_线索质量风险", ["year_month", "duty_dept_no", "duty_dept_name", "report_quality_signal", "report_risk_signal", "watchout_signal", "gambling_scale_signal", "incident_count"], _count_by(rows, ["year_month", "duty_dept_no", "duty_dept_name", "report_quality_signal", "report_risk_signal", "watchout_signal", "gambling_scale_signal"])),
    ]


def _plain_detail_rows(rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for row in rows:
        out.append(
            {
                "警情编号": row.get("case_no", ""),
                "接警时间": row.get("call_time", ""),
                "发案时间": row.get("occur_time", ""),
                "日期": row.get("call_date", ""),
                "月份": row.get("year_month", ""),
                "业务周": row.get("business_week", ""),
                "星期": row.get("weekday", ""),
                "小时": row.get("hour", ""),
                "地区编码": row.get("cmd_id", ""),
                "地区名称": row.get("cmd_name", ""),
                "派出所编码": row.get("duty_dept_no", ""),
                "派出所名称": row.get("duty_dept_name", ""),
                "接警类型编码": row.get("new_recv_type_no", ""),
                "接警类型": row.get("new_recv_type_name", ""),
                "警情等级": row.get("case_level", ""),
                "警情标识": row.get("case_mark", "") or row.get("case_mark_no", ""),
                "确认警情性质编码": row.get("new_chara_subclass_no", ""),
                "确认警情性质": row.get("new_chara_subclass_name", ""),
                "原始警情性质编码": row.get("new_ori_chara_subclass_no", ""),
                "原始警情性质": row.get("new_ori_chara_subclass_name", ""),
                "报警人": row.get("caller_name", ""),
                "报警电话": row.get("caller_phone", ""),
                "警情地址": row.get("occur_address", ""),
                "报警内容": row.get("case_contents", ""),
                "处警情况": row.get("replies", ""),
                "经度": row.get("lng", ""),
                "纬度": row.get("lat", ""),
                "坐标来源": row.get("coord_source", ""),
                "有效性分类": row.get("effective_class", ""),
                "问题信号": row.get("problem_signal", ""),
                "赌博方式": row.get("gambling_way", ""),
                "场所类型": row.get("venue_type", ""),
                "处警结果": row.get("disposal_result", ""),
                "证据情况": row.get("disposal_evidence", ""),
                "证据信号": row.get("evidence_signal", ""),
                "风险信号": row.get("report_risk_signal", ""),
                "线索质量": row.get("report_quality_signal", ""),
                "看风盯梢": row.get("watchout_signal", ""),
                "规模信号": row.get("gambling_scale_signal", ""),
                "赌资/抽水信号": row.get("profit_signal", ""),
                "查处一致性": row.get("consistency_signal", ""),
                "报警内容长度": row.get("content_len", ""),
                "处警情况长度": row.get("reply_len", ""),
            }
        )
    return out


def _build_yoy_sheets(
    current_rows: Sequence[Mapping[str, Any]],
    previous_rows: Sequence[Mapping[str, Any]],
    *,
    begin_date: str,
    end_date: str,
    y2y_start_time: str,
    y2y_end_time: str,
    y2y_errors: Sequence[str],
) -> List[SheetData]:
    if y2y_errors:
        return [(YOY_SUMMARY_SHEET_TITLE, ["error"], [{"error": "；".join(y2y_errors)}])]

    current_period = f"{begin_date} 至 {end_date}"
    previous_period = f"{y2y_start_time} 至 {y2y_end_time}"
    return [
        (
            YOY_SUMMARY_SHEET_TITLE,
            ["indicator", "current_period", "previous_period", "current_count", "previous_count", "increase_count", "yoy_rate_pct", "yoy_desc"],
            _yoy_summary_rows(current_rows, previous_rows, current_period, previous_period),
        ),
        (
            YOY_MONTH_SHEET_TITLE,
            ["month", "current_year_month", "previous_year_month", "current_count", "previous_count", "increase_count", "yoy_rate_pct", "yoy_desc"],
            _yoy_month_rows(current_rows, previous_rows),
        ),
        (
            YOY_DEPT_SHEET_TITLE,
            [
                "duty_dept_no",
                "pcs_code",
                "current_count",
                "previous_count",
                "increase_count",
                "yoy_rate_pct",
                "yoy_desc",
                "current_effective_count",
                "previous_effective_count",
                "current_effective_rate",
                "previous_effective_rate",
                "current_low_value_count",
                "previous_low_value_count",
                "current_low_value_rate",
                "previous_low_value_rate",
            ],
            _yoy_dept_rows(current_rows, previous_rows),
        ),
    ]


def _localize_plain_sheet(title: str, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> SheetData:
    localized_headers = [_plain_header_label(header) for header in headers]
    localized_rows = [
        {
            localized_header: row.get(original_header, "")
            for original_header, localized_header in zip(headers, localized_headers)
        }
        for row in rows
    ]
    return title, localized_headers, localized_rows


def _plain_header_label(header: str) -> str:
    return PLAIN_EXPORT_HEADER_LABELS.get(str(header), str(header))


def _yoy_summary_rows(
    current_rows: Sequence[Mapping[str, Any]],
    previous_rows: Sequence[Mapping[str, Any]],
    current_period: str,
    previous_period: str,
) -> List[Dict[str, Any]]:
    metrics = [
        ("警情总数", lambda row: True),
        ("有效违法警情", lambda row: row.get("effective_class") == "有效违法警情"),
        ("疑似违法但未抓现行", lambda row: row.get("effective_class") == "疑似违法但未抓现行"),
        ("核实无违法行为", lambda row: row.get("effective_class") == "核实无违法行为"),
        ("低质/非赌线索", lambda row: row.get("problem_signal") in LOW_VALUE_SIGNALS),
        ("到场已散/未抓现行", lambda row: row.get("problem_signal") == "到场已散/未抓现行"),
        ("网络/手机赌博", lambda row: row.get("problem_signal") == "网络/手机赌博"),
        ("有人看风/盯梢", lambda row: _has_label(row.get("report_risk_signal"), "有人看风/盯梢") or _has_label(row.get("watchout_signal"), "有看风盯梢")),
        ("多人聚赌", lambda row: _has_label(row.get("report_risk_signal"), "多人聚赌")),
        ("大额赌资/抽水", lambda row: _has_label(row.get("report_risk_signal"), "大额赌资/抽水") or _has_label(row.get("profit_signal"), "大额赌资") or _has_label(row.get("profit_signal"), "抽水营利")),
        ("麻将", lambda row: _has_label(row.get("gambling_way"), "麻将")),
        ("三公", lambda row: _has_label(row.get("gambling_way"), "三公")),
        ("棋牌室/麻将馆", lambda row: _has_label(row.get("venue_type"), "棋牌室/麻将馆")),
        ("山林野外", lambda row: _has_label(row.get("venue_type"), "山林野外")),
    ]
    out: List[Dict[str, Any]] = []
    for indicator, predicate in metrics:
        current_count = sum(1 for row in current_rows if predicate(row))
        previous_count = sum(1 for row in previous_rows if predicate(row))
        out.append(
            {
                "indicator": indicator,
                "current_period": current_period,
                "previous_period": previous_period,
                **_yoy_count_fields(current_count, previous_count),
            }
        )
    return out


def _yoy_month_rows(current_rows: Sequence[Mapping[str, Any]], previous_rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    current_counts = _count_by_month_suffix(current_rows)
    previous_counts = _count_by_month_suffix(previous_rows)
    current_month_labels = _month_label_map(current_rows)
    previous_month_labels = _month_label_map(previous_rows)
    out = []
    for month in sorted(set(current_counts) | set(previous_counts)):
        current_count = current_counts.get(month, 0)
        previous_count = previous_counts.get(month, 0)
        out.append(
            {
                "month": month,
                "current_year_month": current_month_labels.get(month, ""),
                "previous_year_month": previous_month_labels.get(month, ""),
                **_yoy_count_fields(current_count, previous_count),
            }
        )
    return out


def _yoy_dept_rows(current_rows: Sequence[Mapping[str, Any]], previous_rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    current_stats = _dept_yoy_stats(current_rows)
    previous_stats = _dept_yoy_stats(previous_rows)
    out: List[Dict[str, Any]] = []
    for code in sorted(set(current_stats) | set(previous_stats)):
        current = current_stats.get(code, {})
        previous = previous_stats.get(code, {})
        current_count = int(current.get("total_count") or 0)
        previous_count = int(previous.get("total_count") or 0)
        current_effective = int(current.get("effective_count") or 0)
        previous_effective = int(previous.get("effective_count") or 0)
        current_low = int(current.get("low_value_count") or 0)
        previous_low = int(previous.get("low_value_count") or 0)
        out.append(
            {
                "duty_dept_no": code,
                "pcs_code": current.get("pcs_code") or previous.get("pcs_code") or _pcs_code(code),
                **_yoy_count_fields(current_count, previous_count),
                "current_effective_count": current_effective,
                "previous_effective_count": previous_effective,
                "current_effective_rate": _rate(current_effective, current_count),
                "previous_effective_rate": _rate(previous_effective, previous_count),
                "current_low_value_count": current_low,
                "previous_low_value_count": previous_low,
                "current_low_value_rate": _rate(current_low, current_count),
                "previous_low_value_rate": _rate(previous_low, previous_count),
            }
        )
    return sorted(out, key=lambda item: (-int(item["current_count"] or 0), -int(item["increase_count"] or 0), str(item["duty_dept_no"])))


def _count_by_month_suffix(rows: Sequence[Mapping[str, Any]]) -> Dict[str, int]:
    counter: Counter[str] = Counter()
    for row in rows:
        year_month = str(row.get("year_month") or "")
        if len(year_month) >= 7:
            counter[year_month[-2:]] += 1
    return dict(counter)


def _month_label_map(rows: Sequence[Mapping[str, Any]]) -> Dict[str, str]:
    labels = {}
    for row in rows:
        year_month = str(row.get("year_month") or "")
        if len(year_month) >= 7:
            labels.setdefault(year_month[-2:], year_month)
    return labels


def _dept_yoy_stats(rows: Sequence[Mapping[str, Any]]) -> Dict[str, Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        code = str(row.get("duty_dept_no") or "")
        if not code:
            continue
        item = grouped.setdefault(code, {"pcs_code": row.get("pcs_code", ""), "total_count": 0, "effective_count": 0, "low_value_count": 0})
        item["total_count"] += 1
        if row.get("effective_class") == "有效违法警情":
            item["effective_count"] += 1
        if row.get("problem_signal") in LOW_VALUE_SIGNALS:
            item["low_value_count"] += 1
    return grouped


def _yoy_count_fields(current_count: int, previous_count: int) -> Dict[str, Any]:
    return {
        "current_count": current_count,
        "previous_count": previous_count,
        "increase_count": current_count - previous_count,
        "yoy_rate_pct": _yoy_rate(current_count, previous_count),
        "yoy_desc": _yoy_desc(current_count, previous_count),
    }


def _yoy_rate(current_count: int, previous_count: int) -> Any:
    if previous_count == 0:
        return 0 if current_count == 0 else ""
    return round((current_count - previous_count) * 100.0 / previous_count, 1)


def _yoy_desc(current_count: int, previous_count: int) -> str:
    if previous_count == 0:
        return "持平" if current_count == 0 else "去年同期为0，新增"
    rate = _yoy_rate(current_count, previous_count)
    if current_count > previous_count:
        return f"同比上升{abs(rate)}%"
    if current_count < previous_count:
        return f"同比下降{abs(rate)}%"
    return "同比持平"


def _rate(count: int, total: int) -> Any:
    return round(count * 100.0 / total, 2) if total else ""


def _has_label(value: Any, label: str) -> bool:
    labels = {part.strip() for part in str(value or "").replace(",", "、").replace("，", "、").split("、") if part.strip()}
    return label in labels


def _count_by(rows: Sequence[Mapping[str, Any]], keys: Sequence[str]) -> List[Dict[str, Any]]:
    counter: Counter[Tuple[Any, ...]] = Counter()
    for row in rows:
        counter[tuple(row.get(key, "") for key in keys)] += 1
    out = [dict(zip(keys, key_tuple), incident_count=count) for key_tuple, count in counter.items()]
    return sorted(out, key=lambda item: tuple(str(item.get(key, "")) for key in keys) + (-item["incident_count"],))


def _dept_effectiveness_headers(*, plain: bool = False) -> List[str]:
    headers = [
        "duty_dept_no",
        "pcs_code",
        "total_count",
        "effective_count",
        "left_scene_count",
        "noise_count",
        "false_count",
        "competition_count",
        "online_count",
        "other_count",
        "effective_rate",
        "low_value_rate",
    ]
    if plain:
        headers.insert(1, "duty_dept_name")
    return headers


def _dept_effectiveness(rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        code = str(row.get("duty_dept_no") or "")
        item = grouped.setdefault(
            code,
            {
                "duty_dept_no": code,
                "duty_dept_name": row.get("duty_dept_name", ""),
                "pcs_code": row.get("pcs_code", ""),
                "total_count": 0,
            },
        )
        if not item.get("duty_dept_name") and row.get("duty_dept_name"):
            item["duty_dept_name"] = row.get("duty_dept_name", "")
        item["total_count"] += 1
        cls = str(row.get("effective_class") or "")
        signal = str(row.get("problem_signal") or "")
        if cls == "有效违法警情":
            item["effective_count"] = item.get("effective_count", 0) + 1
        elif cls == "疑似违法但未抓现行":
            item["left_scene_count"] = item.get("left_scene_count", 0) + 1
        elif signal == "噪音扰民/非赌诉求":
            item["noise_count"] = item.get("noise_count", 0) + 1
        elif signal == "误报/虚假/核实无":
            item["false_count"] = item.get("false_count", 0) + 1
        elif signal == "同行竞争/纠纷类":
            item["competition_count"] = item.get("competition_count", 0) + 1
        elif signal == "网络/手机赌博":
            item["online_count"] = item.get("online_count", 0) + 1
        else:
            item["other_count"] = item.get("other_count", 0) + 1

    for item in grouped.values():
        total = int(item.get("total_count") or 0)
        for key in ["effective_count", "left_scene_count", "noise_count", "false_count", "competition_count", "online_count", "other_count"]:
            item.setdefault(key, 0)
        low_value = item["noise_count"] + item["false_count"] + item["competition_count"]
        item["effective_rate"] = round(item["effective_count"] * 100.0 / total, 2) if total else 0
        item["low_value_rate"] = round(low_value * 100.0 / total, 2) if total else 0
    return sorted(grouped.values(), key=lambda item: (-item["total_count"], str(item["duty_dept_no"])))


def _repeat_headers(hash_field: str) -> List[str]:
    return [
        hash_field,
        "incident_count",
        "distinct_duty_dept_count",
        "distinct_address_count",
        "distinct_phone_count",
        "effective_count",
        "low_value_count",
        "first_call_time",
        "last_call_time",
    ]


def _repeat_rank(
    rows: Sequence[Mapping[str, Any]],
    field: str,
    *,
    address_field: str = "address_hash",
    phone_field: str = "caller_phone_hash",
) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key = str(row.get(field) or "")
        if not key:
            continue
        item = grouped.setdefault(
            key,
            {
                field: key,
                "incident_count": 0,
                "_dept": set(),
                "_addr": set(),
                "_phone": set(),
                "effective_count": 0,
                "low_value_count": 0,
                "first_call_time": "",
                "last_call_time": "",
            },
        )
        item["incident_count"] += 1
        item["_dept"].add(row.get("duty_dept_no", ""))
        item["_addr"].add(row.get(address_field, ""))
        item["_phone"].add(row.get(phone_field, ""))
        if row.get("effective_class") == "有效违法警情":
            item["effective_count"] += 1
        if row.get("problem_signal") in {"噪音扰民/非赌诉求", "误报/虚假/核实无", "同行竞争/纠纷类"}:
            item["low_value_count"] += 1
        call_time = str(row.get("call_time") or "")
        if call_time:
            if not item["first_call_time"] or call_time < item["first_call_time"]:
                item["first_call_time"] = call_time
            if not item["last_call_time"] or call_time > item["last_call_time"]:
                item["last_call_time"] = call_time

    out: List[Dict[str, Any]] = []
    for item in grouped.values():
        out.append(
            {
                field: item[field],
                "incident_count": item["incident_count"],
                "distinct_duty_dept_count": len({v for v in item["_dept"] if v}),
                "distinct_address_count": len({v for v in item["_addr"] if v}),
                "distinct_phone_count": len({v for v in item["_phone"] if v}),
                "effective_count": item["effective_count"],
                "low_value_count": item["low_value_count"],
                "first_call_time": item["first_call_time"],
                "last_call_time": item["last_call_time"],
            }
        )
    return sorted(out, key=lambda item: (-item["incident_count"], str(item[field])))


def _low_value_rows(rows: Sequence[Mapping[str, Any]], *, plain: bool = False) -> List[Dict[str, Any]]:
    low_signals = {"噪音扰民/非赌诉求", "误报/虚假/核实无", "同行竞争/纠纷类"}
    keys = ["year_month", "duty_dept_no", "problem_signal"]
    if plain:
        keys.insert(2, "duty_dept_name")
    return _count_by([row for row in rows if row.get("problem_signal") in low_signals], keys)


def _geo_point_headers(*, plain: bool = False) -> List[str]:
    if plain:
        return ["case_no", "call_time", "duty_dept_no", "duty_dept_name", "pcs_code", "occur_address", "lng", "lat", "coord_source", "effective_class", "problem_signal"]
    return ["case_hash", "call_time", "duty_dept_no", "pcs_code", "lng", "lat", "coord_source", "effective_class", "problem_signal"]


def _geo_points(rows: Sequence[Mapping[str, Any]], *, plain: bool = False) -> List[Dict[str, Any]]:
    headers = _geo_point_headers(plain=plain)
    return [
        {key: row.get(key, "") for key in headers}
        for row in rows
        if row.get("lng") not in ("", None) and row.get("lat") not in ("", None)
    ]


def _geo_grid_headers() -> List[str]:
    return ["grid_lng", "grid_lat", "duty_dept_no", "incident_count", "effective_count", "low_value_count", "distinct_phone_count", "distinct_address_count"]


def _geo_grid(rows: Sequence[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    grouped: Dict[Tuple[Any, ...], Dict[str, Any]] = {}
    for row in rows:
        if row.get("lng") in ("", None) or row.get("lat") in ("", None):
            continue
        key = (round(float(row["lng"]), 4), round(float(row["lat"]), 4), row.get("duty_dept_no", ""))
        item = grouped.setdefault(
            key,
            {
                "grid_lng": key[0],
                "grid_lat": key[1],
                "duty_dept_no": key[2],
                "incident_count": 0,
                "_phone": set(),
                "_addr": set(),
                "effective_count": 0,
                "low_value_count": 0,
            },
        )
        item["incident_count"] += 1
        item["_phone"].add(row.get("caller_phone") or row.get("caller_phone_hash", ""))
        item["_addr"].add(row.get("occur_address") or row.get("address_hash", ""))
        if row.get("effective_class") == "有效违法警情":
            item["effective_count"] += 1
        if row.get("problem_signal") in {"噪音扰民/非赌诉求", "误报/虚假/核实无", "同行竞争/纠纷类"}:
            item["low_value_count"] += 1
    out = []
    for item in grouped.values():
        out.append(
            {
                "grid_lng": item["grid_lng"],
                "grid_lat": item["grid_lat"],
                "duty_dept_no": item["duty_dept_no"],
                "incident_count": item["incident_count"],
                "effective_count": item["effective_count"],
                "low_value_count": item["low_value_count"],
                "distinct_phone_count": len({v for v in item["_phone"] if v}),
                "distinct_address_count": len({v for v in item["_addr"] if v}),
            }
        )
    return sorted(out, key=lambda item: (-item["incident_count"], str(item["duty_dept_no"])))


def _rising_week_headers(*, plain: bool = False) -> List[str]:
    headers = ["duty_dept_no", "pcs_code", "latest_week", "previous_count", "latest_count", "increase_count", "current_rising_weeks", "current_rising_steps", "trend_series"]
    if plain:
        headers.insert(1, "duty_dept_name")
    return headers


def _weekly_rising(rows: Sequence[Mapping[str, Any]], start_dt: datetime, end_dt: datetime) -> List[Dict[str, Any]]:
    week_starts = _business_week_starts(start_dt.date(), end_dt.date())
    if not week_starts:
        return []
    dept_codes = sorted({str(row.get("duty_dept_no") or "") for row in rows if row.get("duty_dept_no")})
    counts: Dict[Tuple[str, date], int] = defaultdict(int)
    pcs_by_dept = {}
    name_by_dept = {}
    for row in rows:
        code = str(row.get("duty_dept_no") or "")
        call_dt = _parse_datetime(row.get("call_time"))
        if not code or not call_dt:
            continue
        counts[(code, _business_week_start(call_dt.date()))] += 1
        pcs_by_dept[code] = row.get("pcs_code", "")
        if row.get("duty_dept_name"):
            name_by_dept[code] = row.get("duty_dept_name", "")
    latest_week = week_starts[-1]
    out: List[Dict[str, Any]] = []
    for code in dept_codes:
        series = [counts.get((code, week), 0) for week in week_starts]
        steps = 0
        idx = len(series) - 1
        while idx > 0 and series[idx] > series[idx - 1]:
            steps += 1
            idx -= 1
        previous_count = series[-2] if len(series) >= 2 else 0
        latest_count = series[-1]
        out.append(
            {
                "duty_dept_no": code,
                "duty_dept_name": name_by_dept.get(code, ""),
                "pcs_code": pcs_by_dept.get(code, _pcs_code(code)),
                "latest_week": _week_label(latest_week),
                "previous_count": previous_count,
                "latest_count": latest_count,
                "increase_count": latest_count - previous_count,
                "current_rising_weeks": steps + 1 if steps else 0,
                "current_rising_steps": steps,
                "trend_series": " -> ".join(str(value) for value in series[-8:]),
            }
        )
    return sorted(out, key=lambda item: (-item["current_rising_steps"], -item["increase_count"], -item["latest_count"], item["duty_dept_no"]))


def _build_database_sheets(begin_date: str, end_date: str) -> List[SheetData]:
    connection = get_database_connection()
    try:
        specs = [
            ("20_案件来源状态", ["year_month", "pcs_code", "case_type", "case_source", "case_status", "case_count"], CASE_SOURCE_SQL),
            ("21_转案主动查处", CASE_TRANSFER_HEADERS, CASE_TRANSFER_SQL),
            ("22_处罚强度", ENFORCEMENT_HEADERS, ENFORCEMENT_SQL),
            ("23_嫌疑人复涉", ["person_hash", "case_count", "first_month", "last_month"], SUSPECT_REPEAT_SQL),
            ("24_涉案人员复涉", ["person_hash", "case_count", "first_month", "last_month"], INVOLVED_REPEAT_SQL),
        ]
        sheets: List[SheetData] = []
        for title, headers, sql in specs:
            try:
                sheets.append((title, headers, _query_dicts(connection, sql, [begin_date, end_date])))
            except Exception as exc:  # noqa: BLE001
                sheets.append((title, ["error"], [{"error": str(exc)}]))
        return sheets
    finally:
        connection.close()


def _query_dicts(connection: Any, sql: str, params: Sequence[Any]) -> List[Dict[str, Any]]:
    with connection.cursor(cursor_factory=RealDictCursor) as cursor:
        cursor.execute(sql, list(params))
        return [dict(row) for row in cursor.fetchall()]


CASE_SOURCE_SQL = """
WITH params AS (
    SELECT %s::timestamp AS start_time, %s::timestamp AS end_time
),
ay AS (
    SELECT ctc.ay_pattern
    FROM "ywdata"."case_type_config" ctc
    WHERE ctc.leixing = '赌博'
),
case_base AS (
    SELECT
        TO_CHAR(aj.ajxx_lasj, 'YYYY-MM') AS year_month,
        LEFT(aj.ajxx_cbdw_bh_dm, 8) || '0000' AS pcs_code,
        aj.ajxx_ajlx AS case_type,
        CASE
            WHEN aj.ajxx_jqbh IS NOT NULL AND aj.ajxx_jqbh ~ '^[0-9]' THEN 'incident_transfer'
            WHEN aj.ajxx_jqbh IS NOT NULL AND aj.ajxx_jqbh ~ '^[Zz]' THEN 'self_discovered'
            ELSE 'unknown'
        END AS case_source,
        aj.ajxx_ajzt AS case_status,
        aj.ajxx_ajbh
    FROM "ywdata"."zq_zfba_ajxx" aj
    CROSS JOIN params p
    WHERE aj.ajxx_lasj >= p.start_time
      AND aj.ajxx_lasj <  p.end_time
      AND aj.ajxx_aymc IS NOT NULL
      AND EXISTS (
          SELECT 1
          FROM ay
          WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern
      )
)
SELECT year_month, pcs_code, case_type, case_source, case_status, COUNT(DISTINCT ajxx_ajbh) AS case_count
FROM case_base
GROUP BY year_month, pcs_code, case_type, case_source, case_status
ORDER BY year_month, pcs_code, case_count DESC
"""

CASE_TRANSFER_HEADERS = [
    "year_month",
    "pcs_code",
    "case_total",
    "incident_transfer_count",
    "self_discovered_count",
    "unknown_source_count",
    "admin_case_count",
    "criminal_case_count",
]

CASE_TRANSFER_SQL = """
WITH params AS (
    SELECT %s::timestamp AS start_time, %s::timestamp AS end_time
),
ay AS (
    SELECT ctc.ay_pattern
    FROM "ywdata"."case_type_config" ctc
    WHERE ctc.leixing = '赌博'
),
case_base AS (
    SELECT
        TO_CHAR(aj.ajxx_lasj, 'YYYY-MM') AS year_month,
        LEFT(aj.ajxx_cbdw_bh_dm, 8) || '0000' AS pcs_code,
        aj.ajxx_ajlx,
        CASE
            WHEN aj.ajxx_jqbh IS NOT NULL AND aj.ajxx_jqbh ~ '^[0-9]' THEN 'incident_transfer'
            WHEN aj.ajxx_jqbh IS NOT NULL AND aj.ajxx_jqbh ~ '^[Zz]' THEN 'self_discovered'
            ELSE 'unknown'
        END AS case_source,
        aj.ajxx_ajbh
    FROM "ywdata"."zq_zfba_ajxx" aj
    CROSS JOIN params p
    WHERE aj.ajxx_lasj >= p.start_time
      AND aj.ajxx_lasj <  p.end_time
      AND aj.ajxx_aymc IS NOT NULL
      AND EXISTS (
          SELECT 1
          FROM ay
          WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern
      )
)
SELECT
    year_month,
    pcs_code,
    COUNT(DISTINCT ajxx_ajbh) AS case_total,
    COUNT(DISTINCT ajxx_ajbh) FILTER (WHERE case_source = 'incident_transfer') AS incident_transfer_count,
    COUNT(DISTINCT ajxx_ajbh) FILTER (WHERE case_source = 'self_discovered') AS self_discovered_count,
    COUNT(DISTINCT ajxx_ajbh) FILTER (WHERE case_source = 'unknown') AS unknown_source_count,
    COUNT(DISTINCT ajxx_ajbh) FILTER (WHERE ajxx_ajlx = '行政') AS admin_case_count,
    COUNT(DISTINCT ajxx_ajbh) FILTER (WHERE ajxx_ajlx = '刑事') AS criminal_case_count
FROM case_base
GROUP BY year_month, pcs_code
ORDER BY year_month, pcs_code
"""

ENFORCEMENT_HEADERS = [
    "year_month",
    "pcs_code",
    "admin_penalty_count",
    "admin_detention_count",
    "criminal_detention_count",
    "arrest_count",
    "prosecution_person_count",
]

ENFORCEMENT_SQL = """
WITH params AS (
    SELECT %s::timestamp AS start_time, %s::timestamp AS end_time
),
ay AS (
    SELECT ctc.ay_pattern
    FROM "ywdata"."case_type_config" ctc
    WHERE ctc.leixing = '赌博'
),
facts AS (
    SELECT
        'admin_penalty_count' AS metric,
        TO_CHAR(xz.xzcfjds_spsj, 'YYYY-MM') AS year_month,
        LEFT(xz.xzcfjds_cbdw_bh_dm, 8) || '0000' AS pcs_code,
        COUNT(DISTINCT COALESCE(xz.xzcfjds_id, xz.ajxx_ajbh || '_' || xz.xzcfjds_rybh, xz.ajxx_ajbh, xz.xzcfjds_rybh)) AS cnt
    FROM "ywdata"."zq_zfba_xzcfjds" xz
    JOIN "ywdata"."zq_zfba_ajxx" aj ON aj.ajxx_ajbh = xz.ajxx_ajbh
    CROSS JOIN params p
    WHERE xz.xzcfjds_spsj >= p.start_time
      AND xz.xzcfjds_spsj <  p.end_time
      AND aj.ajxx_aymc IS NOT NULL
      AND EXISTS (SELECT 1 FROM ay WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern)
    GROUP BY 1, 2, 3

    UNION ALL

    SELECT
        'admin_detention_count',
        TO_CHAR(xz.xzcfjds_spsj, 'YYYY-MM'),
        LEFT(xz.xzcfjds_cbdw_bh_dm, 8) || '0000',
        COUNT(DISTINCT COALESCE(xz.xzcfjds_id, xz.ajxx_ajbh || '_' || xz.xzcfjds_rybh, xz.ajxx_ajbh, xz.xzcfjds_rybh))
    FROM "ywdata"."zq_zfba_xzcfjds" xz
    JOIN "ywdata"."zq_zfba_ajxx" aj ON aj.ajxx_ajbh = xz.ajxx_ajbh
    CROSS JOIN params p
    WHERE xz.xzcfjds_spsj >= p.start_time
      AND xz.xzcfjds_spsj <  p.end_time
      AND xz.xzcfjds_cfzl IS NOT NULL
      AND xz.xzcfjds_cfzl ~ '拘留'
      AND aj.ajxx_aymc IS NOT NULL
      AND EXISTS (SELECT 1 FROM ay WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern)
    GROUP BY 1, 2, 3

    UNION ALL

    SELECT
        'criminal_detention_count',
        TO_CHAR(jlz.jlz_pzsj, 'YYYY-MM'),
        LEFT(jlz.jlz_cbdw_bh_dm, 8) || '0000',
        COUNT(DISTINCT COALESCE(jlz.jlz_id, jlz.ajxx_ajbh || '_' || jlz.jlz_rybh, jlz.ajxx_ajbh, jlz.jlz_rybh))
    FROM "ywdata"."zq_zfba_jlz" jlz
    LEFT JOIN "ywdata"."zq_zfba_ajxx" aj ON aj.ajxx_ajbh = jlz.ajxx_ajbh
    CROSS JOIN params p
    WHERE jlz.jlz_pzsj >= p.start_time
      AND jlz.jlz_pzsj <  p.end_time
      AND (
          (jlz.jlz_ay_mc IS NOT NULL AND EXISTS (SELECT 1 FROM ay WHERE jlz.jlz_ay_mc SIMILAR TO ay.ay_pattern))
          OR (aj.ajxx_aymc IS NOT NULL AND EXISTS (SELECT 1 FROM ay WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern))
      )
    GROUP BY 1, 2, 3

    UNION ALL

    SELECT
        'arrest_count',
        TO_CHAR(dbz.dbz_pzdbsj, 'YYYY-MM'),
        LEFT(dbz.dbz_cbdw_bh_dm, 8) || '0000',
        COUNT(DISTINCT COALESCE(dbz.dbz_id, dbz.ajxx_ajbh || '_' || dbz.dbz_rybh, dbz.ajxx_ajbh, dbz.dbz_rybh))
    FROM "ywdata"."zq_zfba_dbz" dbz
    LEFT JOIN "ywdata"."zq_zfba_ajxx" aj ON aj.ajxx_ajbh = dbz.ajxx_ajbh
    CROSS JOIN params p
    WHERE dbz.dbz_pzdbsj >= p.start_time
      AND dbz.dbz_pzdbsj <  p.end_time
      AND (
          (dbz.dbz_ay_mc IS NOT NULL AND EXISTS (SELECT 1 FROM ay WHERE dbz.dbz_ay_mc SIMILAR TO ay.ay_pattern))
          OR (aj.ajxx_aymc IS NOT NULL AND EXISTS (SELECT 1 FROM ay WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern))
      )
    GROUP BY 1, 2, 3

    UNION ALL

    SELECT
        'prosecution_person_count',
        TO_CHAR(qs.qsryxx_tfsj, 'YYYY-MM'),
        LEFT(qs.qsryxx_cbdw_bh_dm, 8) || '0000',
        COUNT(DISTINCT COALESCE(qs.qsryxx_id, qs.ajxx_ajbh || '_' || qs.qsryxx_rybh, qs.ajxx_ajbh, qs.qsryxx_rybh))
    FROM "ywdata"."zq_zfba_qsryxx" qs
    LEFT JOIN "ywdata"."zq_zfba_ajxx" aj ON aj.ajxx_ajbh = qs.ajxx_ajbh
    CROSS JOIN params p
    WHERE qs.qsryxx_tfsj >= p.start_time
      AND qs.qsryxx_tfsj <  p.end_time
      AND (
          (qs.ajxx_ay IS NOT NULL AND EXISTS (SELECT 1 FROM ay WHERE qs.ajxx_ay SIMILAR TO ay.ay_pattern))
          OR (aj.ajxx_aymc IS NOT NULL AND EXISTS (SELECT 1 FROM ay WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern))
      )
    GROUP BY 1, 2, 3
)
SELECT
    year_month,
    pcs_code,
    SUM(cnt) FILTER (WHERE metric = 'admin_penalty_count') AS admin_penalty_count,
    SUM(cnt) FILTER (WHERE metric = 'admin_detention_count') AS admin_detention_count,
    SUM(cnt) FILTER (WHERE metric = 'criminal_detention_count') AS criminal_detention_count,
    SUM(cnt) FILTER (WHERE metric = 'arrest_count') AS arrest_count,
    SUM(cnt) FILTER (WHERE metric = 'prosecution_person_count') AS prosecution_person_count
FROM facts
GROUP BY year_month, pcs_code
ORDER BY year_month, pcs_code
"""

SUSPECT_REPEAT_SQL = """
WITH params AS (
    SELECT %s::timestamp AS start_time, %s::timestamp AS end_time
),
ay AS (
    SELECT ctc.ay_pattern
    FROM "ywdata"."case_type_config" ctc
    WHERE ctc.leixing = '赌博'
),
person_cases AS (
    SELECT
        MD5(xyr.xyrxx_sfzh || 'gambling-analysis-export') AS person_hash,
        aj.ajxx_ajbh,
        aj.ajxx_lasj
    FROM "ywdata"."zq_zfba_xyrxx" xyr
    JOIN "ywdata"."zq_zfba_ajxx" aj ON aj.ajxx_ajbh = xyr.ajxx_join_ajxx_ajbh
    CROSS JOIN params p
    WHERE aj.ajxx_lasj >= p.start_time
      AND aj.ajxx_lasj <  p.end_time
      AND xyr.xyrxx_sfzh IS NOT NULL
      AND aj.ajxx_aymc IS NOT NULL
      AND EXISTS (SELECT 1 FROM ay WHERE aj.ajxx_aymc SIMILAR TO ay.ay_pattern)
)
SELECT
    person_hash,
    COUNT(DISTINCT ajxx_ajbh) AS case_count,
    MIN(TO_CHAR(ajxx_lasj, 'YYYY-MM')) AS first_month,
    MAX(TO_CHAR(ajxx_lasj, 'YYYY-MM')) AS last_month
FROM person_cases
GROUP BY person_hash
HAVING COUNT(DISTINCT ajxx_ajbh) >= 1
ORDER BY case_count DESC, person_hash
"""

INVOLVED_REPEAT_SQL = """
WITH params AS (
    SELECT %s::timestamp AS start_time, %s::timestamp AS end_time
),
ay AS (
    SELECT ctc.ay_pattern
    FROM "ywdata"."case_type_config" ctc
    WHERE ctc.leixing = '赌博'
),
person_cases AS (
    SELECT
        MD5(sary.saryxx_sfzh || 'gambling-analysis-export') AS person_hash,
        sary.saryxx_lrsj,
        sary.saryxx_rybh
    FROM "ywdata"."zq_zfba_saryxx" sary
    CROSS JOIN params p
    WHERE sary.saryxx_lrsj >= p.start_time
      AND sary.saryxx_lrsj <  p.end_time
      AND sary.saryxx_sfzh IS NOT NULL
      AND sary.saryxx_ay_mc IS NOT NULL
      AND EXISTS (SELECT 1 FROM ay WHERE sary.saryxx_ay_mc SIMILAR TO ay.ay_pattern)
)
SELECT
    person_hash,
    COUNT(DISTINCT saryxx_rybh) AS case_count,
    MIN(TO_CHAR(saryxx_lrsj, 'YYYY-MM')) AS first_month,
    MAX(TO_CHAR(saryxx_lrsj, 'YYYY-MM')) AS last_month
FROM person_cases
GROUP BY person_hash
HAVING COUNT(DISTINCT saryxx_rybh) >= 1
ORDER BY case_count DESC, person_hash
"""


def _write_info_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    *,
    begin_date: str,
    end_date: str,
    tag_csv: str,
    tag_names: str,
    incident_count: int,
    y2y_start_time: str,
    y2y_end_time: str,
    y2y_incident_count: int,
    y2y_errors: Sequence[str],
    db_errors: Sequence[str],
    desensitized: bool,
) -> None:
    rows = [
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("分析开始时间", begin_date),
        ("分析结束时间", end_date),
        ("同比开始时间", y2y_start_time),
        ("同比结束时间", y2y_end_time),
        ("警情数据源", "http://68.253.2.111/dsjfx/case/list"),
        ("警情类型编码", tag_csv),
        ("警情类型名称", tag_names),
        ("警情记录数", incident_count),
        ("去年同期警情记录数", y2y_incident_count),
        ("脱敏状态", "已脱敏" if desensitized else "未脱敏"),
    ]
    if desensitized:
        rows.append(("脱敏说明", "不导出地区名称、派出所名称、报警人姓名、明文电话、明文地址、原文警情内容或处警反馈。"))
    else:
        rows.append(("附件说明", "导出明文警情编号、地区/派出所名称、报警人、报警电话、警情地址、报警内容和处警情况，可作为分析报告附件使用。"))
    if y2y_errors:
        rows.extend((f"同比提示{idx}", message) for idx, message in enumerate(y2y_errors, 1))
    if db_errors:
        rows.extend((f"业务库提示{idx}", message) for idx, message in enumerate(db_errors, 1))
    for idx, (key, value) in enumerate(rows, 1):
        ws.cell(row=idx, column=1, value=key).font = openpyxl.styles.Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100


def _write_table_sheet(workbook: Workbook, title: str, headers: Sequence[str], rows: Sequence[Mapping[str, Any]]) -> None:
    ws = workbook.create_sheet(title[:31])
    wide_columns = {
        "警情地址": 42,
        "报警内容": 58,
        "处警情况": 80,
        "派出所名称": 24,
        "地区名称": 20,
        "报警电话": 18,
        "警情编号": 24,
        "警情数": 14,
        "occur_address": 42,
        "caller_phone": 18,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
            if header in {"警情地址", "报警内容", "处警情况", "occur_address"}:
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, 1):
        letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[letter].width = wide_columns.get(header, min(max(len(header) + 4, 14), 28))


def _first(row: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return ""


def _hash_value(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    return hashlib.md5((text + HASH_SALT).encode("utf-8")).hexdigest()


def _clean_code(value: Any) -> str:
    return str(value or "").strip()


def _pcs_code(duty_dept_no: str) -> str:
    code = _clean_code(duty_dept_no)
    if len(code) >= 8 and code[:8].isdigit():
        return code[:8] + "0000"
    return code


def _parse_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip().replace("T", " ")
    if not text:
        return None
    candidates = [text]
    if len(text) >= 19:
        candidates.append(text[:19])
    if len(text) >= 16:
        candidates.append(text[:16])
    if len(text) >= 10:
        candidates.append(text[:10])
    for candidate in candidates:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def _format_dt(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def _extract_coordinate(row: Mapping[str, Any]) -> Tuple[Any, Any, str]:
    candidates = (
        ("lngOfCriterion", "latOfCriterion", "criterion"),
        ("lngofcriterion", "latofcriterion", "criterion"),
        ("lngOfLocate", "latOfLocate", "locate"),
        ("lngoflocate", "latoflocate", "locate"),
        ("lngOfCall", "latOfCall", "call"),
        ("lngofcall", "latofcall", "call"),
        ("lng", "lat", "generic"),
        ("longitude", "latitude", "generic"),
    )
    for lng_key, lat_key, label in candidates:
        lng = row.get(lng_key)
        lat = row.get(lat_key)
        try:
            lng_f = float(lng)
            lat_f = float(lat)
        except (TypeError, ValueError):
            continue
        if -180 <= lng_f <= 180 and -90 <= lat_f <= 90:
            return lng_f, lat_f, label
    return "", "", "none"


def _business_week_start(day: date) -> date:
    return FRIDAY_ANCHOR + timedelta(days=((day - FRIDAY_ANCHOR).days // 7) * 7)


def _business_week_label(day: date) -> str:
    return _week_label(_business_week_start(day))


def _week_label(start: date) -> str:
    end = start + timedelta(days=6)
    return f"{start.isoformat()}~{end.isoformat()}"


def _business_week_starts(start: date, end: date) -> List[date]:
    if end <= start:
        return []
    current = _business_week_start(start)
    out: List[date] = []
    while current < end:
        out.append(current)
        current += timedelta(days=7)
    return out
