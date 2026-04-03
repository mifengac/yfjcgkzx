import io
import math
from collections import defaultdict

import openpyxl
from openpyxl import Workbook

from gonggong.service.upstream_jingqing_client import api_client


def haversine_distance(lng1, lat1, lng2, lat2):
    """Calculate the great-circle distance (meters) between two points."""
    try:
        lng1, lat1, lng2, lat2 = map(math.radians, [float(lng1), float(lat1), float(lng2), float(lat2)])
    except (ValueError, TypeError):
        return float("inf")

    dlng = lng2 - lng1
    dlat = lat2 - lat1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
    c = 2 * math.asin(math.sqrt(a))
    return c * 6371000


def fetch_all_case_list(base_payload, max_page_size=2000):
    """Paginate through the case list API and fetch all rows."""
    all_rows = []
    try:
        page_size = int(base_payload.get("pageSize", 100))
    except Exception:
        page_size = 100
    try:
        max_page_size = int(max_page_size)
    except Exception:
        max_page_size = 2000
    max_page_size = max(1, max_page_size)
    page_size = max(1, min(page_size, max_page_size))
    current_page = 1

    payload = base_payload.copy()
    payload["pageSize"] = page_size

    while True:
        payload["pageNum"] = current_page
        result = api_client.get_case_list(payload)

        rows = result.get("rows", [])
        total = result.get("total", 0)

        if not rows:
            break

        all_rows.extend(rows)
        if len(all_rows) >= total:
            break

        current_page += 1

    return all_rows


def fetch_srr_list(payload, trace_id=None):
    """Fetch SRR data directly from upstream API."""
    return api_client.get_srr_list(payload, trace_id=trace_id)


def _normalize_time_bucket_hours(bucket_hours):
    valid = [1, 2, 3, 4, 6, 8, 12]
    try:
        v = int(bucket_hours)
    except Exception:
        v = 3
    if v not in valid:
        v = 3
    return v

def calc_time_hourly_counts(data):
    """Return 24-length hourly count array for local front-end re-bucketing."""
    hourly = [0] * 24
    for row in data:
        call_time = row.get("callTime")
        if not call_time or len(call_time) < 13:
            continue
        try:
            hour = int(call_time[11:13])
        except Exception:
            continue
        if 0 <= hour <= 23:
            hourly[hour] += 1
    return hourly


def calc_time_period(data, bucket_hours=3):
    """Calculate case counts by dynamic time buckets."""
    bucket_hours = _normalize_time_bucket_hours(bucket_hours)

    hourly = calc_time_hourly_counts(data)
    periods = []
    for start in range(0, 24, bucket_hours):
        end = start + bucket_hours
        count = sum(hourly[start:end])
        periods.append((f"{start}-{end}时", count))
    return sorted(periods, key=lambda x: x[1], reverse=True)


def calc_duty_dept(data, top_n=None):
    """Aggregate by duty department and optionally keep top N."""
    dept_counts = defaultdict(int)
    for row in data:
        dept = row.get("dutyDeptName") or "未知"
        dept_counts[dept] += 1

    sorted_items = sorted(dept_counts.items(), key=lambda x: x[1], reverse=True)

    if top_n is None:
        return sorted_items

    try:
        n = int(top_n)
    except Exception:
        return sorted_items

    if n < 1:
        return sorted_items
    return sorted_items[:n]


def calc_repeat_phone(data, min_count=2):
    """Aggregate duplicate caller phones with configurable threshold."""
    try:
        min_count = int(min_count)
    except Exception:
        min_count = 2
    min_count = max(2, min_count)

    phone_counts = defaultdict(int)
    for row in data:
        phone = row.get("callerPhone", "")
        if not phone:
            continue

        cleaned = "".join(c for c in str(phone) if c.isdigit())
        if len(cleaned) < 5:
            continue
        if cleaned == "00000000":
            continue
        phone_counts[cleaned] += 1

    return sorted(
        [(k, v) for k, v in phone_counts.items() if v >= min_count],
        key=lambda x: x[1],
        reverse=True,
    )


def _normalize_radius_meters(radius_meters):
    try:
        radius = int(radius_meters)
    except Exception:
        radius = 50
    radius = max(50, min(500, radius))
    radius = int(round(radius / 50.0) * 50)
    return max(50, min(500, radius))


def _build_spatial_points(data):
    points = []
    for row in sorted(data, key=lambda x: x.get("callTime", "")):
        lng = row.get("lngOfCriterion")
        lat = row.get("latOfCriterion")
        try:
            lng_f = float(lng)
            lat_f = float(lat)
        except (TypeError, ValueError):
            continue

        if abs(lng_f) > 180 or abs(lat_f) > 90:
            continue

        points.append(
            {
                "row": row,
                "lng": lng_f,
                "lat": lat_f,
            }
        )
    return points


def _build_spatial_grid(points, cell_size_meters):
    grid = defaultdict(list)
    for idx, p in enumerate(points):
        lat_rad = math.radians(p["lat"])
        meter_x = p["lng"] * 111320 * max(math.cos(lat_rad), 1e-6)
        meter_y = p["lat"] * 110540
        cell_x = int(math.floor(meter_x / cell_size_meters))
        cell_y = int(math.floor(meter_y / cell_size_meters))
        p["cell"] = (cell_x, cell_y)
        grid[(cell_x, cell_y)].append(idx)
    return grid


def calc_repeat_address(data, radius_meters=50):
    """Cluster repeat addresses by configurable radius with spatial pre-bucketing."""
    radius_meters = _normalize_radius_meters(radius_meters)
    points = _build_spatial_points(data)
    if not points:
        return []

    grid = _build_spatial_grid(points, radius_meters)
    neighbor_cache = {}

    def get_neighbors(index):
        if index in neighbor_cache:
            return neighbor_cache[index]

        p = points[index]
        cx, cy = p["cell"]
        neighbors = []

        for dx in (-1, 0, 1):
            for dy in (-1, 0, 1):
                for candidate_idx in grid.get((cx + dx, cy + dy), []):
                    if candidate_idx == index:
                        continue
                    q = points[candidate_idx]
                    dist = haversine_distance(p["lng"], p["lat"], q["lng"], q["lat"])
                    if dist <= radius_meters:
                        neighbors.append(candidate_idx)

        neighbor_cache[index] = neighbors
        return neighbors

    visited = set()
    clusters = []

    for idx in range(len(points)):
        if idx in visited:
            continue

        queue = [idx]
        visited.add(idx)
        component = [idx]

        while queue:
            cur = queue.pop()
            for nb in get_neighbors(cur):
                if nb in visited:
                    continue
                visited.add(nb)
                queue.append(nb)
                component.append(nb)

        if len(component) >= 2:
            clusters.append(component)

    result = []
    for comp in clusters:
        center_row = points[comp[0]]["row"]
        address = center_row.get("occurAddress") or "未知地址"
        time_str = center_row.get("callTime") or ""
        count = len(comp)
        label = f"{address}:{time_str}({count}次)"
        result.append((label, count))

    return sorted(result, key=lambda x: x[1], reverse=True)


def calc_50m_cluster(data):
    """Backward-compatible wrapper for legacy call sites."""
    return calc_repeat_address(data, radius_meters=50)


def generate_excel_report(analysis_results, all_data, dimensions_selected, analysis_options=None):
    """Generate excel workbook containing summary and raw data."""
    opts = analysis_options or {}

    wb = Workbook()
    wb.remove(wb.active)

    dim_names = {
        "srr": "各地同环比",
        "time": f"时段(每{opts.get('timeBucketHours', 3)}小时)",
        "dept": "派出所",
        "phone": f"重复报警电话(>= {opts.get('repeatPhoneMinCount', 2)}次)",
        "cluster": f"重复报警地址(半径{opts.get('repeatAddrRadiusMeters', 50)}米)",
    }

    raw_headers = [
        "接警号",
        "报警时间",
        "警情级别",
        "涉案地址",
        "报警人电话",
        "管辖单位",
        "警情状态",
        "简要案情",
    ]

    for dim_key, title in dim_names.items():
        if dim_key not in dimensions_selected:
            continue

        ws = wb.create_sheet(title=title[:31])
        dim_data = analysis_results.get(dim_key, [])

        ws.cell(row=1, column=1, value=f"{title}统计表").font = openpyxl.styles.Font(bold=True)
        if dim_key == "srr":
            ws.cell(row=3, column=1, value="单位名称")
            ws.cell(row=3, column=2, value="本期数")
            ws.cell(row=3, column=3, value="同比上期")
            ws.cell(row=3, column=4, value="同比比例")
            ws.cell(row=3, column=5, value="环比上期")
            ws.cell(row=3, column=6, value="环比比例")

            row_idx = 4
            for item in dim_data:
                ws.cell(row=row_idx, column=1, value=item.get("name", ""))
                ws.cell(row=row_idx, column=2, value=item.get("presentCycle", ""))
                ws.cell(row=row_idx, column=3, value=item.get("upperY2yCycle", ""))
                ws.cell(row=row_idx, column=4, value=item.get("y2yProportion", ""))
                ws.cell(row=row_idx, column=5, value=item.get("upperM2mCycle", ""))
                ws.cell(row=row_idx, column=6, value=item.get("m2mProportion", ""))
                row_idx += 1
        else:
            ws.cell(row=3, column=1, value="统计项")
            ws.cell(row=3, column=2, value="数量")

            row_idx = 4
            for item in dim_data:
                ws.cell(row=row_idx, column=1, value=str(item[0]))
                ws.cell(row=row_idx, column=2, value=str(item[1]))
                row_idx += 1

        start_raw_row = row_idx + 3
        ws.cell(row=start_raw_row, column=1, value="分析源数据明细").font = openpyxl.styles.Font(bold=True)

        start_raw_row += 1
        for col_idx, header in enumerate(raw_headers, 1):
            ws.cell(row=start_raw_row, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

        row_idx = start_raw_row + 1
        for raw_row in all_data:
            ws.cell(row=row_idx, column=1, value=raw_row.get("caseNo", ""))
            ws.cell(row=row_idx, column=2, value=raw_row.get("callTime", ""))
            ws.cell(row=row_idx, column=3, value=raw_row.get("caseLevelName", ""))
            ws.cell(row=row_idx, column=4, value=raw_row.get("occurAddress", ""))
            ws.cell(row=row_idx, column=5, value=raw_row.get("callerPhone", ""))
            ws.cell(row=row_idx, column=6, value=raw_row.get("dutyDeptName", ""))
            ws.cell(row=row_idx, column=7, value=raw_row.get("caseState", ""))
            ws.cell(row=row_idx, column=8, value=raw_row.get("caseContents", ""))
            row_idx += 1

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["D"].width = 40
        ws.column_dimensions["H"].width = 80

    if not wb.sheetnames:
        wb.create_sheet("无数据")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out
