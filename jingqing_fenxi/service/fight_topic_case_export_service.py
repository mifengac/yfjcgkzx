from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, Mapping, Sequence

import openpyxl
from openpyxl import Workbook

from jingqing_fenxi.dao.fight_topic_case_export_dao import list_unclosed_fight_cases
from jingqing_fenxi.service.fight_topic_service import default_time_range


UNFINISHED_FIGHT_CASE_HEADERS: Sequence[tuple[str, str]] = (
    ("ay_name", "案由"),
    ("case_name", "案件名称"),
    ("case_status", "案件状态"),
    ("handling_unit", "办案单位"),
    ("filing_time", "立案时间"),
    ("incident_address", "发案地点"),
    ("incident_time", "发案时间"),
    ("summary", "简要案情"),
)


def _normalize_datetime(value: Any) -> str:
    text = str(value or "").strip().replace("T", " ")
    if not text:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text + " 00:00:00"
    if re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}$", text):
        return text + ":00"
    if not re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$", text):
        raise ValueError("时间格式不正确")
    return text


def _resolve_time_range(params: Mapping[str, Any]) -> tuple[str, str]:
    default_start, default_end = default_time_range()
    begin_date = _normalize_datetime(params.get("beginDate") or default_start)
    end_date = _normalize_datetime(params.get("endDate") or default_end)
    start_dt = datetime.strptime(begin_date, "%Y-%m-%d %H:%M:%S")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d %H:%M:%S")
    if end_dt < start_dt:
        raise ValueError("结束时间不能早于开始时间")
    return begin_date, end_date


def _stringify_cell_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def build_unclosed_case_export_filename(begin_date: str, end_date: str, now: datetime | None = None) -> str:
    timestamp = (now or datetime.now()).strftime("%Y%m%d%H%M%S")
    return f"{begin_date[:10]}-{end_date[:10]}打架斗殴未办结案件明细{timestamp}.xlsx"


def generate_unclosed_fight_cases_excel(
    rows: Sequence[Mapping[str, Any]],
    *,
    begin_date: str,
    end_date: str,
) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "未办结案件"

    worksheet.cell(row=1, column=1, value=f"统计时间:{begin_date}-{end_date}").font = openpyxl.styles.Font(bold=True, size=12)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(UNFINISHED_FIGHT_CASE_HEADERS))
    worksheet.cell(row=1, column=1).alignment = openpyxl.styles.Alignment(horizontal="left")

    for col_idx, (_field, header) in enumerate(UNFINISHED_FIGHT_CASE_HEADERS, 1):
        worksheet.cell(row=2, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)

    row_idx = 3
    for raw_row in rows:
        for col_idx, (field, _header) in enumerate(UNFINISHED_FIGHT_CASE_HEADERS, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=_stringify_cell_value(raw_row.get(field, "")))
        row_idx += 1

    column_widths = {
        "A": 18,
        "B": 28,
        "C": 16,
        "D": 24,
        "E": 20,
        "F": 28,
        "G": 20,
        "H": 52,
    }
    for column_name, width in column_widths.items():
        worksheet.column_dimensions[column_name].width = width

    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)
    return out


def export_unclosed_fight_cases(params: Mapping[str, Any]) -> tuple[io.BytesIO, Dict[str, str]]:
    begin_date, end_date = _resolve_time_range(params)
    rows = list_unclosed_fight_cases(begin_date, end_date)
    export_file = generate_unclosed_fight_cases_excel(rows, begin_date=begin_date, end_date=end_date)
    return export_file, {"beginDate": begin_date, "endDate": end_date}