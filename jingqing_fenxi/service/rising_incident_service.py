from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Mapping, Sequence, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from jingqing_fenxi.service.jingqing_api_client import api_client


DATETIME_FMT = "%Y-%m-%d %H:%M:%S"
BUSINESS_WEEK_ANCHOR = date(2000, 1, 7)
PERIOD_TYPE_LABELS = {
    "business_week": "业务周",
    "month": "自然月",
}
CASE_LIST_PAGE_SIZE = 1000
CASE_LIST_MAX_PAGES = 200


@dataclass(frozen=True)
class Period:
    start: datetime
    end: datetime
    label: str


def _normalize_datetime(value: Any) -> str:
    text = str(value or "").strip().replace("T", " ")
    if not text:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text + " 00:00:00"
    if re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}$", text):
        return text + ":00"
    if not re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$", text):
        raise ValueError("时间格式不正确")
    return text


def _parse_datetime(value: Any) -> datetime:
    normalized = _normalize_datetime(value)
    if not normalized:
        raise ValueError("开始时间和结束时间不能为空")
    return datetime.strptime(normalized, DATETIME_FMT)


def _format_datetime(value: datetime) -> str:
    return value.strftime(DATETIME_FMT)


def _parse_int(value: Any, default: int, minimum: int, maximum: int) -> int:
    try:
        number = int(str(value).strip())
    except Exception:
        number = default
    return max(minimum, min(number, maximum))


def _parse_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in ("1", "true", "yes", "on"):
        return True
    if text in ("0", "false", "no", "off"):
        return False
    return default


def _normalize_csv(value: Any) -> str:
    tokens = [token.strip() for token in str(value or "").split(",") if token and token.strip()]
    seen = set()
    out = []
    for token in tokens:
        if token in seen:
            continue
        seen.add(token)
        out.append(token)
    return ",".join(out)


def _to_int(value: Any) -> int:
    try:
        return int(float(str(value or "0").strip()))
    except Exception:
        return 0


def _business_week_start(value: datetime) -> datetime:
    days = (value.date() - BUSINESS_WEEK_ANCHOR).days
    start = BUSINESS_WEEK_ANCHOR + timedelta(days=(days // 7) * 7)
    return datetime(start.year, start.month, start.day)


def _month_start(value: datetime) -> datetime:
    return datetime(value.year, value.month, 1)


def _next_month(value: datetime) -> datetime:
    year = value.year + (1 if value.month == 12 else 0)
    month = 1 if value.month == 12 else value.month + 1
    return datetime(year, month, 1)


def _period_label(period_type: str, boundary_start: datetime, boundary_end: datetime) -> str:
    if period_type == "month":
        return boundary_start.strftime("%Y-%m")
    display_end = boundary_end - timedelta(days=1)
    return f"{boundary_start:%Y-%m-%d}至{display_end:%Y-%m-%d}"


def build_periods(start_dt: datetime, end_dt: datetime, period_type: str) -> List[Period]:
    if end_dt <= start_dt:
        raise ValueError("结束时间必须晚于开始时间")
    if period_type not in PERIOD_TYPE_LABELS:
        raise ValueError("周期类型不正确")

    periods: List[Period] = []
    boundary_start = _month_start(start_dt) if period_type == "month" else _business_week_start(start_dt)
    while boundary_start < end_dt:
        boundary_end = _next_month(boundary_start) if period_type == "month" else boundary_start + timedelta(days=7)
        query_start = max(boundary_start, start_dt)
        query_end = min(boundary_end, end_dt)
        if query_start < query_end:
            periods.append(
                Period(
                    start=query_start,
                    end=query_end,
                    label=_period_label(period_type, boundary_start, boundary_end),
                )
            )
        boundary_start = boundary_end
    return periods


def _build_case_payload(
    *,
    start_dt: datetime,
    end_dt: datetime,
    chara_no: str = "",
    chara_name: str = "",
    page_num: int,
) -> Dict[str, Any]:
    normalized_chara_no = _normalize_csv(chara_no)
    normalized_chara_name = _normalize_csv(chara_name)
    return {
        "params[colArray]": "",
        "beginDate": _format_datetime(start_dt),
        "endDate": _format_datetime(max(start_dt, end_dt - timedelta(seconds=1))),
        "newCaseSourceNo": "",
        "newCaseSource": "全部",
        "dutyDeptNo": "",
        "dutyDeptName": "全部",
        "newCharaSubclassNo": "",
        "newCharaSubclass": "全部",
        "newOriCharaSubclassNo": normalized_chara_no,
        "newOriCharaSubclass": normalized_chara_name or "全部",
        "caseNo": "",
        "callerName": "",
        "callerPhone": "",
        "phoneAddress": "",
        "callerIdentity": "",
        "operatorNo": "",
        "operatorName": "",
        "params[isInvalidCase]": "",
        "occurAddress": "",
        "caseMarkNo": "",
        "caseMark": "全部",
        "params[repetitionCase]": "",
        "params[originalDuplicateCase]": "",
        "params[startTimePeriod]": "",
        "params[endTimePeriod]": "",
        "caseContents": "",
        "replies": "",
        "params[sinceRecord]": "",
        "dossierResult": "",
        "params[isVideo]": "",
        "params[isConversation]": "",
        "pageSize": str(CASE_LIST_PAGE_SIZE),
        "pageNum": str(page_num),
        "orderByColumn": "callTime",
        "isAsc": "desc",
    }


def _parse_total(value: Any) -> int | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(value)
    except Exception:
        return None


def _parse_row_datetime(value: Any) -> datetime | None:
    try:
        return datetime.strptime(_normalize_datetime(value), DATETIME_FMT)
    except Exception:
        return None


def _find_period_index(call_dt: datetime, periods: Sequence[Period]) -> int | None:
    for idx, period in enumerate(periods):
        if period.start <= call_dt < period.end:
            return idx
    return None


def _fetch_period_counts(
    periods: Sequence[Period],
    start_dt: datetime,
    end_dt: datetime,
    *,
    chara_no: str = "",
    chara_name: str = "",
) -> Tuple[Dict[str, Dict[str, Any]], List[Dict[str, Any]]]:
    dept_map: Dict[str, Dict[str, Any]] = {}
    detail_rows: List[Dict[str, Any]] = []
    page_num = 1
    total: int | None = None

    while True:
        result = api_client.get_case_list(
            _build_case_payload(
                start_dt=start_dt,
                end_dt=end_dt,
                chara_no=chara_no,
                chara_name=chara_name,
                page_num=page_num,
            )
        )
        if not isinstance(result, Mapping):
            raise RuntimeError("case/list 响应格式异常")
        if str(result.get("code", "0")).strip() != "0":
            message = result.get("msg") or result.get("message") or "上游接口异常"
            raise RuntimeError(f"case/list 取数失败：{message}")

        page_rows = result.get("rows") or []
        if not isinstance(page_rows, list):
            raise RuntimeError("case/list rows 不是数组")
        if total is None:
            total = _parse_total(result.get("total"))

        for row in page_rows:
            if not isinstance(row, Mapping):
                continue
            call_dt = _parse_row_datetime(row.get("callTime") or row.get("calltime"))
            if not call_dt or not (start_dt <= call_dt < end_dt):
                continue
            period_idx = _find_period_index(call_dt, periods)
            if period_idx is None:
                continue
            dept_code = str(row.get("dutyDeptNo") or row.get("code") or "").strip()
            dept_name = str(row.get("dutyDeptName") or row.get("name") or "").strip()
            if not dept_code and not dept_name:
                continue
            dept_key = dept_code or dept_name
            item = dept_map.setdefault(
                dept_key,
                {
                    "deptCode": dept_code,
                    "deptName": dept_name or dept_code,
                    "counts": [0] * len(periods),
                },
            )
            if dept_code and not item["deptCode"]:
                item["deptCode"] = dept_code
            if dept_name and (not item["deptName"] or item["deptName"] == item["deptCode"]):
                item["deptName"] = dept_name
            item["counts"][period_idx] += 1

        if not page_rows:
            break
        if len(page_rows) < CASE_LIST_PAGE_SIZE:
            break
        if total is not None and page_num * CASE_LIST_PAGE_SIZE >= total:
            break
        if page_num >= CASE_LIST_MAX_PAGES:
            raise RuntimeError("case/list 查询结果过大，请缩短时间范围后重试")
        page_num += 1

    for dept in dept_map.values():
        for period_idx, period in enumerate(periods):
            detail_rows.append(
                {
                    "派出所名称": dept["deptName"],
                    "派出所代码": dept["deptCode"],
                    "周期": period.label,
                    "开始时间": _format_datetime(period.start),
                    "结束时间": _format_datetime(period.end),
                    "警情数": dept["counts"][period_idx],
                }
            )
    return dept_map, detail_rows


def _current_rising_steps(counts: Sequence[int]) -> Tuple[int, int]:
    steps = 0
    for idx in range(len(counts) - 1, 0, -1):
        if counts[idx] > counts[idx - 1]:
            steps += 1
            continue
        break
    return steps, len(counts) - 1


def _longest_rising_steps(counts: Sequence[int]) -> Tuple[int, int]:
    best_steps = 0
    best_end_idx = len(counts) - 1
    current_steps = 0
    for idx in range(1, len(counts)):
        if counts[idx] > counts[idx - 1]:
            current_steps += 1
        else:
            current_steps = 0
        if current_steps > best_steps:
            best_steps = current_steps
            best_end_idx = idx
    return best_steps, best_end_idx


def _risk_level(rising_periods: int, steps: int, latest_count: int, increment: int) -> str:
    if rising_periods >= 5 or steps >= 4 or latest_count >= 30 or increment >= 10:
        return "高风险"
    if latest_count >= 15 and increment >= 5:
        return "高风险"
    if rising_periods >= 4 or latest_count >= 10 or increment >= 3:
        return "中风险"
    return "低风险"


def _build_rising_rows(
    dept_map: Mapping[str, Mapping[str, Any]],
    periods: Sequence[Period],
    *,
    min_periods: int,
    current_only: bool,
    period_type: str,
) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for dept in dept_map.values():
        counts = list(dept.get("counts") or [])
        if len(counts) < 2:
            continue
        steps, end_idx = _current_rising_steps(counts) if current_only else _longest_rising_steps(counts)
        rising_periods = steps + 1 if steps > 0 else 0
        if rising_periods < min_periods:
            continue

        start_idx = end_idx - steps
        series = counts[start_idx : end_idx + 1]
        previous_count = counts[end_idx - 1] if end_idx > 0 else 0
        latest_count = counts[end_idx]
        increment = latest_count - previous_count
        rows.append(
            {
                "派出所名称": dept.get("deptName") or "",
                "派出所代码": dept.get("deptCode") or "",
                "风险等级": _risk_level(rising_periods, steps, latest_count, increment),
                "周期类型": PERIOD_TYPE_LABELS[period_type],
                "最新周期": periods[end_idx].label,
                "上期数量": previous_count,
                "最新数量": latest_count,
                "增量": increment,
                "当前连续上升周期数": rising_periods,
                "当前连续上升次数": steps,
                "趋势序列": " -> ".join(str(item) for item in series),
                "涉及周期范围": f"{periods[start_idx].label} 至 {periods[end_idx].label}",
            }
        )

    rows.sort(
        key=lambda item: (
            -_to_int(item.get("当前连续上升周期数")),
            -_to_int(item.get("增量")),
            -_to_int(item.get("最新数量")),
            str(item.get("风险等级") or ""),
            str(item.get("派出所代码") or ""),
        )
    )
    return rows


def run_rising_incident_analysis(params: Mapping[str, Any]) -> Dict[str, Any]:
    begin_date = _normalize_datetime(params.get("beginDate"))
    end_date = _normalize_datetime(params.get("endDate"))
    start_dt = _parse_datetime(begin_date)
    end_dt = _parse_datetime(end_date)
    period_type = str(params.get("periodType") or "business_week").strip() or "business_week"
    min_periods = _parse_int(params.get("minPeriods"), 3, 2, 24)
    current_only = _parse_bool(params.get("currentOnly"), True)
    chara_no = _normalize_csv(params.get("newOriCharaSubclassNo", ""))
    chara_name = _normalize_csv(params.get("newOriCharaSubclass", ""))

    periods = build_periods(start_dt, end_dt, period_type)
    if len(periods) < min_periods:
        raise ValueError("统计周期数量少于连续上升阈值，请扩大时间范围或降低阈值")

    dept_map, detail_rows = _fetch_period_counts(
        periods,
        start_dt,
        end_dt,
        chara_no=chara_no,
        chara_name=chara_name,
    )
    rows = _build_rising_rows(
        dept_map,
        periods,
        min_periods=min_periods,
        current_only=current_only,
        period_type=period_type,
    )
    return {
        "meta": {
            "beginDate": begin_date,
            "endDate": end_date,
            "periodType": period_type,
            "periodTypeLabel": PERIOD_TYPE_LABELS[period_type],
            "minPeriods": min_periods,
            "currentOnly": current_only,
            "periodCount": len(periods),
            "caseTypeSource": str(params.get("caseTypeSource") or "nature").strip() or "nature",
            "charaNo": chara_no,
            "chara": chara_name or "全部",
        },
        "periods": [
            {
                "label": period.label,
                "startTime": _format_datetime(period.start),
                "endTime": _format_datetime(period.end),
            }
            for period in periods
        ],
        "rows": rows,
        "periodDetails": detail_rows,
    }


def build_export_filename(meta: Mapping[str, Any], now: datetime | None = None) -> str:
    begin = str(meta.get("beginDate") or "")[:10]
    end = str(meta.get("endDate") or "")[:10]
    timestamp = (now or datetime.now()).strftime("%Y%m%d%H%M%S")
    return f"{begin}-{end}警情升势预警{timestamp}.xlsx"


def _write_headers(ws: openpyxl.worksheet.worksheet.Worksheet, headers: Sequence[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(32, len(header) * 2 + 4))


def generate_rising_incident_excel(result: Mapping[str, Any]) -> io.BytesIO:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "升势预警"
    headers = [
        "派出所名称",
        "派出所代码",
        "风险等级",
        "周期类型",
        "最新周期",
        "上期数量",
        "最新数量",
        "增量",
        "当前连续上升周期数",
        "当前连续上升次数",
        "趋势序列",
        "涉及周期范围",
    ]
    _write_headers(summary, headers)
    for row_idx, row in enumerate(result.get("rows") or [], 2):
        for col_idx, header in enumerate(headers, 1):
            summary.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    detail = workbook.create_sheet("周期明细")
    detail_headers = ["派出所名称", "派出所代码", "周期", "开始时间", "结束时间", "警情数"]
    _write_headers(detail, detail_headers)
    for row_idx, row in enumerate(result.get("periodDetails") or [], 2):
        for col_idx, header in enumerate(detail_headers, 1):
            detail.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)
    return out
