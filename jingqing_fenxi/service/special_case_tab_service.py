from __future__ import annotations

import csv
import io
import logging
import re
import uuid
from datetime import datetime
from typing import Any, Callable, Dict, Iterable, List, Sequence, Tuple

import openpyxl
from openpyxl import Workbook

from jingqing_fenxi.service.jingqing_fenxi_service import fetch_all_case_list


logger = logging.getLogger(__name__)


CUSTOM_CASE_MONITOR_LABEL = "自定义警情监测"

BRANCH_CMD_ID_MAP = {
    "市局": "445300000000",
    "云城": "445302000000",
    "云安": "445303000000",
    "罗定": "445381000000",
    "新兴": "445321000000",
    "郁南": "445322000000",
}

SPECIAL_CASE_UPSTREAM_PAGE_SIZE = 5000

EXPORT_HEADERS = [
    ("caseNo", "接警号"),
    ("callTime", "报警时间"),
    ("cmdId", "分局编码"),
    ("dutyDeptName", "管辖单位"),
    ("caseLevelName", "警情级别"),
    ("occurAddress", "涉案地址"),
    ("callerName", "报警人"),
    ("callerPhone", "报警人电话"),
    ("caseContents", "简要案情"),
    ("replies", "反馈内容"),
]

RULE_FIELD_OPTIONS = [
    {"value": "combined_text", "label": "报警内容+反馈内容"},
    {"value": "caseContents", "label": "报警内容"},
    {"value": "replies", "label": "反馈内容"},
    {"value": "occurAddress", "label": "涉案地址"},
    {"value": "dutyDeptName", "label": "管辖单位"},
    {"value": "callerName", "label": "报警人"},
    {"value": "callerPhone", "label": "报警人电话"},
    {"value": "cmdId", "label": "分局编码"},
    {"value": "caseLevelName", "label": "警情级别"},
]

RULE_OPERATOR_OPTIONS = [
    {"value": "contains_any", "label": "包含任一值"},
    {"value": "contains_all", "label": "包含全部值"},
    {"value": "not_contains_any", "label": "不包含任一值"},
    {"value": "equals", "label": "完全等于"},
    {"value": "in_list", "label": "属于值列表"},
    {"value": "regex_any", "label": "正则命中任一条"},
    {"value": "regex_all", "label": "正则同时命中全部"},
]

ALLOWED_RULE_FIELDS = {item["value"] for item in RULE_FIELD_OPTIONS}
ALLOWED_RULE_OPERATORS = {item["value"] for item in RULE_OPERATOR_OPTIONS}
RULE_FIELD_LABEL_MAP = {item["value"]: item["label"] for item in RULE_FIELD_OPTIONS}
HIT_KEYWORD_HEADER = ("hitKeywords", "命中关键字")
_FILENAME_SAFE_PATTERN = re.compile(r'[\\/:*?"<>|\r\n\t]+')
_SHEET_TITLE_SAFE_PATTERN = re.compile(r"[\[\]\*?:/\\]")
_FILTER_PROGRESS_STEP = 100
_REGEX_RULE_OPERATORS = {"regex_any", "regex_all"}
ProgressCallback = Callable[[Dict[str, Any]], None]


def default_time_range() -> Tuple[str, str]:
    now = datetime.now()
    end_dt = datetime(now.year, now.month, now.day, 23, 59, 59)
    start_dt = datetime(now.year, now.month, now.day, 0, 0, 0)
    return start_dt.strftime("%Y-%m-%d %H:%M:%S"), end_dt.strftime("%Y-%m-%d %H:%M:%S")


def field_options() -> List[Dict[str, str]]:
    return [dict(item) for item in RULE_FIELD_OPTIONS]


def operator_options() -> List[Dict[str, str]]:
    return [dict(item) for item in RULE_OPERATOR_OPTIONS]


def sanitize_filename_component(value: Any) -> str:
    text = str(value or "").strip()
    text = _FILENAME_SAFE_PATTERN.sub("_", text)
    text = re.sub(r"\s+", " ", text).strip(" ._")
    return text or "export"


def sanitize_sheet_title(value: Any) -> str:
    text = str(value or "").strip()
    text = _SHEET_TITLE_SAFE_PATTERN.sub("_", text)
    text = re.sub(r"\s+", " ", text).strip(" '")
    return (text or "Sheet")[:31]


def build_export_filename(scheme_name: str, start_time: str, end_time: str, export_format: str) -> str:
    fmt = str(export_format or "xlsx").strip().lower().lstrip(".") or "xlsx"
    start_text = normalize_datetime_text(start_time) if start_time else default_time_range()[0]
    end_text = normalize_datetime_text(end_time) if end_time else default_time_range()[1]
    start_date = start_text.split(" ", 1)[0]
    end_date = end_text.split(" ", 1)[0]
    return f"{sanitize_filename_component(scheme_name)}_{start_date}_{end_date}.{fmt}"


def build_defaults_payload() -> Dict[str, Any]:
    start_time, end_time = default_time_range()
    return {
        "success": True,
        "start_time": start_time,
        "end_time": end_time,
        "branches": branch_options(),
    }


def normalize_datetime_text(value: str) -> str:
    text = str(value or "").strip().replace("T", " ")
    if not text:
        return ""
    if re.match(r"^\d{4}-\d{2}-\d{2}$", text):
        return text + " 00:00:00"
    if re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}$", text):
        return text + ":00"
    if not re.match(r"^\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}$", text):
        raise ValueError("时间格式不正确")
    return text


def normalize_branch_selection(branches: Sequence[str] | None) -> List[str]:
    normalized: List[str] = []
    seen = set()
    for branch in branches or []:
        name = str(branch or "").strip()
        if not name or name not in BRANCH_CMD_ID_MAP or name in seen:
            continue
        seen.add(name)
        normalized.append(name)
    return normalized


def branch_options() -> List[Dict[str, str]]:
    return [
        {"value": label, "label": label, "cmd_id": cmd_id}
        for label, cmd_id in BRANCH_CMD_ID_MAP.items()
    ]


def build_special_case_payload(begin_date: str, end_date: str) -> Dict[str, Any]:
    return {
        "params[colArray]": "",
        "beginDate": begin_date,
        "endDate": end_date,
        "newCaseSourceNo": "",
        "newCaseSource": "全部",
        "dutyDeptNo": "",
        "dutyDeptName": "全部",
        "newCharaSubclassNo": "",
        "newCharaSubclass": "全部",
        "newOriCharaSubclassNo": "",
        "newOriCharaSubclass": "全部",
        "caseNo": "",
        "callerName": "",
        "callerPhone": "",
        "phoneAddress": "",
        "callerIdentity": "",
        "operatorNo": "",
        "operatorName": "",
        "params[isInvalidCase]": "",
        "occurAddress": "",
        "caseMarkNo": "",
        "caseMark": "全部",
        "params[repetitionCase]": "",
        "params[originalDuplicateCase]": "",
        "params[startTimePeriod]": "",
        "params[endTimePeriod]": "",
        "caseContents": "",
        "replies": "",
        "params[sinceRecord]": "",
        "dossierResult": "",
        "params[isVideo]": "",
        "params[isConversation]": "",
        "pageSize": SPECIAL_CASE_UPSTREAM_PAGE_SIZE,
        "pageNum": 1,
        "orderByColumn": "callTime",
        "isAsc": "desc",
    }


def fetch_all_special_case_rows(begin_date: str, end_date: str) -> List[Dict[str, Any]]:
    return fetch_all_case_list(
        build_special_case_payload(begin_date, end_date),
        max_page_size=SPECIAL_CASE_UPSTREAM_PAGE_SIZE,
    )


def normalize_rule_values(values: Any, operator: str | None = None) -> List[str]:
    result: List[str] = []
    seen = set()
    split_pattern = r"[\r\n]+" if operator in _REGEX_RULE_OPERATORS else r"[\r\n,]+"

    def add_token(raw_value: Any) -> None:
        for item in re.split(split_pattern, str(raw_value or "")):
            token = item.strip()
            if not token or token in seen:
                continue
            seen.add(token)
            result.append(token)

    if isinstance(values, (list, tuple)):
        for value in values:
            add_token(value)
    else:
        add_token(values)
    return result


def normalize_rule_group_no(value: Any, default: int = 1) -> int:
    if value in (None, ""):
        return default
    try:
        group_no = int(str(value).strip())
    except Exception as exc:
        raise ValueError("规则组必须是大于等于 1 的整数") from exc
    if group_no < 1:
        raise ValueError("规则组必须是大于等于 1 的整数")
    return group_no


def validate_scheme_rules(rules: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for index, rule in enumerate(rules or [], start=1):
        field_name = str(rule.get("field_name") or "").strip()
        operator = str(rule.get("operator") or "").strip()
        rule_values = normalize_rule_values(rule.get("rule_values"), operator)
        group_no = normalize_rule_group_no(rule.get("group_no"), default=1)
        if field_name not in ALLOWED_RULE_FIELDS:
            raise ValueError(f"规则字段不支持：{field_name}")
        if operator not in ALLOWED_RULE_OPERATORS:
            raise ValueError(f"规则操作符不支持：{operator}")
        if not rule_values:
            raise ValueError("规则值列表不能为空")
        if operator in _REGEX_RULE_OPERATORS:
            for rule_value in rule_values:
                try:
                    re.compile(rule_value)
                except re.error as exc:
                    raise ValueError(f"正则表达式不合法：{rule_value}") from exc
        normalized.append(
            {
                "field_name": field_name,
                "operator": operator,
                "rule_values": rule_values,
                "group_no": group_no,
                "sort_order": int(rule.get("sort_order") or index),
                "is_enabled": bool(rule.get("is_enabled", True)),
            }
        )
    if not normalized:
        raise ValueError("至少需要配置一条规则")
    if not any(rule["is_enabled"] for rule in normalized):
        raise ValueError("至少需要一条启用规则")
    return normalized


def _row_field_text(row: Dict[str, Any], field_name: str) -> str:
    if field_name == "combined_text":
        return (str(row.get("caseContents") or "") + "\n" + str(row.get("replies") or "")).strip()
    return str(row.get(field_name) or "").strip()


def _ordered_unique(values: Iterable[str]) -> List[str]:
    result: List[str] = []
    seen = set()
    for value in values:
        token = str(value or "").strip()
        if not token or token in seen:
            continue
        seen.add(token)
        result.append(token)
    return result


def _compile_rule_patterns(rule: Dict[str, Any]) -> List[re.Pattern[str]]:
    cached = rule.get("_compiled_patterns")
    if isinstance(cached, list):
        return cached

    operator = str(rule.get("operator") or "").strip()
    patterns = [
        re.compile(rule_value)
        for rule_value in normalize_rule_values(rule.get("rule_values"), operator)
    ]
    rule["_compiled_patterns"] = patterns
    return patterns


def _pattern_hit_values(pattern: re.Pattern[str], target: str) -> List[str]:
    hits: List[str] = []
    seen = set()
    for match in pattern.finditer(target):
        hit_text = str(match.group(0) or "").strip()
        if not hit_text or hit_text in seen:
            continue
        seen.add(hit_text)
        hits.append(hit_text)
    return hits


def _regex_rule_hit_values(target: str, rule: Dict[str, Any], *, require_all: bool) -> List[str]:
    collected: List[str] = []
    for pattern in _compile_rule_patterns(rule):
        hits = _pattern_hit_values(pattern, target)
        if require_all and not hits:
            return []
        collected.extend(hits)
    if not require_all and not collected:
        return []
    return _ordered_unique(collected)


def _rule_hit_values(row: Dict[str, Any], rule: Dict[str, Any]) -> List[str]:
    target = _row_field_text(row, rule["field_name"])
    operator = str(rule.get("operator") or "").strip()
    values = normalize_rule_values(rule.get("rule_values"), operator)
    if operator == "regex_any":
        return _regex_rule_hit_values(target, rule, require_all=False)
    if operator == "regex_all":
        return _regex_rule_hit_values(target, rule, require_all=True)
    if operator == "contains_any":
        return [value for value in values if value in target]
    if operator == "contains_all":
        return values if all(value in target for value in values) else []
    if operator == "not_contains_any":
        return []
    if operator == "equals":
        return [value for value in values if target == value]
    if operator == "in_list":
        return [target] if target in set(values) else []
    return []


def _rule_matches_row(row: Dict[str, Any], rule: Dict[str, Any]) -> bool:
    target = _row_field_text(row, rule["field_name"])
    operator = str(rule.get("operator") or "").strip()
    values = normalize_rule_values(rule.get("rule_values"), operator)
    if operator == "regex_any":
        return bool(_regex_rule_hit_values(target, rule, require_all=False))
    if operator == "regex_all":
        return bool(_regex_rule_hit_values(target, rule, require_all=True))
    if operator == "contains_any":
        return any(value in target for value in values)
    if operator == "contains_all":
        return all(value in target for value in values)
    if operator == "not_contains_any":
        return all(value not in target for value in values)
    if operator == "equals":
        return any(target == value for value in values)
    if operator == "in_list":
        return target in set(values)
    return False


def _enabled_rule_groups(rules: Sequence[Dict[str, Any]]) -> List[Tuple[int, List[Dict[str, Any]]]]:
    grouped: Dict[int, List[Dict[str, Any]]] = {}
    for rule in rules or []:
        if not rule.get("is_enabled", True):
            continue
        group_no = normalize_rule_group_no(rule.get("group_no"), default=1)
        grouped.setdefault(group_no, []).append(rule)

    return [
        (
            group_no,
            sorted(
                group_rules,
                key=lambda item: (
                    int(item.get("sort_order") or 0),
                    int(item.get("id") or 0),
                ),
            ),
        )
        for group_no, group_rules in sorted(grouped.items(), key=lambda item: item[0])
    ]


def _matching_rule_groups(row: Dict[str, Any], rules: Sequence[Dict[str, Any]]) -> List[Tuple[int, List[Dict[str, Any]]]]:
    matches: List[Tuple[int, List[Dict[str, Any]]]] = []
    for group_no, group_rules in _enabled_rule_groups(rules):
        if all(_rule_matches_row(row, rule) for rule in group_rules):
            matches.append((group_no, group_rules))
    return matches


def _emit_progress(
    progress_callback: ProgressCallback | None,
    *,
    stage: str,
    message: str,
    stats: Dict[str, int],
) -> None:
    if not progress_callback:
        return
    progress_callback(
        {
            "stage": stage,
            "message": message,
            "stats": dict(stats),
        }
    )


def collect_rule_hit_keywords(row: Dict[str, Any], rules: Sequence[Dict[str, Any]]) -> List[str]:
    keywords: List[str] = []
    seen = set()
    for _group_no, group_rules in _matching_rule_groups(row, rules):
        for rule in group_rules:
            for value in _rule_hit_values(row, rule):
                if value in seen:
                    continue
                seen.add(value)
                keywords.append(value)
    return keywords


def _resolve_hit_field_labels(row: Dict[str, Any], rule: Dict[str, Any], value: str) -> List[str]:
    field_name = str(rule.get("field_name") or "").strip()
    if field_name != "combined_text":
        return [RULE_FIELD_LABEL_MAP.get(field_name, field_name or "命中字")]

    operator = str(rule.get("operator") or "").strip()
    case_text = str(row.get("caseContents") or "").strip()
    reply_text = str(row.get("replies") or "").strip()
    labels: List[str] = []
    if operator in {"contains_any", "contains_all", "regex_any", "regex_all"}:
        if value in case_text:
            labels.append(RULE_FIELD_LABEL_MAP["caseContents"])
        if value in reply_text:
            labels.append(RULE_FIELD_LABEL_MAP["replies"])
    elif operator in {"equals", "in_list"}:
        if case_text == value:
            labels.append(RULE_FIELD_LABEL_MAP["caseContents"])
        if reply_text == value:
            labels.append(RULE_FIELD_LABEL_MAP["replies"])
    if labels:
        return labels
    return [RULE_FIELD_LABEL_MAP["combined_text"]]


def collect_rule_hit_keyword_details(row: Dict[str, Any], rules: Sequence[Dict[str, Any]]) -> List[str]:
    details: List[str] = []
    seen = set()
    for _group_no, group_rules in _matching_rule_groups(row, rules):
        for rule in group_rules:
            for value in _rule_hit_values(row, rule):
                for label in _resolve_hit_field_labels(row, rule, value):
                    detail = f"{label}→{value}"
                    if detail in seen:
                        continue
                    seen.add(detail)
                    details.append(detail)
    return details


def filter_rows_by_rules(
    rows: Iterable[Dict[str, Any]],
    rules: Sequence[Dict[str, Any]],
    *,
    progress_callback: ProgressCallback | None = None,
    progress_step: int = _FILTER_PROGRESS_STEP,
) -> List[Dict[str, Any]]:
    row_list = list(rows or [])
    enabled_rule_groups = _enabled_rule_groups(rules)
    if not enabled_rule_groups:
        _emit_progress(
            progress_callback,
            stage="rule_filtering",
            message="规则过滤完成，未配置启用规则组",
            stats={
                "rule_scanned_count": 0,
                "rule_match_count": 0,
            },
        )
        return []
    filtered: List[Dict[str, Any]] = []
    matched_count = 0
    total = len(row_list)
    for index, row in enumerate(row_list, start=1):
        if _matching_rule_groups(row, rules):
            filtered.append(row)
            matched_count += 1
        if index % progress_step == 0 or index == total:
            _emit_progress(
                progress_callback,
                stage="rule_filtering",
                message=f"规则过滤中：已扫描 {index} 条，命中 {matched_count} 条",
                stats={
                    "rule_scanned_count": index,
                    "rule_match_count": matched_count,
                },
            )
    return filtered


def filter_rows_by_branches(
    rows: Iterable[Dict[str, Any]],
    selected_branches: Sequence[str] | None = None,
    *,
    progress_callback: ProgressCallback | None = None,
    progress_step: int = _FILTER_PROGRESS_STEP,
) -> List[Dict[str, Any]]:
    row_list = list(rows or [])
    branch_names = normalize_branch_selection(selected_branches)
    branch_cmd_ids = {BRANCH_CMD_ID_MAP[name] for name in branch_names}
    if not branch_cmd_ids:
        preserved_count = len(row_list)
        _emit_progress(
            progress_callback,
            stage="branch_filtering",
            message=f"分局过滤完成：未选择分局，保留 {preserved_count} 条",
            stats={
                "branch_scanned_count": preserved_count,
                "branch_filtered_count": preserved_count,
            },
        )
        return row_list

    filtered: List[Dict[str, Any]] = []
    total = len(row_list)
    for index, row in enumerate(row_list, start=1):
        cmd_id = str(row.get("cmdId") or "").strip()
        if cmd_id in branch_cmd_ids:
            filtered.append(row)
        if index % progress_step == 0 or index == total:
            _emit_progress(
                progress_callback,
                stage="branch_filtering",
                message=f"分局过滤中：已扫描 {index} 条，保留 {len(filtered)} 条",
                stats={
                    "branch_scanned_count": index,
                    "branch_filtered_count": len(filtered),
                },
            )
    return filtered


def paginate_rows(rows: Sequence[Dict[str, Any]], page_num: int = 1, page_size: int = 15) -> Dict[str, Any]:
    page_num = max(1, int(page_num or 1))
    page_size = max(1, min(int(page_size or 15), 200))
    total = len(rows or [])
    start = (page_num - 1) * page_size
    end = start + page_size
    return {
        "total": total,
        "page_num": page_num,
        "page_size": page_size,
        "rows": list(rows[start:end]),
    }


def collect_cmd_id_samples(rows: Iterable[Dict[str, Any]], limit: int = 8) -> List[str]:
    samples: List[str] = []
    seen = set()
    for row in rows or []:
        cmd_id = str(row.get("cmdId") or "").strip()
        if not cmd_id or cmd_id in seen:
            continue
        seen.add(cmd_id)
        samples.append(cmd_id)
        if len(samples) >= limit:
            break
    return samples


def summarize_rule_groups(rules: Sequence[Dict[str, Any]]) -> List[str]:
    summaries: List[str] = []
    for group_no, group_rules in _enabled_rule_groups(rules):
        rule_text = " & ".join(f"{rule['field_name']}:{rule['operator']}" for rule in group_rules)
        summaries.append(f"G{group_no}({rule_text})")
    return summaries


def collect_rule_hit_samples(rows: Iterable[Dict[str, Any]], rules: Sequence[Dict[str, Any]], limit: int = 5) -> List[Dict[str, str]]:
    samples: List[Dict[str, str]] = []
    rule_group_summary = summarize_rule_groups(rules)
    for row in rows or []:
        combined_text = _row_field_text(row, "combined_text").replace("\r", " ")
        samples.append(
            {
                "case_no": str(row.get("caseNo") or ""),
                "cmd_id": str(row.get("cmdId") or ""),
                "rules": " | ".join(rule_group_summary),
                "snippet": combined_text[:120],
            }
        )
        if len(samples) >= limit:
            break
    return samples


def query_special_case_records(
    *,
    label: str,
    scheme_id: int,
    scheme_name: str,
    rules: Sequence[Dict[str, Any]],
    start_time: str,
    end_time: str,
    branches: Sequence[str] | None,
    page_num: int,
    page_size: int,
    progress_callback: ProgressCallback | None = None,
    include_hit_keyword_details: bool = False,
) -> Dict[str, Any]:
    trace_id = uuid.uuid4().hex[:10]
    begin_date = normalize_datetime_text(start_time) if start_time else default_time_range()[0]
    end_date = normalize_datetime_text(end_time) if end_time else default_time_range()[1]
    progress_stats = {
        "upstream_row_count": 0,
        "rule_scanned_count": 0,
        "rule_match_count": 0,
        "branch_scanned_count": 0,
        "branch_filtered_count": 0,
    }

    def report(stage: str, message: str, stats: Dict[str, int] | None = None) -> None:
        if stats:
            progress_stats.update(stats)
        _emit_progress(
            progress_callback,
            stage=stage,
            message=message,
            stats=progress_stats,
        )

    report("fetching", "正在拉取警情...", {})
    rows = fetch_all_special_case_rows(begin_date, end_date)
    report(
        "fetching",
        f"警情拉取完成，共 {len(rows)} 条",
        {
            "upstream_row_count": len(rows),
        },
    )
    normalized_branches = normalize_branch_selection(branches)
    rule_filtered_rows = filter_rows_by_rules(
        rows,
        rules,
        progress_callback=lambda payload: report(
            payload.get("stage") or "rule_filtering",
            payload.get("message") or "规则过滤中...",
            payload.get("stats") or {},
        ),
    )
    branch_filtered_rows = filter_rows_by_branches(
        rule_filtered_rows,
        selected_branches=branches,
        progress_callback=lambda payload: report(
            payload.get("stage") or "branch_filtering",
            payload.get("message") or "分局过滤中...",
            payload.get("stats") or {},
        ),
    )
    paged = paginate_rows(branch_filtered_rows, page_num=page_num, page_size=page_size)
    if include_hit_keyword_details:
        paged["rows"] = [
            {
                **row,
                "hitKeywordDetails": "；".join(collect_rule_hit_keyword_details(row, rules)),
            }
            for row in paged["rows"]
        ]
    enabled_rule_groups = _enabled_rule_groups(rules)
    debug_info = {
        "trace_id": trace_id,
        "requested_branches": list(branches or []),
        "normalized_branches": normalized_branches,
        "branch_cmd_ids": [BRANCH_CMD_ID_MAP[name] for name in normalized_branches],
        "upstream_row_count": len(rows),
        "rule_scanned_count": progress_stats["rule_scanned_count"] or len(rows),
        "rule_group_count": len(enabled_rule_groups),
        "rule_match_count": len(rule_filtered_rows),
        "branch_scanned_count": progress_stats["branch_scanned_count"] or len(rule_filtered_rows),
        "branch_filtered_count": len(branch_filtered_rows),
        "page_row_count": len(paged["rows"]),
        "sample_cmd_ids": collect_cmd_id_samples(rows),
        "rule_hit_samples": collect_rule_hit_samples(rule_filtered_rows, rules),
    }
    logger.info(
        "[trace:%s][custom-case-monitor] scheme_id=%s scheme_name=%s start=%s end=%s rules=%s rule_groups=%s req_branches=%s normalized_branches=%s upstream=%s rule_match=%s branch_filtered=%s page_num=%s page_size=%s page_rows=%s sample_cmd_ids=%s",
        trace_id,
        scheme_id,
        scheme_name,
        begin_date,
        end_date,
        len([rule for rule in rules if rule.get("is_enabled", True)]),
        len(enabled_rule_groups),
        list(branches or []),
        normalized_branches,
        len(rows),
        len(rule_filtered_rows),
        len(branch_filtered_rows),
        paged["page_num"],
        paged["page_size"],
        len(paged["rows"]),
        debug_info["sample_cmd_ids"],
    )
    report(
        "done",
        f"查询完成：命中 {len(branch_filtered_rows)} 条",
        {
            "branch_scanned_count": len(rule_filtered_rows),
            "branch_filtered_count": len(branch_filtered_rows),
        },
    )
    return {
        "success": True,
        "label": label,
        "scheme_id": scheme_id,
        "scheme_name": scheme_name,
        "start_time": begin_date,
        "end_time": end_date,
        "branches": normalized_branches,
        "debug": debug_info,
        **paged,
    }


def generate_special_case_excel(
    rows: Sequence[Dict[str, Any]],
    title: str,
    headers: Sequence[Tuple[str, str]] | None = None,
) -> io.BytesIO:
    headers = list(headers or EXPORT_HEADERS)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_sheet_title(title)
    for col_idx, (_, header) in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_idx, value=header).font = openpyxl.styles.Font(bold=True)
    for row_idx, row in enumerate(rows or [], 2):
        for col_idx, (field, _) in enumerate(headers, 1):
            worksheet.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
    out = io.BytesIO()
    workbook.save(out)
    out.seek(0)
    return out


def generate_special_case_csv(
    rows: Sequence[Dict[str, Any]],
    headers: Sequence[Tuple[str, str]] | None = None,
) -> io.BytesIO:
    headers = list(headers or EXPORT_HEADERS)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([header for _, header in headers])
    for row in rows or []:
        writer.writerow([row.get(field, "") for field, _ in headers])
    return io.BytesIO(output.getvalue().encode("utf-8-sig"))


def export_special_case_records(
    *,
    label: str,
    scheme_name: str,
    rules: Sequence[Dict[str, Any]],
    export_format: str,
    start_time: str,
    end_time: str,
    branches: Sequence[str] | None,
    include_hit_keywords: bool = False,
    download_name: str | None = None,
):
    result = query_special_case_records(
        label=label,
        scheme_id=0,
        scheme_name=scheme_name,
        rules=rules,
        start_time=start_time,
        end_time=end_time,
        branches=branches,
        page_num=1,
        page_size=200000,
    )
    rows = result["rows"]
    if export_format == "csv":
        return (
            generate_special_case_csv(rows),
            "text/csv; charset=utf-8",
            download_name or f"{sanitize_filename_component(scheme_name)}.csv",
        )
    headers = list(EXPORT_HEADERS)
    export_rows = list(rows)
    if include_hit_keywords:
        headers.append(HIT_KEYWORD_HEADER)
        export_rows = [
            {**row, HIT_KEYWORD_HEADER[0]: "、".join(collect_rule_hit_keywords(row, rules))}
            for row in rows
        ]
    return (
        generate_special_case_excel(export_rows, scheme_name, headers=headers),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        download_name or f"{sanitize_filename_component(scheme_name)}.xlsx",
    )
