from __future__ import annotations

import io
import re
import secrets
import tempfile
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

from tongyonggongju.dao.background_check_dao import (
    query_dispute_rows,
    query_mental_health_rows,
    query_prior_case_rows,
)


MAX_XLSX_BYTES = 20 * 1024 * 1024
MAX_ID_COUNT = 10000
UPLOAD_TTL_SECONDS = 12 * 60 * 60
UPLOAD_DIR = Path(tempfile.gettempdir()) / "yfjcgkzx_tygj_background"
ID_PATTERN = re.compile(r"^(?:\d{15}|\d{17}[\dX])$")
TOKEN_PATTERN = re.compile(r"^[A-Za-z0-9_-]{16,}$")


def inspect_and_store_workbook(file_bytes: bytes, filename: str | None) -> Dict[str, Any]:
    _validate_upload(file_bytes, filename)
    workbook = _load_workbook(file_bytes)
    worksheet = workbook.active
    columns = _read_header_columns(worksheet)
    if not columns:
        raise ValueError("未读取到表头，请确认第一行是列名")

    token = _store_upload(file_bytes)
    return {
        "token": token,
        "filename": filename or "upload.xlsx",
        "sheet_name": worksheet.title,
        "columns": columns,
    }


def run_background_check(token: str, id_column_index: Any, name_column_index: Any = None) -> Dict[str, Any]:
    extracted = extract_id_numbers(token, id_column_index, name_column_index)
    ids = [person["id_number"] for person in extracted["people"]]
    if not ids:
        raise ValueError("所选列中未识别到有效身份证号")
    if len(ids) > MAX_ID_COUNT:
        raise ValueError(f"单次最多审查 {MAX_ID_COUNT} 个不同身份证号")

    prior_rows = _serialize_rows(query_prior_case_rows(ids))
    dispute_rows = _serialize_rows(query_dispute_rows(ids))
    mental_rows = _serialize_rows(query_mental_health_rows(ids))
    overview = _build_overview(extracted["people"], prior_rows, dispute_rows, mental_rows)
    hit_overview = [row for row in overview if row["命中类型"] != "未命中"]

    hit_people_count = len(hit_overview)
    return {
        "stats": {
            "有效身份证行数": extracted["valid_count"],
            "去重后人数": extracted["unique_count"],
            "重复身份证行数": extracted["duplicate_count"],
            "无效身份证行数": extracted["invalid_count"],
            "命中人数": hit_people_count,
            "前科命中人数": _hit_count(prior_rows),
            "矛盾纠纷命中人数": _hit_count(dispute_rows),
            "精神障碍命中人数": _hit_count(mental_rows),
            "前科记录数": len(prior_rows),
            "矛盾纠纷记录数": len(dispute_rows),
            "精神障碍记录数": len(mental_rows),
        },
        "invalid_samples": extracted["invalid_samples"],
        "overview": hit_overview,
        "details": {
            "prior_case": prior_rows,
            "dispute": dispute_rows,
            "mental_health": mental_rows,
        },
    }


def extract_id_numbers(token: str, id_column_index: Any, name_column_index: Any = None) -> Dict[str, Any]:
    column_index = _parse_column_index(id_column_index, "身份证号")
    parsed_name_column_index = (
        _parse_column_index(name_column_index, "姓名") if name_column_index is not None else None
    )
    if parsed_name_column_index == column_index:
        raise ValueError("姓名列和身份证号列不能相同")

    workbook = _load_workbook(_upload_path(token).read_bytes())
    worksheet = workbook.active
    if column_index > worksheet.max_column:
        raise ValueError("身份证列不存在，请重新选择")
    if parsed_name_column_index and parsed_name_column_index > worksheet.max_column:
        raise ValueError("姓名列不存在，请重新选择")

    unique_people: List[Dict[str, Any]] = []
    seen = set()
    invalid_samples: List[Dict[str, Any]] = []
    valid_count = 0
    invalid_count = 0

    min_col = min(column_index, parsed_name_column_index or column_index)
    max_col = max(column_index, parsed_name_column_index or column_index)
    id_offset = column_index - min_col
    name_offset = parsed_name_column_index - min_col if parsed_name_column_index else None

    for row_number, row in enumerate(
        worksheet.iter_rows(
            min_row=2,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ),
        start=2,
    ):
        raw_value = row[id_offset] if row and len(row) > id_offset else None
        raw_name = row[name_offset] if name_offset is not None and row and len(row) > name_offset else None
        normalized = normalize_id_number(raw_value)
        if not normalized:
            continue
        if not ID_PATTERN.fullmatch(normalized):
            invalid_count += 1
            if len(invalid_samples) < 20:
                invalid_samples.append({"Excel行号": row_number, "原始值": _cell_to_text(raw_value)})
            continue
        valid_count += 1
        if normalized in seen:
            continue
        seen.add(normalized)
        unique_people.append(
            {
                "row_number": row_number,
                "name": normalize_person_name(raw_name),
                "id_number": normalized,
                "source_value": _cell_to_text(raw_value),
            }
        )

    return {
        "people": unique_people,
        "valid_count": valid_count,
        "unique_count": len(unique_people),
        "duplicate_count": max(valid_count - len(unique_people), 0),
        "invalid_count": invalid_count,
        "invalid_samples": invalid_samples,
    }


def normalize_id_number(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    if isinstance(value, int):
        text = str(value)
    elif isinstance(value, float):
        text = str(int(value)) if value.is_integer() else str(value)
    else:
        text = str(value)
    text = text.strip().upper().replace("　", "")
    text = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text


def normalize_person_name(value: Any) -> str:
    if value is None or isinstance(value, bool):
        return ""
    text = _cell_to_text(value).strip().replace("　", "")
    return re.sub(r"\s+", "", text)


def _validate_upload(file_bytes: bytes, filename: str | None) -> None:
    if not file_bytes:
        raise ValueError("请先选择 xlsx 文件")
    if len(file_bytes) > MAX_XLSX_BYTES:
        raise ValueError("xlsx 文件不能超过 20MB")
    if not filename or not filename.lower().endswith(".xlsx"):
        raise ValueError("仅支持上传 .xlsx 文件")


def _load_workbook(file_bytes: bytes):
    try:
        return load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ValueError("xlsx 文件无法解析，请确认文件格式正确") from exc


def _read_header_columns(worksheet) -> List[Dict[str, Any]]:
    rows = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(rows, None) or []
    columns: List[Dict[str, Any]] = []
    for index, value in enumerate(header, start=1):
        title = _cell_to_text(value).strip() or f"未命名列{index}"
        letter = get_column_letter(index)
        columns.append(
            {
                "index": index,
                "letter": letter,
                "label": title,
                "display": f"{letter} - {title}",
            }
        )
    return columns


def _store_upload(file_bytes: bytes) -> str:
    _cleanup_uploads()
    UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    token = secrets.token_urlsafe(18)
    _upload_path(token, must_exist=False).write_bytes(file_bytes)
    return token


def _upload_path(token: str, *, must_exist: bool = True) -> Path:
    text = str(token or "").strip()
    if not TOKEN_PATTERN.fullmatch(text):
        raise ValueError("上传文件标识无效，请重新上传")
    path = UPLOAD_DIR / f"{text}.xlsx"
    if must_exist and not path.exists():
        raise ValueError("上传文件已失效，请重新上传")
    return path


def _cleanup_uploads() -> None:
    if not UPLOAD_DIR.exists():
        return
    cutoff = time.time() - UPLOAD_TTL_SECONDS
    for path in UPLOAD_DIR.glob("*.xlsx"):
        try:
            if path.stat().st_mtime < cutoff:
                path.unlink()
        except OSError:
            continue


def _parse_column_index(value: Any, label: str) -> int:
    try:
        index = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"请选择{label}所在列") from None
    if index < 1:
        raise ValueError(f"请选择{label}所在列")
    return index


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _serialize_rows(rows: Iterable[Mapping[str, Any]]) -> List[Dict[str, Any]]:
    return [{str(key): _serialize_value(value) for key, value in row.items()} for row in rows]


def _serialize_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return str(value)
    return value


def _group_by_id(rows: Iterable[Mapping[str, Any]]) -> Dict[str, List[Mapping[str, Any]]]:
    grouped: Dict[str, List[Mapping[str, Any]]] = defaultdict(list)
    for row in rows:
        id_number = str(row.get("身份证号") or "").strip().upper()
        if id_number:
            grouped[id_number].append(row)
    return grouped


def _hit_count(rows: Iterable[Mapping[str, Any]]) -> int:
    return len({str(row.get("身份证号") or "").strip().upper() for row in rows if row.get("身份证号")})


def _status_for(rows: List[Mapping[str, Any]]) -> str:
    if not rows:
        return "未命中"
    statuses = [str(row.get("管理状态") or "").strip() for row in rows]
    if "在管" in statuses and "撤管" in statuses:
        return "在管/撤管"
    if "在管" in statuses:
        return "在管"
    if "撤管" in statuses:
        return "撤管"
    return "已命中"


def _build_overview(
    people: Iterable[Mapping[str, Any]],
    prior_rows: List[Mapping[str, Any]],
    dispute_rows: List[Mapping[str, Any]],
    mental_rows: List[Mapping[str, Any]],
) -> List[Dict[str, Any]]:
    prior_counts = Counter(str(row.get("身份证号") or "").strip().upper() for row in prior_rows)
    dispute_grouped = _group_by_id(dispute_rows)
    mental_grouped = _group_by_id(mental_rows)
    overview: List[Dict[str, Any]] = []
    for person in people:
        id_number = str(person.get("id_number") or "").strip().upper()
        prior_count = int(prior_counts.get(id_number, 0))
        dispute_count = len(dispute_grouped.get(id_number, []))
        mental_count = len(mental_grouped.get(id_number, []))
        hits = []
        if prior_count:
            hits.append("前科")
        if dispute_count:
            hits.append("矛盾纠纷")
        if mental_count:
            hits.append("精神障碍")
        overview.append(
            {
                "Excel行号": person.get("row_number"),
                "姓名": person.get("name") or "",
                "身份证号": id_number,
                "前科记录数": prior_count,
                "矛盾纠纷状态": _status_for(dispute_grouped.get(id_number, [])),
                "矛盾纠纷记录数": dispute_count,
                "精神障碍状态": _status_for(mental_grouped.get(id_number, [])),
                "精神障碍记录数": mental_count,
                "命中类型": "、".join(hits) if hits else "未命中",
            }
        )
    return overview
